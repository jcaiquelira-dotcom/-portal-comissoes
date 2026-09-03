# -*- coding: utf-8 -*-
"""Quebra as saidas do mes por rubrica — a base do DRE.

Como funciona, e por que assim: a formula de "Debitos" da planilha lista, uma a
uma, as celulas de subtotal que o autor considera despesa do mes. Resolver
essas referencias da a quebra EXATA e ela fecha no total por construcao. Em
agosto/2026 sao 15 componentes que somam R$ 674.635, o mesmo Debitos.

Tentei antes varrer os blocos por cabecalho e nao funciona: eles se empilham na
mesma coluna, a coluna do valor ora e a vizinha ora a seguinte, e o mesmo texto
e titulo num lugar e item em outro. "Cartao" e item no bloco de Diversos (B5,
com valor ao lado) e "Cartoes" e titulo de bloco (N8, com a celula da direita
vazia) — e essa diferenca, celula-da-direita-vazia, e o que separa titulo de
item aqui.

Classificacao de cada rubrica no DRE (confirmada pelo gestor em 01/09/2026):
    CMV        Sucatas (compra do veiculo) + custos de sucata (guincho,
               oficina) — sem eles a peca nao existe pra vender
    Deducoes   Impostos, Devolucoes
    Despesa    Colaboradores, Fixo, Cartoes, Diversos, Transportes, Marketing,
               Embalagem, Nevada Ecopecas
    NAO e despesa:
      Investimentos e Patrimonial  -> viram patrimonio
      Parcelas e socios (P1 Ricardo, P2 Odilon, P3 Caique, P4 Gabriela)
                                   -> distribuicao de lucro
"""
import openpyxl
import re
import unicodedata
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "app"))
from caminhos import caminho  # config/caminhos.json — ver app/caminhos.py


P = str(caminho("fluxo_caixa_xlsx"))


def ach(t):
    t = "".join(c for c in unicodedata.normalize("NFKD", str(t or '').strip().lower())
                if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', t)


# Vocabulario fechado das rubricas. Casar contra ele em vez de pegar "o texto
# mais proximo acima" — o mais proximo costuma ser o ULTIMO ITEM do bloco
# ("Vinicius", "Erp", "P4 ML"), nao o titulo dele.
VOCAB = {
    'impostos': 'Impostos', 'cartoes': 'Cartoes', 'cartao': 'Cartoes',
    'marketing': 'Marketing', 'embalagem': 'Embalagem',
    'sucatas': 'Sucatas', 'sucata': 'Sucatas', 'transportes': 'Transportes',
    'colaboradores': 'Colaboradores', 'imoveis': 'Imoveis',
    'investimentos': 'Investimentos', 'patrimonial': 'Patrimonial',
    'fixo': 'Fixo', 'nevada ecopecas': 'Nevada Ecopecas',
    'devolucao': 'Devolucoes', 'devolucoes': 'Devolucoes',
    'ecomerce': 'Receita e-commerce', 'ecommerce': 'Receita e-commerce',
    'vb': 'Receita balcao', 'v b': 'Receita balcao',
}
# Cabecalho de bloco tem a celula da direita VAZIA; item tem valor nela. E o
# que separa o titulo "Cartoes" (N8, com O8 vazia) do item "Cartao" (B5, com
# C5 = 8.692). Sem isso o bloco de Diversos era rotulado de Cartoes.
CABECALHO_DIREITA_VAZIA = True
# Blocos sem titulo escrito, identificados pela coluna onde o subtotal cai.
# Sao os dois primeiros blocos da planilha, que o autor nunca nomeou porque
# "todo mundo sabe": o de gastos avulsos e o de parcelas/socios.
SEM_TITULO = {'C': 'Diversos', 'I': 'Parcelas e socios', 'L': 'Nevada Ecopecas'}


def rotulo_de(ws, cel):
    """Nome da rubrica daquele subtotal.

    Sobe pela coluna procurando um texto que esteja no vocabulario. Se nao
    achar, cai na tabela por coluna — e so entao desiste, dizendo que desistiu.
    """
    col_letra = ''.join(c for c in cel if c.isalpha())
    col = openpyxl.utils.column_index_from_string(col_letra)
    lin = int(''.join(c for c in cel if c.isdigit()))
    for r in range(lin, max(0, lin - 45), -1):
        for c in (col, col - 1):
            if c < 1:
                continue
            v = ws.cell(row=r, column=c).value
            if not (isinstance(v, str) and ach(v) in VOCAB):
                continue
            direita = ws.cell(row=r, column=c + 1).value
            if isinstance(direita, (int, float)) and direita:
                continue          # e item da lista, nao titulo do bloco
            return VOCAB[ach(v)]
    return SEM_TITULO.get(col_letra, '(sem rotulo)')


# Nomes de vendedor no bloco de devolucoes sao COMISSAO paga, nao devolucao.
# Ficam na mesma coluna porque e onde couberam, nao porque sao a mesma coisa.
VENDEDORES = {"matheus", "brenda", "flavia", "gustavo", "lucas", "vinicius",
              "caique", "nycollas", "rafael"}

# Lancamentos que o rotulo nao revela e que o gestor identificou um a um.
# "M. Ram" aparece duas vezes em agosto: R31 e o motor que voltou (devolucao de
# verdade) e R32 e um cheque caucao que entrou e saiu da empresa. Caucao nao e
# despesa nem receita — nao afeta resultado, so passa pelo caixa. Sem esta
# linha os dois seriam lidos como devolucao, e o mes fecharia R$ 5.000 pior do
# que foi.
OVERRIDES = {
    ("Ago", "R32"): ("Caucao e transito", "Cheque caucao — entrou e saiu"),
}

# Blocos que misturam naturezas diferentes e por isso sao abertos item a item
# em vez de entrarem pelo subtotal.
MISTURADOS = {"Devolucoes"}


def detalhar_bloco(ws, aba, cel_subtotal, rubrica):
    """Abre um bloco misturado item a item, classificando cada linha.

    O bloco de Devolucoes de agosto tem tres naturezas na mesma coluna:
    devolucao de venda, comissao de vendedor e um cheque caucao. Somar tudo
    como "devolucao" jogava R$ 22.832 de comissao pra fora das despesas de
    pessoal e R$ 5.000 de caucao pra dentro do resultado.
    """
    col = openpyxl.utils.column_index_from_string(
        ''.join(c for c in cel_subtotal if c.isalpha()))
    lin = int(''.join(c for c in cel_subtotal if c.isdigit()))
    alvo_val = ws[cel_subtotal].value or 0
    itens, soma = [], 0.0
    for r in range(lin - 1, max(0, lin - 40), -1):
        # Para assim que os itens reconstroem o subtotal. Subir por contagem de
        # linhas passava do inicio do bloco e engolia o de cima — em agosto ia
        # buscar os R$ 232.451 de sucata, que nao tem nada a ver com este.
        if abs(soma - alvo_val) < 1 and itens:
            break
        val = ws.cell(row=r, column=col).value
        rot = ws.cell(row=r, column=col - 1).value
        if not isinstance(val, (int, float)) or not val:
            continue
        cel = ws.cell(row=r, column=col).coordinate
        nome = ach(rot)
        chave = OVERRIDES.get((aba, cel))
        if chave:
            destino, rotulo = chave
        elif nome in VENDEDORES:
            destino, rotulo = "Comissoes", f"Comissao {str(rot).strip()}"
        else:
            destino, rotulo = rubrica, str(rot or "").strip() or "(sem rotulo)"
        itens.append({"celula": cel, "valor": round(float(val), 2),
                      "rubrica": destino, "detalhe": rotulo})
        soma += val
    # So vale se os itens reconstruirem o subtotal. Se nao fecharem, e porque
    # o bloco nao e o que eu penso que e — melhor cair fora e usar o subtotal.
    if abs(soma - alvo_val) > 1:
        return None
    return itens


def itens_do_bloco(ws, cel_subtotal, alvo):
    """Os lancamentos que compoem um subtotal, ou None se nao fecharem.

    Mesma tecnica de `detalhar_bloco`: sobe pela coluna ate os itens
    reconstruirem o subtotal e para ali. Parar pela soma, e nao por contagem de
    linhas, e o que impede de invadir o bloco de cima — em agosto passava do
    inicio e engolia os R$ 232.451 de sucata, que sao de outro bloco.

    Nao fechou, devolve None. Bloco que nao fecha nao e o que eu penso que e, e
    e melhor perder o detalhe do que detalhar errado.
    """
    col = openpyxl.utils.column_index_from_string(
        ''.join(c for c in cel_subtotal if c.isalpha()))
    lin = int(''.join(c for c in cel_subtotal if c.isdigit()))
    itens, soma = [], 0.0
    for r in range(lin - 1, max(0, lin - 45), -1):
        if abs(soma - alvo) < 1 and itens:
            break
        v = ws.cell(row=r, column=col).value
        rot = ws.cell(row=r, column=col - 1).value
        if not isinstance(v, (int, float)) or not v:
            continue
        itens.append({"rotulo": str(rot or "").strip(), "valor": float(v),
                      "celula": ws.cell(row=r, column=col).coordinate})
        soma += v
    if abs(soma - alvo) > 1:
        return None
    return itens


def itens_de_socios(ws, wsf, cel_subtotal, alvo):
    """Abre a coluna das retiradas dos socios, marcando de quem e cada linha.

    A coluna nao e um bloco so: sao quatro faixas de linhas, uma por socio, com
    o preenchimento alternando de tom pra separar visualmente. Em agosto a
    propria planilha escreve os subtotais de cada faixa — I36 = I2..I13 (P1),
    I37 = I14..I25 (P2), I38 = I26..I30 (P3), I39 = I31..I35 (P4) — e em julho
    o Itau de R$ 10.000 cai na linha 26, dentro da faixa do P3, que e o que o
    gestor descreveu de cabeca.

    As linhas de subtotal sao puladas: da pra reconhece-las porque a formula
    delas aponta pra outras celulas da MESMA coluna. Sem isso eu somaria o
    subtotal junto com os itens que ele ja resume, e o mes dobraria.

    Devolve None se os itens nao reconstruirem o subtotal, ou se as faixas nao
    derem exatamente quatro — a atribuicao por socio so vale se a estrutura for
    a que eu penso que e.
    """
    letra = ''.join(c for c in cel_subtotal if c.isalpha())
    col = openpyxl.utils.column_index_from_string(letra)
    lin = int(''.join(c for c in cel_subtotal if c.isdigit()))

    # As linhas que o subtotal de fato soma. Varrer a coluna inteira pegava
    # linha de fora: em julho o total e SUM(I4:I34) e a linha I3 existe com
    # valor, mas nao entra — somar ela dava R$ 1.864 a mais que o mes.
    faixa_total = _linhas_da_formula(wsf[cel_subtotal].value, letra)
    if not faixa_total:
        faixa_total = set(range(1, lin))

    ref_propria = re.compile(r'\b%s\d+' % letra, re.I)
    brutos = []
    for r in sorted(faixa_total):
        c = ws.cell(row=r, column=col)
        if not isinstance(c.value, (int, float)) or not c.value:
            continue
        f = wsf.cell(row=r, column=col).value
        if isinstance(f, str) and f.startswith('=') and ref_propria.search(f):
            continue                      # e subtotal de faixa, nao lancamento
        brutos.append({"linha": r, "valor": float(c.value),
                       "rotulo": str(ws.cell(row=r, column=col - 1).value or "").strip(),
                       "celula": c.coordinate, "tom": _tom(c)})

    if not brutos or abs(sum(b["valor"] for b in brutos) - alvo) > 1:
        return None

    bandas = (_bandas_por_subtotal(ws, wsf, col, letra, lin, faixa_total)
              or _bandas_por_cor(brutos))
    if not bandas:
        return [dict(b, socio=None) for b in brutos]
    return [dict(b, socio=_socio_da_linha(b["linha"], bandas)) for b in brutos]


def _linhas_da_formula(formula, letra) -> set:
    """Linhas da propria coluna que uma formula soma. Aceita SUM(I2:I35) e
    tambem a forma escrita a mao, =I2+I3+I4..., que a planilha usa igual."""
    if not isinstance(formula, str) or not formula.startswith('='):
        return set()
    linhas = set()
    for a, b in re.findall(r'%s(\d+)\s*:\s*%s(\d+)' % (letra, letra),
                           formula, re.I):
        linhas.update(range(int(a), int(b) + 1))
    sem_faixas = re.sub(r'%s\d+\s*:\s*%s\d+' % (letra, letra), '', formula, flags=re.I)
    linhas.update(int(n) for n in re.findall(r'%s(\d+)' % letra, sem_faixas, re.I))
    return linhas


def _bandas_por_subtotal(ws, wsf, col, letra, lin_total, faixa_total):
    """As faixas dos socios lidas dos subtotais que a planilha escreve.

    Em agosto elas estao explicitas: I36 = I2..I13, I37 = I14..I25,
    I38 = I26..I30, I39 = I31..I35. E a evidencia mais forte que existe, e
    funciona ate quando um socio nao teve movimento nenhum no mes — foi o caso
    do P3 em agosto, em que a cor sozinha so mostrava tres faixas.
    """
    achadas = []
    for r in range(1, lin_total):
        if r in faixa_total:
            continue
        f = wsf.cell(row=r, column=col).value
        linhas = _linhas_da_formula(f, letra)
        if linhas and linhas <= faixa_total:
            achadas.append((r, linhas))
    if len(achadas) != 4:
        return None
    achadas.sort(key=lambda x: min(x[1]))
    return [linhas for _r, linhas in achadas]


def _bandas_por_cor(brutos):
    """As faixas pelo preenchimento: tom novo, socio novo.

    Vale quando a planilha nao escreve os subtotais por faixa. So aceita se der
    exatamente quatro — com tres ou cinco eu nao sei qual e qual, e chutar aqui
    poe a retirada de um socio na conta de outro.
    """
    faixas, atual = [], None
    for b in brutos:
        if atual is None or b["tom"] != atual:
            faixas.append(set())
            atual = b["tom"]
        faixas[-1].add(b["linha"])
    return faixas if len(faixas) == 4 else None


def _socio_da_linha(linha, bandas):
    for i, banda in enumerate(bandas, start=1):
        if linha in banda:
            return i
    return None


def _tom(celula) -> str:
    """Assinatura do preenchimento da celula, pra comparar faixas."""
    f = celula.fill
    if not f or f.patternType is None:
        return "-"
    cor = f.fgColor
    tema = getattr(cor, "theme", None)
    if isinstance(tema, int):
        return "t%d/%.2f" % (tema, getattr(cor, "tint", 0) or 0)
    rgb = getattr(cor, "rgb", None)
    return rgb if isinstance(rgb, str) else "?"


def itens_do_mes(ws, partes, wsf=None):
    """Todo lancamento do mes, um a um, com o bloco de onde saiu.

    O bloco viaja junto porque o rotulo sozinho nao basta pra classificar:
    "itau" no bloco de Parcelas e uma parcela de financiamento, e "itau" no
    bloco Fixo e tarifa de conta. Sem o contexto, os dois viram a mesma coisa.

    Subtotal que nao abre entra como uma linha so, marcada — sao Comissoes e
    Devolucoes, que ja sao categoria pura e nao precisam abrir.
    """
    saida = []
    for p in partes:
        bloco = p["rubrica"]
        if "detalhe" in p:            # ja veio item, de um bloco misturado
            saida.append({"rotulo": p["detalhe"], "valor": p["valor"],
                          "bloco": bloco, "celula": p.get("celula", "")})
            continue
        cel = p.get("celula")
        if not cel or cel == "(Veiculos)":
            saida.append({"rotulo": "(Veiculos)", "valor": p["valor"],
                          "bloco": bloco, "celula": cel or ""})
            continue
        # A coluna dos socios abre por faixa, nao por soma de baixo pra cima:
        # ali as linhas de subtotal moram na mesma coluna dos lancamentos, e
        # subir somando pegaria o subtotal em vez dos itens que ele resume.
        if bloco == "Parcelas e socios" and wsf is not None:
            dos_socios = itens_de_socios(ws, wsf, cel, p["valor"])
            if dos_socios is not None:
                for i in dos_socios:
                    saida.append({**i, "bloco": bloco})
                continue
        itens = itens_do_bloco(ws, cel, p["valor"])
        if itens is None:
            saida.append({"rotulo": "(bloco inteiro)", "valor": p["valor"],
                          "bloco": bloco, "celula": cel})
            continue
        for i in itens:
            saida.append({**i, "bloco": bloco})
    return saida


def quebra_do_mes(caminho, aba):
    wv = openpyxl.load_workbook(caminho, data_only=True)
    wf = openpyxl.load_workbook(caminho, data_only=False)
    ws, wsf = wv[aba], wf[aba]
    formula = None
    for linha in ws.iter_rows():
        for c in linha:
            if ach(c.value) == 'debitos':
                for k in (1, 2):
                    alvo = wsf.cell(row=c.row, column=c.column + k)
                    if isinstance(alvo.value, str) and alvo.value.startswith('='):
                        formula = alvo.value
                        break
        if formula:
            break
    if not formula:
        return None, []
    refs = re.findall(r'[A-Z]{1,2}[0-9]{1,4}', formula)
    partes = []
    for ref in refs:
        v = ws[ref].value
        if not isinstance(v, (int, float)) or not v:
            continue
        rub = rotulo_de(ws, ref)
        if rub in MISTURADOS:
            det = detalhar_bloco(ws, aba, ref, rub)
            if det:
                partes.extend(det)
                continue
        partes.append({'celula': ref, 'valor': round(float(v), 2),
                       'rubrica': rub})
    return formula, partes


