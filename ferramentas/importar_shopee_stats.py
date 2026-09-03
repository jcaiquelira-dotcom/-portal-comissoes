"""
Leva o relatorio "shop-stats" da Shopee pro portal (chave shopee_conta).

Usa a aba "Produto Pago" — vendas efetivamente pagas, o mesmo criterio do
Mercado Livre no painel (pagamento aprovado, nao pedido feito).

O Seller Centre muda a granularidade conforme o periodo pedido: intervalo
longo sai com UMA LINHA POR MES (01/01, 01/02, ...), intervalo curto sai DIA A
DIA. Este importador reconhece os dois — se ha mais de uma data no mesmo mes,
e diario. O diario vai pra serie_dia, que o painel soma exato em qualquer
filtro; o mensal vai pra serie_mes, que o painel rateia por dia quando o
filtro corta um mes no meio.

Mes que existe nos dois ganha o diario: o servidor ignora a linha mensal de
qualquer mes com cobertura diaria, entao importar planilhas de periodos
diferentes nunca conta a mesma venda duas vezes.

As duas series ficam no Supabase (chave shopee_conta) e so crescem: cada
importacao atualiza o que veio na planilha e preserva o resto. E assim que o
historico se acumula sem depender de guardar os arquivos.

Uso:
    set DATABASE_URL=postgresql://...
    python scripts/importar_shopee_stats.py "caminho/do/arquivo.xlsx"
"""

import io
import os
import re
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "app"))
import nevada_comum as C  # biblioteca comum — ver app/nevada_comum.py

FUSO = C.FUSO
# "27/08/2026" — data unica. A linha-resumo do topo traz intervalo
# ("01/08/2026-27/08/2026") e por isso nao casa com fullmatch.
PADRAO_DATA = re.compile(r"(\d{2})/(\d{2})/(\d{4})")


def num_br(v) -> float:
    """"1.153,65" -> 1153.65. O relatorio vem sempre no formato brasileiro."""
    t = str(v or "").strip().replace(".", "").replace(",", ".")
    try:
        return float(t)
    except ValueError:
        return 0.0


def ler(caminho: Path):
    """Devolve (formato, serie). formato e "dia" ou "mes"; a chave da serie
    e "AAAA-MM-DD" ou "AAAA-MM" conforme o caso."""
    wb = load_workbook(caminho, data_only=True)
    ws = wb["Produto Pago"]

    linhas = []
    for linha in ws.iter_rows(min_row=5, values_only=True):
        m = PADRAO_DATA.fullmatch(str(linha[0] or "").strip())
        if not m:
            continue
        dia, mes, ano = m.groups()
        linhas.append((f"{ano}-{mes}-{dia}",
                       round(num_br(linha[1]), 2), int(num_br(linha[3]))))
    if not linhas:
        raise SystemExit("nenhuma linha de data encontrada na aba 'Produto Pago'.")

    # Duas datas no mesmo mes so acontecem no relatorio diario; o mensal traz
    # sempre o dia 01 de meses distintos.
    meses = [d[:7] for d, _, _ in linhas]
    diario = len(meses) != len(set(meses)) or any(d[8:] != "01" for d, _, _ in linhas)

    serie = {}
    for data, total, qtd in linhas:
        serie[data if diario else data[:7]] = {"total": total, "qtd": qtd}
    return ("dia" if diario else "mes"), serie


def resumir(formato: str, serie: dict) -> None:
    total = sum(m["total"] for m in serie.values())
    qtd = sum(m["qtd"] for m in serie.values())
    print(f"relatório {formato.upper()}: {len(serie)} registros "
          f"({min(serie)} a {max(serie)}) | R$ {total:,.2f} em {qtd} pedidos pagos")
    por_mes = {}
    for chave, m in serie.items():
        acc = por_mes.setdefault(chave[:7], {"total": 0.0, "qtd": 0})
        acc["total"] = round(acc["total"] + m["total"], 2)
        acc["qtd"] += m["qtd"]
    for mes in sorted(por_mes):
        m = por_mes[mes]
        print(f"  {mes}: R$ {m['total']:>10,.2f} | {m['qtd']:>3} pedidos")


# Sao duas contas Shopee desde 02/09/2026: a nevadaecopecas (loja 1, chave
# historica `shopee_conta`) e a gabrielanevada (loja 2, `shopee_conta_2`). A
# chave da loja 1 NAO mudou de nome de proposito — renomear obrigaria a migrar
# o historico e a mexer em tudo que ja le essa chave.
CHAVES = {"1": "shopee_conta", "2": "shopee_conta_2"}


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    lojas = [a.split("=", 1)[1] for a in sys.argv[1:] if a.startswith("--loja=")]
    loja = lojas[0] if lojas else "1"
    if loja not in CHAVES:
        raise SystemExit(f"--loja deve ser 1 ou 2 (veio {loja!r})")
    chave = CHAVES[loja]
    if not args:
        raise SystemExit("uso: python importar_shopee_stats.py <arquivo.xlsx> [--loja=1|2]")
    caminho = Path(args[0])
    if not caminho.exists():
        raise SystemExit(f"arquivo não encontrado: {caminho}")
    print(f"loja {loja} -> chave {chave}")

    formato, serie = ler(caminho)
    resumir(formato, serie)

    from psycopg2.extras import Json
    conn = C.conexao()
    with conn, conn.cursor() as cur:
        cur.execute("SELECT valor FROM dados_json WHERE chave=%s FOR UPDATE", (chave,))
        linha = cur.fetchone()
        vendas = (linha[0].get("vendas") or {}) if linha else {}
        serie_mes = dict(vendas.get("serie_mes") or {})
        serie_dia = dict(vendas.get("serie_dia") or {})
        # Merge na serie do formato que veio: reimportar um periodo atualiza o
        # que ele cobre e preserva todo o resto do historico.
        (serie_dia if formato == "dia" else serie_mes).update(serie)
        cur.execute("INSERT INTO dados_json (chave, valor) VALUES (%s, %s) "
                    "ON CONFLICT (chave) DO UPDATE SET valor = EXCLUDED.valor",
                    (chave, Json({
                        "gerado_em": datetime.now(FUSO).isoformat(timespec="seconds"),
                        "fonte": caminho.name,
                        "vendas": {"serie_mes": serie_mes, "serie_dia": serie_dia},
                    })))
    conn.close()
    print(f"\n  gravado {chave} ({len(serie_dia)} dias, {len(serie_mes)} meses)")


if __name__ == "__main__":
    C.saida_utf8()
    main()
