# -*- coding: utf-8 -*-
"""Historico da folha de pagamento: planilha "Colaboradores 2026.xlsx" -> rh_folha.

A planilha tem uma aba por mes (Janeiro ... Setembro) com, por pessoa:
  Dia 05 -> vale | Dia 20 -> restante do salario | COMISSAO -> comissao
  Descontos (texto) -> descontos | Observacoes + Google -> obs
e a aba "Meta bonus" com o bonus pago por pessoa e mes -> bonus (dia 10).

So preenche o que o painel nao tem: mes/pessoa que ja tem valor na
rh_folha fica como esta (a folha do painel e quem manda desde 09/2026).
Nomes casam com a ficha do RH por nome ou apelido sem acento; abreviacao
("Vinicius L", "Pedro P", "Marcela") casa por prefixo quando so um bate.

Uso:
    python ferramentas/importar_folha.py          # so compara
    python ferramentas/importar_folha.py --gravar
"""
import os
import re
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "app"))
import openpyxl
import nevada_comum as C  # biblioteca comum — ver app/nevada_comum.py
from caminhos import caminho  # config/caminhos.json — ver app/caminhos.py

MESES = ["janeiro", "fevereiro", "marco", "abril", "maio", "junho", "julho",
         "agosto", "setembro", "outubro", "novembro", "dezembro"]
ANO = 2026


def chato(t) -> str:
    import unicodedata
    return " ".join("".join(c for c in unicodedata.normalize("NFKD", str(t or "").lower())
                            if not unicodedata.combining(c)).split())


def numero(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        m = re.fullmatch(r"\s*R?\$?\s*(\d+(?:[.,]\d+)?)\s*", v)
        if m:
            return float(m.group(1).replace(",", "."))
    return None


def casar(nome, colaboradores, setor=None):
    alvo = chato(nome).rstrip(".")
    if not alvo:
        return None
    exatos = [cid for cid, c in colaboradores.items()
              if chato(c.get("nome")) == alvo or chato(c.get("apelido")) == alvo]
    if len(exatos) == 1:
        return exatos[0]
    # prefixo: "vinicius l" -> "vinicius lyra"; "marcela" -> "marcella"
    def bate(c):
        n, a = chato(c.get("nome")), chato(c.get("apelido"))
        return n.startswith(alvo) or a.startswith(alvo) or alvo.startswith(n) or (a and alvo.startswith(a)) \
            or n.replace("ll", "l") == alvo.replace("ll", "l")
    cands = [cid for cid, c in colaboradores.items() if bate(c)]
    if len(cands) > 1 and setor:
        por_setor = [cid for cid in cands if chato(colaboradores[cid].get("setor")).startswith(chato(setor)[:5])]
        if len(por_setor) == 1:
            return por_setor[0]
    return cands[0] if len(cands) == 1 else None


def ler_mes(ws):
    """Colunas pelo cabecalho da linha 1; para no total (linha sem nome com numero)."""
    cab = {}
    linhas = list(ws.iter_rows(values_only=True))
    for j, v in enumerate(linhas[0]):
        k = chato(v)
        if k == "dia 05":
            cab["vale"] = j
        elif k == "dia 20":
            cab["salario"] = j
        elif k.startswith("comiss"):
            cab["comissao"] = j
        elif k == "descontos":
            cab["descontos"] = j
        elif k.startswith("observa"):
            cab["obs"] = j
        elif k == "google":
            cab["google"] = j
    saida = {}
    for row in linhas[1:]:
        nome = row[0]
        if not isinstance(nome, str) or not nome.strip() or chato(nome).startswith("folha de pagamento"):
            if not isinstance(nome, str):
                break   # linha do total: acabou a lista de gente
            continue
        reg = {}
        for k in ("vale", "salario", "comissao"):
            v = numero(row[cab[k]]) if k in cab and cab[k] < len(row) else None
            if v:
                reg[k] = v
        desc = row[cab["descontos"]] if "descontos" in cab and cab["descontos"] < len(row) else None
        obs = row[cab["obs"]] if "obs" in cab and cab["obs"] < len(row) else None
        goog = row[cab["google"]] if "google" in cab and cab["google"] < len(row) else None
        if desc not in (None, ""):
            reg["descontos"] = str(desc).strip()[:200]
        partes = [str(obs).strip() for obs in (obs,) if obs not in (None, "")]
        if goog not in (None, ""):
            partes.append(f"Google R$ {goog}")
        if partes:
            reg["obs"] = " · ".join(partes)[:200]
        if reg:
            saida[nome.strip()] = reg
    return saida


def ler_meta_bonus(ws):
    """{(setor, nome): {mes: valor}} — aba RESUMO META BONUS, blocos por setor."""
    linhas = list(ws.iter_rows(values_only=True))
    cab = None
    setor = None
    saida = {}
    for row in linhas:
        a = row[0]
        if isinstance(a, str) and chato(a).startswith("nomes"):
            cab = {j: chato(v) for j, v in enumerate(row) if isinstance(v, str) and chato(v) in MESES}
            continue
        if not cab or not isinstance(a, str) or not a.strip():
            continue
        vals = {f"{ANO}-{MESES.index(m) + 1:02d}": numero(row[j]) for j, m in cab.items() if j < len(row)}
        if all(v is None for v in vals.values()) and a.strip().upper() == a.strip():
            setor = a.strip()   # linha de setor: so o nome em caixa alta
            continue
        saida[(setor, a.strip())] = {m: v for m, v in vals.items() if v}
    return saida


def main() -> int:
    gravar = "--gravar" in sys.argv
    # Historico so ate agosto/26: de setembro em diante a folha e lancada no
    # painel e o mes nasce em branco (pedido do gestor, 03/09/2026).
    ate = next((a.split("=", 1)[1] for a in sys.argv if a.startswith("--ate=")), "2026-08")
    os.environ["DATABASE_URL"] = C.url_banco()
    import nucleo  # noqa: E402
    from areas.rh import _rh_gravar  # noqa: E402
    colaboradores = nucleo._rh_ler("colaboradores")
    folha = nucleo._rh_ler("folha")
    arquivo = Path(caminho("colaboradores_planilha"))
    print(f"planilha: {arquivo}")
    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)

    novos, sem_nome, mantidos = [], set(), 0
    def por_mes(mes, nome, reg, setor=None):
        nonlocal mantidos
        if mes > ate:
            return
        cid = casar(nome, colaboradores, setor)
        if not cid:
            sem_nome.add(nome); return
        atual = (folha.get(mes) or {}).get(cid) or {}
        if any(float(atual.get(k) or 0) for k in ("vale", "bonus", "comissao", "salario")):
            mantidos += 1; return   # o painel ja tem: fica
        novos.append((mes, cid, reg))

    for ws in wb.worksheets:
        k = chato(ws.title)
        if k in MESES:
            mes = f"{ANO}-{MESES.index(k) + 1:02d}"
            for nome, reg in ler_mes(ws).items():
                por_mes(mes, nome, reg)
    if "Meta bônus" in wb.sheetnames:
        for (setor, nome), meses in ler_meta_bonus(wb["Meta bônus"]).items():
            for mes, v in meses.items():
                por_mes(mes, nome, {"bonus": v}, setor)

    # junta os pedacos (mes, cid) — folha do mes + bonus da outra aba
    juntos = {}
    for mes, cid, reg in novos:
        juntos.setdefault((mes, cid), {}).update(reg)
    por_mes_qtd = {}
    for (mes, cid) in juntos:
        por_mes_qtd[mes] = por_mes_qtd.get(mes, 0) + 1
    print(f"  a preencher: {len(juntos)} (mes, pessoa) | por mes: {dict(sorted(por_mes_qtd.items()))} | ja no painel (mantidos): {mantidos}")
    if sem_nome:
        print("  nomes da planilha sem ficha no RH (ignorados):", ", ".join(sorted(sem_nome)))
    if not gravar:
        print("(sem --gravar: nada gravado)")
        return 0
    carimbo = C.agora().isoformat(timespec="seconds")
    for (mes, cid), reg in juntos.items():
        folha.setdefault(mes, {})[cid] = {"vale": 0.0, "bonus": 0.0, "comissao": 0.0, "salario": 0.0,
                                          "descontos": "", "obs": "", **reg,
                                          "origem": "planilha Colaboradores 2026", "editado_em": carimbo}
    _rh_gravar("folha", folha)
    print(f"  gravado rh_folha: {len(juntos)} registros")
    return 0


if __name__ == "__main__":
    C.saida_utf8()
    sys.exit(main())
