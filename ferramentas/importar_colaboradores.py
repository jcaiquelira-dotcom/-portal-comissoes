"""
Importa a planilha "Colaboradores 2026.xlsx" pro módulo de RH do portal.

Três abas da planilha, cada uma com um pedaço da verdade:
  Colaboradores → nome, salário bruto, bonificação e VT. A linha
                  "COLABORADORES DESLIGADOS" no meio divide quem está e quem
                  saiu — quem vem depois dela entra como desligado.
  Emails        → e-mail de cada um.
  Meta bônus    → o setor, que aparece como linha-título antes de cada grupo.
                  É a única aba que diz em que setor a pessoa trabalha.

O casamento entre as abas é por nome normalizado (sem acento, sem espaço
sobrando, minúsculo), porque a mesma pessoa está escrita de jeito diferente em
cada uma: "Josias " na folha, "Josias" nos e-mails. Na aba de metas os nomes
vêm abreviados ("Pedro P", "Vinicius L") e por isso existe APELIDOS abaixo.

O que este script NÃO faz de propósito: ligar a ficha ao vendedor do portal.
Existe "Matheus" na planilha no setor de Higienização e "Matheus" vendedor no
portal, e nada garante que seja a mesma pessoa. Ligar errado mostraria a venda
de um no painel do outro — isso é um clique no seletor da ficha, feito por quem
conhece o time.

Uso:
    set DATABASE_URL=postgresql://...
    python scripts/importar_colaboradores.py
    python scripts/importar_colaboradores.py --seco     # só mostra o que faria
"""

import io
import json
import os
import re
import sys
import unicodedata
import uuid
from datetime import datetime, timezone, timedelta
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "app"))
from caminhos import caminho  # config/caminhos.json — ver app/caminhos.py

from openpyxl import load_workbook

PLANILHA = caminho("colaboradores_planilha")

# Como o setor aparece na planilha -> como o portal chama.
SETORES = {
    "DESMONTAGEM": "Desmontagem", "CADASTRO": "Cadastro", "ANÚNCIO": "Anúncios",
    "EXPEDIÇÃO": "Expedição", "HIGIENIZAÇÃO": "Higienização",
    "ESTOQUE": "Estoque", "GERÊNCIA": "Gerência",
}

# Mesma pessoa escrita de dois jeitos nas abas. A chave da esquerda passa a
# valer como a da direita, senao a mesma pessoa vira dois cadastros.
APELIDOS = {
    "vinicius l": "vinicius lyra",
    "pedro p": "pedro paulo",
    "pedro h": "pedro henrique",
    "joao": "joao docinho",       # o "João" da expedição é o João Docinho
    "japa": "vinicius hideki",    # Japa é como chamam o Vinicius Hideki
    "rozemir": "nego",            # Nego é como chamam o Rozemir
    "joao gustavo": "gustavo",    # o Gustavo dos anúncios é o João Gustavo
}

# Quem a empresa trata pelo apelido: a ficha guarda o nome, a lista mostra o
# apelido — é por ele que o gestor procura a pessoa.
IDENTIDADES = {
    "nego": ("Rozemir", "Nego"),
    "vinicius hideki": ("Vinicius Hideki", "Japa"),
    "gustavo": ("João Gustavo", "Gustavo"),
}

# Setor ditado pelo gestor, pessoa por pessoa. Vale mais que a aba "Meta
# bônus": ela cobre só metade do time e estava vencida em quatro casos —
# Gustavo, Júlia, Japa e Matheus mudaram de área depois que ela foi escrita.
# As chaves já são as canônicas, depois de APELIDOS.
SETOR_DITADO = {
    "alison": "Anúncios",          "andreia": "Administrativo",
    "brenda": "Comercial",         "etelmilson": "Higienização",
    "felipe": "Desmontagem",       "flavia": "Comercial",
    "gustavo": "Comercial",        "vinicius hideki": "Estoque",
    "josias": "Gerência",          "joao docinho": "Expedição",
    "julia": "Administrativo",     "luan": "Expedição",
    "lucas": "Comercial",          "marcella": "Anúncios",
    "matheus": "Comercial",        "nego": "Desmontagem",
    "otavio": "Cadastro",          "pedro henrique": "Anúncios",
    "pedro paulo": "Anúncios",     "pietro": "Estoque",
    "vagner": "Higienização",      "vinicius franca": "Comercial",
    "vinicius lyra": "Cadastro",
}

# Quem do comercial tem portal de vendedor. Liga a ficha ao painel de
# desempenho; Lucas e Vinicius França são do comercial mas não têm portal.
VENDEDOR_DO_PORTAL = {
    "brenda": "brenda", "flavia": "flavia",
    "gustavo": "gustavo", "matheus": "matheus",
}

FUSO = timezone(timedelta(hours=-3))


def chave(nome):
    t = unicodedata.normalize("NFKD", str(nome or "").strip().lower())
    t = "".join(c for c in t if not unicodedata.combining(c))
    t = re.sub(r"\s+", " ", t).strip()
    return APELIDOS.get(t, t)


def num(v):
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    t = str(v or "").strip().replace("R$", "").replace(",", ".")
    try:
        return round(float(t), 2)
    except ValueError:
        return None


def ler():
    wb = load_workbook(PLANILHA, data_only=True)
    pessoas, ordem = {}, []

    def registrar(k, nome, origem):
        if k not in pessoas:
            pessoas[k] = {"nome": nome, "situacao": "ativo", "origens": set()}
            ordem.append(k)
        pessoas[k]["origens"].add(origem)
        return pessoas[k]

    # ---- folha, com o divisor de desligados ----
    desligados = False
    for linha in wb["Colaboradores"].iter_rows(min_row=2, values_only=True):
        bruto = str(linha[0] or "").strip()
        if not bruto:
            continue
        if "DESLIGAD" in bruto.upper():
            desligados = True
            continue
        p = registrar(chave(bruto), bruto, "folha")
        p["situacao"] = "desligado" if desligados else "ativo"
        p["salario"] = num(linha[1])
        p["bonificacao"] = num(linha[4])
        p["vt"] = num(linha[5])

    # ---- e-mails ----
    for linha in wb["Emails"].iter_rows(min_row=2, values_only=True):
        nome, email = str(linha[0] or "").strip(), str(linha[1] or "").strip()
        if nome and "@" in email:
            registrar(chave(nome), nome, "emails")["email"] = email

    # ---- setor ----
    setor = None
    for linha in wb["Meta bônus"].iter_rows(min_row=3, values_only=True):
        primeira = str(linha[0] or "").strip()
        if not primeira:
            continue
        # Linha-título: só a primeira célula preenchida e em maiúscula.
        if not [c for c in linha[1:] if c not in (None, "")] and primeira.isupper():
            setor = SETORES.get(primeira.strip())
            continue
        if setor:
            registrar(chave(primeira), primeira, "metas")["setor"] = setor

    # Quem nao aparece na folha veio so de uma aba lateral: pode ser apelido de
    # alguem que ja esta na lista com outro nome. Marca em vez de adivinhar.
    for p in pessoas.values():
        if "folha" not in p["origens"]:
            de_onde = " e ".join(sorted(p["origens"]))
            p["obs"] = (f"Importado só da aba {de_onde} — sem linha na folha. "
                        "Confira se não é a mesma pessoa de outro cadastro.")
        p["origens"] = sorted(p["origens"])
    for k, (nome, apelido) in IDENTIDADES.items():
        if k in pessoas:
            pessoas[k]["nome"], pessoas[k]["apelido"] = nome, apelido
    for k, p in pessoas.items():
        if k in SETOR_DITADO:
            p["setor"] = SETOR_DITADO[k]
        if k in VENDEDOR_DO_PORTAL:
            p["vendedor_id"] = VENDEDOR_DO_PORTAL[k]
    faltando = [k for k in SETOR_DITADO if k not in pessoas]
    if faltando:
        print("AVISO: setor ditado pra quem não está na planilha:", ", ".join(faltando))
    return [pessoas[k] for k in ordem]


def resumir(pessoas):
    ativos = [p for p in pessoas if p["situacao"] == "ativo"]
    print(f"{len(pessoas)} pessoas | {len(ativos)} ativas | "
          f"{len(pessoas) - len(ativos)} desligadas")
    sem_setor = [p["nome"] for p in ativos if not p.get("setor")]
    print(f"\ncom setor: {sum(1 for p in ativos if p.get('setor'))} de {len(ativos)} ativos")
    if sem_setor:
        print("sem setor na planilha:", ", ".join(sem_setor))
    print(f"com e-mail: {sum(1 for p in pessoas if p.get('email'))}")
    print(f"com salário: {sum(1 for p in pessoas if p.get('salario'))}")
    incompletos = [p["nome"] for p in pessoas if "folha" not in p.get("origens", ["folha"])]
    if incompletos:
        print("sem linha na folha (marcados pra conferência):", ", ".join(incompletos))
    print("\n  " + f"{'nome':22} {'situação':10} {'setor':14} {'salário':>9} {'VT':>7} {'bônus':>7}")
    for p in pessoas:
        print(f"  {p['nome'][:20]:22} {p['situacao']:10} {(p.get('setor') or '—')[:12]:14} "
              f"{p.get('salario') or '—':>9} {p.get('vt') or '—':>7} {p.get('bonificacao') or '—':>7}")


def gravar(pessoas, url):
    import psycopg2
    from psycopg2.extras import Json

    agora = datetime.now(FUSO).isoformat(timespec="seconds")
    conn = psycopg2.connect(url)
    with conn, conn.cursor() as cur:
        cur.execute("SELECT valor FROM dados_json WHERE chave = 'rh_colaboradores'")
        linha = cur.fetchone()
        atual = linha[0] if linha else {}

        # Casa pelo nome pra não duplicar quem já foi cadastrado à mão.
        por_nome = {chave(v.get("nome")): k for k, v in atual.items()}
        novos = atualizados = 0
        for p in pessoas:
            k = por_nome.get(chave(p["nome"]))
            if k:
                atualizados += 1
            else:
                k = uuid.uuid4().hex[:12]
                novos += 1
            antigo = atual.get(k, {})
            atual[k] = {
                **{campo: antigo.get(campo, "") for campo in
                   ("cargo", "contrato", "admissao", "nascimento", "telefone",
                    "endereco", "emergencia", "cpf", "rg",
                    "desligamento", "motivo_desligamento")},
                **antigo,
                "nome": p["nome"],
                "apelido": p.get("apelido") or antigo.get("apelido", ""),
                "vendedor_id": p.get("vendedor_id") or antigo.get("vendedor_id", ""),
                "situacao": p["situacao"],
                "setor": p.get("setor") or antigo.get("setor", ""),
                "email": p.get("email") or antigo.get("email", ""),
                "salario": p.get("salario"),
                "vt": p.get("vt"),
                "bonificacao": p.get("bonificacao"),
                "obs": p.get("obs") or antigo.get("obs", ""),
                "criado_em": antigo.get("criado_em") or agora,
                "editado_em": agora,
            }
        cur.execute(
            "INSERT INTO dados_json (chave, valor) VALUES (%s, %s) "
            "ON CONFLICT (chave) DO UPDATE SET valor = EXCLUDED.valor",
            ("rh_colaboradores", Json(atual)))
    conn.close()
    print(f"\n  {novos} cadastrados, {atualizados} atualizados — {len(atual)} no total")


def main():
    if not PLANILHA.exists():
        raise SystemExit(f"planilha não encontrada: {PLANILHA}")
    pessoas = ler()
    resumir(pessoas)
    if "--seco" in sys.argv:
        print("\n(--seco: nada foi gravado)")
        return
    url = os.environ.get("DATABASE_URL")
    if not url:
        raise SystemExit("\nDATABASE_URL não definida.")
    gravar(pessoas, url)


if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    main()
