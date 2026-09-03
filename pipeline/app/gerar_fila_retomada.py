"""
Fila de retomada: os melhores clientes pra cada vendedor ligar de volta.

ATENCAO -- contem dados pessoais (nome e telefone). Fica fora do git de proposito.
Entregue direto ao vendedor responsavel, nao publique.

Quem decide agora e a leitura da conversa pela IA (tabela classificacao_ia), nao
mais palavra-chave. A diferenca que importa: a heuristica so sabia dizer "o
vendedor nunca escreveu 'nao tenho'"; a IA responde se a peca existia e POR QUE
a venda nao saiu. Sem isso a lista enchia de gente que nunca teve chance.

Quem ENTRA (todas as condicoes):
  1. NEM a leitura NEM a heuristica indicam venda. Os dois sinais tem que
     concordar: a leitura exige a compra visivel na conversa e por isso perde as
     que fecham no balcao (acha 236 das 352 vendas reais do portal, 67%), enquanto
     a heuristica de palavra-chave acha 340, 97%. Ligar pra quem ja comprou custa
     mais caro que perder um candidato duvidoso, entao vale o mais abrangente.
  2. a leitura diz que TINHAMOS a peca ('sim') ou algo proximo ('parcial')
  3. o motivo da perda e recuperavel -- quem so pesquisava preco, ou pediu peca
     que nao servia no carro dele, fica de fora
  4. da pra saber QUAL peca ele queria
  5. o cliente nao avisou que ja comprou / desistiu, e nao e alguem querendo
     vender carro pra gente

O que NAO deu certo: cruzar os candidatos contra as vendas reais do portal-comissoes
pra excluir quem comprou. O portal registra so produto, valor, data e vendedor -- nao
tem chave de cliente. Casar por nome de peca gera falso match demais ("capa de antena
de TETO" casa com "luz cortesia TETO solar"), porque um desmonte vende dezenas de
pecas do mesmo carro pra pessoas diferentes. Pra fechar esse furo o portal precisaria
guardar quem comprou.

Depois disso cada um recebe uma nota e so os 30 melhores de cada vendedor entram
na planilha -- a ideia e uma lista curta que da pra trabalhar de verdade num dia,
nao um cadastro de mil linhas que ninguem abre.
"""

import json
import re
import sqlite3
import sys
import time
import unicodedata
import urllib.parse
import urllib.request
from collections import defaultdict, deque
from datetime import datetime, timedelta, timezone
from pathlib import Path
from caminhos import caminho, portal  # config/caminhos.json — ver app/caminhos.py

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import env

ROOT = caminho("dados")   # pasta de dados desta maquina (config/caminhos.json)
TOKEN = env("TOTALK_TOKEN")
BASE = "https://api.wts.chat"
HOJE = datetime.now(timezone.utc)  # "dias parado" conta ate hoje, nao ate o fim do sync
PAUSA = 0.4
POR_VENDEDOR = 30
SAIDA = ROOT / "Fila_Retomada_CONFIDENCIAL.xlsx"

# A mesma fila tambem sai em JSON, pra alimentar o painel de CRM dentro do
# portal-comissoes (app/static/retomada.html la). O painel leva bem mais gente que
# a planilha: o teto de 30 existia porque ninguem trabalha uma planilha de mil
# linhas, mas na tela o vendedor ve os pendentes de maior nota primeiro e a fila
# so precisa nao secar no meio do mes.
POR_VENDEDOR_CRM = 200
# Só entra no painel quem parou de falar com a gente nos últimos DIAS_MAX_CRM
# dias: lead frio de mês passado já resolveu a vida em outro lugar, e uma fila
# curta e recente é a que o vendedor consegue trabalhar de verdade no dia.
# Baixado de 10 pra 5 em 28/08/2026 a pedido do gestor: com 10 dias a fila
# tinha 445 clientes (~148 por vendedor) e ninguém trabalhava até o fim; com 5
# ficam ~66 por vendedor, e quem sumiu há menos de uma semana ainda lembra da
# conversa. Roda todo dia às 07:30 pelo pipeline_diario.bat.
DIAS_MAX_CRM = 5
SAIDA_CRM = ROOT / "Fila_CRM_CONFIDENCIAL.json"
# nome que o Totalk usa pro atendente -> id do vendedor no portal-comissoes
# "Gustavo" -> "lucas" NAO e erro: o historico e do Gustavo (desligado em
# 31/08/2026), mas a FILA e trabalho de quem pode ligar — e quem assumiu o
# assento e os clientes dele foi o Lucas. Fila na mao de um usuario bloqueado
# e cliente que ninguem chama.
ID_PORTAL = {"Flávia": "flavia", "Gustavo": "lucas", "Matheus": "matheus",
             "Lucas": "lucas"}

# motivos onde nao ha o que reverter numa segunda tentativa
MOTIVO_MORTO = {"so_pesquisando", "peca_errada"}

# como cada motivo vira abordagem, e quanto vale a ligacao
MOTIVO = {
    "sem_resposta":  ("A conversa parou do nosso lado", 38),
    "cliente_sumiu": ("Respondemos tudo e ele sumiu", 30),
    "preco":         ("Achou caro — cabe negociar", 24),
    "frete":         ("Travou no frete ou no prazo", 18),
    "pagamento":     ("Travou na forma de pagamento", 18),
    "outro":         ("Conversou e não fechou", 10),
}


def limpa(t):
    """Minuscula e sem acento. Tirar os combining chars e obrigatorio: sem isso
    'ja comprei' nunca casa com o texto decomposto e o filtro deixa passar."""
    t = unicodedata.normalize("NFKD", (t or "").lower())
    return "".join(c for c in t if not unicodedata.combining(c))


JA_COMPROU = [limpa(x) for x in [
    "já comprei", "já consegui", "já achei", "comprei em outro", "consegui em outro",
    "achei em outro", "já resolvi", "não precisa mais", "já peguei", "desisti",
]]
QUER_VENDER = [limpa(x) for x in [
    "vcs compram carro", "vocês compram carro", "compram carro batido", "vendo meu carro",
    "quero vender meu", "comprar meu carro",
]]
# O que aparece no lugar da fala quando o cliente mandou midia em vez de texto.
# BUTTONS fica de fora de proposito: e escolha de menu do bot, nao fala dele.
MIDIA = {"IMAGE": "[foto]", "AUDIO": "[áudio]", "VIDEO": "[vídeo]",
         "DOCUMENT": "[arquivo]", "STICKER": "[figurinha]", "LOCATION": "[localização]"}
# Quem chega pelo site ou pelo anuncio manda um texto pronto ("Ola NEVADA
# ECOPECAS, gostaria de ajuda com uma compra no site" + link). Isso e o
# formulario falando, nao o cliente -- na caixinha do painel so ocupa a linha
# que deveria mostrar o que ele realmente disse. Sao ~1.900 mensagens.
ENTRADA_SITE = "ola nevada ecopecas"
RE_URL = re.compile(r"https?://\S+")
RE_PRECO = re.compile(r"r\$\s*\d|(?<![\d,.])\d{2,5}[,.]00\b|\bpix\b", re.IGNORECASE)


def api(path, params):
    q = "&".join(f"{k}={urllib.parse.quote(str(v))}" for k, v in params.items())
    req = urllib.request.Request(f"{BASE}{path}?{q}",
                                 headers={"Authorization": f"Bearer {TOKEN}"})
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())


def buscar_contatos(desde):
    """Nome/telefone nao foram salvos no sync -- rebusca as sessoes so pra isso."""
    print("buscando contatos na API...")
    contatos, pagina = {}, 1
    while True:
        d = api("/chat/v2/session", {
            "CreatedAt.After": desde, "PageNumber": pagina, "PageSize": 100,
            "IncludeDetails": "ContactDetails", "OrderBy": "createdat",
            "OrderDirection": "ASCENDING",
        })
        for s in d.get("items") or []:
            cd = s.get("contactDetails")
            if cd:
                contatos[s["id"]] = {
                    "nome": (cd.get("name") or "").strip() or "(sem nome)",
                    "fone": cd.get("phonenumberFormatted") or "",
                    "tags": ", ".join(cd.get("tagsName") or []),
                }
        if pagina % 30 == 0:
            print(f"  pagina {pagina}/{d.get('totalPages')}")
        if not d.get("hasMorePages"):
            break
        pagina += 1
        time.sleep(PAUSA)
    print(f"  {len(contatos)} contatos obtidos")
    return contatos


def fala_do_cliente(tipo, txt):
    """Uma fala pronta pra aparecer no painel, ou "" quando nao ha o que mostrar.

    Tira o link inteiro: uma URL de produto ocupa a linha toda e nao diz nada
    que a coluna da peca ja nao diga. Se a mensagem era so o link, sobra a
    marca de que ele mandou um."""
    if tipo != "TEXT":
        return MIDIA.get(tipo, "")
    t = (txt or "").strip()
    if not t or limpa(t).startswith(ENTRADA_SITE):
        return ""
    t = RE_URL.sub("", t).strip()
    return (t or "[link do produto]")[:200]


def exportar_json(fila, datas):
    """Mesma fila da planilha, em JSON, separada por vendedor.

    Um arquivo por vendedor (e não um bloco único) porque é assim que o
    portal-comissoes guarda dado de vendedor -- cada um carrega só o seu.

    O período que sai aqui é o da fila, não o que a IA leu: com o corte de
    DIAS_MAX_CRM dias eles deixaram de ser a mesma coisa, e é o da fila que o
    vendedor vê na tela.
    """
    inicio = (HOJE - timedelta(days=DIAS_MAX_CRM)).date().isoformat()
    saida = {"gerado_em": HOJE.isoformat(), "de": max(inicio, datas[0]),
             "ate": datas[-1], "vendedores": {}}
    print("\njson pro painel de CRM:")
    # Dois nomes podem desaguar no MESMO vid — "Gustavo" e "Lucas" apontam os
    # dois pra fila do lucas desde a troca de assento. Atribuir direto fazia o
    # segundo sobrescrever o primeiro: os 11 clientes do Gustavo sumiam da fila
    # em vez de mudar de dono. Junta primeiro, corta o teto depois.
    por_vid = {}
    for vend, vid in ID_PORTAL.items():
        # O corte por data vem ANTES do teto de POR_VENDEDOR_CRM: filtrar depois
        # deixaria de fora um cliente recente só porque a nota dele não entrou
        # entre as maiores do período inteiro.
        todos = [it for it in fila.get(vend, []) if it["dias"] <= DIAS_MAX_CRM]
        alvo = por_vid.setdefault(vid, {"nome": vend, "itens": [], "de": 0})
        alvo["itens"].extend(todos)
        alvo["de"] += len(todos)
        # o nome mostrado e o de quem TRABALHA a fila — o dono do vid
        if vend.lower().startswith(vid[:4]):
            alvo["nome"] = vend
    for vid, bloco in por_vid.items():
        itens = sorted(bloco["itens"], key=lambda x: -x["nota"])[:POR_VENDEDOR_CRM]
        saida["vendedores"][vid] = {"nome": bloco["nome"], "itens": itens}
        print(f"  {bloco['nome']} ({vid}): {len(itens)} de {bloco['de']} elegíveis "
              f"nos últimos {DIAS_MAX_CRM} dias")
    SAIDA_CRM.write_text(json.dumps(saida, ensure_ascii=False), encoding="utf-8")
    print(f"salvo: {SAIDA_CRM.name} ({SAIDA_CRM.stat().st_size/1024:.0f} KB)")


def main():
    conn = sqlite3.connect(ROOT / "vendas.db")
    dados = json.loads((ROOT / "dataset.json").read_text(encoding="utf-8"))
    ids = json.loads((ROOT / "session_ids.json").read_text(encoding="utf-8"))
    por_id = dict(zip(ids, dados))

    # trava de seguranca: se o pareamento id<->dataset sair de ordem, isso estoura aqui
    # em vez de gerar uma lista com o cliente errado no vendedor errado.
    for sid, created_at in conn.execute("SELECT id, created_at FROM sessoes LIMIT 200"):
        d = por_id.get(sid)
        assert d and d["d"] == created_at[:10], f"pareamento furado em {sid}"
    print("pareamento dataset<->banco: OK")

    ia = {r[0]: {"venda": r[1], "motivo": r[2], "peca": (r[3] or "").strip(),
                 "tinha": r[4], "tipo": r[5], "resumo": r[6]}
          for r in conn.execute(
              "SELECT session_id, virou_venda, motivo_nao_venda, peca_procurada, "
              "tinhamos_a_peca, tipo_cliente, resumo FROM classificacao_ia")}
    if not ia:
        raise SystemExit("classificacao_ia esta vazia — rode app/classificar_ia.py antes")
    datas = sorted(por_id[s]["d"] for s in ia if s in por_id)
    print(f"conversas lidas pela IA: {len(ia):,} ({datas[0]} a {datas[-1]})")

    ja_comprou, quer_vender, tem_preco, foto_cliente = set(), set(), set(), set()
    # Só as 3 últimas falas do cliente: é o suficiente pro vendedor lembrar da
    # conversa sem abrir o Totalk, e um deque com teto evita carregar 156 mil
    # mensagens na memória pra jogar quase todas fora depois.
    ultimas = defaultdict(lambda: deque(maxlen=3))
    for sid, direcao, tipo, txt, raw in conn.execute(
        "SELECT session_id, direction, type, text, raw FROM mensagens ORDER BY created_at ASC"
    ):
        if sid not in ia:
            continue
        tl = limpa(txt) if txt else ""
        if direcao == "TO_HUB":
            if txt and json.loads(raw).get("userId") and RE_PRECO.search(txt):
                tem_preco.add(sid)
        else:
            if tipo in ("IMAGE", "VIDEO"):
                foto_cliente.add(sid)
            fala = fala_do_cliente(tipo, txt)
            if fala:
                ultimas[sid].append(fala)
            if txt:
                if any(k in tl for k in JA_COMPROU):
                    ja_comprou.add(sid)
                if any(k in tl for k in QUER_VENDER):
                    quer_vender.add(sid)

    contatos = buscar_contatos(datas[0] + "T00:00:00Z")
    CANAL = {"S": "Site orgânico", "G": "Google Ads", "AX": "Anúncio s/ rastreio",
             "AF": "Facebook Ads", "AI": "Instagram Ads", "AO": "Meta (outro)",
             "I": "Instagram bio", "D": "Direto"}
    ALTA_INTENCAO = {"S", "D", "G"}

    fila = defaultdict(list)
    descartes = defaultdict(int)
    for sid, created_at, raw in conn.execute("SELECT id, created_at, raw FROM sessoes"):
        a = ia.get(sid)
        if not a:
            descartes["fora do período lido pela IA"] += 1
            continue
        m = por_id[sid]
        if a["venda"]:
            descartes["a leitura confirma que comprou"] += 1
            continue
        # A leitura sozinha nao basta. Ela exige a compra visivel na conversa, e boa
        # parte das vendas fecha fora dela (o cliente some do chat e aparece no balcao):
        # contra o portal de comissoes a leitura acha 236 das 352 vendas reais, 67%.
        # A heuristica de palavra-chave acha 340, 97%. Entao quem a heuristica marca
        # como venda fica de fora mesmo que a leitura discorde -- ligar pra quem ja
        # comprou custa mais caro que perder um candidato duvidoso.
        if m["cv"] == "P":
            descartes["a heurística indica compra (fechou fora do chat)"] += 1
            continue
        if a["tinha"] not in ("sim", "parcial"):
            descartes[f"não tínhamos a peça ({a['tinha']})"] += 1
            continue
        if a["motivo"] in MOTIVO_MORTO:
            descartes[f"motivo sem volta ({a['motivo']})"] += 1
            continue
        if not a["peca"]:
            descartes["não dá pra saber que peça queria"] += 1
            continue
        if sid in ja_comprou:
            descartes["cliente disse que já comprou"] += 1
            continue
        if sid in quer_vender:
            descartes["queria vender carro pra gente"] += 1
            continue

        dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
        dias = max(0, (HOJE - dt).days)
        c = contatos.get(sid, {})
        tags = c.get("tags", "")
        gancho, peso_motivo = MOTIVO.get(a["motivo"], MOTIVO["outro"])

        # nota: o que faz valer a ligacao
        nota = peso_motivo
        nota += 20 if a["tinha"] == "sim" else 8     # tinha em maos > tinha parecida
        nota += 18 if sid in tem_preco else 0        # ja teve preco na mesa
        nota += 12 if a["tipo"] == "oficina" else 0  # oficina recompra
        nota += 10 if sid in foto_cliente else 0     # cliente mandou foto = interesse real
        nota += 10 if m["c"] in ALTA_INTENCAO else 0
        nota += 10 if "oportunidade" in limpa(tags) else 0
        nota += max(0, 15 - dias // 3)               # quanto mais recente, melhor
        nota += min(m["nm"] // 3, 8)                 # conversa que andou de verdade

        fila[m["u"]].append({
            "sid": sid,
            "nota": nota,
            "prio": "ALTA" if a["motivo"] == "sem_resposta" else "MÉDIA",
            "data": dt.strftime("%d/%m/%Y"),
            "dias": dias,
            "nome": c.get("nome", "(não encontrado)"),
            "fone": c.get("fone", ""),
            "peca": a["peca"][:180],
            "gancho": gancho,
            "tinha": "Sim" if a["tinha"] == "sim" else "Parecida",
            "tipo": {"oficina": "Oficina", "consumidor": "Consumidor"}.get(a["tipo"], "—"),
            "resumo": (a["resumo"] or "")[:220],
            "canal": CANAL.get(m["c"], m["c"]),
            "ultimas": list(ultimas.get(sid, ())),
            "link": json.loads(raw).get("previewUrl") or "",
        })

    print("\ndescartados:")
    for k, v in sorted(descartes.items(), key=lambda x: -x[1]):
        print(f"  {v:5d}  {k}")

    exportar_json(fila, datas)
    if "--somente-json" in sys.argv:
        return

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Como usar"
    ws0.column_dimensions["A"].width = 108
    periodo = (f"{datas[0][8:10]}/{datas[0][5:7]} a {datas[-1][8:10]}/{datas[-1][5:7]}"
               f"/{datas[-1][:4]}")
    INSTRU = [
        ("Fila de retomada — clientes pra ligar", "titulo"),
        (f"Gerado em {HOJE.strftime('%d/%m/%Y')} · atendimentos de {periodo}", "sub"),
        ("", ""),
        ("O que é esta lista", "h"),
        ("Os 30 clientes de cada vendedor com maior chance de virar venda numa segunda tentativa. "
         "Cada aba tem o nome de um vendedor — são os atendimentos dele.", "p"),
        ("Cada conversa foi lida inteira, uma por uma. Não é filtro de palavra-chave: a leitura "
         "responde se a peça existia no nosso estoque e por que a venda não saiu.", "p"),
        ("", ""),
        ("Todo mundo aqui passou por estes filtros", "h"),
        ("• Não comprou — nem a leitura da conversa nem a detecção automática indicam venda. "
         "Os dois critérios tiveram que concordar, justamente pra tirar da lista quem fechou "
         "a compra fora do WhatsApp.", "p"),
        ("• Nós TÍNHAMOS a peça (ou uma bem parecida) — quem ouviu \"não tenho\" ficou de fora", "p"),
        ("• O motivo da perda dá pra reverter — quem só pesquisava preço, ou pediu peça que não "
         "servia no carro dele, ficou de fora", "p"),
        ("• O cliente não avisou que já comprou em outro lugar nem que desistiu", "p"),
        ("• Dá pra saber qual peça ele queria — está na coluna \"Peça que o cliente procurava\"", "p"),
        ("", ""),
        ("Como ler a prioridade", "h"),
        ("ALTA (vermelho) — a conversa parou do nosso lado. Quase sempre houve resposta antes; o que "
         "faltou foi o último passo. Ninguém disse não pra ele, então é o mais fácil de "
         "recuperar. Comece por aqui.", "p"),
        ("MÉDIA (amarelo) — conversou de verdade e não fechou. A coluna \"Por que ligar\" diz onde travou.", "p"),
        ("", ""),
        ("Dica pra abordagem", "h"),
        ("A coluna \"Peça que o cliente procurava\" traz a peça com veículo e ano. Use isso na abertura — "
         "retomar citando exatamente o que a pessoa pediu funciona melhor que um \"oi, tudo bem?\" genérico.", "p"),
        ("\"O que aconteceu\" resume o atendimento em uma linha, pra você chegar na ligação sabendo onde parou.", "p"),
        ("A última coluna abre a conversa original no Totalk.", "p"),
        ("", ""),
        ("Uma ressalva honesta", "h"),
        ("Ainda pode aparecer aqui alguém que comprou no balcão ou pelo site, porque o registro "
         "de vendas não guarda quem foi o cliente — só o produto. Não dá pra cruzar as duas "
         "pontas. Se você reconhecer alguém que já levou a peça, é só pular.", "p"),
        ("", ""),
        ("Documento confidencial — contém nome e telefone de cliente. Não encaminhe para fora da equipe.", "aviso"),
    ]
    for texto, tipo in INSTRU:
        ws0.append([texto])
        c = ws0.cell(row=ws0.max_row, column=1)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if tipo == "titulo":
            c.font = Font(bold=True, size=16)
        elif tipo == "sub":
            c.font = Font(size=10, color="7A7166")
        elif tipo == "h":
            c.font = Font(bold=True, size=11, color="B4501A")
        elif tipo == "aviso":
            c.font = Font(bold=True, size=10, color="A33232")
        else:
            c.font = Font(size=10)
            ws0.row_dimensions[ws0.max_row].height = 28

    COLS = [("#", 4), ("Prioridade", 11), ("Data", 11), ("Dias", 6), ("Cliente", 24),
            ("WhatsApp", 17), ("Peça que o cliente procurava", 52), ("Temos?", 10),
            ("Por que ligar", 27), ("Perfil", 12), ("O que aconteceu", 58),
            ("Canal", 13), ("Conversa", 10)]
    CORES = {"ALTA": "F8D7DA", "MÉDIA": "FFF9E6"}
    QUEBRA = {7, 11}  # peca e resumo precisam de wrap

    # Gustavo continua na lista de proposito: sessoes dele anteriores ao corte
    # de 31/08 ainda podem estar na janela de 5 dias, e cliente nao evapora
    # porque o vendedor saiu — a fila dele morre sozinha quando a janela andar.
    for vend in ["Flávia", "Gustavo", "Matheus", "Lucas"]:
        todos = fila.get(vend, [])
        itens = sorted(todos, key=lambda x: -x["nota"])[:POR_VENDEDOR]
        if not itens:
            continue
        ws = wb.create_sheet(vend)
        ws.append([c[0] for c in COLS])
        for i, (_, larg) in enumerate(COLS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = larg
            cel = ws.cell(row=1, column=i)
            cel.font = Font(bold=True, color="FFFFFF", size=10)
            cel.fill = PatternFill("solid", fgColor="1A1611")
        ws.freeze_panes = "A2"
        for i, it in enumerate(itens, start=1):
            ws.append([i, it["prio"], it["data"], it["dias"], it["nome"], it["fone"],
                       it["peca"], it["tinha"], it["gancho"], it["tipo"], it["resumo"],
                       it["canal"], ""])
            r = ws.max_row
            fill = PatternFill("solid", fgColor=CORES[it["prio"]])
            for c in range(1, len(COLS) + 1):
                cel = ws.cell(row=r, column=c)
                cel.fill = fill
                cel.alignment = Alignment(vertical="top", wrap_text=(c in QUEBRA))
            ws.cell(row=r, column=2).font = Font(bold=True, size=9)
            ws.cell(row=r, column=11).font = Font(size=9, color="5C5348")
            if it["link"]:
                cel = ws.cell(row=r, column=len(COLS))
                cel.value = "abrir"
                cel.hyperlink = it["link"]
                cel.font = Font(color="0563C1", underline="single", size=9)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"
        alta = sum(1 for i in itens if i["prio"] == "ALTA")
        print(f"  {vend}: {len(itens)} de {len(todos)} elegíveis "
              f"(alta {alta}, média {len(itens)-alta}) — nota {itens[-1]['nota']} a {itens[0]['nota']}")

    # O Excel trava o arquivo enquanto esta aberto. Perder a rodada inteira (que
    # rebusca 9 mil contatos na API) por causa disso nao faz sentido: grava ao lado.
    destino = SAIDA
    try:
        wb.save(destino)
    except PermissionError:
        destino = SAIDA.with_name(
            f"{SAIDA.stem}_{HOJE.strftime('%d-%m_%Hh%M')}{SAIDA.suffix}")
        wb.save(destino)
        print(f"\n[aviso] {SAIDA.name} esta aberto no Excel e nao pode ser sobrescrito.")
    print(f"\nsalvo: {destino.name} ({destino.stat().st_size/1024:.0f} KB)")


if __name__ == "__main__":
    main()
