# -*- coding: utf-8 -*-
"""Gera a planilha de fluxo de caixa pro socio preencher.

Quem alimenta o controle e um dos socios, com quase 60 anos, e a planilha atual
e dele. Entao isto nao e uma planilha nova: e a MESMA planilha, com as
categorias certas.

O formato dele, que fica igual:
    - blocos de tres colunas — Dia | Tipo | Valor — LADO A LADO
    - o nome do bloco escrito em cima da coluna do meio
    - subtotal no pe de cada bloco
    - uma aba por mes, com os mesmos nomes (Jan..Maio..Dez)
    - as palavras dele: Creditos, Debitos, Saldo
    - uma aba de ano, no lugar da "Geral"

O que muda, que e a razao de existir:
    - cada bloco e uma CATEGORIA do DRE, e nao uma caixa generica como
      "Diversos", "Cartoes" ou "Fixo". A posicao ja diz a categoria: ele
      escreve embaixo do titulo certo e acabou. Nao ha lista pra escolher nem
      nome pra digitar — e por isso que a agencia Beelieve nao tem como virar
      seis grafias diferentes de novo.
    - some o bloco "Cartao": cartao e como se paga, nao o que se compra. A
      compra vai na coluna do que foi comprado, venha do cartao ou nao. Sao
      R$ 327 mil do ano que hoje entram como "Cartao" e nao dizem nada.
    - as retiradas dos socios sao quatro colunas, uma por socio, no lugar da
      cor de fundo. Mesma informacao, sem depender de lembrar qual cinza e.
    - o subtotal, o resumo do mes e o DRE se fazem por formula. Ninguem soma.

Uso:
    python scripts/gerar_planilha_modelo.py --saida "C:/.../Fluxo 2026.xlsx" --ano 2026
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "app"))

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

MESES = ["Jan", "Fev", "Mar", "Abr", "Maio", "Jun",
         "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
# Linhas por bloco. O maior movimento de uma categoria num mes foi frete, com
# cerca de 16 lancamentos; 18 deixa folga sem esticar a tela pra baixo.
LINHAS_POR_BLOCO = 18
# Duas faixas de blocos, e so. Cada faixa a mais sao 21 linhas a mais de altura,
# e altura e exatamente o que o gestor pediu pra nao ter.
FAIXAS = 2
PRIMEIRA_FAIXA = 4
# O mes inteiro cabe na tela com um zoom menor. 80% mostra ~56 linhas num
# monitor comum, e o mes ocupa 45 — ele abre e ve tudo, sem rolar.
ZOOM = 80

# Onde o resumo do mes mora. Fixo aqui porque a aba do ano e o teste apontam
# pra ca: quando o resumo saiu de B4/B5 pra ca, a aba do ano continuou lendo
# B4 e passou a mostrar mes vazio sem reclamar de nada.
CEL_CREDITOS = "D1"
CEL_DEBITOS = "G1"
CEL_SALDO = "J1"
CEL_GERAIS = "M1"

AZUL = "1F3864"
CINZA_CLARO = "F2F2F2"
CINZA_LINHA = "BFBFBF"
VERDE = "1E7B4D"
VERMELHO = "B03A2E"
BRANCO = "FFFFFF"

DINHEIRO = 'R$ #,##0.00'
FINA = Side(style="thin", color=CINZA_LINHA)
GROSSA = Side(style="medium", color=AZUL)
BORDA = Border(left=FINA, right=FINA, top=FINA, bottom=FINA)

# Uma cor por familia, pra ele achar o bloco pela cor como ja faz hoje.
COR_DA_FAMILIA = {
    "Receita bruta": "1E7B4D",
    "Deducoes da receita": "8E6A00",
    "Custo da peca vendida": "A0522D",
    "Despesas comerciais": "1F6F8B",
    "Pessoal": "5B3E8E",
    "Ocupacao e estrutura": "2E5C9A",
    "Administrativas": "556B2F",
    "Frota propria": "7A5C29",
    "Financeiro": "8B3A62",
    "Sai do caixa, nao e despesa": "6B6B6B",
}


def _familias():
    """As familias do plano, cada uma com suas contas, na ordem do DRE."""
    import server
    fora = []
    for bloco in server.PLANO_DE_CONTAS:
        contas = [server.CONTAS_POR_CODIGO[c[0]] for c in bloco["contas"]]
        fora.append({"grupo": bloco["grupo"], "dre": bloco["dre"],
                     "entrada": bool(bloco.get("entrada")), "contas": contas})
    return fora


def _titulo(ws, celula, texto, tamanho=14, cor=None):
    ws[celula] = texto
    ws[celula].font = Font(bold=True, size=tamanho, color=cor or AZUL)


def layout_dos_blocos(familias):
    """Onde cada bloco cai na aba do mes.

    Devolve {codigo: {"col_dia","col_desc","col_valor","linha_1","linha_n",
                      "linha_total"}} mais as faixas pra desenhar.

    O eixo aqui e LARGURA, nao altura. A planilha dele poe os blocos um do lado
    do outro e cabe na tela sem rolar pra baixo: e assim que ele enxerga o mes
    inteiro de uma vez. Empilhar uma familia embaixo da outra dava 307 linhas —
    tecnicamente organizado e inutil na pratica, porque ele perde a visao do
    mes.

    Sao 57 blocos contra os 9 dele, entao nao ha como caber na MESMA largura.
    A troca escolhida: poucas faixas, muita largura — rolar pro lado, que e o
    que ele ja faz hoje, em vez de rolar pra baixo, que e o que ele reclamou.
    """
    # Duas faixas, equilibradas por numero de blocos. Mais faixas e mais altura;
    # menos faixas e uma largura que ninguem navega.
    total_blocos = sum(len(f["contas"]) for f in familias)
    alvo = total_blocos / FAIXAS
    grupos, atual, acumulado = [], [], 0
    for fam in familias:
        atual.append(fam)
        acumulado += len(fam["contas"])
        if acumulado >= alvo and len(grupos) < FAIXAS - 1:
            grupos.append(atual)
            atual, acumulado = [], 0
    if atual:
        grupos.append(atual)

    pos, faixas = {}, []
    linha = PRIMEIRA_FAIXA
    for grupo in grupos:
        cab_familia = linha
        cab_conta = linha + 1
        primeira = linha + 2
        ultima = primeira + LINHAS_POR_BLOCO - 1
        total = ultima + 1
        col = 1
        blocos_da_familia = []
        for fam in grupo:
            col_inicio = col
            for conta in fam["contas"]:
                pos[conta["codigo"]] = {
                    "col_dia": col, "col_desc": col + 1, "col_valor": col + 2,
                    "linha_1": primeira, "linha_n": ultima, "linha_total": total,
                }
                col += 3
            blocos_da_familia.append({"familia": fam, "col_inicio": col_inicio,
                                      "col_fim": col - 1})
        faixas.append({"familias": blocos_da_familia, "cab_familia": cab_familia,
                       "cab_conta": cab_conta, "linha_1": primeira,
                       "linha_n": ultima, "linha_total": total,
                       "colunas": col - 1})
        linha = total + 2                      # uma linha de respiro so
    return pos, faixas, linha


def aba_mes(wb, mes, familias, ano, pos, faixas):
    ws = wb.create_sheet(mes)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = ZOOM

    _titulo(ws, "A1", "%s de %d" % (mes, ano), 16)

    # --- resumo do mes, deitado numa linha so ------------------------------
    # Em pe ele comia seis linhas do alto da tela. Deitado cabe tudo em duas, e
    # cada linha economizada aqui e uma linha a mais de lancamento visivel.
    entradas, saidas = [], []
    for fam in familias:
        for c in fam["contas"]:
            p = pos[c["codigo"]]
            ref = "%s%d" % (get_column_letter(p["col_valor"]), p["linha_total"])
            (entradas if fam["entrada"] else saidas).append(ref)
    soma = lambda refs: "=" + ("+".join(refs) if refs else "0")

    p_ger = pos["7.07"]
    ref_ger = "%s%d" % (get_column_letter(p_ger["col_valor"]), p_ger["linha_total"])
    resumo = [
        (CEL_CREDITOS, "Créditos (entrou)", soma(entradas), DINHEIRO, VERDE),
        (CEL_DEBITOS, "Débitos (saiu)", soma(saidas), DINHEIRO, VERMELHO),
        (CEL_SALDO, "Saldo", "=%s-%s" % (CEL_CREDITOS, CEL_DEBITOS), DINHEIRO, AZUL),
        (CEL_GERAIS, "Despesas gerais",
         "=IF(${0}=0,0,{1}/${0})".format(CEL_DEBITOS, ref_ger), "0.0%", "808080"),
    ]
    for cel, rot, formula, fmt, cor in resumo:
        v = ws[cel]
        v.value = formula
        v.number_format = fmt
        v.font = Font(bold=True, size=13, color=cor)
        r = ws.cell(row=2, column=v.column - 1, value=rot)
        r.font = Font(size=9, color="808080")
        r.alignment = Alignment(horizontal="right")
        ws.merge_cells(start_row=2, start_column=v.column - 1,
                       end_row=2, end_column=v.column)
    ws.conditional_formatting.add(CEL_SALDO, CellIsRule(
        operator="lessThan", formula=["0"], font=Font(bold=True, color=VERMELHO)))
    ws.conditional_formatting.add(CEL_GERAIS, CellIsRule(
        operator="greaterThan", formula=["0.02"],
        font=Font(bold=True, color=VERMELHO)))
    ws.cell(row=2, column=16,
            value="Escreva embaixo do título certo — a coluna já é a categoria.").font = Font(
        size=9, italic=True, color="A0A0A0")

    # --- as faixas ----------------------------------------------------------
    dv = DataValidation(type="decimal", operator="greaterThan", formula1="0",
                        allow_blank=True, errorTitle="Valor",
                        error="O valor precisa ser maior que zero. Devolução "
                              "tem bloco próprio, em Deduções.")
    ws.add_data_validation(dv)

    for faixa in faixas:
        for bloco in faixa["familias"]:
            fam = bloco["familia"]
            cor = COR_DA_FAMILIA.get(fam["grupo"], AZUL)

            ws.merge_cells(start_row=faixa["cab_familia"],
                           start_column=bloco["col_inicio"],
                           end_row=faixa["cab_familia"],
                           end_column=bloco["col_fim"])
            t = ws.cell(row=faixa["cab_familia"], column=bloco["col_inicio"],
                        value=fam["grupo"].upper())
            t.font = Font(bold=True, size=11, color=BRANCO)
            t.fill = PatternFill("solid", fgColor=cor)
            t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[faixa["cab_familia"]].height = 20

            for conta in fam["contas"]:
                p = pos[conta["codigo"]]
                cd, cdesc, cv = p["col_dia"], p["col_desc"], p["col_valor"]

                ws.merge_cells(start_row=faixa["cab_conta"], start_column=cd,
                               end_row=faixa["cab_conta"], end_column=cv)
                h = ws.cell(row=faixa["cab_conta"], column=cd, value=conta["nome"])
                h.font = Font(bold=True, size=9, color=BRANCO)
                h.fill = PatternFill("solid", fgColor=cor)
                h.alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)
                if conta["ajuda"]:
                    h.comment = Comment("%s\n\n%s" % (conta["codigo"], conta["ajuda"]),
                                        "Plano de contas")
                ws.row_dimensions[faixa["cab_conta"]].height = 28

                for r in range(p["linha_1"], p["linha_n"] + 1):
                    for c_, fmt in ((cd, "0"), (cdesc, None), (cv, DINHEIRO)):
                        c = ws.cell(row=r, column=c_)
                        c.border = BORDA
                        c.font = Font(size=9)
                        if fmt:
                            c.number_format = fmt
                        if r % 2 == 0:
                            c.fill = PatternFill("solid", fgColor=CINZA_CLARO)
                    ws.cell(row=r, column=cd).font = Font(size=8, color="808080")
                dv.add("%s%d:%s%d" % (get_column_letter(cv), p["linha_1"],
                                      get_column_letter(cv), p["linha_n"]))

                rt = p["linha_total"]
                lab = ws.cell(row=rt, column=cdesc, value="Total")
                lab.font = Font(bold=True, size=8, color=cor)
                lab.alignment = Alignment(horizontal="right")
                tot = ws.cell(row=rt, column=cv, value="=SUM(%s%d:%s%d)" % (
                    get_column_letter(cv), p["linha_1"],
                    get_column_letter(cv), p["linha_n"]))
                tot.number_format = DINHEIRO
                tot.font = Font(bold=True, size=9, color=cor)
                tot.border = Border(top=GROSSA, bottom=GROSSA, left=FINA, right=FINA)

    # Colunas estreitas: cabe mais bloco na tela sem apertar o texto, porque a
    # descricao e curta ("Goiaba", "Correios", "Aluguel do patio").
    largura = max(f["colunas"] for f in faixas)
    for c in range(1, largura + 1):
        resto = (c - 1) % 3
        ws.column_dimensions[get_column_letter(c)].width = (
            4 if resto == 0 else 18 if resto == 1 else 12)
    ws.freeze_panes = "A3"
    return ws


def aba_como_usar(wb, ano):
    ws = wb.create_sheet("Como usar")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 100

    _titulo(ws, "B2", "Fluxo de caixa %d — como preencher" % ano, 18)
    passos = [
        ("É a mesma planilha de sempre.",
         "Blocos de Dia, Tipo e Valor, um do lado do outro, com o total no pé "
         "de cada um. Uma aba por mês, como sempre foi."),
        ("A diferença: cada bloco agora é uma categoria.",
         "Antes existiam blocos como «Diversos», «Cartão» e «Fixo», que não "
         "dizem o que foi comprado. Agora o bloco é «Aluguel», «Guincho», "
         "«Embalagem» — e é isso que faz o DRE sair sozinho no fim."),
        ("Escreva embaixo do título certo.",
         "Não tem lista pra escolher nem nome pra digitar: a coluna já é a "
         "categoria. Achou o bloco, escreveu o dia, o que foi e quanto."),
        ("Cartão não é bloco.",
         "Cartão é como se paga, não o que se compra. Um guincho pago no "
         "cartão vai no bloco Guincho. A forma de pagamento pode ir no Tipo, "
         "junto com o nome de quem recebeu."),
        ("Retirada de sócio: quatro blocos, um por sócio.",
         "P1 Ricardo, P2 Odilon, P3 Caique, P4 Gabriela — no lugar da cor de "
         "fundo. É a mesma informação de sempre."),
        ("Passe o mouse no título do bloco.",
         "Aparece um lembrete do que entra ali. Ex.: em Manutenção predial, "
         "«reparo — obra nova vai em Investimento»."),
        ("", ""),
        ("Não precisa somar nada.",
         "O total de cada bloco, o resumo do mês e a aba «DRE do ano» são "
         "calculados na hora, por fórmula."),
        ("Não sabe onde lançar?",
         "Use o bloco «Despesas gerais», em Administrativas. Mas ele tem "
         "limite: se passar de 2% do mês o número fica vermelho — é sinal de "
         "que tem coisa ali que pertence a outro bloco."),
    ]
    r = 4
    for titulo, texto in passos:
        if titulo:
            ws.cell(row=r, column=2, value=titulo).font = Font(
                bold=True, size=11, color=AZUL)
            r += 1
        if texto:
            c = ws.cell(row=r, column=2, value=texto)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 32 if len(texto) > 90 else 16
            r += 1
        r += 1
    return ws


def aba_categorias(wb, familias):
    """Lista de consulta: o que entra em cada bloco."""
    ws = wb.create_sheet("O que vai em cada bloco")
    ws.sheet_view.showGridLines = False
    for i, t in enumerate(["Onde fica", "Bloco", "O que entra aqui"], start=1):
        c = ws.cell(row=1, column=i, value=t)
        c.font = Font(bold=True, color=BRANCO)
        c.fill = PatternFill("solid", fgColor=AZUL)
    ws.row_dimensions[1].height = 22
    r = 2
    for fam in familias:
        for conta in fam["contas"]:
            ws.cell(row=r, column=1, value=fam["grupo"]).font = Font(
                size=9, color=COR_DA_FAMILIA.get(fam["grupo"], AZUL))
            ws.cell(row=r, column=2, value=conta["nome"]).font = Font(bold=True, size=10)
            ws.cell(row=r, column=3, value=conta["ajuda"]).font = Font(size=10)
            if r % 2 == 0:
                for col in range(1, 4):
                    ws.cell(row=r, column=col).fill = PatternFill(
                        "solid", fgColor=CINZA_CLARO)
            r += 1
    for col, larg in (("A", 26), ("B", 32), ("C", 60)):
        ws.column_dimensions[col].width = larg
    ws.freeze_panes = "A2"
    return ws


def aba_ano(wb, ano):
    ws = wb.create_sheet("Resumo do ano")
    ws.sheet_view.showGridLines = False
    _titulo(ws, "A1", "Resumo de %d" % ano, 16)
    for i, t in enumerate(["Mês", "Créditos", "Débitos", "Saldo"], start=1):
        c = ws.cell(row=3, column=i, value=t)
        c.font = Font(bold=True, color=BRANCO)
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center")
    for i, m in enumerate(MESES, start=4):
        ws.cell(row=i, column=1, value=m).font = Font(bold=True)
        ws.cell(row=i, column=2,
                value="='%s'!%s" % (m, CEL_CREDITOS)).number_format = DINHEIRO
        ws.cell(row=i, column=3,
                value="='%s'!%s" % (m, CEL_DEBITOS)).number_format = DINHEIRO
        c = ws.cell(row=i, column=4, value="=B%d-C%d" % (i, i))
        c.number_format = DINHEIRO
        c.font = Font(bold=True)
        if i % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=i, column=col).fill = PatternFill(
                    "solid", fgColor=CINZA_CLARO)
    r = len(MESES) + 4
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True, size=12, color=AZUL)
    for col in range(2, 5):
        L = get_column_letter(col)
        c = ws.cell(row=r, column=col, value="=SUM(%s4:%s%d)" % (L, L, r - 1))
        c.number_format = DINHEIRO
        c.font = Font(bold=True, size=12, color=AZUL)
    ws.conditional_formatting.add("D4:D%d" % r, CellIsRule(
        operator="lessThan", formula=["0"], font=Font(bold=True, color=VERMELHO)))
    for col, larg in (("A", 14), ("B", 18), ("C", 18), ("D", 18)):
        ws.column_dimensions[col].width = larg
    ws.freeze_panes = "A4"
    return ws


def aba_dre(wb, familias, ano, pos):
    """DRE do ano, por formula, apontando pro subtotal de cada bloco.

    Nenhum rotulo comeca com "=". O Excel le celula que comeca com igual como
    formula: "= LUCRO BRUTO" aparecia como #NAME? na tela, com o numero certo
    do lado. As linhas de resultado se distinguem por fundo e negrito.
    """
    ws = wb.create_sheet("DRE do ano")
    ws.sheet_view.showGridLines = False
    _titulo(ws, "A1", "DRE %d — monta sozinho, não preencha nada aqui" % ano, 16)

    col_mes = {m: 3 + i for i, m in enumerate(MESES)}
    col_total = 15

    for texto, col in ([("Conta", 1)] + list(col_mes.items())
                       + [("Ano", col_total)]):
        c = ws.cell(row=3, column=col, value=texto)
        c.font = Font(bold=True, color=BRANCO)
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center")

    def ref(mes, codigo):
        p = pos[codigo]
        return "'%s'!%s%d" % (mes, get_column_letter(p["col_valor"]),
                              p["linha_total"])

    def linha_de_conta(r, conta):
        ws.cell(row=r, column=1, value=conta["nome"]).font = Font(size=10)
        for m, c in col_mes.items():
            cel = ws.cell(row=r, column=c, value="=" + ref(m, conta["codigo"]))
            cel.number_format = DINHEIRO
            cel.font = Font(size=10)
        t = ws.cell(row=r, column=col_total, value="=SUM(C%d:N%d)" % (r, r))
        t.number_format = DINHEIRO
        t.font = Font(size=10, bold=True)

    def linha_titulo(r, texto, fundo, formula_por_mes=None):
        ws.cell(row=r, column=1, value=texto).font = Font(
            bold=True, size=11, color=AZUL)
        for col in range(1, col_total + 1):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=fundo)
        if formula_por_mes:
            for m, col in col_mes.items():
                cel = ws.cell(row=r, column=col, value=formula_por_mes(col))
                cel.number_format = DINHEIRO
                cel.font = Font(bold=True, size=10)
            t = ws.cell(row=r, column=col_total, value="=SUM(C%d:N%d)" % (r, r))
            t.number_format = DINHEIRO
            t.font = Font(bold=True, size=10)

    ORDEM = [
        ("receita", "RECEITA BRUTA"),
        ("deducoes", "(−) DEDUÇÕES"),
        ("cmv", "(−) CUSTO DA PEÇA VENDIDA"),
        (None, "LUCRO BRUTO"),
        ("despesas", "(−) DESPESAS OPERACIONAIS"),
        (None, "RESULTADO OPERACIONAL"),
        ("investimento", "(−) INVESTIMENTO E OBRA"),
        ("socios", "(−) DISTRIBUIÇÃO DE LUCRO"),
        ("nao_resultado", "(−) SÓ PASSOU PELO CAIXA"),
        (None, "SOBRA DE CAIXA"),
    ]
    por_dre = {}
    for fam in familias:
        for conta in fam["contas"]:
            por_dre.setdefault(conta["dre"], []).append(conta)

    r, marcos = 4, {}
    for chave, rotulo in ORDEM:
        if chave is None:
            marcos[rotulo] = r
            linha_titulo(r, rotulo, "CFE0F5")
            r += 1
            continue
        contas = por_dre.get(chave, [])
        if not contas:
            continue
        inicio, fim = r + 1, r + len(contas)
        linha_titulo(r, rotulo, "E8EEF7",
                     formula_por_mes=lambda col, i=inicio, f=fim:
                     "=SUM(%s%d:%s%d)" % (get_column_letter(col), i,
                                          get_column_letter(col), f))
        marcos[rotulo] = r
        r += 1
        for conta in contas:
            linha_de_conta(r, conta)
            r += 1

    def preencher(rotulo, expressao):
        rr = marcos[rotulo]
        for m, col in col_mes.items():
            c = ws.cell(row=rr, column=col,
                        value=expressao(get_column_letter(col)))
            c.number_format = DINHEIRO
            c.font = Font(bold=True, size=10, color=AZUL)
        t = ws.cell(row=rr, column=col_total, value="=SUM(C%d:N%d)" % (rr, rr))
        t.number_format = DINHEIRO
        t.font = Font(bold=True, size=11, color=AZUL)

    preencher("LUCRO BRUTO", lambda L: "={0}{1}-{0}{2}-{0}{3}".format(
        L, marcos["RECEITA BRUTA"], marcos["(−) DEDUÇÕES"],
        marcos["(−) CUSTO DA PEÇA VENDIDA"]))
    preencher("RESULTADO OPERACIONAL", lambda L: "={0}{1}-{0}{2}".format(
        L, marcos["LUCRO BRUTO"], marcos["(−) DESPESAS OPERACIONAIS"]))

    def sobra(L):
        expr = "={0}{1}".format(L, marcos["RESULTADO OPERACIONAL"])
        for k in ("(−) INVESTIMENTO E OBRA", "(−) DISTRIBUIÇÃO DE LUCRO",
                  "(−) SÓ PASSOU PELO CAIXA"):
            if k in marcos:
                expr += "-%s%d" % (L, marcos[k])
        return expr
    preencher("SOBRA DE CAIXA", sobra)

    ws.column_dimensions["A"].width = 34
    for col in range(3, col_total + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = "C4"
    return ws


def gerar(saida: str, ano: int) -> str:
    familias = _familias()
    pos, faixas, _ultima = layout_dos_blocos(familias)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    aba_como_usar(wb, ano)
    for m in MESES:
        aba_mes(wb, m, familias, ano, pos, faixas)
    aba_ano(wb, ano)
    aba_dre(wb, familias, ano, pos)
    aba_categorias(wb, familias)

    Path(saida).parent.mkdir(parents=True, exist_ok=True)
    wb.save(saida)
    return saida


def main() -> int:
    arg = sys.argv
    if "--saida" not in arg:
        print("faltou --saida com o caminho do .xlsx")
        return 2
    saida = arg[arg.index("--saida") + 1]
    ano = int(arg[arg.index("--ano") + 1]) if "--ano" in arg else 2026
    caminho = gerar(saida, ano)
    familias = _familias()
    pos, faixas, ultima = layout_dos_blocos(familias)
    print("gerado: %s" % caminho)
    print("  %d blocos em %d faixas, %d linhas por bloco"
          % (len(pos), len(faixas), LINHAS_POR_BLOCO))
    print("  aba do mes: ate a linha %d, %d colunas de largura"
          % (ultima, max(f["colunas"] for f in faixas)))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
