# -*- coding: utf-8 -*-
"""Quebra as saidas do mes por rubrica — a base do DRE.

Como funciona, e por que assim: a formula de "Debitos" da planilha lista, uma a
uma, as celulas de subtotal que o autor considera despesa do mes. Resolver
essas referencias da a quebra EXATA e ela fecha no total por construcao. Em
agosto/2026 sao 15 componentes que somam R$ 674.635, o mesmo Debitos.

Tentei antes varrer os blocos por cabecalho e nao funciona: eles se empilham na
mesma coluna, a coluna do valor ora e a vizinha ora a seguinte, e o mesmo texto
e titulo num lugar e item em outro. "Cartao" e item no bloco de Diversos (B5,
com valor ao lado) e "Cartoes" e titulo de bloco (N8, com a celula da direita
vazia) — e essa diferenca, celula-da-direita-vazia, e o que separa titulo de
item aqui.

Classificacao de cada rubrica no DRE (confirmada pelo gestor em 01/09/2026):
    CMV        Sucatas (compra do veiculo) + custos de sucata (guincho,
               oficina) — sem eles a peca nao existe pra vender
    Deducoes   Impostos, Devolucoes
    Despesa    Colaboradores, Fixo, Cartoes, Diversos, Transportes, Marketing,
               Embalagem, Nevada Ecopecas
    NAO e despesa:
      Investimentos e Patrimonial  -> viram patrimonio
      Parcelas e socios (P1 Ricardo, P2 Odilon, P3 Caique, P4 Gabriela)
                                   -> distribuicao de lucro
"""
import openpyxl
import re
import unicodedata

P = r'C:\Users\José Caique\Downloads\Fluxo Julho 26.xlsx'


def ach(t):
    t = "".join(c for c in unicodedata.normalize("NFKD", str(t or '').strip().lower())
                if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', t)


# Vocabulario fechado das rubricas. Casar contra ele em vez de pegar "o texto
# mais proximo acima" — o mais proximo costuma ser o ULTIMO ITEM do bloco
# ("Vinicius", "Erp", "P4 ML"), nao o titulo dele.
VOCAB = {
    'impostos': 'Impostos', 'cartoes': 'Cartoes', 'cartao': 'Cartoes',
    'marketing': 'Marketing', 'embalagem': 'Embalagem',
    'sucatas': 'Sucatas', 'sucata': 'Sucatas', 'transportes': 'Transportes',
    'colaboradores': 'Colaboradores', 'imoveis': 'Imoveis',
    'investimentos': 'Investimentos', 'patrimonial': 'Patrimonial',
    'fixo': 'Fixo', 'nevada ecopecas': 'Nevada Ecopecas',
    'devolucao': 'Devolucoes', 'devolucoes': 'Devolucoes',
    'ecomerce': 'Receita e-commerce', 'ecommerce': 'Receita e-commerce',
    'vb': 'Receita balcao', 'v b': 'Receita balcao',
}
# Cabecalho de bloco tem a celula da direita VAZIA; item tem valor nela. E o
# que separa o titulo "Cartoes" (N8, com O8 vazia) do item "Cartao" (B5, com
# C5 = 8.692). Sem isso o bloco de Diversos era rotulado de Cartoes.
CABECALHO_DIREITA_VAZIA = True
# Blocos sem titulo escrito, identificados pela coluna onde o subtotal cai.
# Sao os dois primeiros blocos da planilha, que o autor nunca nomeou porque
# "todo mundo sabe": o de gastos avulsos e o de parcelas/socios.
SEM_TITULO = {'C': 'Diversos', 'I': 'Parcelas e socios', 'L': 'Nevada Ecopecas'}


def rotulo_de(ws, cel):
    """Nome da rubrica daquele subtotal.

    Sobe pela coluna procurando um texto que esteja no vocabulario. Se nao
    achar, cai na tabela por coluna — e so entao desiste, dizendo que desistiu.
    """
    col_letra = ''.join(c for c in cel if c.isalpha())
    col = openpyxl.utils.column_index_from_string(col_letra)
    lin = int(''.join(c for c in cel if c.isdigit()))
    for r in range(lin, max(0, lin - 45), -1):
        for c in (col, col - 1):
            if c < 1:
                continue
            v = ws.cell(row=r, column=c).value
            if not (isinstance(v, str) and ach(v) in VOCAB):
                continue
            direita = ws.cell(row=r, column=c + 1).value
            if isinstance(direita, (int, float)) and direita:
                continue          # e item da lista, nao titulo do bloco
            return VOCAB[ach(v)]
    return SEM_TITULO.get(col_letra, '(sem rotulo)')


# Nomes de vendedor no bloco de devolucoes sao COMISSAO paga, nao devolucao.
# Ficam na mesma coluna porque e onde couberam, nao porque sao a mesma coisa.
VENDEDORES = {"matheus", "brenda", "flavia", "gustavo", "lucas", "vinicius",
              "caique", "nycollas", "rafael"}

# Lancamentos que o rotulo nao revela e que o gestor identificou um a um.
# "M. Ram" aparece duas vezes em agosto: R31 e o motor que voltou (devolucao de
# verdade) e R32 e um cheque caucao que entrou e saiu da empresa. Caucao nao e
# despesa nem receita — nao afeta resultado, so passa pelo caixa. Sem esta
# linha os dois seriam lidos como devolucao, e o mes fecharia R$ 5.000 pior do
# que foi.
OVERRIDES = {
    ("Ago", "R32"): ("Caucao e transito", "Cheque caucao — entrou e saiu"),
}

# Blocos que misturam naturezas diferentes e por isso sao abertos item a item
# em vez de entrarem pelo subtotal.
MISTURADOS = {"Devolucoes"}


def detalhar_bloco(ws, aba, cel_subtotal, rubrica):
    """Abre um bloco misturado item a item, classificando cada linha.

    O bloco de Devolucoes de agosto tem tres naturezas na mesma coluna:
    devolucao de venda, comissao de vendedor e um cheque caucao. Somar tudo
    como "devolucao" jogava R$ 22.832 de comissao pra fora das despesas de
    pessoal e R$ 5.000 de caucao pra dentro do resultado.
    """
    col = openpyxl.utils.column_index_from_string(
        ''.join(c for c in cel_subtotal if c.isalpha()))
    lin = int(''.join(c for c in cel_subtotal if c.isdigit()))
    alvo_val = ws[cel_subtotal].value or 0
    itens, soma = [], 0.0
    for r in range(lin - 1, max(0, lin - 40), -1):
        # Para assim que os itens reconstroem o subtotal. Subir por contagem de
        # linhas passava do inicio do bloco e engolia o de cima — em agosto ia
        # buscar os R$ 232.451 de sucata, que nao tem nada a ver com este.
        if abs(soma - alvo_val) < 1 and itens:
            break
        val = ws.cell(row=r, column=col).value
        rot = ws.cell(row=r, column=col - 1).value
        if not isinstance(val, (int, float)) or not val:
            continue
        cel = ws.cell(row=r, column=col).coordinate
        nome = ach(rot)
        chave = OVERRIDES.get((aba, cel))
        if chave:
            destino, rotulo = chave
        elif nome in VENDEDORES:
            destino, rotulo = "Comissoes", f"Comissao {str(rot).strip()}"
        else:
            destino, rotulo = rubrica, str(rot or "").strip() or "(sem rotulo)"
        itens.append({"celula": cel, "valor": round(float(val), 2),
                      "rubrica": destino, "detalhe": rotulo})
        soma += val
    # So vale se os itens reconstruirem o subtotal. Se nao fecharem, e porque
    # o bloco nao e o que eu penso que e — melhor cair fora e usar o subtotal.
    if abs(soma - alvo_val) > 1:
        return None
    return itens


def quebra_do_mes(caminho, aba):
    wv = openpyxl.load_workbook(caminho, data_only=True)
    wf = openpyxl.load_workbook(caminho, data_only=False)
    ws, wsf = wv[aba], wf[aba]
    formula = None
    for linha in ws.iter_rows():
        for c in linha:
            if ach(c.value) == 'debitos':
                for k in (1, 2):
                    alvo = wsf.cell(row=c.row, column=c.column + k)
                    if isinstance(alvo.value, str) and alvo.value.startswith('='):
                        formula = alvo.value
                        break
        if formula:
            break
    if not formula:
        return None, []
    refs = re.findall(r'[A-Z]{1,2}[0-9]{1,4}', formula)
    partes = []
    for ref in refs:
        v = ws[ref].value
        if not isinstance(v, (int, float)) or not v:
            continue
        rub = rotulo_de(ws, ref)
        if rub in MISTURADOS:
            det = detalhar_bloco(ws, aba, ref, rub)
            if det:
                partes.extend(det)
                continue
        partes.append({'celula': ref, 'valor': round(float(v), 2),
                       'rubrica': rub})
    return formula, partes


