"""
Leva a planilha "Carros para chegar" pro portal, na área do gestor.

O que a planilha responde e nenhum outro lugar responde: quanto dinheiro está
parado em carro já comprado que ainda não chegou no pátio, e há quanto tempo.

A coluna SITUAÇÃO é texto livre — "Pago", "Agendado Mafia", "Agendado Mafia
20/08", "Chegou 19/01", "No pátio do Rick". Não dá pra contar por valor
distinto, então aqui ela é classificada em quatro estados; o texto original vai
junto e aparece na tela, pra ninguém perder informação na normalização.

Uso:
    set DATABASE_URL=postgresql://...
    python scripts/sincronizar_carros.py
    python scripts/sincronizar_carros.py --seco     # só confere
"""

import io
import os
import re
import sys
import unicodedata
from datetime import datetime, date
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "app"))
import nevada_comum as C  # biblioteca comum — ver app/nevada_comum.py
from caminhos import caminho  # config/caminhos.json — ver app/caminhos.py

from openpyxl import load_workbook

PLANILHA = caminho("carros_planilha")

# Nome da coluna na planilha (sem acento, minúsculo) -> nome do campo aqui.
# Casado por nome e não por posição: coluna nova no meio não quebra a leitura.
COLUNAS = {
    "data": "data",
    "veiculo": "veiculo",
    "leilao": "leilao",
    "valor": "valor",
    "n leilao": "num_leilao",
    "lote": "lote",
    "situacao": "situacao",
    "endereco": "endereco",
    "placa": "placa",
    "chave": "chave",
    "data do agendamento": "agendamento",
    "data de chegada": "chegada",
    "guinchos resposaveis": "guincho",
    "guinchos responsaveis": "guincho",
    "contato": "contato",
}


def normalizar(t):
    t = unicodedata.normalize("NFKD", str(t or "").strip().lower())
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9 ]", "", t).strip()


def iso(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def classificar(situacao, chegada):
    """Quatro estados, na ordem em que um vence o outro: data de chegada
    preenchida é a prova mais forte de que o carro chegou."""
    s = normalizar(situacao)
    if chegada or "chegou" in s or "no patio" in s:
        return "chegou"
    if "agendad" in s:
        return "agendado"
    if s:
        return "comprado"
    return "sem_situacao"


def ler():
    wb = load_workbook(PLANILHA, data_only=True)
    ws = wb["SUCATA"]
    cab = [normalizar(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    indices = {}
    for i, nome in enumerate(cab):
        if nome in COLUNAS:
            indices[COLUNAS[nome]] = i

    faltando = {"data", "veiculo", "leilao", "situacao"} - set(indices)
    if faltando:
        raise SystemExit(f"colunas não encontradas na planilha: {faltando}")

    hoje = date.today()
    carros = []
    for linha in ws.iter_rows(min_row=2, values_only=True):
        def campo(nome):
            i = indices.get(nome)
            return linha[i] if i is not None and i < len(linha) else None

        veiculo = str(campo("veiculo") or "").strip()
        if not veiculo:
            continue                       # linha vazia da planilha

        compra = iso(campo("data"))
        chegada = iso(campo("chegada"))
        valor = campo("valor")
        valor = float(valor) if isinstance(valor, (int, float)) else None

        # Dias parados só faz sentido pra quem não chegou; pra quem chegou o que
        # importa é quanto tempo levou.
        dias_parado = dias_ate_chegar = None
        if compra:
            d0 = date.fromisoformat(compra)
            if chegada:
                dias_ate_chegar = (date.fromisoformat(chegada) - d0).days
            else:
                dias_parado = (hoje - d0).days

        carros.append({
            "data": compra,
            "veiculo": veiculo,
            "leilao": str(campo("leilao") or "").strip(),
            "valor": valor,
            "lote": str(campo("lote") or "").strip().replace(".0", ""),
            "num_leilao": str(campo("num_leilao") or "").strip().replace(".0", ""),
            "situacao": str(campo("situacao") or "").strip(),
            "estado": classificar(campo("situacao"), chegada),
            "endereco": str(campo("endereco") or "").strip(),
            "placa": str(campo("placa") or "").strip(),
            "agendamento": iso(campo("agendamento")),
            "chegada": chegada,
            "guincho": str(campo("guincho") or "").strip(),
            "contato": str(campo("contato") or "").strip(),
            "dias_parado": dias_parado,
            "dias_ate_chegar": dias_ate_chegar,
        })
    return carros


def resumir(carros):
    from collections import Counter
    print(f"{len(carros)} carros")
    for est, n in Counter(c["estado"] for c in carros).most_common():
        parados = [c for c in carros if c["estado"] == est]
        valor = sum(c["valor"] or 0 for c in parados)
        print(f"  {est:14} {n:>4}  R$ {valor:>12,.2f}")
    pendentes = [c for c in carros if c["estado"] != "chegou" and c["dias_parado"] is not None]
    if pendentes:
        pendentes.sort(key=lambda c: -c["dias_parado"])
        print(f"\n  parado ha mais tempo: {pendentes[0]['veiculo']} "
              f"({pendentes[0]['dias_parado']} dias, {pendentes[0]['leilao']})")
        print(f"  valor parado total  : R$ {sum(c['valor'] or 0 for c in pendentes):,.2f}")
    chegaram = [c["dias_ate_chegar"] for c in carros if c["dias_ate_chegar"] is not None]
    if chegaram:
        chegaram.sort()
        print(f"  tempo ate chegar    : mediana {chegaram[len(chegaram)//2]} dias "
              f"({len(chegaram)} com data de chegada)")


def gravar(carros, url):
    import psycopg2
    from psycopg2.extras import Json

    corpo = {"gerado_em": datetime.now().astimezone().isoformat(timespec="seconds"),
             "carros": carros}
    conn = psycopg2.connect(url)
    with conn, conn.cursor() as cur:
        cur.execute("CREATE TABLE IF NOT EXISTS dados_json ("
                    "chave TEXT PRIMARY KEY, valor JSONB NOT NULL)")
        cur.execute("INSERT INTO dados_json (chave, valor) VALUES (%s, %s) "
                    "ON CONFLICT (chave) DO UPDATE SET valor = EXCLUDED.valor",
                    ("carros_chegar", Json(corpo)))
    conn.close()
    print("\n  gravado carros_chegar")


def main():
    if not PLANILHA.exists():
        raise SystemExit(f"planilha não encontrada: {PLANILHA}")
    carros = ler()
    resumir(carros)
    if "--seco" in sys.argv:
        print("\n(--seco: nada foi gravado)")
        return
    url = os.environ.get("DATABASE_URL")
    if not url:
        raise SystemExit("\nDATABASE_URL não definida.")
    gravar(carros, url)


if __name__ == "__main__":
    C.saida_utf8()
    main()
