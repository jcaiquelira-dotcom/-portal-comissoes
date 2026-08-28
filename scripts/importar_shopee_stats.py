"""
Leva o relatorio "shop-stats" da Shopee pro portal (chave shopee_conta).

A planilha que o Seller Centre exporta e MENSAL (uma linha por mes) e usa a
aba "Produto Pago" — vendas efetivamente pagas, o mesmo criterio do Mercado
Livre no painel (pagamento aprovado, nao pedido feito). Por isso a serie aqui
e serie_mes, nao serie_dia: o resumo do gestor soma mes cheio exato e rateia
por dia quando o filtro corta um mes no meio, igual ja faz com Meta e agencia.

Quando a API oficial da Shopee entrar (Partner ID + Key), o sincronizador dela
substitui este importador gravando serie_dia na mesma chave — o painel nem
percebe a troca.

Uso:
    set DATABASE_URL=postgresql://...
    python scripts/importar_shopee_stats.py "caminho\do\arquivo.xlsx"
"""

import io
import os
import re
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

FUSO = timezone(timedelta(hours=-3))


def num_br(v) -> float:
    t = str(v or "").strip().replace(".", "").replace(",", ".")
    try:
        return float(t)
    except ValueError:
        return 0.0


def ler(caminho: Path) -> dict:
    wb = load_workbook(caminho, data_only=True)
    ws = wb["Produto Pago"]
    serie = {}
    for linha in ws.iter_rows(min_row=5, values_only=True):
        bruto = str(linha[0] or "").strip()
        # linha mensal: "01/07/2026". A linha-resumo do topo tem intervalo
        # ("01/01/2026-31/07/2026") e nao casa com o padrao.
        m = re.fullmatch(r"01/(\d{2})/(\d{4})", bruto)
        if not m:
            continue
        mes = f"{m.group(2)}-{m.group(1)}"
        serie[mes] = {"total": round(num_br(linha[1]), 2), "qtd": int(num_br(linha[3]))}
    if not serie:
        raise SystemExit("nenhuma linha mensal encontrada na aba 'Produto Pago'.")
    return serie


def main():
    if len(sys.argv) < 2:
        raise SystemExit("uso: python importar_shopee_stats.py <arquivo.xlsx>")
    caminho = Path(sys.argv[1])
    if not caminho.exists():
        raise SystemExit(f"arquivo não encontrado: {caminho}")

    serie = ler(caminho)
    total = sum(m["total"] for m in serie.values())
    qtd = sum(m["qtd"] for m in serie.values())
    print(f"{len(serie)} meses ({min(serie)} a {max(serie)}) | "
          f"R$ {total:,.2f} em {qtd} pedidos pagos")
    for mes in sorted(serie):
        m = serie[mes]
        print(f"  {mes}: R$ {m['total']:>10,.2f} | {m['qtd']:>3} pedidos")

    url = os.environ.get("DATABASE_URL")
    if not url:
        raise SystemExit("DATABASE_URL não definida.")
    import psycopg2
    from psycopg2.extras import Json
    conn = psycopg2.connect(url)
    with conn, conn.cursor() as cur:
        # merge por mes: reimportar uma planilha mais nova atualiza os meses
        # dela e preserva os que so existiam na anterior.
        cur.execute("SELECT valor FROM dados_json WHERE chave='shopee_conta' FOR UPDATE")
        linha = cur.fetchone()
        antiga = ((linha[0].get("vendas") or {}).get("serie_mes") or {}) if linha else {}
        pacote = {
            "gerado_em": datetime.now(FUSO).isoformat(timespec="seconds"),
            "fonte": caminho.name,
            "vendas": {"serie_mes": {**antiga, **serie}},
        }
        cur.execute("INSERT INTO dados_json (chave, valor) VALUES (%s, %s) "
                    "ON CONFLICT (chave) DO UPDATE SET valor = EXCLUDED.valor",
                    ("shopee_conta", Json(pacote)))
    conn.close()
    print("\n  gravado shopee_conta")


if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    main()
