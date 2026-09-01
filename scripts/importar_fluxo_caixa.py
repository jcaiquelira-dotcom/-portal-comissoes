# -*- coding: utf-8 -*-
"""Le a planilha de fluxo de caixa da Nevada e monta o dado do portal.

A planilha ("Fluxo Julho 26.xlsx") tem 9 anos de historia e e a unica fonte
desses numeros. Ela funciona, mas guarda a estrutura na cabeca de quem montou:
as rubricas existem — Fixo, Impostos, Sucatas, Colaboradores, Marketing,
Embalagem, Ecommerce — so que espalhadas em blocos de tres colunas que mudam de
lugar a cada mes, com subtotais soltos e sem nenhuma soma que amarre tudo.

O que este script NAO faz: adivinhar. Ele extrai o que a planilha AFIRMA e
confere contra o que ela mesma calcula. Onde as duas contas discordam, ele
denuncia em vez de escolher uma — numero de fluxo de caixa que ninguem
consegue refazer nao vale nada.

Duas leituras que sustentam o resto, verificadas mes a mes:
    Saidas do mes = Debitos + Veiculos
    Saldo do mes  = Entradas - Saidas
Batem exato em Jan, Fev, Mar, Maio, Jun e Jul de 2026. Abr e Ago divergem, e a
divergencia vai gravada junto com o numero, nao escondida.

A aba "Geral" tem Entradas/Saidas/Saldo de 2018 a 2026, mas as Saidas de
2026 pararam de ser preenchidas em marco — o que faz o "saldo" de la mostrar a
entrada inteira, como se o mes nao tivesse tido despesa nenhuma. Aqui as abas
mensais completam esse buraco.

Uso:
    set DATABASE_URL=postgresql://...
    python scripts/importar_fluxo_caixa.py --arquivo "C:\\...\\Fluxo Julho 26.xlsx" --seco
    python scripts/importar_fluxo_caixa.py --arquivo "..." --gravar
"""

import json
import re
import sys
import unicodedata
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "app"))

MESES = ["Jan", "Fev", "Mar", "Abr", "Maio", "Jun", "Jul", "Ago",
         "Set", "Out", "Nov", "Dez"]
NUM_MES = {n: i + 1 for i, n in enumerate(MESES)}
# A aba Geral escreve "Mai" e, em alguns anos, "Lun" no lugar de "Jun" (erro de
# digitacao que se repetiu por copia). Reconhecer os dois evita perder o mes.
APELIDOS = {"mai": 5, "lun": 6, "jun": 6, "mar": 3, "abr": 4, "jan": 1,
            "fev": 2, "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11,
            "dez": 12}


def ach(t) -> str:
    t = "".join(c for c in unicodedata.normalize("NFKD", str(t or "").strip().lower())
                if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", t)


def num(v):
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def ler_geral(wb) -> dict:
    """Entradas/Saidas/Saldo por ano-mes, da aba Geral.

    A aba tem varios quadros de 4 colunas espalhados, um por ano, com o ano
    escrito numa celula solta acima. Percorre procurando esse padrao em vez de
    depender de posicao fixa — a planilha ganhou anos novos em lugares
    diferentes conforme o tempo passou.
    """
    ws = wb["Geral"]
    g = [list(r) for r in ws.iter_rows(values_only=True)]
    fora = {}
    for ri, linha in enumerate(g):
        for ci, cel in enumerate(linha):
            if ach(cel) != "entradas":
                continue
            # o ano fica na linha de cima, em alguma coluna proxima
            ano = None
            for k in range(max(0, ci - 2), min(ci + 4, len(g[ri - 1]) if ri else 0)):
                v = num(g[ri - 1][k]) if ri else None
                if v and 2000 < v < 2100:
                    ano = int(v)
            if not ano:
                continue
            for rr in g[ri + 1: ri + 14]:
                if ci - 1 < 0 or ci - 1 >= len(rr):
                    continue
                m = APELIDOS.get(ach(rr[ci - 1]))
                if not m:
                    continue
                ent = num(rr[ci]) if ci < len(rr) else None
                sai = num(rr[ci + 1]) if ci + 1 < len(rr) else None
                if ent is None and sai is None:
                    continue
                fora[f"{ano:04d}-{m:02d}"] = {
                    "entradas": round(ent or 0, 2),
                    "saidas": round(sai or 0, 2),
                    "fonte": "Geral",
                }
    return fora


# A planilha escreve o mesmo total de jeitos diferentes conforme o mes:
# "Credito" em agosto, "Creditos" em abril, "D.Total" no lugar de "Debito T.".
# Aceitar so uma grafia fazia o mes cair no numero da aba Geral, que e digitado
# na mao — foi assim que abril ficou R$ 11.623 acima do que ele mesmo soma.
ROTULOS = ("debitos", "veiculos", "saldo", "credito", "creditos",
           "debito t.", "d.total", "d. total")


def ler_mes(wb, aba: str, wbf=None) -> dict:
    """Os totais que a propria aba do mes calcula: Debitos, Veiculos, Saldo.

    Le tambem a FORMULA de cada total (por isso o wbf, o mesmo arquivo aberto
    sem data_only). E como se descobre dupla contagem: em agosto/2026 a formula
    de Debitos somava R12 e U12 — exatamente as duas celulas que formam
    Veiculos — e o Debito Total somava os dois de novo. R$ 251.541 entrando
    duas vezes contra o mes.
    """
    if aba not in wb.sheetnames:
        return {}
    achados, formulas = {}, {}
    ws, wsf = wb[aba], (wbf[aba] if wbf else None)
    for linha in ws.iter_rows():
        for ci, cel in enumerate(linha):
            chave = ach(cel.value)
            if chave not in ROTULOS:
                continue
            for k in range(1, 3):
                alvo = ws.cell(row=cel.row, column=cel.column + k)
                v = num(alvo.value)
                if v:
                    achados.setdefault(chave, round(v, 2))
                    if wsf and chave not in formulas:
                        f = wsf.cell(row=alvo.row, column=alvo.column).value
                        if isinstance(f, str) and f.startswith("="):
                            formulas[chave] = f
                    break
    achados["_formulas"] = formulas
    return achados


def dupla_contagem(wb, aba: str, formulas: dict) -> tuple:
    """Quanto de Veiculos ja esta somado dentro de Debitos.

    Compara as celulas citadas nas duas formulas. O que aparece nas duas foi
    contado duas vezes, e o valor dessa sobreposicao volta como desconto — com
    a lista de celulas, pra conferencia na planilha.
    """
    fd, fv = formulas.get("debitos"), formulas.get("veiculos")
    if not fd or not fv:
        return 0.0, []
    refs = lambda f: set(re.findall(r"[A-Z]{1,2}[0-9]{1,4}", f))
    comum = sorted(refs(fd) & refs(fv))
    if not comum:
        return 0.0, []
    ws = wb[aba]
    total = sum(num(ws[c].value) or 0 for c in comum)
    return round(total, 2), comum


# Onde cada rubrica entra no DRE. E aqui que a planilha vira demonstrativo: as
# tres ultimas linhas NAO sao despesa, e hoje somam junto com elas — e por isso
# que o mes parece pior do que foi.
DRE = {
    "Sucatas":          ("cmv",          "Compra e preparo de sucata"),
    "Impostos":         ("deducoes",     "Impostos"),
    "Devolucoes":       ("deducoes",     "Devolucoes e comissoes"),
    "Colaboradores":    ("despesas",     "Pessoal"),
    "Fixo":             ("despesas",     "Fixas"),
    "Cartoes":          ("despesas",     "Cartoes"),
    "Diversos":         ("despesas",     "Diversos"),
    "Transportes":      ("despesas",     "Fretes e transportes"),
    "Marketing":        ("despesas",     "Marketing"),
    "Embalagem":        ("despesas",     "Embalagem"),
    "Nevada Ecopecas":  ("despesas",     "Obrigacoes Nevada"),
    "Investimentos":    ("investimento", "Investimentos"),
    "Patrimonial":      ("investimento", "Patrimonial"),
    "Imoveis":          ("investimento", "Imoveis"),
    "Parcelas e socios": ("socios",      "Parcelas e socios"),
}


def montar_dre(partes: list) -> dict:
    """Agrupa as parcelas do mes nas linhas do DRE.

    Parcela sem rubrica reconhecida vai pra "nao classificado" e aparece na
    tela. Somar o desconhecido dentro de "despesas" faria o numero parecer
    completo quando nao e.
    """
    grupos = {}
    for p in partes:
        grupo, rotulo = DRE.get(p["rubrica"], ("nao_classificado", p["rubrica"]))
        d = grupos.setdefault(grupo, {})
        item = d.setdefault(rotulo, {"valor": 0.0, "celulas": []})
        item["valor"] = round(item["valor"] + p["valor"], 2)
        item["celulas"].append(p["celula"])
    return grupos


def main() -> int:
    arg = sys.argv
    if "--arquivo" not in arg:
        sys.exit("faltou --arquivo com o caminho do .xlsx")
    caminho = arg[arg.index("--arquivo") + 1]
    ano = int(arg[arg.index("--ano") + 1]) if "--ano" in arg else 2026
    gravar = "--gravar" in arg

    import openpyxl
    wb = openpyxl.load_workbook(caminho, data_only=True)

    wbf = openpyxl.load_workbook(caminho, data_only=False)   # so pelas formulas
    import server
    mes_corrente = server.hoje_br().isoformat()[:7]
    fluxo = ler_geral(wb)
    avisos = []

    # As abas mensais mandam no ano corrente: elas tem o detalhe, a Geral so o
    # resumo — e o resumo parou de ser preenchido em marco.
    for aba in MESES:
        d = ler_mes(wb, aba, wbf)
        if not d:
            continue
        chave = f"{ano:04d}-{NUM_MES[aba]:02d}"
        repetido, celulas = dupla_contagem(wb, aba, d.get("_formulas") or {})
        saidas = round(d.get("debitos", 0) + d.get("veiculos", 0) - repetido, 2)
        if not saidas:
            continue
        atual = fluxo.get(chave, {})
        # Aba de mes que ainda nao chegou carrega o dado do ANO PASSADO — a
        # planilha e reaproveitada e so sobrescrita quando o mes acontece.
        # Setembro a dezembro de 2026 traziam, byte a byte, as saidas de 2025.
        # Sem esta trava o portal mostraria quatro meses de prejuizo inventado.
        # O sinal de que o mes existe e a Geral ter entrada dele.
        # Mes que ainda nao fechou nao entra. A planilha e reaproveitada de um
        # ano pro outro, e a aba de um mes futuro ainda carrega os numeros do
        # ano passado — setembro a dezembro de 2026 traziam 2025. Testar "tem
        # entrada?" nao bastava, porque o resto do ano passado TEM entrada.
        # Mes fechado e um fato de calendario, e nao depende de heuristica.
        if chave >= mes_corrente:
            avisos.append(f"{chave}: aba {aba} ignorada — o mes ainda nao fechou "
                          f"(o que esta la e resto do ano anterior)")
            continue
        cred_mes = d.get("credito") or d.get("creditos")
        if not cred_mes and not atual.get("entradas"):
            avisos.append(f"{chave}: aba {aba} ignorada — sem entrada registrada")
            continue
        entradas = cred_mes or atual.get("entradas") or 0
        if cred_mes and atual.get("entradas") and abs(cred_mes - atual["entradas"]) > 1:
            # A aba do mes SOMA as linhas; a Geral e digitada na mao. Quando
            # discordam, vale a que da pra refazer.
            avisos.append(f"{chave}: a aba do mes soma entradas de "
                          f"R$ {cred_mes:,.2f}, a aba Geral traz R$ "
                          f"{atual['entradas']:,.2f} digitado "
                          f"(diferenca R$ {atual['entradas'] - cred_mes:,.2f}). "
                          f"Usei a do mes.")
        registro = {
            "entradas": round(entradas, 2),
            "saidas": saidas,
            "veiculos": d.get("veiculos", 0),
            "debitos": d.get("debitos", 0),
            "fonte": f"aba {aba}",
        }
        if repetido:
            registro["dupla_contagem"] = repetido
            registro["celulas_repetidas"] = celulas
            avisos.append(
                f"{chave}: R$ {repetido:,.2f} contados DUAS vezes — as celulas "
                f"{', '.join(celulas)} estao na formula de Debitos e tambem "
                f"formam Veiculos. Descontei uma vez. Saidas da planilha: "
                f"R$ {d.get('debitos', 0) + d.get('veiculos', 0):,.2f}; corrigidas: "
                f"R$ {saidas:,.2f}.")
        # Conferencia contra o saldo que a aba escreve. Divergencia fica no
        # dado, visivel na tela: nao da pra corrigir o que ninguem ve.
        saldo_aba = d.get("saldo")
        if saldo_aba is not None and entradas:
            calculado = round(entradas - saidas, 2)
            if abs(calculado - saldo_aba) > 1:
                registro["divergencia"] = round(calculado - saldo_aba, 2)
                registro["saldo_planilha"] = saldo_aba
                avisos.append(f"{chave}: entradas-saidas = {calculado:,.2f} mas a "
                              f"planilha escreve saldo {saldo_aba:,.2f} "
                              f"(diferenca {calculado - saldo_aba:,.2f})")
        if atual.get("saidas") and abs(atual["saidas"] - saidas) > 1:
            avisos.append(f"{chave}: aba Geral diz saidas {atual['saidas']:,.2f}, "
                          f"a aba do mes soma {saidas:,.2f}")
        # Quebra por rubrica, pela formula da propria planilha.
        try:
            sys.path.insert(0, str(Path(__file__).resolve().parent))
            import rubricas_fluxo
            _f, partes = rubricas_fluxo.quebra_do_mes(caminho, aba)
            if partes:
                soma = round(sum(p["valor"] for p in partes), 2)
                # A compra de veiculo so esta dentro da formula de Debitos em
                # agosto — e la por engano, que e a dupla contagem. Nos outros
                # meses ela e uma linha a parte, entao entra aqui como CMV.
                # Sem isto julho aparecia sem custo de mercadoria nenhum, num
                # negocio que vive de comprar carro pra desmontar.
                falta = round(saidas - soma, 2)
                veic = d.get("veiculos", 0)
                if falta > 1 and veic and abs(falta - veic) < 2:
                    partes.append({"celula": "(Veiculos)", "valor": veic,
                                   "rubrica": "Sucatas"})
                    soma = round(soma + veic, 2)
                registro["rubricas"] = partes
                registro["dre"] = montar_dre(partes)
                # A quebra TEM que fechar nas SAIDAS do mes. Se nao fechar, o
                # DRE estaria mostrando um mes que nao existe.
                if abs(soma - saidas) > 1:
                    avisos.append(f"{chave}: a quebra por rubrica soma "
                                  f"R$ {soma:,.2f} mas as saidas sao "
                                  f"R$ {saidas:,.2f} (faltam R$ "
                                  f"{saidas - soma:,.2f}) — DRE marcado como "
                                  f"incompleto")
                    registro["dre_incompleto"] = True
        except Exception as e:
            avisos.append(f"{chave}: nao consegui quebrar por rubrica ({e})")

        fluxo[chave] = registro

    for k, v in fluxo.items():
        v["saldo"] = round(v["entradas"] - v["saidas"], 2)

    meses = sorted(fluxo)
    print(f"{len(meses)} meses lidos: {meses[0]} a {meses[-1]}\n")
    print(f"  {'mes':8} {'entradas':>13} {'saidas':>13} {'saldo':>13}  fonte")
    for k in meses[-14:]:
        v = fluxo[k]
        print(f"  {k:8} {v['entradas']:>13,.0f} {v['saidas']:>13,.0f} "
              f"{v['saldo']:>13,.0f}  {v['fonte']}")
    if avisos:
        print("\nDIVERGENCIAS (vao gravadas junto com o numero):")
        for a in avisos:
            print("   " + a)

    pacote = {"gerado_em": None, "meses": fluxo, "avisos": avisos}
    if not gravar:
        print("\n(--seco: nada gravado)")
        return 0

    pacote["gerado_em"] = server.agora_br().isoformat(timespec="seconds")
    server.escrever_json(server.resolver_pasta_dados() / "financeiro_fluxo.json", pacote)
    print(f"\ngravado financeiro_fluxo com {len(fluxo)} meses.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
