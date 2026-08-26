"""
Fila de retomada: os melhores clientes pra cada vendedor ligar de volta.

ATENCAO -- contem dados pessoais (nome e telefone). Fica fora do git de proposito.
Entregue direto ao vendedor responsavel, nao publique.

Quem ENTRA (todas as condicoes):
  1. nao comprou (sem sinal de pagamento + entrega na conversa)
  2. teve conversa de verdade (6+ respostas do vendedor) ou perguntou e ficou sem resposta
  3. o vendedor NUNCA disse que nao tinha a peca -- se nao temos, ligar nao adianta
  4. o cliente nao avisou que ja comprou ou desistiu
  5. da pra saber QUAL peca ele queria
  6. nao e alguem querendo vender carro pra gente

Depois disso, cada um recebe uma nota e so os 30 melhores de cada vendedor entram
na planilha -- a ideia e uma lista curta que da pra trabalhar de verdade num dia,
nao um cadastro de mil linhas que ninguem abre.
"""

import json
import re
import sqlite3
import time
import unicodedata
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
TOKEN = "pn_ZnbNaPSAB1ldG6AhmNhXxAjaq9F2UPmFVw4EXk6Z9s4"
BASE = "https://api.wts.chat"
INICIO = "2026-07-07T00:00:00Z"
HOJE = datetime.now(timezone.utc)  # "dias parado" conta ate hoje, nao ate o fim do sync
PAUSA = 0.4
POR_VENDEDOR = 30
SAIDA = ROOT / "Fila_Retomada_CONFIDENCIAL.xlsx"


def limpa(t):
    """Minuscula e sem acento. Tirar os combining chars e obrigatorio: sem isso
    'nao temos' nunca casa com o texto decomposto e o filtro deixa passar."""
    t = unicodedata.normalize("NFKD", (t or "").lower())
    return "".join(c for c in t if not unicodedata.combining(c))


SEM_ESTOQUE = [limpa(x) for x in [
    "não tenho", "não temos", "não vou ter", "não vamos ter", "já vendi",
    "não tem disponível", "esgotado", "infelizmente não", "não trabalho com",
    "não trabalhamos", "não possuo", "não ficou", "só tenho do", "não achei",
]]
JA_COMPROU = [limpa(x) for x in [
    "já comprei", "já consegui", "já achei", "comprei em outro", "consegui em outro",
    "achei em outro", "já resolvi", "não precisa mais", "já peguei", "desisti",
]]
QUER_VENDER = [limpa(x) for x in [
    "vcs compram carro", "vocês compram carro", "compram carro batido", "vendo meu carro",
    "quero vender meu", "comprar meu carro",
]]
RUIDO = [limpa(x) for x in [
    "tenho interesse e queria mais informa", "posso ter mais informa", "vim do instagram",
    "gostaria de ajuda com uma compra no site", "gostaria de informações sobre esse produto",
    "não há preferência", "olá nevada ecopeças",
]]
SAUDACAO = {limpa(x) for x in [
    "oi", "ola", "olá", "bom dia", "boa tarde", "boa noite", "opa", "blz", "beleza",
    "tudo bem", "sim", "não", "ok", "obrigado", "obrigada", "boa", "certo", "e ai",
]}
PECAS = [limpa(x) for x in [
    "motor", "cambio", "câmbio", "farol", "lanterna", "porta", "parachoque", "para-choque",
    "pára-choque", "paralama", "capo", "capô", "roda", "banco", "volante", "painel",
    "cabecote", "cabeçote", "turbina", "bico", "injetor", "radiador", "compressor",
    "alternador", "bomba", "modulo", "módulo", "chicote", "retrovisor", "vidro", "teto",
    "amortecedor", "suspensao", "suspensão", "escapamento", "catalisador", "diferencial",
    "embreagem", "virabrequim", "coletor", "intercooler", "multimidia", "multimídia",
    "airbag", "fechadura", "maçaneta", "coxim", "bandeja", "manga de eixo", "carter",
    "cárter", "tanque", "caixa", "sensor", "central", "peça", "kit", "coluna", "eixo",
]]
VEICULOS = [limpa(x) for x in [
    "golf", "jetta", "polo", "tiguan", "nivus", "virtus", "gol", "saveiro", "amarok",
    "audi", "a3", "a4", "q3", "q5", "bmw", "320i", "x1", "x5", "x6", "mercedes",
    "civic", "hrv", "hr-v", "fit", "corolla", "hilux", "rav4", "onix", "cruze", "tracker",
    "s10", "montana", "ka", "fiesta", "focus", "ranger", "fusion", "hb20", "creta", "ix35",
    "azera", "sportage", "cerato", "renegade", "compass", "toro", "argo", "cronos", "pulse",
    "kwid", "duster", "sandero", "logan", "captur", "peugeot", "citroen", "c3", "c4",
    "land rover", "evoque", "discovery", "freelander", "volvo", "jeep", "nissan", "kicks",
    "frontier", "versa", "march", "up", "fox", "voyage", "cobalt", "spin", "tcross",
    "t-cross", "taos", "stilo", "palio", "uno", "strada", "doblo", "gti", "gts", "gli",
]]
RE_PRECO = re.compile(r"r\$\s*\d|(?<![\d,.])\d{2,5}[,.]00\b|\bpix\b", re.IGNORECASE)


def api(path, params):
    q = "&".join(f"{k}={urllib.parse.quote(str(v))}" for k, v in params.items())
    req = urllib.request.Request(f"{BASE}{path}?{q}",
                                 headers={"Authorization": f"Bearer {TOKEN}"})
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())


def buscar_contatos():
    """Nome/telefone nao foram salvos no sync -- rebusca as sessoes so pra isso."""
    print("buscando contatos na API...")
    contatos, pagina = {}, 1
    while True:
        d = api("/chat/v2/session", {
            "CreatedAt.After": INICIO, "PageNumber": pagina, "PageSize": 100,
            "IncludeDetails": "ContactDetails", "OrderBy": "createdat",
            "OrderDirection": "ASCENDING",
        })
        for s in d.get("items") or []:
            cd = s.get("contactDetails")
            if cd:
                contatos[s["id"]] = {
                    "nome": (cd.get("name") or "").strip() or "(sem nome)",
                    "fone": cd.get("phonenumberFormatted") or "",
                    "tags": ", ".join(cd.get("tagsName") or []),
                }
        if pagina % 30 == 0:
            print(f"  pagina {pagina}/{d.get('totalPages')}")
        if not d.get("hasMorePages"):
            break
        pagina += 1
        time.sleep(PAUSA)
    print(f"  {len(contatos)} contatos obtidos")
    return contatos


def peca_procurada(msgs):
    """A frase em que o cliente diz o que quer, escolhida por conteudo e nao por
    tamanho -- senao 'chega em quantos dias?' ganha da descricao da peca."""
    melhor, melhor_score, tem_carro, tem_ano = "", 0, False, False
    for t in msgs:
        tl = limpa(t)
        if any(r in tl for r in RUIDO) or "http" in tl:
            continue
        if tl.strip().strip(".!?,") in SAUDACAO or len(t.strip()) < 8:
            continue
        limpo = " ".join(t.split())
        n_peca = sum(1 for p in PECAS if p in tl)
        n_carro = sum(1 for v in VEICULOS if re.search(rf"\b{re.escape(v)}\b", tl))
        ano = bool(re.search(r"\b(19|20)\d{2}\b", tl))
        score = 3 * n_peca + 2 * n_carro + (1 if ano else 0)
        if score == 0:
            continue
        score += min(len(limpo) / 100, 1)
        if score > melhor_score:
            melhor, melhor_score = limpo, score
            tem_carro, tem_ano = n_carro > 0, ano
    return melhor[:180], tem_carro, tem_ano


def main():
    conn = sqlite3.connect("vendas.db")
    dados = json.loads((ROOT / "dataset.json").read_text(encoding="utf-8"))
    ids = json.loads((ROOT / "session_ids.json").read_text(encoding="utf-8"))
    por_id = dict(zip(ids, dados))

    # trava de seguranca: se o pareamento id<->dataset sair de ordem, isso estoura aqui
    # em vez de gerar uma lista com o cliente errado no vendedor errado.
    for sid, created_at in conn.execute("SELECT id, created_at FROM sessoes LIMIT 200"):
        d = por_id.get(sid)
        assert d and d["d"] == created_at[:10], f"pareamento furado em {sid}"
    print("pareamento dataset<->banco: OK")

    sem_estoque, ja_comprou, quer_vender = set(), set(), set()
    msgs_cliente = defaultdict(list)
    tem_preco, foto_cliente = set(), set()
    for sid, direcao, tipo, txt, raw in conn.execute(
        "SELECT session_id, direction, type, text, raw FROM mensagens ORDER BY created_at ASC"
    ):
        tl = limpa(txt) if txt else ""
        if direcao == "TO_HUB":
            if txt and json.loads(raw).get("userId"):
                if any(k in tl for k in SEM_ESTOQUE):
                    sem_estoque.add(sid)
                if RE_PRECO.search(txt):
                    tem_preco.add(sid)
        else:
            if tipo in ("IMAGE", "VIDEO"):
                foto_cliente.add(sid)
            if txt:
                if any(k in tl for k in JA_COMPROU):
                    ja_comprou.add(sid)
                if any(k in tl for k in QUER_VENDER):
                    quer_vender.add(sid)
                if tipo == "TEXT":
                    msgs_cliente[sid].append(txt)

    contatos = buscar_contatos()
    CANAL = {"S": "Site", "AF": "Facebook Ads", "AI": "Instagram Ads",
             "I": "Instagram bio", "D": "Direto"}
    ALTA_INTENCAO = {"S", "D"}

    fila = defaultdict(list)
    descartes = defaultdict(int)
    for sid, created_at, raw in conn.execute("SELECT id, created_at, raw FROM sessoes"):
        m = por_id[sid]
        if m["cv"] == "P":
            descartes["já comprou"] += 1
            continue
        esperando = m["ab"]
        if not esperando and m["nm"] < 6:
            descartes["conversa rasa demais"] += 1
            continue
        if sid in sem_estoque:
            descartes["não tínhamos a peça"] += 1
            continue
        if sid in ja_comprou:
            descartes["cliente disse que já comprou"] += 1
            continue
        if sid in quer_vender:
            descartes["queria vender carro pra gente"] += 1
            continue
        peca, tem_carro, tem_ano = peca_procurada(msgs_cliente.get(sid, []))
        if not peca:
            descartes["não dá pra saber que peça queria"] += 1
            continue

        dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
        dias = max(0, (HOJE - dt).days)
        c = contatos.get(sid, {})
        tags = c.get("tags", "")

        # nota: o que faz valer a ligacao
        nota = 0
        nota += 30 if esperando else 0            # ninguem disse nao, so ninguem voltou
        nota += 20 if sid in tem_preco else 0     # ja teve preco na mesa
        nota += 12 if sid in foto_cliente else 0  # cliente mandou foto = interesse real
        nota += 10 if (tem_carro and tem_ano) else (5 if tem_carro else 0)
        nota += 10 if m["c"] in ALTA_INTENCAO else 0
        nota += 10 if "oportunidade" in limpa(tags) else 0
        nota += max(0, 15 - dias // 3)            # quanto mais recente, melhor
        nota += min(m["nm"] // 3, 8)              # conversa que andou de verdade

        fila[m["u"]].append({
            "nota": nota,
            "prio": "ALTA" if esperando else "MÉDIA",
            "data": dt.strftime("%d/%m/%Y"),
            "dias": dias,
            "nome": c.get("nome", "(não encontrado)"),
            "fone": c.get("fone", ""),
            "peca": peca,
            "gancho": ("Perguntou e ficou sem resposta" if esperando
                       else ("Negociou preço e não fechou" if sid in tem_preco
                             else "Conversou e não fechou")),
            "canal": CANAL.get(m["c"], m["c"]),
            "etiquetas": tags,
            "link": json.loads(raw).get("previewUrl") or "",
        })

    print("\ndescartados:")
    for k, v in sorted(descartes.items(), key=lambda x: -x[1]):
        print(f"  {v:5d}  {k}")

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Como usar"
    ws0.column_dimensions["A"].width = 108
    INSTRU = [
        ("Fila de retomada — clientes pra ligar", "titulo"),
        (f"Gerado em {HOJE.strftime('%d/%m/%Y')} · atendimentos de 07/07 a 21/08/2026", "sub"),
        ("", ""),
        ("O que é esta lista", "h"),
        ("Os 30 clientes de cada vendedor com maior chance de virar venda numa segunda tentativa. "
         "Cada aba tem o nome de um vendedor — são os atendimentos dele.", "p"),
        ("", ""),
        ("Todo mundo aqui passou por estes filtros", "h"),
        ("• Não comprou — nenhum sinal de pagamento e entrega na conversa", "p"),
        ("• Nós TÍNHAMOS a peça — quem ouviu \"não tenho\" ficou de fora, ligar não adianta", "p"),
        ("• O cliente não avisou que já comprou em outro lugar nem que desistiu", "p"),
        ("• Dá pra saber qual peça ele queria — está na coluna \"Peça que o cliente procurava\"", "p"),
        ("• A conversa andou de verdade (ou ele perguntou algo e ficou sem resposta)", "p"),
        ("", ""),
        ("Como ler a prioridade", "h"),
        ("ALTA (vermelho) — o cliente perguntou alguma coisa e ninguém respondeu. Ninguém disse não pra "
         "ele: só ficou no vácuo. É o mais fácil de recuperar, comece por aqui.", "p"),
        ("MÉDIA (amarelo) — conversou de verdade e não fechou. Muitos já tinham preço na mesa.", "p"),
        ("", ""),
        ("Dica pra abordagem", "h"),
        ("A coluna \"Peça que o cliente procurava\" traz a frase dele. Use isso na abertura — retomar "
         "citando exatamente o que a pessoa pediu funciona melhor que um \"oi, tudo bem?\" genérico.", "p"),
        ("A última coluna abre a conversa original no Totalk, pra você ver onde parou antes de ligar.", "p"),
        ("", ""),
        ("Documento confidencial — contém nome e telefone de cliente. Não encaminhe para fora da equipe.", "aviso"),
    ]
    for texto, tipo in INSTRU:
        ws0.append([texto])
        c = ws0.cell(row=ws0.max_row, column=1)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if tipo == "titulo":
            c.font = Font(bold=True, size=16)
        elif tipo == "sub":
            c.font = Font(size=10, color="7A7166")
        elif tipo == "h":
            c.font = Font(bold=True, size=11, color="B4501A")
        elif tipo == "aviso":
            c.font = Font(bold=True, size=10, color="A33232")
        else:
            c.font = Font(size=10)
            ws0.row_dimensions[ws0.max_row].height = 28

    COLS = [("#", 4), ("Prioridade", 11), ("Data", 11), ("Dias", 6), ("Cliente", 26),
            ("WhatsApp", 17), ("Peça que o cliente procurava", 60),
            ("Por que ligar", 29), ("Canal", 13), ("Conversa", 10)]
    CORES = {"ALTA": "F8D7DA", "MÉDIA": "FFF9E6"}

    for vend in ["Flávia", "Gustavo", "Matheus"]:
        todos = fila.get(vend, [])
        itens = sorted(todos, key=lambda x: -x["nota"])[:POR_VENDEDOR]
        if not itens:
            continue
        ws = wb.create_sheet(vend)
        ws.append([c[0] for c in COLS])
        for i, (_, larg) in enumerate(COLS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = larg
            cel = ws.cell(row=1, column=i)
            cel.font = Font(bold=True, color="FFFFFF", size=10)
            cel.fill = PatternFill("solid", fgColor="1A1611")
        ws.freeze_panes = "A2"
        for i, it in enumerate(itens, start=1):
            ws.append([i, it["prio"], it["data"], it["dias"], it["nome"], it["fone"],
                       it["peca"], it["gancho"], it["canal"], ""])
            r = ws.max_row
            fill = PatternFill("solid", fgColor=CORES[it["prio"]])
            for c in range(1, len(COLS) + 1):
                cel = ws.cell(row=r, column=c)
                cel.fill = fill
                cel.alignment = Alignment(vertical="top", wrap_text=(c == 7))
            ws.cell(row=r, column=2).font = Font(bold=True, size=9)
            if it["link"]:
                cel = ws.cell(row=r, column=len(COLS))
                cel.value = "abrir"
                cel.hyperlink = it["link"]
                cel.font = Font(color="0563C1", underline="single", size=9)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"
        alta = sum(1 for i in itens if i["prio"] == "ALTA")
        print(f"  {vend}: {len(itens)} de {len(todos)} elegíveis "
              f"(alta {alta}, média {len(itens)-alta}) — nota {itens[-1]['nota']} a {itens[0]['nota']}")

    wb.save(SAIDA)
    print(f"\nsalvo: {SAIDA.name} ({SAIDA.stat().st_size/1024:.0f} KB)")


if __name__ == "__main__":
    main()
