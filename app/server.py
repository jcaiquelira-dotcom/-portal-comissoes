import calendar
import hashlib
import io
import json
import os
import re
import secrets
import sqlite3
import subprocess
import sys
import urllib.error
import urllib.request
import unicodedata
import uuid
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from flask import (
    Flask, g, has_request_context, jsonify, redirect, request, send_file,
    send_from_directory, session,
)
from openpyxl import Workbook

# Servidor (Render) roda em UTC — Brasil não tem horário de verão desde 2019,
# então um offset fixo de -3h é sempre correto, sem depender de tzdata/IANA.
FUSO_BRASILIA = timezone(timedelta(hours=-3))


def agora_br() -> datetime:
    return datetime.now(FUSO_BRASILIA)


def hoje_br() -> date:
    return agora_br().date()


def parse_dt_tolerante(valor: str) -> datetime:
    """Converte uma string ISO em datetime timezone-aware. Registros gravados
    antes da correção de fuso ficaram sem offset, no horário UTC do servidor
    (Render) — nesse caso assumimos UTC. Registros novos já vêm com -03:00."""
    dt = datetime.fromisoformat(valor)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt

ROOT = Path(__file__).resolve().parent.parent
CONFIG_PATH = ROOT / "config.json"
STATIC_DIR = Path(__file__).resolve().parent / "static"

# Credenciais, senhas e vendas ficam sempre locais, na pasta segredos/dados.
SEGREDOS_DIR = ROOT / "segredos"
VENDEDORES_FILE = SEGREDOS_DIR / "vendedores.json"
CREDENCIAIS_FILE = SEGREDOS_DIR / "credenciais.json"
SECRET_KEY_FILE = SEGREDOS_DIR / "secret_key.txt"
MESES_FECHADOS_FILE = SEGREDOS_DIR / "meses_fechados.json"
LOG_ACESSOS_FILE = SEGREDOS_DIR / "log_acessos.json"
LOG_ACOES_FILE = SEGREDOS_DIR / "log_acoes.json"

DATA_DIR_NAME = "data"
DIAS_MAXIMOS_RETROATIVOS = 7
LIBERACAO_RETROATIVA_MINUTOS = 5
MAX_LOG_ACESSOS = 500
MAX_LOG_ACOES = 500
LOGIN_MAX_TENTATIVAS = 5
LOGIN_JANELA_MINUTOS = 15
PRODUCAO = os.environ.get("PORTAL_PRODUCAO") == "1"

FOTOS_DIR = STATIC_DIR / "fotos"
EXTENSOES_FOTO_PERMITIDAS = {"jpg", "jpeg", "png", "webp"}
METAS_FILE_NAME = "metas.json"

# Em produção (Render), não há disco persistente: os dados ficam num banco
# Postgres (Supabase) e as fotos num bucket, escolhidos por variável de ambiente.
# Sem essas variáveis (uso local), tudo continua indo pra arquivo, como sempre foi.
DATABASE_URL = os.environ.get("DATABASE_URL")
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY")


def carregar_config() -> dict:
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def resolver_pasta_dados() -> Path:
    config = carregar_config()
    data_dir = Path(config["data_dir"])
    if not data_dir.is_absolute():
        data_dir = ROOT / data_dir
    data_dir.mkdir(parents=True, exist_ok=True)
    return data_dir


def _cache_requisicao():
    """Memória de curta duração, válida só durante uma requisição. Várias
    funções carregam o mesmo dado (ex.: `carregar_vendedores()` é chamada 2-3x
    por requisição) — sem isso, cada chamada dessas era uma ida ao banco."""
    if not has_request_context():
        return None
    if not hasattr(g, "_cache_db"):
        g._cache_db = {}
    return g._cache_db


def _chave_de(caminho: Path) -> str:
    """Deriva uma chave curta e única a partir do nome do arquivo (sem extensão) —
    todos os arquivos do projeto já têm nomes únicos (vendedores, vendas_brenda, etc.)."""
    return Path(caminho).stem


if DATABASE_URL:
    import psycopg2
    from psycopg2.extras import Json as _PgJson

    # Reaproveita uma única conexão em vez de abrir uma nova a cada leitura/
    # escrita. Antes, um clique em "Adicionar venda" abria ~17 conexões novas
    # (cada uma com handshake TLS completo com o Supabase), o que deixava o
    # portal lento a ponto do vendedor achar que travou e clicar de novo —
    # gerando venda duplicada. Detalhe importante: `with psycopg2.connect(...)`
    # NÃO fecha a conexão (só encerra a transação), então o modelo antigo ainda
    # dependia do coletor de lixo pra liberar cada conexão.
    _conn_cache = {"conn": None}

    def _db_descartar_conexao(conn) -> None:
        """Some com uma conexão que deu erro, pra próxima operação abrir uma
        nova e limpa em vez de reaproveitar uma transação abortada."""
        try:
            conn.rollback()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
        if _conn_cache.get("conn") is conn:
            _conn_cache["conn"] = None
        cache = _cache_requisicao()
        if cache is not None:
            cache.pop("_conexao_ok", None)

    def _db_conectar():
        conn = _conn_cache["conn"]
        cache = _cache_requisicao()
        if conn is not None and conn.closed == 0:
            # A conexão reaproveitada pode ter caído entre uma requisição e
            # outra (Supabase derruba conexões ociosas), então confirmamos que
            # ainda está viva — mas só uma vez por requisição, senão esse
            # "SELECT 1" viraria uma ida ao banco a cada leitura.
            if cache is not None and cache.get("_conexao_ok"):
                return conn
            try:
                with conn.cursor() as cur:
                    cur.execute("SELECT 1")
                if cache is not None:
                    cache["_conexao_ok"] = True
                return conn
            except Exception:
                try:
                    conn.close()
                except Exception:
                    pass
        conn = psycopg2.connect(DATABASE_URL)
        _conn_cache["conn"] = conn
        if cache is not None:
            cache["_conexao_ok"] = True
        return conn

    def _db_preparar_tabela() -> None:
        conn = _db_conectar()
        with conn.cursor() as cur:
            cur.execute(
                "CREATE TABLE IF NOT EXISTS dados_json ("
                "chave TEXT PRIMARY KEY, valor JSONB NOT NULL)"
            )
            conn.commit()

    _db_preparar_tabela()

    def _db_ler(chave: str, padrao):
        cache = _cache_requisicao()
        if cache is not None and chave in cache:
            return cache[chave]
        conn = _db_conectar()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT valor FROM dados_json WHERE chave = %s", (chave,))
                linha = cur.fetchone()
                valor = linha[0] if linha else padrao
        except Exception:
            # Sem o rollback, a conexão reaproveitada ficaria travada em
            # "transação abortada" e derrubaria as próximas consultas também.
            _db_descartar_conexao(conn)
            raise
        if cache is not None:
            cache[chave] = valor
        return valor

    def _db_escrever(chave: str, dados) -> None:
        conn = _db_conectar()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO dados_json (chave, valor) VALUES (%s, %s) "
                    "ON CONFLICT (chave) DO UPDATE SET valor = EXCLUDED.valor",
                    (chave, _PgJson(dados)),
                )
                conn.commit()
        except Exception:
            _db_descartar_conexao(conn)
            raise
        cache = _cache_requisicao()
        if cache is not None:
            cache[chave] = dados

    def ler_json(caminho: Path, padrao):
        return _db_ler(_chave_de(caminho), padrao)

    def escrever_json(caminho: Path, dados) -> None:
        _db_escrever(_chave_de(caminho), dados)

else:
    def ler_json(caminho: Path, padrao):
        if caminho.exists():
            with open(caminho, "r", encoding="utf-8") as f:
                return json.load(f)
        return padrao

    def escrever_json(caminho: Path, dados) -> None:
        caminho.parent.mkdir(parents=True, exist_ok=True)
        with open(caminho, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=2)


# ============================================================
# Simulador de desconto e parcelamento
# ============================================================
# Auto-contido aqui dentro (antigo projeto irmão portal-simulador,
# incorporado): o catálogo ERP e as regras vivem em data/simulador/, geridos
# pela área do gestor em Configurações. scripts/etl_simulador.py reimporta
# as planilhas do ERP — grava em SQLite local sem DATABASE_URL, ou direto no
# Postgres de produção quando DATABASE_URL está setada (mesmo rodando local,
# pra empurrar uma atualização pra produção).
_SIMULADOR_DB_LOCAL = ROOT / "data" / "simulador" / "simulador.db"
_unaccent_disponivel_simulador = {"valor": False}

if DATABASE_URL:
    def _simulador_preparar_schema() -> None:
        with _db_conectar() as conn, conn.cursor() as cur:
            try:
                cur.execute("CREATE EXTENSION IF NOT EXISTS unaccent")
                conn.commit()
                _unaccent_disponivel_simulador["valor"] = True
            except Exception:
                conn.rollback()
                _unaccent_disponivel_simulador["valor"] = False
            cur.execute("""
                CREATE TABLE IF NOT EXISTS catalogo_erp (
                    id SERIAL PRIMARY KEY,
                    cod_peca TEXT NOT NULL UNIQUE,
                    nome_produto TEXT NOT NULL,
                    etiqueta TEXT,
                    preco DOUBLE PRECISION,
                    qtd_disponivel INTEGER NOT NULL DEFAULT 0,
                    condicao TEXT,
                    localizacao TEXT,
                    codigo_veiculo TEXT,
                    categoria_auto TEXT NOT NULL,
                    tipo_peca_auto TEXT NOT NULL,
                    tipo_peca_rotulo TEXT NOT NULL,
                    apelido_veiculo TEXT,
                    data_compra_veiculo TEXT,
                    tempo_estoque_conhecido INTEGER NOT NULL DEFAULT 0,
                    classe_abc TEXT NOT NULL DEFAULT 'C',
                    ingestido_em TEXT NOT NULL
                )
            """)
            cur.execute("CREATE INDEX IF NOT EXISTS idx_catalogo_disponivel ON catalogo_erp(qtd_disponivel)")
            cur.execute("""
                CREATE TABLE IF NOT EXISTS overrides_estoque (
                    cod_peca TEXT PRIMARY KEY,
                    data_entrada TEXT NOT NULL,
                    definido_por TEXT,
                    definido_em TEXT NOT NULL
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS etl_execucoes (
                    id SERIAL PRIMARY KEY,
                    executado_em TEXT NOT NULL,
                    linhas_catalogo INTEGER,
                    linhas_disponiveis INTEGER,
                    linhas_com_tempo_estoque INTEGER,
                    duracao_seg DOUBLE PRECISION
                )
            """)
            conn.commit()

    _simulador_preparar_schema()


def _simulador_db_local():
    conn = sqlite3.connect(_SIMULADOR_DB_LOCAL)
    conn.row_factory = sqlite3.Row
    return conn


def buscar_pecas_simulador(q: str) -> list:
    if DATABASE_URL:
        termo = f"%{q}%"
        campos = (
            "cod_peca, nome_produto, etiqueta, preco, classe_abc, apelido_veiculo, "
            "tempo_estoque_conhecido, tipo_peca_rotulo"
        )
        if _unaccent_disponivel_simulador["valor"]:
            condicao = (
                "(unaccent(nome_produto) ILIKE unaccent(%s) OR unaccent(cod_peca) ILIKE unaccent(%s) "
                "OR unaccent(coalesce(etiqueta, '')) ILIKE unaccent(%s))"
            )
        else:
            condicao = "(nome_produto ILIKE %s OR cod_peca ILIKE %s OR coalesce(etiqueta, '') ILIKE %s)"
        with _db_conectar() as conn, conn.cursor() as cur:
            cur.execute(
                f"SELECT {campos} FROM catalogo_erp WHERE qtd_disponivel > 0 AND {condicao} LIMIT 30",
                (termo, termo, termo),
            )
            colunas = [d[0] for d in cur.description]
            return [dict(zip(colunas, linha)) for linha in cur.fetchall()]

    conn = _simulador_db_local()
    termo = f"%{q}%"
    linhas = conn.execute(
        "SELECT cod_peca, nome_produto, etiqueta, preco, classe_abc, apelido_veiculo, "
        "tempo_estoque_conhecido, tipo_peca_rotulo FROM catalogo_erp WHERE qtd_disponivel > 0 AND "
        "(nome_produto LIKE ? OR cod_peca LIKE ? OR etiqueta LIKE ?) LIMIT 30",
        (termo, termo, termo),
    ).fetchall()
    conn.close()
    return [dict(r) for r in linhas]


def obter_peca_simulador(cod_peca: str):
    if DATABASE_URL:
        with _db_conectar() as conn, conn.cursor() as cur:
            cur.execute("SELECT * FROM catalogo_erp WHERE cod_peca = %s", (cod_peca,))
            linha = cur.fetchone()
            if not linha:
                return None
            colunas = [d[0] for d in cur.description]
            return dict(zip(colunas, linha))
    conn = _simulador_db_local()
    peca = conn.execute("SELECT * FROM catalogo_erp WHERE cod_peca = ?", (cod_peca,)).fetchone()
    conn.close()
    return dict(peca) if peca else None


def definir_data_entrada_simulador(cod_peca: str, data_entrada_str: str, definido_por: str) -> bool:
    """data_entrada_str já formatada como 'YYYY-MM-DD 00:00:00'. Retorna
    False se a peça não existe no catálogo."""
    agora = agora_br().isoformat()
    if DATABASE_URL:
        with _db_conectar() as conn, conn.cursor() as cur:
            cur.execute("SELECT 1 FROM catalogo_erp WHERE cod_peca = %s", (cod_peca,))
            if not cur.fetchone():
                return False
            cur.execute(
                "INSERT INTO overrides_estoque (cod_peca, data_entrada, definido_por, definido_em) "
                "VALUES (%s,%s,%s,%s) ON CONFLICT (cod_peca) DO UPDATE SET "
                "data_entrada=EXCLUDED.data_entrada, definido_por=EXCLUDED.definido_por, "
                "definido_em=EXCLUDED.definido_em",
                (cod_peca, data_entrada_str, definido_por, agora),
            )
            cur.execute(
                "UPDATE catalogo_erp SET data_compra_veiculo=%s, tempo_estoque_conhecido=1 WHERE cod_peca=%s",
                (data_entrada_str, cod_peca),
            )
            conn.commit()
        return True

    conn = _simulador_db_local()
    peca = conn.execute("SELECT * FROM catalogo_erp WHERE cod_peca = ?", (cod_peca,)).fetchone()
    if not peca:
        conn.close()
        return False
    conn.execute(
        "INSERT INTO overrides_estoque (cod_peca, data_entrada, definido_por, definido_em) "
        "VALUES (?,?,?,?) ON CONFLICT(cod_peca) DO UPDATE SET "
        "data_entrada=excluded.data_entrada, definido_por=excluded.definido_por, definido_em=excluded.definido_em",
        (cod_peca, data_entrada_str, definido_por, agora),
    )
    conn.execute(
        "UPDATE catalogo_erp SET data_compra_veiculo=?, tempo_estoque_conhecido=1 WHERE cod_peca=?",
        (data_entrada_str, cod_peca),
    )
    conn.commit()
    conn.close()
    return True


_REGRAS_SIMULADOR_FILE = ROOT / "data" / "simulador" / "regras.json"


def carregar_regras_simulador():
    if DATABASE_URL:
        return _db_ler("simulador_regras", None)
    if not _REGRAS_SIMULADOR_FILE.exists():
        return None
    return json.loads(_REGRAS_SIMULADOR_FILE.read_text(encoding="utf-8"))


def salvar_regras_simulador(regras) -> None:
    if DATABASE_URL:
        _db_escrever("simulador_regras", regras)
        return
    _REGRAS_SIMULADOR_FILE.parent.mkdir(parents=True, exist_ok=True)
    _REGRAS_SIMULADOR_FILE.write_text(json.dumps(regras, indent=2, ensure_ascii=False), encoding="utf-8")


def status_simulador() -> dict:
    if DATABASE_URL:
        with _db_conectar() as conn, conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM catalogo_erp")
            total = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM catalogo_erp WHERE qtd_disponivel > 0")
            disponiveis = cur.fetchone()[0]
            cur.execute(
                "SELECT executado_em, linhas_catalogo, linhas_disponiveis, linhas_com_tempo_estoque, duracao_seg "
                "FROM etl_execucoes ORDER BY id DESC LIMIT 1"
            )
            linha = cur.fetchone()
            ultima = None
            if linha:
                ultima = {
                    "executado_em": linha[0], "linhas_catalogo": linha[1],
                    "linhas_disponiveis": linha[2], "linhas_com_tempo_estoque": linha[3],
                    "duracao_seg": linha[4],
                }
        return {"total_pecas": total, "pecas_disponiveis": disponiveis, "ultima_importacao": ultima}

    conn = _simulador_db_local()
    total = conn.execute("SELECT COUNT(*) FROM catalogo_erp").fetchone()[0]
    disponiveis = conn.execute("SELECT COUNT(*) FROM catalogo_erp WHERE qtd_disponivel > 0").fetchone()[0]
    linha = conn.execute(
        "SELECT executado_em, linhas_catalogo, linhas_disponiveis, linhas_com_tempo_estoque, duracao_seg "
        "FROM etl_execucoes ORDER BY id DESC LIMIT 1"
    ).fetchone()
    conn.close()
    return {
        "total_pecas": total, "pecas_disponiveis": disponiveis,
        "ultima_importacao": dict(linha) if linha else None,
    }


def _faixa_tempo_de_simulador(dias, regras):
    for faixa in regras["faixas_tempo"]:
        if dias >= faixa["min_dias"] and (faixa["max_dias"] is None or dias <= faixa["max_dias"]):
            return faixa
    return regras["faixas_tempo"][-1]


def _faixa_valor_de_simulador(valor, regras):
    for faixa in regras["faixas_valor"]:
        if valor >= faixa["min_valor"] and (faixa["max_valor"] is None or valor <= faixa["max_valor"]):
            return faixa
    return regras["faixas_valor"][-1]


def montar_simulacao(valor_base, curva, dias_em_estoque, desconto_escolhido_pct, regras):
    """Desconto vale só pra dinheiro/PIX à vista. Cartão de crédito NUNCA
    tem desconto, nem em 1x — todas as opções de cartão são sobre o valor
    cheio."""
    faixa_tempo = (
        _faixa_tempo_de_simulador(dias_em_estoque, regras)
        if dias_em_estoque is not None else regras["faixas_tempo"][0]
    )
    desconto_max_pct = regras["desconto_max_pct"][curva][faixa_tempo["id"]]
    nivel_flex = regras["nivel_flexibilidade"][curva][faixa_tempo["id"]]

    desconto_pct = desconto_escolhido_pct if desconto_escolhido_pct is not None else desconto_max_pct
    desconto_pct = max(0, min(desconto_pct, desconto_max_pct))
    preco_dinheiro_pix = round(valor_base * (1 - desconto_pct / 100), 2)

    parcelas_max = 1
    if valor_base >= regras["valor_minimo_parcelamento"]:
        faixa_valor = _faixa_valor_de_simulador(valor_base, regras)
        parcelas_max = regras["parcelas_max"][faixa_valor["id"]][str(nivel_flex)]
    opcoes_cartao = [
        {"parcelas": n, "valor_parcela": round(valor_base / n, 2), "valor_total": valor_base}
        for n in range(1, parcelas_max + 1)
    ]

    return {
        "valor_base": valor_base,
        "dias_em_estoque": dias_em_estoque,
        "faixa_tempo": faixa_tempo,
        "curva": curva,
        "desconto_max_pct": desconto_max_pct,
        "desconto_aplicado_pct": desconto_pct,
        "nivel_flexibilidade": nivel_flex,
        "preco_dinheiro_pix": preco_dinheiro_pix,
        "opcoes_cartao": opcoes_cartao,
    }


def calcular_simulacao_peca(peca, valor_override, desconto_escolhido_pct, regras):
    valor_base = valor_override if valor_override is not None else (peca["preco"] or 0)

    if peca["tempo_estoque_conhecido"]:
        data_compra = datetime.fromisoformat(peca["data_compra_veiculo"])
        dias_em_estoque = (agora_br().date() - data_compra.date()).days
        dias_em_estoque = max(dias_em_estoque, 0)
    else:
        dias_em_estoque = None

    curva = peca["classe_abc"]
    resultado = montar_simulacao(valor_base, curva, dias_em_estoque, desconto_escolhido_pct, regras)
    resultado["tempo_estoque_conhecido"] = bool(peca["tempo_estoque_conhecido"])
    return resultado


def carregar_vendedores() -> dict:
    return ler_json(VENDEDORES_FILE, {})


def salvar_vendedores(vendedores: dict) -> None:
    escrever_json(VENDEDORES_FILE, vendedores)


def carregar_credenciais() -> dict:
    dados = ler_json(CREDENCIAIS_FILE, None)
    if dados is None:
        dados = {"admin_senha": "troque-esta-senha"}
        escrever_json(CREDENCIAIS_FILE, dados)
    return dados


def arquivo_vendas(vendedor_id: str) -> Path:
    """Cada vendedor tem seu próprio arquivo — assim cada um só grava no que é dele."""
    return resolver_pasta_dados() / f"vendas_{vendedor_id}.json"


def carregar_vendas_vendedor(vendedor_id: str) -> dict:
    return ler_json(arquivo_vendas(vendedor_id), {})


def salvar_vendas_vendedor(vendedor_id: str, vendas: dict) -> None:
    escrever_json(arquivo_vendas(vendedor_id), vendas)


def carregar_vendas_para_comissao(vendedor_id: str, vendedores: dict) -> dict:
    """Carrega só as vendas que a comissão desse vendedor realmente usa: as
    dele e as de quem ele tem override. Antes carregava as de todo mundo (~4.400
    registros) em toda atualização de tela, mesmo pra quem não tem override."""
    info = vendedores.get(vendedor_id, {})
    ids = {vendedor_id}
    for over in info.get("overrides", []):
        outro = over.get("vendedor_id")
        if outro in vendedores:
            ids.add(outro)
    todas = {}
    for vid in ids:
        todas.update(carregar_vendas_vendedor(vid))
    return todas


def carregar_vendas_todos(vendedores: dict) -> dict:
    """Junta as vendas de todos os vendedores — usado só para comissão com
    overrides e para o resumo do gestor, que precisam da visão completa."""
    todas = {}
    for vid in vendedores:
        todas.update(carregar_vendas_vendedor(vid))
    return todas


def carregar_meses_fechados() -> list:
    return ler_json(MESES_FECHADOS_FILE, [])


def salvar_meses_fechados(meses: list) -> None:
    escrever_json(MESES_FECHADOS_FILE, sorted(set(meses)))


def mes_esta_fechado(data_venda: str) -> bool:
    return data_venda[:7] in carregar_meses_fechados()


def arquivo_confirmacoes(vendedor_id: str) -> Path:
    return resolver_pasta_dados() / f"confirmacoes_{vendedor_id}.json"


def carregar_confirmacoes(vendedor_id: str) -> dict:
    return ler_json(arquivo_confirmacoes(vendedor_id), {})


def salvar_confirmacoes(vendedor_id: str, confirmacoes: dict) -> None:
    escrever_json(arquivo_confirmacoes(vendedor_id), confirmacoes)


def limpar_confirmacao(vendedor_id: str, mes: str) -> None:
    """Uma confirmação vira inválida se o vendedor mexer nos dados daquele mês depois."""
    confirmacoes = carregar_confirmacoes(vendedor_id)
    if mes in confirmacoes:
        del confirmacoes[mes]
        salvar_confirmacoes(vendedor_id, confirmacoes)


def arquivo_metas() -> Path:
    return resolver_pasta_dados() / METAS_FILE_NAME


def carregar_metas() -> dict:
    return ler_json(arquivo_metas(), {"grupo": {"diaria": 0, "semanal": 0, "mensal": 0}, "vendedores": {}})


def salvar_metas(metas: dict) -> None:
    escrever_json(arquivo_metas(), metas)


def metas_vendedor(vendedor_id: str, metas: dict) -> dict:
    return metas.get("vendedores", {}).get(vendedor_id, {"diaria": 0, "semanal": 0, "mensal": 0})


def registrar_acesso(tipo: str, sucesso: bool, vendedor_id: str = None, nome: str = None) -> None:
    log = ler_json(LOG_ACESSOS_FILE, [])
    log.append({
        "quando": agora_br().isoformat(timespec="seconds"),
        "tipo": tipo,
        "vendedor_id": vendedor_id,
        "nome": nome,
        "sucesso": sucesso,
        "ip": request.remote_addr,
    })
    escrever_json(LOG_ACESSOS_FILE, log[-MAX_LOG_ACESSOS:])


def registrar_acao(vendedor_id: str, nome: str, acao: str, produto: str, valor: float, detalhe: str = None) -> None:
    """Histórico de edição/exclusão/devolução de vendas, pra o gestor conseguir
    revisar o que cada vendedor mexeu — inclusive vendas já excluídas."""
    log = ler_json(LOG_ACOES_FILE, [])
    log.append({
        "quando": agora_br().isoformat(timespec="seconds"),
        "vendedor_id": vendedor_id,
        "nome": nome,
        "acao": acao,
        "produto": produto,
        "valor": valor,
        "detalhe": detalhe,
    })
    escrever_json(LOG_ACOES_FILE, log[-MAX_LOG_ACOES:])


def excedeu_tentativas_login(tipo: str, vendedor_id: str = None) -> bool:
    """Bloqueia login depois de várias senhas erradas seguidas, pra dificultar
    tentativa de adivinhação por força bruta (importante agora que fica na internet)."""
    log = ler_json(LOG_ACESSOS_FILE, [])
    limite = agora_br() - timedelta(minutes=LOGIN_JANELA_MINUTOS)
    falhas = 0
    for item in reversed(log):
        try:
            quando = parse_dt_tolerante(item["quando"])
        except (KeyError, ValueError):
            continue
        if quando < limite:
            break
        if item.get("tipo") == tipo and item.get("vendedor_id") == vendedor_id:
            if item.get("sucesso"):
                break
            falhas += 1
            if falhas >= LOGIN_MAX_TENTATIVAS:
                return True
    return False


def obter_secret_key() -> str:
    if DATABASE_URL:
        chave = _db_ler("secret_key", None)
        if not chave:
            chave = secrets.token_hex(32)
            _db_escrever("secret_key", chave)
        return chave
    if not SECRET_KEY_FILE.exists():
        SEGREDOS_DIR.mkdir(parents=True, exist_ok=True)
        SECRET_KEY_FILE.write_text(secrets.token_hex(32), encoding="utf-8")
    return SECRET_KEY_FILE.read_text(encoding="utf-8").strip()


app = Flask(__name__, static_folder=str(STATIC_DIR), static_url_path="")
app.secret_key = obter_secret_key()
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=PRODUCAO,
    PERMANENT_SESSION_LIFETIME=timedelta(hours=2),
)


@app.route("/")
def index():
    return send_from_directory(STATIC_DIR, "index.html")


@app.route("/admin.html")
def admin_page():
    return send_from_directory(STATIC_DIR, "admin.html")


@app.route("/painel.html")
def painel_page():
    return send_from_directory(STATIC_DIR, "painel.html")


# ---------- Login do vendedor ----------

@app.route("/api/vendedores-publico")
def api_vendedores_publico():
    vendedores = carregar_vendedores()
    return jsonify([{"id": vid, "nome": v["nome"]} for vid, v in sorted(vendedores.items(), key=lambda kv: kv[1]["nome"])])


@app.route("/api/login", methods=["POST"])
def api_login():
    body = request.get_json(force=True)
    vendedor_id = (body.get("vendedor_id") or "").strip()
    senha = body.get("senha") or ""
    if excedeu_tentativas_login("vendedor", vendedor_id):
        return jsonify({"erro": f"Muitas tentativas erradas. Aguarde {LOGIN_JANELA_MINUTOS} minutos e tente de novo."}), 429
    vendedores = carregar_vendedores()
    v = vendedores.get(vendedor_id)
    if not v or v.get("senha") != senha:
        registrar_acesso("vendedor", False, vendedor_id, v["nome"] if v else None)
        return jsonify({"erro": "Vendedor ou senha inválidos."}), 401
    session.clear()
    session["vendedor_id"] = vendedor_id
    registrar_acesso("vendedor", True, vendedor_id, v["nome"])
    return jsonify({"ok": True, "nome": v["nome"]})


@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/ambiente")
def api_ambiente():
    """Diz se este servidor é o de produção ou uma cópia local.

    Sem isso as duas telas são idênticas, e já aconteceu três vezes de alguém
    olhar o portal local — com dados congelados e sem os arquivos de
    atendimento — e concluir que produção estava quebrada. Público de propósito:
    o aviso precisa aparecer antes do login, que é onde a confusão começa."""
    # A versao e o carimbo dos arquivos da tela. A pagina guarda o valor que
    # recebeu ao abrir e reconfere de tempos em tempos: mudou, e porque saiu
    # publicacao nova e aquela aba esta velha. Sem isso, quem deixa a janela
    # aberta o dia todo (atalho do Chrome em modo app) fica vendo a versao
    # antiga e concluindo que o portal esta com defeito.
    try:
        marcas = [(STATIC_DIR / nome).stat().st_mtime
                  for nome in ("index.html", "admin.html", "portal-nav.js")
                  if (STATIC_DIR / nome).exists()]
        versao = str(int(max(marcas))) if marcas else "0"
    except OSError:
        versao = "0"
    return jsonify({"local": not bool(DATABASE_URL), "versao": versao})


@app.route("/api/me")
def api_me():
    vendedor_id = session.get("vendedor_id")
    if not vendedor_id:
        return jsonify({"logado": False}), 401
    vendedores = carregar_vendedores()
    v = vendedores.get(vendedor_id)
    if not v:
        session.clear()
        return jsonify({"logado": False}), 401
    return jsonify({
        "logado": True,
        "id": vendedor_id,
        "nome": v["nome"],
        "percentual": v.get("percentual", 0),
    })


def mes_para_intervalo(mes: str) -> tuple[str, str]:
    return f"{mes}-01", f"{mes}-31"


def valor_liquido(v: dict) -> float:
    """Valor de uma venda descontando o que foi devolvido, sem apagar o histórico original."""
    devolucao = v.get("devolucao")
    if not devolucao:
        return v["valor"]
    if devolucao.get("tipo") == "total":
        return 0.0
    return max(0.0, v["valor"] - float(devolucao.get("valor_devolvido", 0)))


def total_vendido(vendedor_id: str, de: str, ate: str, vendas: dict, tipo: str = "venda") -> float:
    total = sum(
        valor_liquido(v)
        for v in vendas.values()
        if v["vendedor_id"] == vendedor_id and de <= v["data"] <= ate and v.get("tipo", "venda") == tipo
    )
    return round(total, 2)


def calcular_comissao(vendedor_id: str, de: str, ate: str, vendedores: dict, vendas: dict):
    info = vendedores[vendedor_id]
    proprio = total_vendido(vendedor_id, de, ate, vendas)
    percentual = float(info.get("percentual", 0))
    comissao = proprio * percentual / 100

    overrides_detalhe = []
    for over in info.get("overrides", []):
        outro_id = over.get("vendedor_id")
        outro_percentual = float(over.get("percentual", 0))
        if outro_id not in vendedores:
            continue
        outro_total = total_vendido(outro_id, de, ate, vendas)
        valor_over = round(outro_total * outro_percentual / 100, 2)
        comissao += valor_over
        overrides_detalhe.append({
            "vendedor_id": outro_id,
            "nome": vendedores[outro_id]["nome"],
            "percentual": outro_percentual,
            "total_vendido": outro_total,
            "valor": valor_over,
        })

    return {
        "total_vendido": proprio,
        "percentual": percentual,
        "comissao_propria": round(proprio * percentual / 100, 2),
        "overrides": overrides_detalhe,
        "comissao": round(comissao, 2),
        "total_bonus": total_vendido(vendedor_id, de, ate, vendas, tipo="bonus"),
    }


@app.route("/api/minha-comissao")
def api_minha_comissao():
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401
    mes = request.args.get("mes", hoje_br().isoformat()[:7])
    de, ate = mes_para_intervalo(mes)
    vendedores = carregar_vendedores()
    vendas = carregar_vendas_para_comissao(vendedor_id, vendedores)
    return jsonify(calcular_comissao(vendedor_id, de, ate, vendedores, vendas))


def perfil_de(vendedor_id: str) -> str:
    """"vendedor" (padrao) ou "expedicao".

    O portal nasceu so com vendedores; a expedicao entrou depois e reaproveita
    o mesmo login. Quem e da expedicao nao vende — some das listas de venda,
    comissao e meta — e em troca ve a fila de pedidos pra liberar.
    """
    v = carregar_vendedores().get(vendedor_id) or {}
    return v.get("perfil") or "vendedor"


def exigir_expedicao():
    """Id de quem esta logado, se for da expedicao ou o gestor."""
    vid = exigir_vendedor()
    if vid and perfil_de(vid) == "expedicao":
        return vid
    if exigir_admin():
        return "gestor"
    return None


def exigir_vendedor():
    vendedor_id = session.get("vendedor_id")
    if not vendedor_id:
        return None
    vendedores = carregar_vendedores()
    if vendedor_id not in vendedores:
        return None
    return vendedor_id


def _atd_resolvidos() -> dict:
    """{id da sessao: carimbo da ultima mensagem quando foi marcada}."""
    return ler_json(resolver_pasta_dados() / "atendimento_resolvido.json", None) or {}


def _atd_pendentes(conversas, resolvidos):
    """Tira da lista o que foi resolvido, com DUAS validades:

    1. O cliente nao pode ter falado depois. Se mandou mensagem nova, a
       conversa volta — esconder quem voltou a falar seria pior que nao ter
       alerta nenhum.
    2. A marcacao vale so pelo dia em que foi feita. "Ja resolvi" quer dizer
       "estou tratando isso agora", nao "esquece pra sempre": se amanhecer e a
       conversa continuar sem resposta da loja, ela volta pra fila. Foi o que o
       gestor pediu em 28/08/2026 — nada pode cair no esquecimento.
    """
    hoje = hoje_br().isoformat()
    saida = []
    for c in conversas:
        marca = resolvidos.get(c["id"])
        if isinstance(marca, str):        # formato antigo: so o carimbo
            marca = {"ultima_em": marca, "em": marca[:10]}
        if (marca and marca.get("em") == hoje
                and c.get("ultima_em") and c["ultima_em"] <= marca.get("ultima_em", "")):
            continue
        saida.append(c)
    return saida


@app.route("/api/atendimento/resolver", methods=["POST"])
def api_atendimento_resolver():
    """Marca uma conversa como resolvida. Vale pro vendedor e pro gestor: os
    dois veem a mesma fila, entao resolver num lugar resolve no outro."""
    if not (exigir_vendedor() or exigir_admin()):
        return jsonify({"erro": "Não autenticado."}), 401
    corpo = request.get_json(silent=True) or {}
    sessao = (corpo.get("id") or "").strip()
    if not sessao:
        return jsonify({"erro": "Sessão não informada."}), 400

    resolvidos = _atd_resolvidos()
    if corpo.get("desfazer"):
        resolvidos.pop(sessao, None)
    else:
        # Guarda o carimbo da ultima fala do cliente, nao a hora do clique:
        # e assim que a conversa reaparece se ele voltar a falar.
        resolvidos[sessao] = {
            "ultima_em": (corpo.get("ultima_em")
                          or agora_br().isoformat(timespec="seconds")),
            # A data do clique e o que faz a marcacao expirar amanha.
            "em": hoje_br().isoformat(),
        }
    # Marcacao de ontem pra tras ja nao vale — nao precisa ficar no arquivo.
    hoje = hoje_br().isoformat()
    resolvidos = {k: v for k, v in resolvidos.items()
                  if (v.get("em") if isinstance(v, dict) else str(v)[:10]) == hoje}
    escrever_json(resolver_pasta_dados() / "atendimento_resolvido.json", resolvidos)
    return jsonify({"ok": True, "resolvidas": len(resolvidos)})


@app.route("/api/meu-atendimento")
def api_meu_atendimento():
    """Os clientes que estao esperando ESTE vendedor responder.

    Mesma fonte do painel do gestor (atendimento_alerta), filtrada pelo id de
    quem esta logado: cada um ve so a propria fila. Quem age e o vendedor, por
    isso o alerta vive aqui e nao so na area do gestor.
    """
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401
    d = ler_json(resolver_pasta_dados() / "atendimento_alerta.json", None)
    if not d:
        return jsonify({"sem_dados": True})
    resolvidos = _atd_resolvidos()
    minhas = _atd_pendentes([c for c in (d.get("conversas") or [])
                             if c.get("vendedor_id") == vendedor_id], resolvidos)
    return jsonify({
        "gerado_em": d.get("gerado_em"),
        "limites": d.get("limites"),
        "total": len(minhas),
        "critico": sum(1 for c in minhas if c["nivel"] == "critico"),
        "urgente": sum(1 for c in minhas if c["nivel"] == "urgente"),
        "conversas": minhas,
    })


@app.route("/api/meu-painel")
def api_meu_painel():
    """Tudo que o painel do vendedor mostra, numa chamada só — evita a tela
    disparar cinco requisições e ficar montando aos pedaços."""
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401

    mes = request.args.get("mes", hoje_br().isoformat()[:7])
    de, ate = mes_para_intervalo(mes)
    vendedores = carregar_vendedores()
    vendas_comissao = carregar_vendas_para_comissao(vendedor_id, vendedores)
    comissao = calcular_comissao(vendedor_id, de, ate, vendedores, vendas_comissao)

    minhas = [
        v for v in vendas_comissao.values()
        if v["vendedor_id"] == vendedor_id and de <= v["data"] <= ate
        and v.get("tipo", "venda") == "venda"
    ]

    hoje = hoje_br()
    inicio_semana = hoje - timedelta(days=hoje.weekday())
    metas = metas_vendedor(vendedor_id, carregar_metas())

    def soma(itens):
        return round(sum(valor_liquido(v) for v in itens), 2)

    # Só conta o dia/semana se o mês em tela for o corrente — senão "vendi hoje"
    # apareceria zerado enquanto o vendedor revisa um mês passado.
    mes_corrente = mes == hoje.isoformat()[:7]
    total_hoje = soma([v for v in minhas if v["data"] == hoje.isoformat()]) if mes_corrente else None
    total_semana = soma([v for v in minhas if v["data"] >= inicio_semana.isoformat()]) if mes_corrente else None

    total_mes = comissao["total_vendido"]
    qtd = len(minhas)
    devolvidas = [v for v in minhas if v.get("devolucao")]

    # Evolução dia a dia, pro gráfico de linha
    por_dia = {}
    for v in minhas:
        por_dia[v["data"]] = round(por_dia.get(v["data"], 0) + valor_liquido(v), 2)

    def ranking(campo, rotulo_vazio):
        acumulado = {}
        for v in minhas:
            chave = (v.get(campo) or "").strip() or rotulo_vazio
            acumulado[chave] = round(acumulado.get(chave, 0) + valor_liquido(v), 2)
        top = sorted(acumulado.items(), key=lambda kv: kv[1], reverse=True)[:8]
        return [{"nome": k, "valor": val} for k, val in top]

    dias_no_mes = calendar.monthrange(int(mes[:4]), int(mes[5:7]))[1]
    dias_restantes = max(1, dias_no_mes - hoje.day + 1) if mes_corrente else 1
    meta_mensal = float(metas.get("mensal", 0))
    falta = max(0.0, meta_mensal - total_mes)

    fila = carregar_fila_retomada(vendedor_id)
    status_retomada = carregar_status_retomada(vendedor_id)
    resumo_retomada = _resumo_retomada(fila.get("itens", []), status_retomada) if fila else None
    if resumo_retomada:
        # Os tres primeiros ja no painel: e a primeira tela que ele abre, e ali o
        # follow-up vira trabalho a fazer em vez de mais um link no menu.
        resumo_retomada["topo"] = _topo_retomada(fila.get("itens", []), status_retomada)

    return jsonify({
        "mes": mes,
        "mes_corrente": mes_corrente,
        "total_mes": total_mes,
        "total_hoje": total_hoje,
        "total_semana": total_semana,
        "qtd_vendas": qtd,
        "ticket_medio": round(total_mes / qtd, 2) if qtd else 0,
        "comissao": comissao,
        "devolucoes": {
            "quantidade": len(devolvidas),
            "valor": round(sum(v["valor"] - valor_liquido(v) for v in devolvidas), 2),
        },
        "metas": {
            "diaria": float(metas.get("diaria", 0)),
            "semanal": float(metas.get("semanal", 0)),
            "mensal": meta_mensal,
            "falta_no_mes": round(falta, 2),
            "necessario_por_dia": round(falta / dias_restantes, 2) if falta else 0,
            "dias_restantes": dias_restantes if mes_corrente else 0,
        },
        "evolucao": [{"data": d, "valor": por_dia[d]} for d in sorted(por_dia)],
        "top_produtos": ranking("produto", "Sem descrição"),
        "top_canais": ranking("canal", "Sem canal"),
        "retomada": resumo_retomada,
    })


# ---------- Vendas do vendedor logado ----------

@app.route("/api/vendas", methods=["GET"])
def api_listar_vendas():
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401
    todos = request.args.get("todos") == "1"
    mes = request.args.get("mes", hoje_br().isoformat()[:7])
    vendas = carregar_vendas_vendedor(vendedor_id)
    minhas = [
        {**v, "id": vid}
        for vid, v in vendas.items()
        if (todos or v["data"][:7] == mes) and v.get("tipo", "venda") == "venda"
    ]
    minhas.sort(key=lambda v: v["data"], reverse=True)
    return jsonify(minhas)


# O canal era texto livre e virou 19 grafias pra meia duzia de canais reais —
# "ITAU" de quatro jeitos, "B" pra balcao. Normaliza na entrada, que e a unica
# porta: o historico foi corrigido uma vez por script, e daqui pra frente toda
# venda ja chega com o nome canonico. Grafia desconhecida passa como veio,
# porque inventar mapeamento e pior que fragmentar.
CANAIS_CANONICOS = {
    "b": "Balcão", "balcao": "Balcão",
    "ml": "Mercado Livre", "mercado livre": "Mercado Livre",
    "itau": "Itaú",
    "cred": "Crediário", "crediario": "Crediário",
    "din": "Dinheiro", "dinh": "Dinheiro", "dinheiro": "Dinheiro",
    "site": "Site", "loja integrada": "Site",
    "itau / cred": "Itaú / Crediário",
}


def _achatar_canal(x: str) -> str:
    """minusculo, sem acento, espacos colapsados. O NFKD pode CRIAR espaco —
    o acento morto (U+00B4, tecla morta do ABNT) vira espaco + combinante —
    entao colapsa de novo no fim, senao "ML" + acento morto vira "ml " e nao
    casa com nada (e ja rendeu um 500 no lancamento)."""
    x = "".join(c for c in unicodedata.normalize("NFKD", x.lower())
                if not unicodedata.combining(c))
    return " ".join(x.split())


def normalizar_canal(texto) -> str:
    t = " ".join(str(texto or "").split())
    if not t:
        return ""
    chave = _achatar_canal(t)
    if chave in CANAIS_CANONICOS:
        return CANAIS_CANONICOS[chave]
    # "Cred 4x" -> "Crediário 4x": a primeira palavra e o canal, o resto e
    # detalhe (parcelamento) que nao se joga fora. Divide o texto REAL, nao a
    # chave achatada — os dois podem ter contagens de espaco diferentes.
    partes = t.split(" ", 1)
    if len(partes) == 2 and _achatar_canal(partes[0]) in CANAIS_CANONICOS:
        return CANAIS_CANONICOS[_achatar_canal(partes[0])] + " " + partes[1]
    return t


def validar_valor_produto(body: dict) -> tuple[float, str, str, str]:
    try:
        valor = round(float(body.get("valor")), 2)
    except (TypeError, ValueError):
        raise ValueError("Valor inválido.")
    if valor <= 0:
        raise ValueError("Valor deve ser maior que zero.")
    produto = (body.get("produto") or "").strip()
    if not produto:
        raise ValueError("Informe o que foi vendido.")
    canal = normalizar_canal(body.get("canal"))
    sku = (body.get("sku") or "").strip()
    return valor, produto, canal, sku


def validar_data_venda(data_venda: str, ignorar_limite: bool = False) -> None:
    """Confere se a data é válida e está dentro da janela permitida de lançamento."""
    try:
        data_obj = date.fromisoformat(data_venda)
    except ValueError:
        raise ValueError("Data inválida.")
    hoje = hoje_br()
    if data_obj > hoje:
        raise ValueError("Não é possível usar uma data futura.")
    if not ignorar_limite and (hoje - data_obj).days > DIAS_MAXIMOS_RETROATIVOS:
        raise ValueError(
            f"Essa data é de mais de {DIAS_MAXIMOS_RETROATIVOS} dias atrás. "
            "Fale com o gestor para lançar vendas retroativas além desse prazo."
        )


def montar_venda(vendedor_id: str, body: dict, ignorar_limite_retroativo: bool = False) -> dict:
    """Valida os campos de uma venda e retorna o dict pronto para salvar.
    Lança ValueError com a mensagem de erro em caso de dado inválido."""
    valor, produto, canal, sku = validar_valor_produto(body)
    data_venda = (body.get("data") or hoje_br().isoformat()).strip()
    validar_data_venda(data_venda, ignorar_limite=ignorar_limite_retroativo)
    if mes_esta_fechado(data_venda):
        raise ValueError("Esse mês já foi fechado pelo gestor e não aceita mais lançamentos.")

    venda = {
        "vendedor_id": vendedor_id,
        "data": data_venda,
        "valor": valor,
        "produto": produto,
        "tipo": "venda",
        "criado_em": agora_br().isoformat(timespec="seconds"),
    }
    if canal:
        venda["canal"] = canal
    if sku:
        venda["sku"] = sku
    return venda


def venda_igual_no_mes(vendas: dict, produto: str, valor: float, mes: str):
    """Devolve uma venda do mesmo mês com produto e valor idênticos, se houver.
    Serve pro aviso de "esse produto já foi lançado" na hora do lançamento —
    é só um alerta, porque duas peças iguais de carros iguais são possíveis."""
    produto_norm = produto.strip().lower()
    for v in vendas.values():
        if v.get("tipo", "venda") != "venda":
            continue
        if v["data"][:7] != mes:
            continue
        if v["produto"].strip().lower() == produto_norm and v["valor"] == valor:
            return v
    return None


def retroativo_ativo(vendedor: dict) -> bool:
    """Verifica se a liberação temporária de lançamento retroativo do gestor
    ainda está dentro da janela de tempo (não é mais de uso único)."""
    ate = vendedor.get("liberacao_retroativa_ate")
    if not ate:
        return False
    try:
        return agora_br() < parse_dt_tolerante(ate)
    except ValueError:
        return False


@app.route("/api/vendas", methods=["POST"])
def api_criar_venda():
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401
    liberado = retroativo_ativo(carregar_vendedores().get(vendedor_id, {}))
    body = request.get_json(force=True)
    try:
        venda = montar_venda(vendedor_id, body, ignorar_limite_retroativo=liberado)
    except ValueError as e:
        return jsonify({"erro": str(e)}), 400

    vendas = carregar_vendas_vendedor(vendedor_id)

    # Proteção contra duplicidade. O `envio_id` é gerado pelo navegador uma vez
    # por lançamento: se a mesma tentativa chegar de novo (clique duplo, conexão
    # lenta que o vendedor achou que travou, refresh no meio do envio), a gente
    # devolve a venda que já foi salva em vez de criar outra. É mais confiável
    # que a checagem por tempo, que falhava justamente no caso ruim — servidor
    # demorando pra responder e vendedor tentando de novo depois de 8 segundos.
    envio_id = (body.get("envio_id") or "").strip()
    if envio_id:
        for vid_existente, v in vendas.items():
            if v.get("envio_id") == envio_id:
                return jsonify({"ok": True, "id": vid_existente, "ja_existia": True})
        venda["envio_id"] = envio_id

    # Aviso (não bloqueio): já existe venda igual em produto e valor no mesmo
    # mês? Pode ser legítimo — duas peças iguais de carros iguais — então só
    # perguntamos. O vendedor reenvia com `confirmar_duplicata` pra confirmar.
    if not body.get("confirmar_duplicata"):
        igual = venda_igual_no_mes(vendas, venda["produto"], venda["valor"], venda["data"][:7])
        if igual:
            return jsonify({
                "confirmar_duplicata": True,
                "existente": {"data": igual["data"], "produto": igual["produto"], "valor": igual["valor"]},
            }), 409

    novo_id = uuid.uuid4().hex[:12]
    vendas[novo_id] = venda
    salvar_vendas_vendedor(vendedor_id, vendas)
    limpar_confirmacao(vendedor_id, venda["data"][:7])
    return jsonify({"ok": True, "id": novo_id})


@app.route("/api/vendas/lote", methods=["POST"])
def api_criar_vendas_lote():
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401
    liberado = retroativo_ativo(carregar_vendedores().get(vendedor_id, {}))
    body = request.get_json(force=True)
    linhas = body.get("vendas", [])
    if not isinstance(linhas, list) or not linhas:
        return jsonify({"erro": "Nenhuma linha para salvar."}), 400

    vendas = carregar_vendas_vendedor(vendedor_id)

    # Mesma proteção do lançamento avulso: um `envio_id` por clique em "Salvar
    # tudo". Aqui ela é ainda mais importante, porque um lote pode ter linhas
    # legitimamente iguais (duas peças iguais vendidas no mesmo dia), então não
    # dá pra deduplicar comparando produto/valor/data como no avulso.
    envio_id = (body.get("envio_id") or "").strip()
    if envio_id:
        ja_salvas = [vid for vid, v in vendas.items() if v.get("envio_id") == envio_id]
        if ja_salvas:
            return jsonify({"ok": True, "salvas": len(ja_salvas), "erros": [], "ja_existia": True})

    salvas = 0
    erros = []
    linhas_salvas = []
    meses_afetados = set()
    for idx, linha in enumerate(linhas, start=1):
        try:
            venda = montar_venda(vendedor_id, linha, ignorar_limite_retroativo=liberado)
        except ValueError as e:
            erros.append({"linha": idx, "erro": str(e)})
            continue
        if envio_id:
            venda["envio_id"] = envio_id
        vendas[uuid.uuid4().hex[:12]] = venda
        salvas += 1
        linhas_salvas.append(idx)
        meses_afetados.add(venda["data"][:7])

    if salvas:
        salvar_vendas_vendedor(vendedor_id, vendas)
        for mes in meses_afetados:
            limpar_confirmacao(vendedor_id, mes)
    # `linhas_salvas` deixa o navegador apagar da planilha só as linhas que
    # realmente entraram, mantendo as que deram erro. Sem isso, quando parte do
    # lote falhava a planilha continuava inteira na tela e o vendedor corrigia
    # e salvava tudo de novo — duplicando o que já tinha sido salvo.
    return jsonify({"ok": True, "salvas": salvas, "erros": erros, "linhas_salvas": linhas_salvas})


@app.route("/api/vendas/<venda_id>", methods=["DELETE"])
def api_remover_venda(venda_id):
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401
    vendas = carregar_vendas_vendedor(vendedor_id)
    if venda_id not in vendas:
        return jsonify({"erro": "Venda não encontrada."}), 404
    if mes_esta_fechado(vendas[venda_id]["data"]):
        return jsonify({"erro": "Esse mês já foi fechado pelo gestor e não aceita mais alterações."}), 403
    mes_afetado = vendas[venda_id]["data"][:7]
    removida = vendas[venda_id]
    del vendas[venda_id]
    salvar_vendas_vendedor(vendedor_id, vendas)
    limpar_confirmacao(vendedor_id, mes_afetado)
    nome = carregar_vendedores().get(vendedor_id, {}).get("nome", vendedor_id)
    registrar_acao(vendedor_id, nome, "excluiu", removida["produto"], removida["valor"], removida["data"])
    return jsonify({"ok": True})


@app.route("/api/vendas/<venda_id>", methods=["PUT"])
def api_editar_venda(venda_id):
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401
    vendas = carregar_vendas_vendedor(vendedor_id)
    atual = vendas.get(venda_id)
    if not atual:
        return jsonify({"erro": "Venda não encontrada."}), 404
    if atual.get("tipo", "venda") != "venda":
        return jsonify({"erro": "Não é possível editar esse registro."}), 400
    if mes_esta_fechado(atual["data"]):
        return jsonify({"erro": "Esse mês já foi fechado pelo gestor e não aceita mais alterações."}), 403

    liberado = retroativo_ativo(carregar_vendedores().get(vendedor_id, {}))
    body = request.get_json(force=True)
    try:
        valor, produto, canal, sku = validar_valor_produto(body)
        nova_data = (body.get("data") or atual["data"]).strip()
        if nova_data != atual["data"]:
            validar_data_venda(nova_data, ignorar_limite=liberado)
            if mes_esta_fechado(nova_data):
                raise ValueError("Esse mês já foi fechado pelo gestor.")
    except ValueError as e:
        return jsonify({"erro": str(e)}), 400

    mes_antigo = atual["data"][:7]
    atualizada = {
        **atual,
        "data": nova_data,
        "valor": valor,
        "produto": produto,
        "editado_em": agora_br().isoformat(timespec="seconds"),
    }
    if canal:
        atualizada["canal"] = canal
    else:
        atualizada.pop("canal", None)
    # canal_original documenta a grafia que a migracao corrigiu. Se o vendedor
    # trocou o canal de verdade, a grafia antiga nao descreve mais esta venda.
    if canal != atual.get("canal"):
        atualizada.pop("canal_original", None)
    if sku:
        atualizada["sku"] = sku
    else:
        atualizada.pop("sku", None)
    vendas[venda_id] = atualizada
    salvar_vendas_vendedor(vendedor_id, vendas)

    limpar_confirmacao(vendedor_id, mes_antigo)
    if nova_data[:7] != mes_antigo:
        limpar_confirmacao(vendedor_id, nova_data[:7])

    mudancas = []
    if atual["valor"] != valor:
        mudancas.append(f"valor {atual['valor']:.2f} → {valor:.2f}")
    if atual["produto"] != produto:
        mudancas.append(f"produto \"{atual['produto']}\" → \"{produto}\"")
    if atual["data"] != nova_data:
        mudancas.append(f"data {atual['data']} → {nova_data}")
    nome = carregar_vendedores().get(vendedor_id, {}).get("nome", vendedor_id)
    registrar_acao(vendedor_id, nome, "editou", produto, valor, "; ".join(mudancas) or None)
    return jsonify({"ok": True})


@app.route("/api/vendas/<venda_id>/devolucao", methods=["POST"])
def api_marcar_devolucao(venda_id):
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401
    vendas = carregar_vendas_vendedor(vendedor_id)
    atual = vendas.get(venda_id)
    if not atual:
        return jsonify({"erro": "Venda não encontrada."}), 404
    if atual.get("tipo", "venda") != "venda":
        return jsonify({"erro": "Não é possível marcar devolução nesse registro."}), 400

    body = request.get_json(force=True)
    tipo = (body.get("tipo") or "").strip()
    if tipo not in ("parcial", "total"):
        return jsonify({"erro": "Tipo de devolução inválido."}), 400

    if tipo == "total":
        valor_devolvido = atual["valor"]
    else:
        try:
            valor_devolvido = round(float(body.get("valor")), 2)
        except (TypeError, ValueError):
            return jsonify({"erro": "Valor devolvido inválido."}), 400
        if valor_devolvido <= 0:
            return jsonify({"erro": "Valor devolvido deve ser maior que zero."}), 400
        if valor_devolvido > atual["valor"]:
            return jsonify({"erro": "Valor devolvido não pode ser maior que o valor da venda."}), 400

    vendas[venda_id] = {
        **atual,
        "devolucao": {
            "tipo": tipo,
            "valor_devolvido": valor_devolvido,
            "marcado_em": agora_br().isoformat(timespec="seconds"),
        },
    }
    salvar_vendas_vendedor(vendedor_id, vendas)
    limpar_confirmacao(vendedor_id, atual["data"][:7])
    nome = carregar_vendedores().get(vendedor_id, {}).get("nome", vendedor_id)
    acao = "marcou devolução total" if tipo == "total" else "marcou devolução parcial"
    registrar_acao(vendedor_id, nome, acao, atual["produto"], valor_devolvido)
    return jsonify({"ok": True})


@app.route("/api/vendas/<venda_id>/devolucao", methods=["DELETE"])
def api_remover_devolucao(venda_id):
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401
    vendas = carregar_vendas_vendedor(vendedor_id)
    atual = vendas.get(venda_id)
    if not atual:
        return jsonify({"erro": "Venda não encontrada."}), 404
    if "devolucao" not in atual:
        return jsonify({"erro": "Essa venda não tem devolução marcada."}), 400

    nova = dict(atual)
    nova.pop("devolucao")
    vendas[venda_id] = nova
    salvar_vendas_vendedor(vendedor_id, vendas)
    limpar_confirmacao(vendedor_id, atual["data"][:7])
    nome = carregar_vendedores().get(vendedor_id, {}).get("nome", vendedor_id)
    registrar_acao(vendedor_id, nome, "desfez devolução", atual["produto"], atual["valor"])
    return jsonify({"ok": True})


@app.route("/api/confirmar-mes", methods=["POST"])
def api_confirmar_mes():
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401
    body = request.get_json(force=True)
    mes = (body.get("mes") or "").strip()
    if len(mes) != 7 or mes[4] != "-":
        return jsonify({"erro": "Mês inválido."}), 400
    confirmacoes = carregar_confirmacoes(vendedor_id)
    confirmacoes[mes] = agora_br().isoformat(timespec="seconds")
    salvar_confirmacoes(vendedor_id, confirmacoes)
    return jsonify({"ok": True, "confirmado_em": confirmacoes[mes]})


@app.route("/api/minha-confirmacao")
def api_minha_confirmacao():
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401
    mes = request.args.get("mes", hoje_br().isoformat()[:7])
    confirmacoes = carregar_confirmacoes(vendedor_id)
    return jsonify({"mes": mes, "confirmado_em": confirmacoes.get(mes)})


# ---------- Painel público de ranking (sem login, pensado pra ficar numa TV/monitor) ----------

@app.route("/api/metas")
def api_metas():
    # So o gestor consome esta rota (tela de configuracao). Aberta, ela
    # entregava a meta individual de cada vendedor a qualquer um sem login —
    # dado que o resto do portal esconde de proposito.
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    return jsonify(carregar_metas())


@app.route("/api/painel/ranking")
def api_painel_ranking():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    vendedores = carregar_vendedores()
    vendas = carregar_vendas_todos(vendedores)
    metas = carregar_metas()

    hoje = hoje_br()
    inicio_semana = hoje - timedelta(days=hoje.weekday())
    inicio_mes = hoje.replace(day=1)

    resultado = []
    grupo_hoje = grupo_semana = grupo_mes = 0.0

    for vid, info in vendedores.items():
        m = metas_vendedor(vid, metas)
        hoje_v = total_vendido(vid, hoje.isoformat(), hoje.isoformat(), vendas)
        semana_v = total_vendido(vid, inicio_semana.isoformat(), hoje.isoformat(), vendas)
        mes_v = total_vendido(vid, inicio_mes.isoformat(), hoje.isoformat(), vendas)
        # Conta oculta so aparece na TV se tiver movimento — dinheiro na tela
        # sempre, cartao vazio de conta de teste nunca.
        if info.get("oculto") and not (hoje_v or semana_v or mes_v):
            continue
        grupo_hoje += hoje_v
        grupo_semana += semana_v
        grupo_mes += mes_v
        resultado.append({
            "id": vid,
            "nome": info["nome"],
            "foto": info.get("foto"),
            "avatar": info.get("avatar", ""),
            "hoje": hoje_v,
            "semana": semana_v,
            "mes": mes_v,
            "meta_diaria": float(m.get("diaria", 0)),
            "meta_semanal": float(m.get("semanal", 0)),
            "meta_mensal": float(m.get("mensal", 0)),
        })

    resultado.sort(key=lambda v: v["mes"], reverse=True)
    grupo_metas = metas.get("grupo", {})

    return jsonify({
        "agora": agora_br().isoformat(timespec="seconds"),
        "grupo": {
            "hoje": round(grupo_hoje, 2),
            "semana": round(grupo_semana, 2),
            "mes": round(grupo_mes, 2),
            "meta_diaria": float(grupo_metas.get("diaria", 0)),
            "meta_semanal": float(grupo_metas.get("semanal", 0)),
            "meta_mensal": float(grupo_metas.get("mensal", 0)),
        },
        "vendedores": resultado,
    })


# ---------- Área do gestor ----------

@app.route("/api/admin/meses-fechados", methods=["GET"])
def api_admin_listar_meses_fechados():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    return jsonify(carregar_meses_fechados())


@app.route("/api/admin/meses-fechados", methods=["POST"])
def api_admin_alterar_mes_fechado():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    body = request.get_json(force=True)
    mes = (body.get("mes") or "").strip()
    fechar = bool(body.get("fechar"))
    if len(mes) != 7 or mes[4] != "-":
        return jsonify({"erro": "Mês inválido."}), 400

    meses = set(carregar_meses_fechados())
    if fechar:
        meses.add(mes)
    else:
        meses.discard(mes)
    salvar_meses_fechados(list(meses))
    return jsonify({"ok": True, "meses_fechados": sorted(meses)})


@app.route("/api/admin/log-acessos")
def api_admin_log_acessos():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    log = ler_json(LOG_ACESSOS_FILE, [])
    return jsonify(list(reversed(log))[:200])


@app.route("/api/admin/log-acoes")
def api_admin_log_acoes():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    log = ler_json(LOG_ACOES_FILE, [])
    return jsonify(list(reversed(log))[:200])


@app.route("/api/admin/login", methods=["POST"])
def api_admin_login():
    body = request.get_json(force=True)
    senha = body.get("senha") or ""
    if excedeu_tentativas_login("admin"):
        return jsonify({"erro": f"Muitas tentativas erradas. Aguarde {LOGIN_JANELA_MINUTOS} minutos e tente de novo."}), 429
    cred = carregar_credenciais()
    if senha != cred.get("admin_senha"):
        registrar_acesso("admin", False)
        return jsonify({"erro": "Senha incorreta."}), 401
    session["admin"] = True
    registrar_acesso("admin", True)
    return jsonify({"ok": True})


def _hash_codigo(codigo: str) -> str:
    return hashlib.sha256(codigo.encode()).hexdigest()


@app.route("/api/admin/gerar-codigo-recuperacao", methods=["POST"])
def api_admin_gerar_codigo_recuperacao():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    body = request.get_json(silent=True) or {}
    codigo_escolhido = (body.get("codigo") or "").strip().upper()
    if codigo_escolhido:
        if len(codigo_escolhido) < 6:
            return jsonify({"erro": "O código precisa ter pelo menos 6 caracteres."}), 400
        codigo = codigo_escolhido
    else:
        codigo = "-".join(secrets.token_hex(2).upper() for _ in range(2))
    cred = carregar_credenciais()
    cred["recuperacao_hash"] = _hash_codigo(codigo)
    cred["recuperacao_gerado_em"] = agora_br().isoformat(timespec="seconds")
    escrever_json(CREDENCIAIS_FILE, cred)
    return jsonify({"codigo": codigo})


@app.route("/api/recuperar-senha-admin", methods=["POST"])
def api_recuperar_senha_admin():
    if excedeu_tentativas_login("admin_recuperacao"):
        return jsonify({"erro": f"Muitas tentativas erradas. Aguarde {LOGIN_JANELA_MINUTOS} minutos e tente de novo."}), 429

    body = request.get_json(force=True)
    codigo = (body.get("codigo") or "").strip().upper()
    nova_senha = body.get("nova_senha") or ""

    cred = carregar_credenciais()
    hash_salvo = cred.get("recuperacao_hash")
    if not hash_salvo or _hash_codigo(codigo) != hash_salvo:
        registrar_acesso("admin_recuperacao", False)
        return jsonify({"erro": "Código inválido."}), 401

    if len(nova_senha) < 4:
        return jsonify({"erro": "A nova senha precisa ter pelo menos 4 caracteres."}), 400

    cred["admin_senha"] = nova_senha
    cred.pop("recuperacao_hash", None)
    cred.pop("recuperacao_gerado_em", None)
    escrever_json(CREDENCIAIS_FILE, cred)
    registrar_acesso("admin_recuperacao", True)
    return jsonify({"ok": True})


@app.route("/api/admin/logout", methods=["POST"])
def api_admin_logout():
    session.pop("admin", None)
    return jsonify({"ok": True})


def exigir_admin():
    return bool(session.get("admin"))


@app.route("/api/admin/me")
def api_admin_me():
    return jsonify({"logado": exigir_admin()})


@app.route("/api/admin/resumo")
def api_admin_resumo():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401

    hoje = hoje_br().isoformat()
    de = request.args.get("de", f"{hoje[:7]}-01")
    ate = request.args.get("ate", f"{hoje[:7]}-31")
    filtro_vendedor = request.args.get("vendedor_id") or None

    vendedores = carregar_vendedores()
    vendas = carregar_vendas_todos(vendedores)

    ids_alvo = [filtro_vendedor] if filtro_vendedor in vendedores else list(vendedores.keys())
    mes_unico = de[:7] if de[:7] == ate[:7] else None

    por_vendedor = {vid: [] for vid in vendedores}
    for vid_venda, v in vendas.items():
        if not (de <= v["data"] <= ate):
            continue
        vendedor_id = v["vendedor_id"]
        if vendedor_id not in por_vendedor:
            por_vendedor[vendedor_id] = []
        por_vendedor[vendedor_id].append({**v, "id": vid_venda})

    metas_todas = carregar_metas()
    resultado = []
    total_geral = 0.0
    comissao_geral = 0.0
    qtd_vendas_geral = 0
    serie_por_mes = {}
    serie_por_dia = {}

    for vid in sorted(ids_alvo, key=lambda x: vendedores[x]["nome"]):
        info = vendedores[vid]
        # Conta oculta (gestor/teste) sai das listas — MAS so enquanto nao tem
        # venda no periodo. Se um dia tiver, aparece: dinheiro nunca some da tela.
        if (info.get("oculto") and not por_vendedor.get(vid)
                and not info.get("overrides") and vid != filtro_vendedor):
            continue
        # Quem e da expedicao nao vende: fora das listas de venda e comissao,
        # a nao ser que tenha lancamento (dinheiro nunca some de tela).
        if info.get("perfil") == "expedicao" and not por_vendedor.get(vid):
            continue
        lista_vendas = [v for v in por_vendedor.get(vid, []) if v.get("tipo", "venda") == "venda"]
        lista_vendas.sort(key=lambda v: v["data"], reverse=True)
        lista_bonus = [v for v in por_vendedor.get(vid, []) if v.get("tipo") == "bonus"]
        lista_bonus.sort(key=lambda v: v["data"], reverse=True)

        for v in lista_vendas:
            chave = v["data"][:7]
            serie_por_mes[chave] = serie_por_mes.get(chave, 0.0) + valor_liquido(v)
            serie_por_dia[v["data"]] = serie_por_dia.get(v["data"], 0.0) + valor_liquido(v)

        calc = calcular_comissao(vid, de, ate, vendedores, vendas)
        total_geral += calc["total_vendido"]
        comissao_geral += calc["comissao"]
        qtd_vendas_geral += len(lista_vendas)

        confirmado_em = None
        if mes_unico:
            confirmado_em = carregar_confirmacoes(vid).get(mes_unico)

        resultado.append({
            "id": vid,
            "nome": info["nome"],
            "percentual": calc["percentual"],
            "total_vendido": calc["total_vendido"],
            "comissao_propria": calc["comissao_propria"],
            "overrides": calc["overrides"],
            "comissao": calc["comissao"],
            "total_bonus": calc["total_bonus"],
            "qtd_vendas": len(lista_vendas),
            "vendas": lista_vendas,
            "bonus": lista_bonus,
            "confirmado_em": confirmado_em,
            "meta_mensal": float(metas_vendedor(vid, metas_todas).get("mensal", 0) or 0),
        })

    resultado.sort(key=lambda r: r["total_vendido"], reverse=True)
    serie_mensal = [
        {"mes": mes, "total_vendido": round(valor, 2)}
        for mes, valor in sorted(serie_por_mes.items())
    ]
    # Num periodo dentro de um mes so, o grafico por mes vira uma barra unica e
    # nao diz nada — nesse caso mandamos o dia a dia.
    serie_diaria = [
        {"data": dia, "total_vendido": round(valor, 2)}
        for dia, valor in sorted(serie_por_dia.items())
    ] if mes_unico else []
    ticket_medio = round(total_geral / qtd_vendas_geral, 2) if qtd_vendas_geral else 0.0

    # Ritmo do mes: sem isso um vendedor com 60% da meta no dia 10 parece
    # atrasado, quando na verdade esta muito a frente.
    ritmo = None
    if mes_unico:
        dias_no_mes = calendar.monthrange(int(mes_unico[:4]), int(mes_unico[5:7]))[1]
        hoje_data = hoje_br()
        if mes_unico == hoje_data.isoformat()[:7]:
            dias_corridos = hoje_data.day
        else:
            dias_corridos = dias_no_mes          # mes fechado
        ritmo = {"dias_no_mes": dias_no_mes,
                 "dias_corridos": dias_corridos,
                 "pct_do_mes": round(100 * dias_corridos / dias_no_mes)}

    # Vendas de marketplace somadas no MESMO periodo do filtro, a partir da
    # serie diaria que o sincronizador acumula. Lista, nao campo unico: quando
    # a Shopee chegar, e mais um item aqui e o painel ja mostra. Filtrando um
    # vendedor especifico o bloco sai — marketplace nao e de ninguem do time.
    marketplaces = []
    # Canal que EXISTE no sistema mas nao tem numero no periodo pedido entra
    # aqui, nao no silencio: sumir da tela faz o gestor achar que o total ja
    # inclui aquele canal. Foi o que aconteceu com a Shopee em agosto.
    marketplaces_ausentes = []
    if not filtro_vendedor:
        ml = ler_json(resolver_pasta_dados() / "ml_conta.json", None) or {}
        serie_ml = (ml.get("vendas") or {}).get("serie_dia") or {}
        desde_ml = (ml.get("vendas") or {}).get("serie_desde") or "9999"
        soma = round(sum(x.get("total", 0) for d, x in serie_ml.items() if de <= d <= ate), 2)
        qtd = sum(x.get("qtd", 0) for d, x in serie_ml.items() if de <= d <= ate)

        # Venda que o time lanca com canal "Mercado Livre" foi paga pelo
        # checkout do ML — ela JA esta nos pagamentos da serie acima. Conta uma
        # vez so: fica inteira no comercial (comissao e meta do vendedor nao
        # mudam) e sai da fatia do ML no consolidado. Confirmado pelo gestor em
        # 28/08/2026.
        dup_total, dup_qtd = 0.0, 0
        for v_ in vendas.values():
            if (v_.get("tipo", "venda") == "venda" and de <= v_["data"] <= ate
                    and str(v_.get("canal") or "").startswith("Mercado Livre")):
                dup_total += valor_liquido(v_)
                dup_qtd += 1
        dup_total = round(dup_total, 2)

        if qtd:
            marketplaces.append({
                "id": "mercado_livre", "nome": "Mercado Livre",
                "total": round(max(0.0, soma - dup_total), 2),
                "qtd": max(0, qtd - dup_qtd),
                "descontado_comercial": {"total": dup_total, "qtd": dup_qtd},
                # Serie comeca em serie_desde: periodo pedido antes disso vem
                # incompleto, e a tela avisa em vez de fingir que ML nao vendia.
                "cobre_periodo": desde_ml <= de,
            })

        # Site proprio: serie diaria, mesmo tratamento do ML. Venda direta,
        # sem passar por vendedor — por isso nao ha o que descontar aqui.
        st = ler_json(resolver_pasta_dados() / "site_conta.json", None) or {}
        serie_st = (st.get("vendas") or {}).get("serie_dia") or {}
        soma_t = round(sum(x.get("total", 0) for d, x in serie_st.items() if de <= d <= ate), 2)
        qtd_t = sum(x.get("qtd", 0) for d, x in serie_st.items() if de <= d <= ate)
        if qtd_t:
            marketplaces.append({
                "id": "site", "nome": "Site próprio",
                "total": soma_t, "qtd": qtd_t,
                "cobre_periodo": ((st.get("vendas") or {}).get("serie_desde") or "9999") <= de,
            })

        # Shopee vem do relatorio mensal do Seller Centre (importar_shopee_stats):
        # mes cheio dentro do filtro soma exato; mes cortado no meio rateia por
        # dia, o mesmo criterio do Meta e da agencia. Quando a API entrar com
        # serie diaria, este bloco muda de fonte e o painel nem percebe.
        shp = ler_json(resolver_pasta_dados() / "shopee_conta.json", None) or {}
        serie_shp = (shp.get("vendas") or {}).get("serie_mes") or {}
        serie_shp_dia = (shp.get("vendas") or {}).get("serie_dia") or {}
        soma_s, qtd_s, rateado = 0.0, 0.0, False
        d_de, d_ate = date.fromisoformat(de), date.fromisoformat(ate)

        # Dia a dia primeiro (exportacao de periodo curto): soma exata, sem
        # rateio. O mes que tem cobertura diaria ignora a linha mensal logo
        # abaixo — senao a mesma venda entraria duas vezes.
        meses_com_dia = set()
        for dia_chave, dd in serie_shp_dia.items():
            meses_com_dia.add(dia_chave[:7])
            if de <= dia_chave <= ate:
                soma_s += dd.get("total", 0)
                qtd_s += dd.get("qtd", 0)

        for mes_chave, mm in serie_shp.items():
            if mes_chave in meses_com_dia:
                continue
            ano, mes_n = int(mes_chave[:4]), int(mes_chave[5:7])
            dias_mes = calendar.monthrange(ano, mes_n)[1]
            ini = max(d_de, date(ano, mes_n, 1))
            fim = min(d_ate, date(ano, mes_n, dias_mes))
            if ini > fim:
                continue
            dentro = (fim - ini).days + 1
            fracao = dentro / dias_mes
            if fracao < 1:
                rateado = True
            soma_s += mm.get("total", 0) * fracao
            qtd_s += mm.get("qtd", 0) * fracao
        if qtd_s >= 0.5:
            marketplaces.append({
                "id": "shopee", "nome": "Shopee",
                "total": round(soma_s, 2), "qtd": int(round(qtd_s)),
                "rateado": rateado,
                "cobre_periodo": (min({**serie_shp, **{k[:7]: 1 for k in serie_shp_dia}})
                                  <= de[:7] and ate[:7]
                                  <= max({**serie_shp, **{k[:7]: 1 for k in serie_shp_dia}})),
            })
        elif serie_shp or serie_shp_dia:
            # A serie diaria e mais precisa que a mensal pra dizer ate quando o
            # dado vai; misturar as duas gerava "ate 08/2026" quando na verdade
            # havia dado diario ate 27/08.
            if serie_shp_dia:
                ult = max(serie_shp_dia)
                ate_quando = f"{ult[8:10]}/{ult[5:7]}"
            else:
                ult = max(serie_shp)
                ate_quando = f"{ult[5:7]}/{ult[:4]}"
            marketplaces_ausentes.append({
                "id": "shopee", "nome": "Shopee",
                "motivo": (f"sem venda neste período" if de <= ult
                           else f"planilha importada — dados até {ate_quando}"),
            })

        # Mesma regra pros canais de serie diaria: existe historico, mas nada
        # dentro do periodo filtrado.
        # "Nao vendeu" e "ainda nao sincronizamos esse dia" sao coisas
        # diferentes, e confundir as duas faz o gestor achar que o dia foi
        # zerado quando na verdade o dado nao chegou. O ultimo dia da serie
        # decide qual das duas e.
        if not qtd and serie_ml:
            ultimo_ml = max(serie_ml)
            marketplaces_ausentes.append({
                "id": "mercado_livre", "nome": "Mercado Livre",
                "motivo": ("sem venda registrada neste período"
                           if de <= ultimo_ml
                           else f"ainda não sincronizado — dados até {ultimo_ml[8:10]}/{ultimo_ml[5:7]}"),
            })
        if not qtd_t and serie_st:
            ultimo_st = max(serie_st)
            marketplaces_ausentes.append({
                "id": "site", "nome": "Site próprio",
                "motivo": ("sem pedido pago neste período"
                           if de <= ultimo_st
                           else f"leitura manual — dados até {ultimo_st[8:10]}/{ultimo_st[5:7]}"),
            })

    return jsonify({
        "de": de,
        "ate": ate,
        "mes_unico": mes_unico,
        "meta_grupo": float((metas_todas.get("grupo") or {}).get("mensal", 0) or 0),
        "ritmo": ritmo,
        "vendedor_id": filtro_vendedor,
        "vendedores": resultado,
        "total_geral": round(total_geral, 2),
        "comissao_geral": round(comissao_geral, 2),
        "qtd_vendas_geral": qtd_vendas_geral,
        "ticket_medio": ticket_medio,
        "serie_mensal": serie_mensal,
        "serie_diaria": serie_diaria,
        "marketplaces": marketplaces,
        "marketplaces_ausentes": marketplaces_ausentes,
    })


def _nome_aba_excel(nome: str) -> str:
    """Aba do Excel não aceita \\ / ? * [ ] : nem mais de 31 caracteres."""
    limpo = re.sub(r'[\\/?*\[\]:]', "-", nome)
    return limpo[:31] or "Vendedor"


# ============================================================
# Auditoria de comissões — conferência das vendas contra o caixa
# ============================================================
# O que este painel NÃO faz: bater venda com extrato automaticamente. O portal
# não tem dado bancário nenhum — nem extrato, nem conciliação, e `canal` está
# vazio em mais da metade das vendas. Fingir um "confere/não confere"
# automático seria inventar uma certeza que não existe.
#
# O que ele faz: escolher o que vale a pena conferir na mão, guardar o
# veredito e mostrar quanto do faturamento já passou por conferência.
#
# A amostra é estável de propósito: sorteio semeado pelo id da venda e pelo mês,
# então recarregar a tela devolve exatamente as mesmas vendas. Auditoria em que
# a amostra muda a cada F5 não é auditoria — dá pra ficar re-sorteando até vir
# um conjunto confortável.

STATUS_AUDITORIA = {
    "conferida": "Confere com o caixa",
    "divergente": "Não bate",
    "nao_achei": "Não encontrei no caixa",
}

# Peso de cada sinal na hora de decidir o que entra na amostra. Não é
# probabilidade de fraude — é "isto merece um olhar antes daquilo".
SINAIS_AUDITORIA = {
    "duplicata": ("Possível duplicata", 5),
    "lancada_tarde": ("Lançada dias depois da data", 4),
    "editada": ("Editada depois de criada", 3),
    "valor_alto": ("Entre as maiores do mês", 2),
    "fim_de_semana": ("Lançada em fim de semana", 1),
}
# `sem_canal` foi removido da pontuação de propósito: dispara em 56% das vendas.
# Sinal que acende na maioria dos casos não separa nada — só faria a amostra
# virar "quase tudo" e o gestor parar de olhar. Vira indicador de qualidade do
# cadastro, que é o problema real ali.


def _chave_sorteio(venda_id: str, mes: str) -> int:
    """Ordem estável: mesma venda, mesmo mês, mesma posição — sempre."""
    return int(hashlib.sha256(f"{mes}:{venda_id}".encode()).hexdigest()[:12], 16)


def carregar_auditoria() -> dict:
    return ler_json(resolver_pasta_dados() / "auditoria.json", None) or {}


def _sinais_da_venda(v, contagem_dup, corte_alto):
    sinais = []
    chave = (v["data"], (v.get("produto") or "").strip().lower(), round(v["valor"], 2))
    if contagem_dup.get(chave, 0) > 1:
        sinais.append("duplicata")
    criado = v.get("criado_em")
    if criado:
        try:
            dias = (parse_dt_tolerante(criado).date() - date.fromisoformat(v["data"])).days
            if dias > 3:
                sinais.append("lancada_tarde")
        except (ValueError, TypeError):
            pass
    # Só `editado_em` marca edição. O log de ações não guarda o id da venda,
    # então não dá pra cruzar de volta — inventar esse vínculo por produto e
    # valor acertaria umas e erraria outras.
    if v.get("editado_em"):
        sinais.append("editada")
    if corte_alto and v["valor"] >= corte_alto:
        sinais.append("valor_alto")
    if date.fromisoformat(v["data"]).weekday() >= 5:
        sinais.append("fim_de_semana")
    return sinais


@app.route("/api/admin/auditoria")
def api_admin_auditoria():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401

    # Periodo por data, nao por mes: conferencia raramente casa com o calendario
    # — e mais comum querer "a semana passada" ou "do dia 10 ao 20". `mes` ainda
    # e aceito pra nao quebrar link salvo.
    mes = request.args.get("mes") or ""
    de = request.args.get("de") or ""
    ate = request.args.get("ate") or ""
    if not (de and ate):
        base = mes or hoje_br().isoformat()[:7]
        ultimo = calendar.monthrange(int(base[:4]), int(base[5:7]))[1]
        de, ate = f"{base}-01", f"{base}-{ultimo:02d}"

    try:
        tamanho = max(5, min(100, int(request.args.get("tamanho") or 20)))
    except ValueError:
        tamanho = 20
    filtro_vendedor = request.args.get("vendedor") or ""
    busca = (request.args.get("busca") or "").strip().lower()

    vendedores = carregar_vendedores()
    nomes_vend = {vid: v["nome"] for vid, v in vendedores.items()}

    def casa_busca(v):
        """Procura no que a pessoa lembra: nome da peca, do carro, SKU, canal,
        vendedor — ou o valor. Numero digitado casa tanto por texto ("1.200")
        quanto por comparacao, entao '1200' acha 1.200,00 e 1.200,50."""
        if not busca:
            return True
        valor = float(v.get("valor") or 0)
        alvo = _achatar_canal(" ".join(str(x) for x in (
            v.get("produto"), v.get("sku"), v.get("canal"),
            # A grafia antiga do canal continua achavel: quem sempre digitou
            # "ITAU" nao pode perder a venda porque ela virou "Itaú".
            v.get("canal_original"),
            nomes_vend.get(v.get("vendedor_id"), ""), v.get("data"),
            # Tres jeitos de escrever o mesmo valor. Quem procura digita como le
            # na tela ("1.200,00"), como pensa ("1200") ou como o teclado dá.
            f"{valor:.2f}".replace(".", ","), f"{valor:,.2f}".replace(",", "."),
            f"{int(valor)}",
        ) if x))
        # Todas as palavras precisam aparecer (sem acento dos dois lados):
        # "farol sorento" nao traz farol de outro carro, "itau" acha "Itaú".
        return all(termo in alvo for termo in _achatar_canal(busca).split())

    todas = carregar_vendas_todos(vendedores)
    vendas = [{**v, "id": vid} for vid, v in todas.items()
              if v.get("tipo", "venda") == "venda" and de <= v["data"] <= ate
              and (not filtro_vendedor or v["vendedor_id"] == filtro_vendedor)
              and casa_busca(v)]
    if not vendas:
        return jsonify({
            "mes": mes, "de": de, "ate": ate, "busca": busca, "vazio": True,
            "modo": request.args.get("modo") or "amostra", "foco": "",
            "amostra": [], "revisadas": [], "total_lista": [],
            "rotulos": STATUS_AUDITORIA, "sinais": {k: x[0] for k, x in SINAIS_AUDITORIA.items()},
            "total": {"vendas": 0, "valor": 0.0},
            "cobertura": {"conferidas": 0, "divergentes": 0, "valor_conferido": 0.0,
                          "pct_qtd": 0, "pct_valor": 0},
            "com_sinal": 0,
            "qualidade": {"sem_canal": 0, "sem_sku": 0, "sem_criado_em": 0},
            "nomes": {k: v["nome"] for k, v in vendedores.items()},
            "filtro": {"vendedor": filtro_vendedor, "tamanho": tamanho},
            "vendedores": [{"id": k, "nome": v["nome"]}
                           for k, v in sorted(vendedores.items(),
                                              key=lambda kv: kv[1]["nome"])]})

    contagem_dup = {}
    for v in vendas:
        chave = (v["data"], (v.get("produto") or "").strip().lower(), round(v["valor"], 2))
        contagem_dup[chave] = contagem_dup.get(chave, 0) + 1

    ordenados = sorted(vendas, key=lambda v: -v["valor"])
    corte_alto = ordenados[max(0, len(ordenados) // 10 - 1)]["valor"] if len(ordenados) >= 10 else None

    for v in vendas:
        v["sinais"] = _sinais_da_venda(v, contagem_dup, corte_alto)
        v["risco"] = sum(SINAIS_AUDITORIA[s][1] for s in v["sinais"])

    marcas = carregar_auditoria()
    for v in vendas:
        m = marcas.get(v["id"]) or {}
        v["status"] = m.get("status")
        v["obs"] = m.get("obs")
        v["conferida_em"] = m.get("em")

    # Quem tem sinal entra antes; dentro do mesmo risco, o sorteio estável
    # decide. Assim a amostra cobre o que chama atenção sem virar uma lista só
    # dos casos estranhos — venda normal também precisa ser conferida, senão a
    # auditoria não diz nada sobre o conjunto.
    ja_marcadas = [v for v in vendas if v["status"]]
    candidatas = [v for v in vendas if not v["status"]]
    candidatas.sort(key=lambda v: (-v["risco"], _chave_sorteio(v["id"], de + ate)))
    com_sinal = [v for v in candidatas if v["risco"] > 0]
    sem_sinal = [v for v in candidatas if v["risco"] == 0]

    metade = max(1, tamanho // 2)
    amostra = com_sinal[:metade] + sem_sinal[:tamanho - min(metade, len(com_sinal))]
    amostra.sort(key=lambda v: (-v["risco"], v["data"]))

    def enxuto(v):
        return {k: v.get(k) for k in ("id", "data", "produto", "valor", "canal", "sku",
                                      "vendedor_id", "criado_em", "sinais", "risco",
                                      "status", "obs", "conferida_em")}

    conferidas = [v for v in vendas if v["status"] == "conferida"]
    divergentes = [v for v in vendas if v["status"] in ("divergente", "nao_achei")]
    total_mes = round(sum(v["valor"] for v in vendas), 2)
    valor_conferido = round(sum(v["valor"] for v in conferidas), 2)

    # Modo total: a planilha inteira do mês, sem sorteio. Serve pra fechar o mês
    # de ponta a ponta; a amostra serve pra rodar rápido no meio do mês. Os dois
    # gravam no mesmo lugar, então o que for conferido num aparece no outro.
    #
    # `foco` é o clique num dos números do topo: mostra exatamente aquelas
    # vendas. Vem do servidor e não da tela porque no modo amostra a tela só tem
    # as vendas sorteadas — filtrar ali devolveria menos do que o número promete,
    # e um contador que não bate com a lista é pior do que não ter contador.
    FOCOS = {
        "conferidas": lambda v: v["status"] == "conferida",
        "divergentes": lambda v: v["status"] in ("divergente", "nao_achei"),
        "sinal": lambda v: v["risco"] > 0,
    }
    foco = request.args.get("foco") or ""
    lista_total = None
    if foco in FOCOS:
        lista_total = [enxuto(v) for v in sorted(
            (x for x in vendas if FOCOS[foco](x)), key=lambda v: (v["data"], -v["valor"]))]
    elif request.args.get("modo") == "total":
        lista_total = [enxuto(v) for v in sorted(
            vendas, key=lambda v: (v["data"], -v["valor"]))]
    elif request.args.get("modo") == "pendentes":
        # O que ainda nao passou por ninguem, do maior valor pro menor. E a
        # pergunta de quem abre a tela no meio do mes: "o que falta conferir?".
        # Ordem por valor, nao por data: se o tempo acabar, o que ficou de fora
        # e o que menos importa.
        lista_total = [enxuto(v) for v in sorted(
            (x for x in vendas if not x["status"]),   # sem marca = ninguem olhou
            key=lambda v: -v["valor"])]

    return jsonify({
        "mes": mes,
        "de": de,
        "ate": ate,
        "busca": busca,
        "modo": request.args.get("modo") or "amostra",
        "foco": foco if foco in FOCOS else "",
        "total_lista": lista_total,
        "rotulos": STATUS_AUDITORIA,
        "sinais": {k: v[0] for k, v in SINAIS_AUDITORIA.items()},
        "vendedores": [{"id": k, "nome": v["nome"]}
                       for k, v in sorted(vendedores.items(), key=lambda kv: kv[1]["nome"])],
        "nomes": {k: v["nome"] for k, v in vendedores.items()},
        "filtro": {"vendedor": filtro_vendedor, "tamanho": tamanho},
        "total": {"vendas": len(vendas), "valor": total_mes},
        "cobertura": {
            "conferidas": len(conferidas),
            "divergentes": len(divergentes),
            "valor_conferido": valor_conferido,
            "pct_qtd": round(100 * len(conferidas) / len(vendas), 1),
            "pct_valor": round(100 * valor_conferido / total_mes, 1) if total_mes else 0,
        },
        "com_sinal": len(com_sinal) + sum(1 for v in ja_marcadas if v["risco"] > 0),
        "qualidade": {
            "sem_canal": sum(1 for v in vendas if not (v.get("canal") or "").strip()),
            "sem_sku": sum(1 for v in vendas if not (v.get("sku") or "").strip()),
            "sem_criado_em": sum(1 for v in vendas if not v.get("criado_em")),
        },
        "amostra": [enxuto(v) for v in amostra],
        "revisadas": [enxuto(v) for v in sorted(ja_marcadas, key=lambda v: v.get("conferida_em") or "",
                                                reverse=True)[:40]],
    })


@app.route("/api/admin/auditoria/<venda_id>", methods=["POST"])
def api_admin_auditoria_marcar(venda_id):
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    corpo = request.get_json(silent=True) or {}
    novo = (corpo.get("status") or "").strip()
    if novo and novo not in STATUS_AUDITORIA:
        return jsonify({"erro": "status inválido"}), 400

    # Só aceita venda que existe: id chutado viraria uma marca órfã que conta
    # como conferida e infla a cobertura.
    vendedores = carregar_vendedores()
    if venda_id not in carregar_vendas_todos(vendedores):
        return jsonify({"erro": "Venda não encontrada."}), 404

    marcas = carregar_auditoria()
    if not novo:
        marcas.pop(venda_id, None)
    else:
        marcas[venda_id] = {"status": novo,
                            "obs": (corpo.get("obs") or "").strip()[:400],
                            "em": agora_br().isoformat(timespec="seconds")}
    escrever_json(resolver_pasta_dados() / "auditoria.json", marcas)
    return jsonify({"ok": True})


# ============================================================
# Recursos humanos
# ============================================================
# Quatro coleções separadas em vez de uma ficha gigante por pessoa: ausência,
# documento e ocorrência acontecem o tempo todo e em volume diferente do
# cadastro. Juntas num objeto só, cada anotação reescreveria a ficha inteira —
# e duas pessoas mexendo ao mesmo tempo perderiam trabalho.
#
# O cadastro guarda CPF e salário por decisão do gestor. Ficam aqui os dois
# cuidados que isso exige: não saem em nenhuma rota que não seja a do gestor, e
# a tela avisa que são dados restritos. O elo fraco continua sendo a senha única
# do portal — vale trocar por senha por pessoa quando der.

# Setores como a empresa se organiza de fato, na ordem da planilha de folha.
SETORES_RH = ["Comercial", "Anúncios", "Cadastro", "Desmontagem", "Expedição",
              "Estoque", "Higienização", "Pátio", "Gerência", "Administrativo",
              "Marketing", "Outro"]
CONTRATOS_RH = ["CLT", "PJ", "Estágio", "Temporário", "Sócio"]
SITUACOES_RH = {"ativo": "Ativo", "afastado": "Afastado", "desligado": "Desligado"}
TIPOS_AUSENCIA = {"ferias": "Férias", "falta": "Falta", "atestado": "Atestado",
                  "licenca": "Licença", "folga": "Folga"}
TIPOS_DOCUMENTO = ["ASO", "CNH", "Contrato", "Certificação", "Exame periódico", "Outro"]
TIPOS_OCORRENCIA = {"elogio": "Elogio", "advertencia": "Advertência",
                    "feedback": "Conversa de feedback", "suspensao": "Suspensão",
                    "nota": "Anotação"}

RH_COLECOES = ("colaboradores", "ausencias", "documentos", "ocorrencias")


def _rh_ler(nome: str) -> dict:
    return ler_json(resolver_pasta_dados() / f"rh_{nome}.json", None) or {}


def _rh_gravar(nome: str, dados: dict) -> None:
    escrever_json(resolver_pasta_dados() / f"rh_{nome}.json", dados)


def _rh_texto(v, limite=200) -> str:
    return str(v or "").strip()[:limite]


def _rh_data(v) -> str:
    """Aceita só data ISO. Campo de data vazio é normal (nem todo mundo tem
    tudo preenchido); data mal formada não entra e vira vazio."""
    t = _rh_texto(v, 10)
    try:
        date.fromisoformat(t)
        return t
    except ValueError:
        return ""


def _rh_num(v):
    """Campo de dinheiro vazio é normal e vira None, não zero: zero diz
    "ganha zero", vazio diz "não sei"."""
    t = str(v if v is not None else "").strip()
    if t in ("", "None"):
        return None
    try:
        return round(float(t.replace(",", ".")), 2)
    except ValueError:
        return None


def _idade_ou_tempo(desde: str):
    """Anos completos entre uma data e hoje. Serve pra idade e pra tempo de
    casa, que é a mesma conta."""
    if not desde:
        return None
    try:
        d = date.fromisoformat(desde)
    except ValueError:
        return None
    hoje = hoje_br()
    anos = hoje.year - d.year - ((hoje.month, hoje.day) < (d.month, d.day))
    return max(0, anos)


def _dias_ate(quando: str):
    if not quando:
        return None
    try:
        return (date.fromisoformat(quando) - hoje_br()).days
    except ValueError:
        return None


@app.route("/api/admin/rh")
def api_admin_rh():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401

    colaboradores = _rh_ler("colaboradores")
    ausencias = _rh_ler("ausencias")
    documentos = _rh_ler("documentos")
    ocorrencias = _rh_ler("ocorrencias")
    hoje = hoje_br().isoformat()
    mes_atual = hoje[5:7]

    lista = []
    for cid, c in colaboradores.items():
        item = {**c, "id": cid}
        item["tempo_casa"] = _idade_ou_tempo(c.get("admissao"))
        item["idade"] = _idade_ou_tempo(c.get("nascimento"))
        # Aniversário do mês corrente, independente do ano de nascimento.
        item["faz_aniversario_no_mes"] = bool(c.get("nascimento")) and c["nascimento"][5:7] == mes_atual
        item["fora_hoje"] = any(
            a["colaborador_id"] == cid and a.get("de", "") <= hoje <= (a.get("ate") or a.get("de", ""))
            for a in ausencias.values())
        lista.append(item)
    # Ordena pelo apelido quando existe: e o nome que a lista mostra, e procurar
    # "Nego" na letra R nao ajuda ninguem.
    lista.sort(key=lambda c: ((c.get("situacao") or "ativo") != "ativo",
                              (c.get("apelido") or c.get("nome") or "").lower()))

    nomes = {cid: c.get("nome", "?") for cid, c in colaboradores.items()}

    def enfeitar(colecao, extra=None):
        saida = []
        for xid, x in colecao.items():
            item = {**x, "id": xid, "colaborador": nomes.get(x.get("colaborador_id"), "(removido)")}
            if extra:
                extra(item)
            saida.append(item)
        return saida

    lista_ausencias = enfeitar(ausencias)
    lista_ausencias.sort(key=lambda a: a.get("de") or "", reverse=True)

    def marcar_vencimento(d):
        d["dias"] = _dias_ate(d.get("vence_em"))
    lista_documentos = enfeitar(documentos, marcar_vencimento)
    # Vencido e a vencer primeiro: documento com prazo longo não precisa de
    # atenção, e ordenar por data joga justamente os urgentes pro fim.
    lista_documentos.sort(key=lambda d: (d["dias"] is None, d["dias"] if d["dias"] is not None else 0))

    lista_ocorrencias = enfeitar(ocorrencias)
    lista_ocorrencias.sort(key=lambda o: o.get("data") or "", reverse=True)

    ativos = [c for c in lista if (c.get("situacao") or "ativo") == "ativo"]
    por_setor = {}
    for c in ativos:
        por_setor[c.get("setor") or "Sem setor"] = por_setor.get(c.get("setor") or "Sem setor", 0) + 1

    # Custo mensal direto: o que sai do caixa por cada pessoa todo mes. Nao e a
    # folha contabil (falta encargo, ferias, 13o), e o rodape da tela diz isso.
    def custo(c):
        return sum(float(c.get(campo) or 0) for campo in ("salario", "vt", "bonificacao"))
    custo_folha = sum(custo(c) for c in ativos)
    sem_salario = [c["nome"] for c in ativos if not c.get("salario")]

    tempos = [c["tempo_casa"] for c in ativos if c["tempo_casa"] is not None]
    ano = hoje[:4]
    desligados_ano = [c for c in lista
                      if c.get("situacao") == "desligado" and (c.get("desligamento") or "").startswith(ano)]
    admitidos_ano = [c for c in lista if (c.get("admissao") or "").startswith(ano)]

    return jsonify({
        "colaboradores": lista,
        "ausencias": lista_ausencias,
        "documentos": lista_documentos,
        "ocorrencias": lista_ocorrencias,
        "opcoes": {
            "setores": SETORES_RH, "contratos": CONTRATOS_RH,
            "situacoes": SITUACOES_RH, "tipos_ausencia": TIPOS_AUSENCIA,
            "tipos_documento": TIPOS_DOCUMENTO, "tipos_ocorrencia": TIPOS_OCORRENCIA,
        },
        "vendedores": [{"id": vid, "nome": v["nome"]}
                       for vid, v in sorted(carregar_vendedores().items(),
                                            key=lambda kv: kv[1]["nome"])],
        "indicadores": {
            "ativos": len(ativos),
            "afastados": sum(1 for c in lista if c.get("situacao") == "afastado"),
            "desligados": sum(1 for c in lista if c.get("situacao") == "desligado"),
            "por_setor": sorted(({"setor": k, "qtd": v} for k, v in por_setor.items()),
                                key=lambda x: -x["qtd"]),
            "tempo_medio_casa": round(sum(tempos) / len(tempos), 1) if tempos else None,
            "custo_folha": round(custo_folha, 2),
            "sem_salario": sem_salario,
            "fora_hoje": [c["nome"] for c in lista if c["fora_hoje"]],
            "aniversariantes": sorted(
                ({"nome": c["nome"], "dia": c["nascimento"][8:10]} for c in lista
                 if c["faz_aniversario_no_mes"] and (c.get("situacao") or "ativo") != "desligado"),
                key=lambda x: x["dia"]),
            "documentos_vencendo": sum(1 for d in lista_documentos
                                       if d["dias"] is not None and d["dias"] <= 30),
            "admitidos_no_ano": len(admitidos_ano),
            "desligados_no_ano": len(desligados_ano),
            # Turnover simples: desligamentos sobre o quadro medio do ano. Com
            # equipe pequena um desligamento move muito o numero — por isso a
            # tela mostra o valor absoluto junto.
            "turnover_ano": (round(100 * len(desligados_ano) / max(1, len(ativos) + len(desligados_ano)), 1)
                             if lista else 0),
        },
    })


@app.route("/api/admin/rh/colaboradores", methods=["POST"])
def api_admin_rh_salvar():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    corpo = request.get_json(silent=True) or {}
    nome = _rh_texto(corpo.get("nome"), 80)
    if not nome:
        return jsonify({"erro": "Informe o nome."}), 400

    dados = _rh_ler("colaboradores")
    cid = _rh_texto(corpo.get("id"), 12) or uuid.uuid4().hex[:12]
    antigo = dados.get(cid, {})

    situacao = corpo.get("situacao") if corpo.get("situacao") in SITUACOES_RH else "ativo"
    ficha = {
        "nome": nome,
        "apelido": _rh_texto(corpo.get("apelido"), 40),
        "cargo": _rh_texto(corpo.get("cargo"), 60),
        "setor": _rh_texto(corpo.get("setor"), 40),
        "contrato": _rh_texto(corpo.get("contrato"), 20),
        "situacao": situacao,
        "admissao": _rh_data(corpo.get("admissao")),
        "desligamento": _rh_data(corpo.get("desligamento")) if situacao == "desligado" else "",
        "motivo_desligamento": (_rh_texto(corpo.get("motivo_desligamento"), 200)
                                if situacao == "desligado" else ""),
        "nascimento": _rh_data(corpo.get("nascimento")),
        "telefone": _rh_texto(corpo.get("telefone"), 30),
        "email": _rh_texto(corpo.get("email"), 80),
        "endereco": _rh_texto(corpo.get("endereco"), 160),
        "emergencia": _rh_texto(corpo.get("emergencia"), 120),
        "cpf": _rh_texto(corpo.get("cpf"), 20),
        "rg": _rh_texto(corpo.get("rg"), 20),
        "salario": _rh_num(corpo.get("salario")),
        "vt": _rh_num(corpo.get("vt")),
        "bonificacao": _rh_num(corpo.get("bonificacao")),
        # Liga a ficha ao vendedor do portal: com isso a tela de RH consegue
        # abrir o desempenho da pessoa sem cadastro duplicado.
        "vendedor_id": _rh_texto(corpo.get("vendedor_id"), 40),
        "obs": _rh_texto(corpo.get("obs"), 500),
        "criado_em": antigo.get("criado_em") or agora_br().isoformat(timespec="seconds"),
        "editado_em": agora_br().isoformat(timespec="seconds"),
    }
    dados[cid] = ficha
    _rh_gravar("colaboradores", dados)
    return jsonify({"ok": True, "id": cid})


@app.route("/api/admin/rh/colaboradores/<cid>", methods=["DELETE"])
def api_admin_rh_remover(cid):
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    dados = _rh_ler("colaboradores")
    if cid not in dados:
        return jsonify({"erro": "Colaborador não encontrado."}), 404
    dados.pop(cid)
    _rh_gravar("colaboradores", dados)
    # Ausência, documento e ocorrência de quem saiu do cadastro viram lixo
    # órfão: aparecem como "(removido)" e sujam os indicadores pra sempre.
    for colecao in ("ausencias", "documentos", "ocorrencias"):
        itens = _rh_ler(colecao)
        sobra = {k: v for k, v in itens.items() if v.get("colaborador_id") != cid}
        if len(sobra) != len(itens):
            _rh_gravar(colecao, sobra)
    return jsonify({"ok": True})


@app.route("/api/admin/rh/<colecao>", methods=["POST"])
def api_admin_rh_registro(colecao):
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    if colecao not in ("ausencias", "documentos", "ocorrencias"):
        return jsonify({"erro": "Coleção inválida."}), 404

    corpo = request.get_json(silent=True) or {}
    cid = _rh_texto(corpo.get("colaborador_id"), 12)
    if cid not in _rh_ler("colaboradores"):
        return jsonify({"erro": "Escolha o colaborador."}), 400

    if colecao == "ausencias":
        de = _rh_data(corpo.get("de"))
        if not de:
            return jsonify({"erro": "Informe a data de início."}), 400
        ate = _rh_data(corpo.get("ate")) or de
        if ate < de:
            return jsonify({"erro": "A data final é anterior à inicial."}), 400
        registro = {"colaborador_id": cid,
                    "tipo": corpo.get("tipo") if corpo.get("tipo") in TIPOS_AUSENCIA else "falta",
                    "de": de, "ate": ate,
                    "dias": (date.fromisoformat(ate) - date.fromisoformat(de)).days + 1,
                    "obs": _rh_texto(corpo.get("obs"), 200)}
    elif colecao == "documentos":
        registro = {"colaborador_id": cid,
                    "tipo": _rh_texto(corpo.get("tipo"), 40) or "Outro",
                    "numero": _rh_texto(corpo.get("numero"), 40),
                    "vence_em": _rh_data(corpo.get("vence_em")),
                    "obs": _rh_texto(corpo.get("obs"), 200)}
    else:
        registro = {"colaborador_id": cid,
                    "tipo": corpo.get("tipo") if corpo.get("tipo") in TIPOS_OCORRENCIA else "nota",
                    "data": _rh_data(corpo.get("data")) or hoje_br().isoformat(),
                    "texto": _rh_texto(corpo.get("texto"), 800)}
        if not registro["texto"]:
            return jsonify({"erro": "Escreva o que aconteceu."}), 400

    registro["criado_em"] = agora_br().isoformat(timespec="seconds")
    itens = _rh_ler(colecao)
    rid = uuid.uuid4().hex[:12]
    itens[rid] = registro
    _rh_gravar(colecao, itens)
    return jsonify({"ok": True, "id": rid})


@app.route("/api/admin/rh/<colecao>/<rid>", methods=["DELETE"])
def api_admin_rh_remover_registro(colecao, rid):
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    if colecao not in ("ausencias", "documentos", "ocorrencias"):
        return jsonify({"erro": "Coleção inválida."}), 404
    itens = _rh_ler(colecao)
    if rid not in itens:
        return jsonify({"erro": "Registro não encontrado."}), 404
    itens.pop(rid)
    _rh_gravar(colecao, itens)
    return jsonify({"ok": True})


# ============================================================
# Meta Bônus
# ============================================================
# O lancamento agora acontece AQUI: o dado bruto (pessoas, lancamentos diarios,
# veiculos desmontados) mora em metas_bonus_dados — a historia veio importada do
# painel-metas, que lancava local. A agregacao por mes que o sincronizador fazia
# roda no proprio portal, entao lancou, apareceu.

SETORES_META = {"anunciante": "Anunciantes", "cadastrador": "Cadastradores"}
# setor da pessoa -> tipo do lancamento, nomes herdados do painel-metas.
TIPO_META = {"anunciante": "anuncio", "cadastrador": "cadastro"}


def _mb_bruto() -> dict:
    d = ler_json(resolver_pasta_dados() / "metas_bonus_dados.json", None) or {}
    d.setdefault("pessoas", {})
    d.setdefault("lancamentos", {})
    d.setdefault("veiculos", {})
    d.setdefault("meta_veiculos", {"meta": 0, "meta_bonus": 0})
    return d


def _mb_gravar(dados: dict) -> None:
    escrever_json(resolver_pasta_dados() / "metas_bonus_dados.json", dados)


def _mb_agregar(dados: dict) -> dict:
    """Producao por pessoa/mes contra meta e bonus — o mesmo agregado que o
    sincronizador antigo montava, agora calculado do dado vivo."""
    producao = {}
    for setor, tipo in TIPO_META.items():
        for l in (dados["lancamentos"].get(tipo) or {}).values():
            chave = (setor, l.get("pessoa_id"), (l.get("data") or "")[:7])
            producao[chave] = producao.get(chave, 0.0) + float(l.get("quantidade") or 0)

    meses_set = ({c[2] for c in producao}
                 | {(v.get("data") or "")[:7] for v in dados["veiculos"].values() if v.get("data")})
    meta_veic = dados["meta_veiculos"]

    por_mes = {}
    for mes in sorted(m for m in meses_set if m):
        setores = {}
        for setor, gente in dados["pessoas"].items():
            linhas = []
            for pid, p in gente.items():
                total = round(producao.get((setor, pid, mes), 0), 2)
                meta = float(p.get("meta") or 0)
                bonus = float(p.get("meta_bonus") or 0)
                # Quem nao lancou nada no mes nao entra: apareceria como 0% e
                # pareceria alguem que trabalhou e nao produziu, quando na
                # verdade nao estava no time naquele mes.
                if total <= 0:
                    continue
                linhas.append({
                    "nome": p.get("nome") or "(sem nome)",
                    "total": total, "meta": meta, "meta_bonus": bonus,
                    "pct": round(100 * total / meta, 1) if meta else None,
                    "bateu_meta": bool(meta) and total >= meta,
                    "bateu_bonus": bool(bonus) and total >= bonus,
                })
            linhas.sort(key=lambda x: -x["total"])
            if linhas:
                setores[setor] = linhas

        do_mes = [v for v in dados["veiculos"].values() if (v.get("data") or "")[:7] == mes]
        por_mes[mes] = {
            "setores": setores,
            "veiculos": {
                "carros": len(do_mes),
                "pecas": round(sum(float(v.get("pecas") or 0) for v in do_mes), 2),
                "meta": float(meta_veic.get("meta") or 0),
                "meta_bonus": float(meta_veic.get("meta_bonus") or 0),
                "lista": sorted(({"data": v.get("data"), "carro": v.get("carro"),
                                  "codigo": v.get("codigo"), "pecas": v.get("pecas") or 0}
                                 for v in do_mes), key=lambda v: v["data"] or ""),
            },
        }
    return por_mes


@app.route("/api/admin/ml-conta")
def api_admin_ml_conta():
    """Saude da conta Mercado Livre: reputacao oficial, pos-venda e Product
    Ads, empurrados por scripts/sincronizar_ml.py a partir do ml-dashboard."""
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    d = ler_json(resolver_pasta_dados() / "ml_conta.json", None)
    if not d:
        return jsonify({"sem_dados": True})
    return jsonify(d)


@app.route("/api/admin/site-conta", methods=["POST"])
def api_admin_site_gravar():
    """Recebe a serie diaria do site proprio (vaapt). O painel de la so filtra
    um dia por vez e nao expoe API de pedidos, entao a leitura das paginas e
    feita no navegador do gestor, ja logado, e o resultado chega aqui.
    Merge por dia: releitura atualiza os dias lidos e preserva os antigos."""
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    corpo = request.get_json(silent=True) or {}
    serie = corpo.get("serie_dia")
    if not isinstance(serie, dict) or not serie:
        return jsonify({"erro": "Série vazia."}), 400

    limpa = {}
    for dia, val in serie.items():
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(dia)):
            return jsonify({"erro": f"Data inválida: {dia}"}), 400
        try:
            limpa[dia] = {"total": round(float(val.get("total") or 0), 2),
                          "qtd": int(val.get("qtd") or 0)}
        except (TypeError, ValueError, AttributeError):
            return jsonify({"erro": f"Valor inválido em {dia}"}), 400

    atual = ler_json(resolver_pasta_dados() / "site_conta.json", None) or {}
    antiga = (atual.get("vendas") or {}).get("serie_dia") or {}
    juntas = {**antiga, **limpa}
    escrever_json(resolver_pasta_dados() / "site_conta.json", {
        "gerado_em": agora_br().isoformat(timespec="seconds"),
        "fonte": corpo.get("fonte") or "painel do site (pedidos pagos)",
        "vendas": {"serie_dia": juntas, "serie_desde": min(juntas)},
    })
    return jsonify({"ok": True, "dias": len(juntas)})


@app.route("/api/admin/site-conta")
def api_admin_site_conta():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    d = ler_json(resolver_pasta_dados() / "site_conta.json", None)
    if not d or not (d.get("vendas") or {}).get("serie_dia"):
        return jsonify({"sem_dados": True})
    return jsonify(d)


# ============================================================
# Pedidos de expedição — o vendedor lança, a expedição libera
# ============================================================
# Nasceu em 30/08/2026: o vendedor precisa avisar a expedição que uma peça vai
# sair, com quem vai buscar e se já foi paga. A pergunta que a tela da
# expedição responde é uma só — "posso entregar essa peça pra esse motoboy?" —
# então pagamento e status ficam no topo do registro.
#
# Peça NÃO PAGA não trava a liberação, mas quem liberou fica registrado: em
# loja de autopeças pagar na entrega é comum, e uma trava dura faria o time
# contornar o sistema, que é pior que registrar.

FORMAS_PAGAMENTO = ["Pix", "Dinheiro", "Cartão", "Boleto", "Na entrega", "Faturado"]
STATUS_PEDIDO = {"aguardando": "Aguardando separação",
                 "separado": "Separado, aguardando retirada",
                 "liberado": "Liberado / entregue",
                 "cancelado": "Cancelado"}


def _exp_ler() -> dict:
    return ler_json(resolver_pasta_dados() / "expedicao_pedidos.json", None) or {}


def _exp_gravar(dados: dict) -> None:
    escrever_json(resolver_pasta_dados() / "expedicao_pedidos.json", dados)


def _exp_visivel(p: dict, dias=3) -> bool:
    """Pedido fechado some da fila depois de alguns dias; o aberto fica sempre.
    Sem isso a tela da expedição viraria histórico e ninguém acharia o que
    precisa separar agora."""
    if p.get("status") in ("aguardando", "separado"):
        return True
    corte = (agora_br() - timedelta(days=dias)).isoformat(timespec="seconds")
    return (p.get("liberado_em") or p.get("criado_em") or "") >= corte


def _exp_texto(v, limite=120) -> str:
    return str(v or "").strip()[:limite]


@app.route("/api/expedicao/pedidos")
def api_exp_listar():
    """A fila da expedição. Vendedor vê os próprios; expedição e gestor veem
    todos."""
    vid = exigir_vendedor()
    if not vid and not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401

    ve_tudo = bool(exigir_admin()) or (vid and perfil_de(vid) == "expedicao")
    nomes = {k: x["nome"] for k, x in carregar_vendedores().items()}
    itens = []
    for pid, p in _exp_ler().items():
        if not ve_tudo and p.get("criado_por") != vid:
            continue
        if not _exp_visivel(p):
            continue
        itens.append({**p, "id": pid,
                      "vendedor": nomes.get(p.get("criado_por"), "?"),
                      "liberado_por_nome": nomes.get(p.get("liberado_por"),
                                                     p.get("liberado_por") or "")})
    # Aguardando primeiro, e dentro disso o mais antigo no topo: a fila é FIFO,
    # quem chegou antes espera menos.
    ordem = {"aguardando": 0, "separado": 1, "liberado": 2, "cancelado": 3}
    itens.sort(key=lambda p: (ordem.get(p.get("status"), 9), p.get("criado_em") or ""))
    abertos = [p for p in itens if p["status"] in ("aguardando", "separado")]
    return jsonify({
        "pedidos": itens,
        "abertos": len(abertos),
        "nao_pagos_abertos": sum(1 for p in abertos if not p.get("pago")),
        "opcoes": {"pagamento": FORMAS_PAGAMENTO, "status": STATUS_PEDIDO},
        "sou_expedicao": ve_tudo,
    })


@app.route("/api/expedicao/pedidos", methods=["POST"])
def api_exp_criar():
    vid = exigir_vendedor()
    if not vid:
        return jsonify({"erro": "Não autenticado."}), 401
    corpo = request.get_json(silent=True) or {}
    peca = _exp_texto(corpo.get("peca"), 200)
    if not peca:
        return jsonify({"erro": "Informe a peça."}), 400
    quem_retira = _exp_texto(corpo.get("quem_retira"), 80)
    if not quem_retira:
        return jsonify({"erro": "Informe quem vai retirar."}), 400
    forma = _exp_texto(corpo.get("forma_pagamento"), 30)
    if forma and forma not in FORMAS_PAGAMENTO:
        return jsonify({"erro": "Forma de pagamento inválida."}), 400
    try:
        valor = round(float(str(corpo.get("valor") or 0).replace(",", ".")), 2)
    except ValueError:
        return jsonify({"erro": "Valor inválido."}), 400

    dados = _exp_ler()
    pid = uuid.uuid4().hex[:12]
    dados[pid] = {
        "criado_em": agora_br().isoformat(timespec="seconds"),
        "criado_por": vid,
        "peca": peca,
        "sku": _exp_texto(corpo.get("sku"), 40),
        "cliente": _exp_texto(corpo.get("cliente"), 80),
        "quem_retira": quem_retira,
        "telefone_retira": _exp_texto(corpo.get("telefone_retira"), 30),
        "forma_pagamento": forma,
        "pago": bool(corpo.get("pago")),
        "valor": valor,
        "obs": _exp_texto(corpo.get("obs"), 300),
        "status": "aguardando",
    }
    _exp_gravar(dados)
    return jsonify({"ok": True, "id": pid})


@app.route("/api/expedicao/pedidos/<pid>", methods=["POST"])
def api_exp_atualizar(pid):
    """Muda o status ou marca como pago. Vendedor mexe só no próprio pedido e
    só enquanto ele não saiu; a expedição mexe em qualquer um."""
    vid = exigir_vendedor()
    eh_exp = bool(exigir_admin()) or (vid and perfil_de(vid) == "expedicao")
    if not vid and not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401

    dados = _exp_ler()
    p = dados.get(pid)
    if not p:
        return jsonify({"erro": "Pedido não encontrado."}), 404
    if not eh_exp and p.get("criado_por") != vid:
        return jsonify({"erro": "Esse pedido não é seu."}), 403

    corpo = request.get_json(silent=True) or {}
    if "pago" in corpo:
        p["pago"] = bool(corpo["pago"])
        p["pago_marcado_por"] = vid or "gestor"
        p["pago_marcado_em"] = agora_br().isoformat(timespec="seconds")

    novo = corpo.get("status")
    if novo:
        if novo not in STATUS_PEDIDO:
            return jsonify({"erro": "Status inválido."}), 400
        if novo in ("separado", "liberado") and not eh_exp:
            return jsonify({"erro": "Só a expedição pode separar ou liberar."}), 403
        p["status"] = novo
        if novo == "liberado":
            p["liberado_em"] = agora_br().isoformat(timespec="seconds")
            p["liberado_por"] = vid or "gestor"
            # Saiu sem pagamento: fica gravado no proprio pedido, nao so no log.
            p["liberado_sem_pagamento"] = not p.get("pago")
    _exp_gravar(dados)
    return jsonify({"ok": True})


@app.route("/api/expedicao/pedidos/<pid>", methods=["DELETE"])
def api_exp_apagar(pid):
    """Só quem criou (e só antes de sair) ou o gestor."""
    vid = exigir_vendedor()
    dados = _exp_ler()
    p = dados.get(pid)
    if not p:
        return jsonify({"erro": "Pedido não encontrado."}), 404
    if not exigir_admin():
        if p.get("criado_por") != vid:
            return jsonify({"erro": "Esse pedido não é seu."}), 403
        if p.get("status") == "liberado":
            return jsonify({"erro": "Pedido já liberado não pode ser apagado."}), 400
    dados.pop(pid)
    _exp_gravar(dados)
    return jsonify({"ok": True})


@app.route("/api/admin/atendimento-alerta")
def api_admin_atendimento_alerta():
    """Quem esta esperando a loja responder agora. Alimentado de poucos em
    poucos minutos por scripts/monitorar_sem_resposta.py — nao depende da
    bolinha de nao-lida do Totalk, que o bot zera sozinho."""
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    d = ler_json(resolver_pasta_dados() / "atendimento_alerta.json", None)
    if not d:
        return jsonify({"sem_dados": True})
    resolvidos = _atd_resolvidos()
    conversas = _atd_pendentes(d.get("conversas") or [], resolvidos)
    # por_vendedor tambem precisa descontar os resolvidos: vinha do monitor,
    # calculado antes do filtro, e o gestor via um numero maior que o do portal
    # do proprio vendedor.
    por_vendedor = {}
    for vid, info in (d.get("por_vendedor") or {}).items():
        minhas = [c for c in conversas if c.get("vendedor_id") == vid]
        por_vendedor[vid] = {"nome": info.get("nome"), "total": len(minhas),
                             "critico": sum(1 for c in minhas if c["nivel"] == "critico")}
    d = {**d, "conversas": conversas, "total": len(conversas),
         "resolvidas": len(d.get("conversas") or []) - len(conversas),
         "por_vendedor": por_vendedor,
         "contagem": {n: sum(1 for c in conversas if c["nivel"] == n)
                      for n in ("critico", "urgente", "atencao")}}
    return jsonify(d)


@app.route("/api/admin/shopee-conta")
def api_admin_shopee_conta():
    """Serie mensal da Shopee (importar_shopee_stats). Endpoint separado do
    resumo porque o card dela e analise propria: mesmo quando o periodo
    filtrado nao tem Shopee, a evolucao historica continua na tela."""
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    d = ler_json(resolver_pasta_dados() / "shopee_conta.json", None)
    if not d or not (d.get("vendas") or {}).get("serie_mes"):
        return jsonify({"sem_dados": True})
    return jsonify(d)


@app.route("/api/admin/metas-bonus")
def api_admin_metas_bonus():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401

    dados_brutos = _mb_bruto()
    tem_bruto = any(dados_brutos["lancamentos"].values()) or dados_brutos["veiculos"]
    if tem_bruto:
        meses = _mb_agregar(dados_brutos)
    else:
        # Sem dado bruto, vale o agregado que o sincronizador antigo empurrou.
        bruto = ler_json(resolver_pasta_dados() / "metas_bonus.json", None) or {}
        meses = bruto.get("meses") or {}
    if not meses:
        return jsonify({"sem_dados": True})

    disponiveis = sorted(meses)
    mes = request.args.get("mes") or disponiveis[-1]
    if mes not in meses:
        mes = disponiveis[-1]
    atual = meses[mes]

    # Quantos bateram bonus mes a mes — e a serie que mostra se a meta esta
    # calibrada. Ninguem batendo nunca, ou todo mundo batendo sempre, sao os
    # dois jeitos de uma meta nao significar nada.
    historico = []
    for m in disponiveis:
        d = meses[m]
        pessoas = [p for linhas in d["setores"].values() for p in linhas]
        historico.append({
            "mes": m,
            "pessoas": len(pessoas),
            "na_meta": sum(1 for p in pessoas if p["bateu_meta"]),
            "no_bonus": sum(1 for p in pessoas if p["bateu_bonus"]),
            "pecas": d["veiculos"]["pecas"],
            "carros": d["veiculos"]["carros"],
        })

    ritmo = None
    hoje = hoje_br()
    if mes == hoje.isoformat()[:7]:
        dias = calendar.monthrange(hoje.year, hoje.month)[1]
        ritmo = {"dias_no_mes": dias, "dias_corridos": hoje.day,
                 "pct_do_mes": round(100 * hoje.day / dias)}

    return jsonify({
        "lancar": {
            "pessoas": {setor: sorted(({"id": pid, "nome": p.get("nome") or "?"}
                                       for pid, p in gente.items()), key=lambda x: x["nome"])
                        for setor, gente in dados_brutos["pessoas"].items()},
            "ativo": True,
        },
        "gerado_em": (agora_br().isoformat(timespec="seconds") if tem_bruto else bruto.get("gerado_em")),
        "mes": mes,
        "meses": disponiveis,
        "rotulos": SETORES_META,
        "setores": atual["setores"],
        "veiculos": atual["veiculos"],
        "historico": historico,
        "ritmo": ritmo,
    })


# ============================================================
# Carros pra chegar
# ============================================================
# Vem da planilha "Carros para chegar" (scripts/sincronizar_carros.py).
#
# Um cuidado que muda a leitura inteira: a coluna "Data de Chegada" só começou a
# ser preenchida em maio/2026. Carro comprado antes disso e sem chegada não está
# parado — está sem registro. Tratar os dois como a mesma coisa mostraria
# milhões "travados" que na verdade chegaram e ninguém anotou, e o gestor
# tomaria decisão em cima de um número falso. Por isso o painel separa a janela
# em que o acompanhamento existe do histórico anterior.

def _carros_bruto():
    return ler_json(resolver_pasta_dados() / "carros_chegar.json", None) or {}


def _inicio_acompanhamento(carros):
    """Primeiro dia em que alguém registrou chegada ou agendamento. Antes disso
    a ausência de chegada não significa nada."""
    marcos = [c[campo] for c in carros for campo in ("chegada", "agendamento") if c.get(campo)]
    return min(marcos) if marcos else None



@app.route("/api/admin/metas-bonus/dia")
def api_mb_dia():
    """O dia como uma chamada: cada pessoa com o que ja foi lancado nessa data.
    E o que a tela mostra pra preencher todo mundo de uma vez."""
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    data = (request.args.get("data") or "").strip() or hoje_br().isoformat()
    dados = _mb_bruto()
    setores = {}
    for setor, tipo in TIPO_META.items():
        do_dia = {}
        for l in (dados["lancamentos"].get(tipo) or {}).values():
            if l.get("data") == data:
                pid = l.get("pessoa_id")
                do_dia[pid] = do_dia.get(pid, 0.0) + float(l.get("quantidade") or 0)
        setores[setor] = sorted((
            {"id": pid, "nome": p.get("nome") or "?",
             "meta": p.get("meta") or 0, "meta_bonus": p.get("meta_bonus") or 0,
             "quantidade": do_dia.get(pid) or ""}
            # Inativo (desligado ou de funcao trocada) sai da chamada, mas
            # continua no dicionario: o historico dele segue agregando.
            for pid, p in dados["pessoas"].get(setor, {}).items()
            if not p.get("inativo")),
            key=lambda x: x["nome"])
    veiculos = sorted((
        {"id": vid, "carro": v.get("carro"), "codigo": v.get("codigo"),
         "pecas": v.get("pecas") or 0}
        for vid, v in dados["veiculos"].items() if v.get("data") == data),
        key=lambda x: x["carro"] or "")
    return jsonify({"data": data, "setores": setores, "veiculos": veiculos})


@app.route("/api/admin/metas-bonus/pessoas", methods=["POST"])
def api_mb_pessoa_salvar():
    """Cria ou edita uma pessoa da chamada. Mudar de funcao nao move o
    historico: a entrada antiga vira inativa (os meses lancados continuam
    contando la) e nasce uma entrada ativa na funcao nova."""
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    corpo = request.get_json(silent=True) or {}
    nome = (corpo.get("nome") or "").strip()[:60]
    if not nome:
        return jsonify({"erro": "Informe o nome."}), 400
    setor = corpo.get("setor")
    if setor not in TIPO_META:
        return jsonify({"erro": "Função inválida."}), 400

    def num(v):
        try:
            return max(0.0, float(str(v or 0).replace(",", ".")))
        except ValueError:
            return 0.0
    meta, bonus = num(corpo.get("meta")), num(corpo.get("meta_bonus"))

    dados = _mb_bruto()
    pid = (corpo.get("id") or "").strip()
    setor_atual = next((st for st, gente in dados["pessoas"].items() if pid in gente), None)

    if not pid or not setor_atual:
        # Mesmo nome de alguem inativo nesta funcao? Entao e a mesma pessoa
        # voltando (ou um engano sendo desfeito): reativa e o historico volta
        # junto, em vez de nascer um duplicado sem passado.
        igual = next((x for x, p in dados["pessoas"].get(setor, {}).items()
                      if p.get("inativo") and _achatar_canal(p.get("nome") or "") == _achatar_canal(nome)),
                     None)
        if igual:
            dados["pessoas"][setor][igual].update(
                nome=nome, meta=meta, meta_bonus=bonus, inativo=False)
        else:
            dados["pessoas"].setdefault(setor, {})[uuid.uuid4().hex[:12]] = {
                "nome": nome, "meta": meta, "meta_bonus": bonus}
    elif setor_atual == setor:
        dados["pessoas"][setor][pid].update(nome=nome, meta=meta, meta_bonus=bonus)
    else:
        antigo = dados["pessoas"][setor_atual][pid]
        tipo_antigo = TIPO_META[setor_atual]
        tem_historico = any(l.get("pessoa_id") == pid
                            for l in (dados["lancamentos"].get(tipo_antigo) or {}).values())
        if tem_historico:
            antigo["nome"] = nome
            antigo["inativo"] = True
        else:
            dados["pessoas"][setor_atual].pop(pid)
        dados["pessoas"].setdefault(setor, {})[pid] = {
            "nome": nome, "meta": meta, "meta_bonus": bonus}
    _mb_gravar(dados)
    return jsonify({"ok": True})


@app.route("/api/admin/metas-bonus/pessoas/<setor>/<pid>", methods=["DELETE"])
def api_mb_pessoa_remover(setor, pid):
    """Tira da chamada. So apaga de verdade quem nunca lancou nada — apagar
    alguem com historico faria os meses antigos dele sumirem do painel."""
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    dados = _mb_bruto()
    pessoa = dados["pessoas"].get(setor, {}).get(pid)
    if not pessoa:
        return jsonify({"erro": "Pessoa não encontrada."}), 404
    tipo = TIPO_META[setor]
    tem_historico = any(l.get("pessoa_id") == pid
                        for l in (dados["lancamentos"].get(tipo) or {}).values())
    if tem_historico:
        pessoa["inativo"] = True
        modo = "inativada"
    else:
        dados["pessoas"][setor].pop(pid)
        modo = "apagada"
    _mb_gravar(dados)
    return jsonify({"ok": True, "modo": modo})


@app.route("/api/admin/metas-bonus/dia", methods=["POST"])
def api_mb_dia_salvar():
    """Grava o dia inteiro de uma vez. Upsert por (pessoa, data): o campo da
    tela E o valor do dia — vazio ou zero apaga, numero substitui. Assim a
    grade pode ser corrigida e salva de novo sem duplicar nada."""
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    corpo = request.get_json(silent=True) or {}
    data = (corpo.get("data") or "").strip()
    try:
        parse_dt_tolerante(data)
    except Exception:
        return jsonify({"erro": "Data inválida."}), 400
    itens = corpo.get("itens")
    if not isinstance(itens, list):
        return jsonify({"erro": "Nada para salvar."}), 400

    dados = _mb_bruto()
    carimbo = agora_br().isoformat(timespec="seconds")
    gravados = apagados = 0
    for item in itens:
        tipo = item.get("tipo")
        if tipo not in ("anuncio", "cadastro"):
            return jsonify({"erro": f"Tipo inválido: {tipo}"}), 400
        setor = next(st for st, t in TIPO_META.items() if t == tipo)
        pid = (item.get("pessoa_id") or "").strip()
        if pid not in dados["pessoas"].get(setor, {}):
            return jsonify({"erro": "Pessoa desconhecida."}), 400
        bruto_qtd = str(item.get("quantidade") or "").strip().replace(",", ".")
        try:
            quantidade = float(bruto_qtd) if bruto_qtd else 0.0
        except ValueError:
            nome = dados["pessoas"][setor][pid].get("nome") or "?"
            return jsonify({"erro": f"Quantidade inválida para {nome}."}), 400
        if quantidade < 0:
            nome = dados["pessoas"][setor][pid].get("nome") or "?"
            return jsonify({"erro": f"Quantidade negativa para {nome}."}), 400

        lanc = dados["lancamentos"].setdefault(tipo, {})
        havia = [lid for lid, l in lanc.items()
                 if l.get("pessoa_id") == pid and l.get("data") == data]
        for lid in havia:
            lanc.pop(lid)
        if quantidade > 0:
            lanc[uuid.uuid4().hex[:12]] = {
                "pessoa_id": pid, "data": data, "quantidade": quantidade,
                "lancado_em": carimbo,
            }
            gravados += 1
        elif havia:
            apagados += 1
    _mb_gravar(dados)
    return jsonify({"ok": True, "gravados": gravados, "apagados": apagados})


@app.route("/api/admin/metas-bonus/lancamentos", methods=["POST"])
def api_mb_lancar():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    corpo = request.get_json(silent=True) or {}
    tipo = corpo.get("tipo")
    if tipo not in ("anuncio", "cadastro"):
        return jsonify({"erro": "Tipo inválido."}), 400
    dados = _mb_bruto()
    setor = next(st for st, t in TIPO_META.items() if t == tipo)
    pessoa_id = (corpo.get("pessoa_id") or "").strip()
    if pessoa_id not in dados["pessoas"].get(setor, {}):
        return jsonify({"erro": "Escolha a pessoa."}), 400
    data = (corpo.get("data") or "").strip() or hoje_br().isoformat()
    try:
        parse_dt_tolerante(data)
    except Exception:
        return jsonify({"erro": "Data inválida."}), 400
    try:
        quantidade = float(str(corpo.get("quantidade")).replace(",", "."))
    except (TypeError, ValueError):
        return jsonify({"erro": "Quantidade inválida."}), 400
    if quantidade <= 0:
        return jsonify({"erro": "Quantidade deve ser maior que zero."}), 400

    dados["lancamentos"].setdefault(tipo, {})[uuid.uuid4().hex[:12]] = {
        "pessoa_id": pessoa_id, "data": data, "quantidade": quantidade,
        "lancado_em": agora_br().isoformat(timespec="seconds"),
    }
    _mb_gravar(dados)
    return jsonify({"ok": True})


@app.route("/api/admin/metas-bonus/lancamentos/<tipo>/<lid>", methods=["DELETE"])
def api_mb_apagar(tipo, lid):
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    dados = _mb_bruto()
    if lid not in dados["lancamentos"].get(tipo, {}):
        return jsonify({"erro": "Lançamento não encontrado."}), 404
    dados["lancamentos"][tipo].pop(lid)
    _mb_gravar(dados)
    return jsonify({"ok": True})


@app.route("/api/admin/metas-bonus/veiculos", methods=["POST"])
def api_mb_veiculo():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    corpo = request.get_json(silent=True) or {}
    carro = (corpo.get("carro") or "").strip()
    if not carro:
        return jsonify({"erro": "Informe o carro."}), 400
    data = (corpo.get("data") or "").strip() or hoje_br().isoformat()
    try:
        parse_dt_tolerante(data)
    except Exception:
        return jsonify({"erro": "Data inválida."}), 400
    try:
        pecas = float(str(corpo.get("pecas")).replace(",", "."))
    except (TypeError, ValueError):
        return jsonify({"erro": "Quantidade de peças inválida."}), 400
    if pecas <= 0:
        return jsonify({"erro": "Peças deve ser maior que zero."}), 400

    dados = _mb_bruto()
    dados["veiculos"][uuid.uuid4().hex[:12]] = {
        "data": data, "carro": carro[:80],
        "codigo": (corpo.get("codigo") or "").strip()[:20], "pecas": pecas,
        "lancado_em": agora_br().isoformat(timespec="seconds"),
    }
    _mb_gravar(dados)
    return jsonify({"ok": True})


@app.route("/api/admin/metas-bonus/veiculos/<vid>", methods=["DELETE"])
def api_mb_veiculo_apagar(vid):
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    dados = _mb_bruto()
    if vid not in dados["veiculos"]:
        return jsonify({"erro": "Veículo não encontrado."}), 404
    dados["veiculos"].pop(vid)
    _mb_gravar(dados)
    return jsonify({"ok": True})


@app.route("/api/admin/metas-bonus/recentes")
def api_mb_recentes():
    """Ultimos lancamentos, pra conferir e desfazer erro de digitacao."""
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    dados = _mb_bruto()
    nomes = {pid: p.get("nome") or "?"
             for gente in dados["pessoas"].values() for pid, p in gente.items()}
    itens = []
    for tipo, lanc in dados["lancamentos"].items():
        for lid, l in lanc.items():
            itens.append({"id": lid, "tipo": tipo, "data": l.get("data"),
                          "nome": nomes.get(l.get("pessoa_id"), "?"),
                          "quantidade": l.get("quantidade"),
                          "ordem": l.get("lancado_em") or l.get("data") or ""})
    for vid, v in dados["veiculos"].items():
        itens.append({"id": vid, "tipo": "veiculo", "data": v.get("data"),
                      "nome": v.get("carro"), "codigo": v.get("codigo"),
                      "quantidade": v.get("pecas"),
                      "ordem": v.get("lancado_em") or v.get("data") or ""})
    itens.sort(key=lambda x: x["ordem"], reverse=True)
    return jsonify({"itens": itens[:12]})


@app.route("/api/admin/carros")
def api_admin_carros():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401

    bruto = _carros_bruto()
    carros = bruto.get("carros", [])
    if not carros:
        return jsonify({"sem_dados": True})

    inicio = _inicio_acompanhamento(carros)
    filtro_estado = request.args.get("estado") or ""
    filtro_leilao = request.args.get("leilao") or ""
    de = request.args.get("de") or ""
    ate = request.args.get("ate") or ""

    def dentro(c):
        if de and (not c["data"] or c["data"] < de):
            return False
        if ate and (not c["data"] or c["data"] > ate):
            return False
        if filtro_leilao and c["leilao"] != filtro_leilao:
            return False
        if filtro_estado and c["estado"] != filtro_estado:
            return False
        return True

    visiveis = [c for c in carros if dentro(c)]

    # Acompanhados = comprados depois que o controle de chegada passou a existir.
    acompanhados = [c for c in visiveis
                    if inicio and c["data"] and c["data"] >= inicio]
    antigos = [c for c in visiveis
               if not (inicio and c["data"] and c["data"] >= inicio)]

    def somar(lista):
        return {"qtd": len(lista),
                "valor": round(sum(c["valor"] or 0 for c in lista), 2)}

    pendentes = [c for c in acompanhados if c["estado"] != "chegou"]
    pendentes.sort(key=lambda c: -(c["dias_parado"] or 0))
    chegaram = [c for c in acompanhados if c["estado"] == "chegou"]

    tempos = sorted(c["dias_ate_chegar"] for c in chegaram
                    if c["dias_ate_chegar"] is not None)
    mediana = tempos[len(tempos) // 2] if tempos else None

    def agrupar(lista, campo):
        d = {}
        for c in lista:
            chave = c.get(campo) or "—"
            item = d.setdefault(chave, {"qtd": 0, "valor": 0.0, "pendentes": 0})
            item["qtd"] += 1
            item["valor"] += c["valor"] or 0
            if c["estado"] != "chegou":
                item["pendentes"] += 1
        return sorted(({campo: k, "qtd": v["qtd"], "valor": round(v["valor"], 2),
                        "pendentes": v["pendentes"]} for k, v in d.items()),
                      key=lambda x: -x["qtd"])

    return jsonify({
        "gerado_em": bruto.get("gerado_em"),
        "inicio_acompanhamento": inicio,
        "filtro": {"estado": filtro_estado, "leilao": filtro_leilao, "de": de, "ate": ate},
        "leiloes": sorted({c["leilao"] for c in carros if c["leilao"]}),
        "estados": ["comprado", "agendado", "chegou", "sem_situacao"],
        "acompanhados": {
            **somar(acompanhados),
            "pendentes": somar(pendentes),
            "chegaram": somar(chegaram),
            "mediana_dias": mediana,
        },
        "historico_sem_registro": somar([c for c in antigos if c["estado"] != "chegou"]),
        "lista_pendentes": pendentes,
        # A planilha inteira, nao so os pendentes: o painel respondia "o que
        # esta atrasado?" mas nao "cade o carro tal?". Ordem por compra, mais
        # recente em cima; a busca acontece na tela, que ja tem tudo em maos.
        "lista": sorted(visiveis, key=lambda c: (c["data"] or "", c["veiculo"] or ""),
                        reverse=True),
        "por_leilao": agrupar(acompanhados, "leilao"),
        "por_estado": agrupar(acompanhados, "estado"),
        "total_geral": somar(visiveis),
    })


# ============================================================
# Marketing
# ============================================================
# Os números vêm do vendas-insights (espelho do Totalk + investimento de mídia
# do Windsor), sincronizados no grão diário por `sincronizar_marketing.py`. O
# recorte é feito aqui, na leitura: assim período, canal e vendedor mudam sem
# precisar regerar nada do outro lado.
#
# Vendedor vê só as linhas dele. Gestor vê tudo, incluindo investimento — que é
# de conta, não de pessoa, e por isso nunca aparece na tela do vendedor.

def _marketing_bruto():
    leads = ler_json(resolver_pasta_dados() / "marketing_leads.json", None) or {}
    gasto = ler_json(resolver_pasta_dados() / "marketing_gasto.json", None) or {}
    return leads, gasto


def _recortar(linhas, de, ate, campo_data="data"):
    return [l for l in linhas if de <= l[campo_data] <= ate]


def _agregar_marketing(linhas, vendedores_nome=None):
    """Soma um conjunto de linhas de lead em vários cortes de uma vez."""
    total = {"leads": 0, "sinal": 0, "provavel": 0}
    por_canal, por_dia, por_vendedor = {}, {}, {}
    for l in linhas:
        total["leads"] += l["leads"]
        total["sinal"] += l["sinal"]
        total["provavel"] += l.get("provavel", 0)
        for destino, chave in ((por_canal, l["canal"]), (por_dia, l["data"]),
                               (por_vendedor, l["vendedor"] or "sem_atendente")):
            d = destino.setdefault(chave, {"leads": 0, "sinal": 0, "provavel": 0})
            d["leads"] += l["leads"]
            d["sinal"] += l["sinal"]
            d["provavel"] += l.get("provavel", 0)

    def lista(dic, rotulo, ordenar_por_chave=False):
        itens = [{rotulo: k, **v} for k, v in dic.items()]
        itens.sort(key=(lambda x: x[rotulo]) if ordenar_por_chave
                   else (lambda x: -x["leads"]))
        return itens

    saida = {
        "total": total,
        "por_canal": lista(por_canal, "canal"),
        "por_dia": lista(por_dia, "data", ordenar_por_chave=True),
        "por_vendedor": lista(por_vendedor, "vendedor"),
    }
    if vendedores_nome:
        for item in saida["por_vendedor"]:
            item["nome"] = vendedores_nome.get(item["vendedor"], "Sem atendente")
    return saida


def _vendas_no_periodo(vendedores, de, ate, so_vendedor=None, universo=None):
    """Vendas lançadas no portal — é o que fecha a conta do marketing: leads de
    um lado, venda de verdade do outro.

    `universo` limita a conta a quem realmente aparece no dado de leads. Sem
    isso a Brenda entraria: ela não atende pelo Totalk, então as vendas dela
    somariam no numerador sem nenhum lead no denominador e a conversão sairia
    inflada."""
    vendas = carregar_vendas_todos(vendedores)
    itens = [v for v in vendas.values()
             if v.get("tipo", "venda") == "venda" and de <= v["data"] <= ate
             and (so_vendedor is None or v["vendedor_id"] == so_vendedor)
             and (universo is None or v["vendedor_id"] in universo)]
    total = round(sum(valor_liquido(v) for v in itens), 2)
    por_vendedor = {}
    for v in itens:
        d = por_vendedor.setdefault(v["vendedor_id"], {"qtd": 0, "total": 0.0})
        d["qtd"] += 1
        d["total"] += valor_liquido(v)
    return {
        "qtd": len(itens),
        "total": total,
        "ticket": round(total / len(itens), 2) if itens else 0.0,
        "por_vendedor": {k: {"qtd": d["qtd"], "total": round(d["total"], 2)}
                         for k, d in por_vendedor.items()},
    }


def _periodo_pedido():
    hoje = hoje_br().isoformat()
    de = request.args.get("de") or f"{hoje[:7]}-01"
    ate = request.args.get("ate") or hoje
    return de, ate


def _janela_comparavel(de, ate, cobertura):
    """Leads param na data em que o Totalk foi sincronizado; as vendas seguem
    até hoje. Comparar as duas coisas na janela cheia faz a conversão parecer
    melhor do que é — então a conta usa só o pedaço em que os dois lados
    existem."""
    if not cobertura:
        return de, ate
    return max(de, cobertura["de"]), min(ate, cobertura["ate"])


def _cobertura(linhas):
    """Até quando o espelho do Totalk foi sincronizado. Sem isso, um período
    que passa dessa data mostra queda de leads que é só falta de dado."""
    if not linhas:
        return None
    datas = [l["data"] for l in linhas]
    return {"de": min(datas), "ate": max(datas)}


def _google_por_objetivo(linhas):
    """Separa o gasto do Google entre trazer gente pra LOJA e vender no SITE.

    Linha antiga (antes de 30/08/2026) nao tem o campo `objetivo` — cai em
    "site", que era o unico tipo de campanha que existia ate entao.
    """
    saida = {}
    for g in linhas:
        alvo = g.get("objetivo") or "site"
        d = saida.setdefault(alvo, {"investimento": 0.0, "clicks": 0,
                                    "ligacoes": 0, "conversoes": 0.0})
        d["investimento"] += g.get("spend", 0)
        d["clicks"] += g.get("clicks", 0)
        d["ligacoes"] += g.get("ligacoes", 0)
        d["conversoes"] += g.get("conversoes", 0)
    for d in saida.values():
        d["investimento"] = round(d["investimento"], 2)
        d["conversoes"] = round(d["conversoes"], 2)
    return saida


def _google_por_campanha(linhas):
    """Uma linha por campanha, com tipo e objetivo — pro gestor ver onde o
    dinheiro esta e o que cada uma entrega."""
    saida = {}
    for g in linhas:
        nome = g.get("campanha") or "—"
        d = saida.setdefault(nome, {"campanha": nome, "investimento": 0.0,
                                    "clicks": 0, "ligacoes": 0, "conversoes": 0.0,
                                    "tipo_nome": g.get("tipo_nome") or "",
                                    "objetivo": g.get("objetivo") or "site"})
        d["investimento"] += g.get("spend", 0)
        d["clicks"] += g.get("clicks", 0)
        d["ligacoes"] += g.get("ligacoes", 0)
        d["conversoes"] += g.get("conversoes", 0)
    for d in saida.values():
        d["investimento"] = round(d["investimento"], 2)
        d["conversoes"] = round(d["conversoes"], 2)
        d["cpc"] = round(d["investimento"] / d["clicks"], 2) if d["clicks"] else None
    return sorted(saida.values(), key=lambda x: -x["investimento"])


@app.route("/api/admin/marketing")
def api_marketing_gestor():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    de, ate = _periodo_pedido()
    filtro_canal = request.args.get("canal") or ""
    filtro_vendedor = request.args.get("vendedor") or ""

    leads_bruto, gasto_bruto = _marketing_bruto()
    vendedores = carregar_vendedores()
    # So quem vende entra na tela de marketing. A expedicao usa o mesmo login,
    # mas nao atende lead nenhum — se entrasse aqui apareceria eternamente na
    # lista de "fora do Totalk", como se estivesse deixando de atender.
    nomes = {vid: v["nome"] for vid, v in vendedores.items()
             if (v.get("perfil") or "vendedor") != "expedicao"}

    cobertura = _cobertura(leads_bruto.get("linhas", []))
    ef_de, ef_ate = _janela_comparavel(de, ate, cobertura)

    linhas = _recortar(leads_bruto.get("linhas", []), ef_de, ef_ate)
    # Quem de fato tem lead no periodo — e o universo que pode entrar na
    # conversao. Fora dele a venda nao tem lead correspondente.
    universo = {l["vendedor"] for l in linhas if l["vendedor"]}
    fora = sorted(nomes[vid] for vid in nomes
                  if vid not in universo and not vendedores[vid].get("oculto"))
    if filtro_canal:
        linhas = [l for l in linhas if l["canal"] == filtro_canal]
    if filtro_vendedor:
        linhas = [l for l in linhas if l["vendedor"] == filtro_vendedor]
    agregado = _agregar_marketing(linhas, nomes)

    # canais disponíveis saem do período inteiro, não do recorte — senão filtrar
    # por um canal apagaria os outros da lista e não daria pra voltar
    todos_canais = sorted({l["canal"] for l in _recortar(leads_bruto.get("linhas", []), ef_de, ef_ate)})

    # O investimento segue a janela pedida: gasto de midia nao depende do Totalk.
    gasto_linhas = _recortar(gasto_bruto.get("linhas", []), de, ate)
    investimento = round(sum(g["spend"] for g in gasto_linhas), 2)
    cliques = sum(g["clicks"] for g in gasto_linhas)
    impressoes = sum(g["impressions"] for g in gasto_linhas)

    por_campanha = {}
    for g in gasto_linhas:
        d = por_campanha.setdefault(g["campanha"], {"spend": 0.0, "clicks": 0, "impressions": 0})
        d["spend"] += g["spend"]
        d["clicks"] += g["clicks"]
        d["impressions"] += g["impressions"]
    campanhas = sorted(
        [{"campanha": k, "spend": round(v["spend"], 2), "clicks": v["clicks"],
          "impressions": v["impressions"],
          "cpc": round(v["spend"] / v["clicks"], 2) if v["clicks"] else None,
          "ctr": round(100 * v["clicks"] / v["impressions"], 2) if v["impressions"] else None}
         for k, v in por_campanha.items()],
        key=lambda x: -x["spend"])

    gasto_dia = {}
    for g in gasto_linhas:
        gasto_dia[g["data"]] = round(gasto_dia.get(g["data"], 0.0) + g["spend"], 2)

    # O Meta vem agregado do periodo inteiro do relatorio, nao por dia. Exigir
    # que a janela cobrisse o relatorio inteiro pra somar deixava o investimento
    # dele fora da conta em qualquer mes — que e justamente como o painel e
    # aberto. Entao rateia por dia: o gasto do relatorio dividido pelos dias que
    # ele cobre, multiplicado pelos dias que caem na janela pedida.
    #
    # E aproximacao, e a tela diz isso. Gasto de midia nao e uniforme dia a dia,
    # mas num recorte de semanas o erro e pequeno perto de simplesmente ignorar
    # metade do investimento.
    meta = gasto_bruto.get("meta")
    meta_rateio = None

    # Desde 28/08/2026 o Meta vem do Windsor com serie por dia: soma exata do
    # periodo filtrado, sem rateio. O bloco de rateio abaixo continua pro caso
    # de a fonte voltar a ser o CSV do Gerenciador, que so traz o agregado.
    serie_meta = (meta or {}).get("serie_dia") or {}
    gasto_meta_periodo = 0.0   # nome estavel pra comparacao com o periodo anterior
    if serie_meta:
        dentro = {d: v for d, v in serie_meta.items() if de <= d <= ate}
        if dentro:
            gasto_meta = round(sum(v["spend"] for v in dentro.values()), 2)
            meta_rateio = {
                "de": min(dentro), "ate": max(dentro),
                "dias_dentro": len(dentro), "dias_relatorio": len(dentro),
                "spend": gasto_meta,
                "impressions": sum(v["impressions"] for v in dentro.values()),
                "conversas": sum(v["conversas"] for v in dentro.values()),
                "integral": True,
                "por_dia": True,
            }
            # Facebook x Instagram. A quebra vem dentro de cada dia, entao ela
            # acompanha o filtro de periodo em vez de depender de um total ja
            # fechado. Alcance fica de fora de proposito: e a unica metrica que
            # nao pode ser somada entre as duas, porque a mesma pessoa aparece
            # nas duas e o total viraria gente que nao existe.
            plataformas = {}
            for v_dia in dentro.values():
                for nome, p in (v_dia.get("plataforma") or {}).items():
                    # O Meta devolve um balde "unknown" que vem sempre zerado.
                    # Filtrar pelo nome nao basta — o que importa e nao ter
                    # movimento nenhum, entao a checagem e por valor.
                    if not (p.get("spend") or p.get("conversas")
                            or p.get("clicks")):
                        continue
                    acc = plataformas.setdefault(nome, {"spend": 0.0, "clicks": 0,
                                                        "impressions": 0, "conversas": 0})
                    acc["spend"] = round(acc["spend"] + p.get("spend", 0), 2)
                    acc["clicks"] += p.get("clicks", 0)
                    acc["impressions"] += p.get("impressions", 0)
                    acc["conversas"] += p.get("conversas", 0)
            if plataformas:
                for nome, p in plataformas.items():
                    p["nome"] = {"facebook": "Facebook",
                                 "instagram": "Instagram"}.get(nome, nome.title())
                    p["custo_por_conversa"] = (round(p["spend"] / p["conversas"], 2)
                                               if p["conversas"] else None)
                    p["fatia"] = (round(100 * p["spend"] / gasto_meta, 1)
                                  if gasto_meta else 0)
                meta_rateio["plataformas"] = sorted(
                    plataformas.values(), key=lambda p: -p["spend"])
            investimento = round(investimento + gasto_meta, 2)
            gasto_meta_periodo = gasto_meta
            impressoes += meta_rateio["impressions"]
    elif meta and meta.get("de") and meta.get("ate"):
        ini = max(de, meta["de"])
        fim = min(ate, meta["ate"])
        if ini <= fim:
            dias_relatorio = (date.fromisoformat(meta["ate"])
                              - date.fromisoformat(meta["de"])).days + 1
            dias_dentro = (date.fromisoformat(fim) - date.fromisoformat(ini)).days + 1
            fatia = dias_dentro / dias_relatorio if dias_relatorio else 0
            meta_rateio = {
                "de": ini, "ate": fim,
                "dias_dentro": dias_dentro, "dias_relatorio": dias_relatorio,
                "spend": round(meta["spend"] * fatia, 2),
                "impressions": int(meta.get("impressions", 0) * fatia),
                "conversas": int(meta.get("conversas", 0) * fatia),
                "integral": dias_dentro == dias_relatorio,
            }
            investimento = round(investimento + meta_rateio["spend"], 2)
            gasto_meta_periodo = meta_rateio["spend"]
            impressoes += meta_rateio["impressions"]

    # Product Ads do Mercado Livre, com o MESMO rateio por dia do Meta: o
    # sincronizador entrega uma janela fechada de 30 dias, e daqui sai a fatia
    # que cai no periodo pedido. Aproximacao — e a tela diz isso.
    ml_conta = ler_json(resolver_pasta_dados() / "ml_conta.json", None) or {}
    ml_ads = ml_conta.get("ads")
    ml_rateio = None
    if ml_ads and ml_ads.get("de") and ml_ads.get("ate") and ml_ads.get("investido"):
        ini = max(de, ml_ads["de"])
        fim = min(ate, ml_ads["ate"])
        if ini <= fim:
            dias_relatorio = (date.fromisoformat(ml_ads["ate"])
                              - date.fromisoformat(ml_ads["de"])).days + 1
            dias_dentro = (date.fromisoformat(fim) - date.fromisoformat(ini)).days + 1
            fatia = dias_dentro / dias_relatorio if dias_relatorio else 0
            ml_rateio = {
                "de": ini, "ate": fim,
                "dias_dentro": dias_dentro, "dias_relatorio": dias_relatorio,
                "spend": round(float(ml_ads["investido"]) * fatia, 2),
                "cliques": int((ml_ads.get("cliques") or 0) * fatia),
                "impressions": int((ml_ads.get("impressoes") or 0) * fatia),
                "receita": round(float(ml_ads.get("receita_atribuida") or 0) * fatia, 2),
                "acos": ml_ads.get("acos"),
                "integral": dias_dentro == dias_relatorio,
            }
            investimento = round(investimento + ml_rateio["spend"], 2)
            impressoes += ml_rateio["impressions"]

    # A agencia que gerencia as campanhas custa fixo por mes e entra no
    # investimento, porque gestao e custo de midia tanto quanto o clique.
    # NAO rateia por dia: a nota vem cheia todo mes, independente de quantos
    # dias voce esta olhando. Rateando, um filtro de 01 a 30/08 mostrava
    # R$ 1.838,71 — um valor que nunca foi pago e nao existe em lugar nenhum.
    # Aqui o periodo conta os meses que ele toca, cada um pelo valor cheio.
    # Valor unico declarado aqui: mudou o contrato, muda esta linha.
    AGENCIA_MENSAL = 1900.0
    d_ini, d_fim = date.fromisoformat(de), date.fromisoformat(ate)
    meses_agencia = sorted({d_ini.strftime("%Y-%m"), d_fim.strftime("%Y-%m")}
                           | {(d_ini.replace(day=1) + timedelta(days=32 * k)).strftime("%Y-%m")
                              for k in range(1, 60)
                              if (d_ini.replace(day=1) + timedelta(days=32 * k)) <= d_fim})
    agencia = round(AGENCIA_MENSAL * len(meses_agencia), 2)
    investimento = round(investimento + agencia, 2)

    vendas = _vendas_no_periodo(vendedores, ef_de, ef_ate,
                                so_vendedor=filtro_vendedor or None,
                                universo=universo)
    for item in agregado["por_vendedor"]:
        item["vendas"] = vendas["por_vendedor"].get(item["vendedor"], {"qtd": 0, "total": 0.0})
        item["conversao"] = (round(100 * item["vendas"]["qtd"] / item["leads"], 1)
                             if item["leads"] else None)

    total_leads = agregado["total"]["leads"]

    # Receita provavel por canal. A venda nao carrega o canal de origem — o
    # chat esta no Totalk, a venda esta no portal, e nada liga um ao outro.
    # Entao o unico jeito de saber que canal traz dinheiro e estimar: quantas
    # conversas daquele canal chegaram ao fechamento, vezes o ticket real.
    #
    # O ticket sai das vendas REAIS do proprio periodo, nao de um valor fixo:
    # assim ele acompanha o mes em vez de envelhecer. So cai num padrao quando
    # o periodo nao tem venda lancada suficiente pra sustentar uma media.
    TICKET_PADRAO = 968.0   # calibrado em ago/2026 contra 519 vendas reais
    ticket = (round(vendas["total"] / vendas["qtd"], 2)
              if vendas.get("qtd") else TICKET_PADRAO)
    for item in agregado["por_canal"]:
        item["receita_provavel"] = round(item.get("provavel", 0) * ticket, 2)
    agregado["total"]["receita_provavel"] = round(
        agregado["total"].get("provavel", 0) * ticket, 2)

    # ---- comparacao com o periodo anterior ----
    # A janela anterior tem o mesmo NUMERO DE DIAS COBERTOS, nao o mes
    # calendario anterior. Agosto ainda esta aberto: comparar 28 dias de agosto
    # com 31 de julho inventaria uma queda de 10% que e so calendario. Por isso
    # a base e ef_de/ef_ate, que ja e a janela que tem dado de verdade.
    dias_janela = ((date.fromisoformat(ef_ate) - date.fromisoformat(ef_de)).days + 1
                   if ef_de and ef_ate else 0)
    anterior = None
    if dias_janela > 0:
        ant_ate = (date.fromisoformat(ef_de) - timedelta(days=1)).isoformat()
        ant_de = (date.fromisoformat(ant_ate)
                  - timedelta(days=dias_janela - 1)).isoformat()
        linhas_ant = _recortar(leads_bruto.get("linhas", []), ant_de, ant_ate)
        if filtro_canal:
            linhas_ant = [l for l in linhas_ant if l["canal"] == filtro_canal]
        if filtro_vendedor:
            linhas_ant = [l for l in linhas_ant if l["vendedor"] == filtro_vendedor]
        if linhas_ant:
            ag_ant = _agregar_marketing(linhas_ant, nomes)
            vendas_ant = _vendas_no_periodo(vendedores, ant_de, ant_ate,
                                            so_vendedor=filtro_vendedor or None,
                                            universo=universo)
            # O ticket do periodo anterior e o DELE, nao o de agora: senao a
            # receita "mudaria" so porque o ticket medio mudou.
            ticket_ant = (round(vendas_ant["total"] / vendas_ant["qtd"], 2)
                          if vendas_ant.get("qtd") else ticket)
            for item in ag_ant["por_canal"]:
                item["receita_provavel"] = round(
                    item.get("provavel", 0) * ticket_ant, 2)
            ag_ant["total"]["receita_provavel"] = round(
                ag_ant["total"].get("provavel", 0) * ticket_ant, 2)
            anterior = {
                "de": ant_de, "ate": ant_ate, "dias": dias_janela,
                "total": ag_ant["total"],
                "por_canal": {c["canal"]: c for c in ag_ant["por_canal"]},
                "vendas": vendas_ant,
                "ticket": ticket_ant,
            }

    def _var(agora, antes):
        """Variacao percentual. Sem base anterior devolve None — nao existe
        'subiu 100%' partindo do zero, isso so engana quem le."""
        if not antes:
            return None
        return round(100 * (agora - antes) / antes, 1)

    if anterior:
        a = anterior["total"]
        agregado["total"]["var"] = {
            "leads": _var(agregado["total"]["leads"], a.get("leads")),
            "sinal": _var(agregado["total"]["sinal"], a.get("sinal")),
            "provavel": _var(agregado["total"].get("provavel", 0), a.get("provavel")),
            "receita_provavel": _var(agregado["total"]["receita_provavel"],
                                     a.get("receita_provavel")),
        }
        for item in agregado["por_canal"]:
            ant = anterior["por_canal"].get(item["canal"])
            item["var"] = {
                "leads": _var(item["leads"], (ant or {}).get("leads")),
                "sinal": _var(item["sinal"], (ant or {}).get("sinal")),
                "receita_provavel": _var(item["receita_provavel"],
                                         (ant or {}).get("receita_provavel")),
            } if ant else None
        anterior["var_vendas"] = {
            "qtd": _var(vendas["qtd"], anterior["vendas"]["qtd"]),
            "total": _var(vendas["total"], anterior["vendas"]["total"]),
        }

        # Conversao e taxa, e taxa se compara em PONTO PERCENTUAL, nao em
        # variacao relativa. De 6,3% pra 5,9% a queda e de 0,4 p.p.; dizer
        # "-6,3%" seria verdade aritmetica e leitura errada — parece que
        # despencou quando andou meio ponto. Os dois vao juntos: o p.p. e o
        # numero pra decidir, o relativo fica de contexto.
        leads_ant = anterior["total"].get("leads") or 0
        conv_ant = (round(100 * anterior["vendas"]["qtd"] / leads_ant, 1)
                    if leads_ant else None)
        conv_agora = (round(100 * vendas["qtd"] / total_leads, 1)
                      if total_leads else None)
        anterior["conversao"] = conv_ant
        anterior["var_conversao"] = (
            {"pp": round(conv_agora - conv_ant, 2),
             "relativo": _var(conv_agora, conv_ant)}
            if conv_ant is not None and conv_agora is not None else None)

        # Investimento do periodo anterior. Refeito com as mesmas fontes do
        # atual — Google por dia, Meta por dia, agencia pelos meses tocados —
        # senao comparar gasto contra gasto estaria comparando bases diferentes.
        ant_google = round(sum(g["spend"] for g in _recortar(
            gasto_bruto.get("linhas", []), anterior["de"], anterior["ate"])), 2)
        ant_meta = round(sum(
            v_dia["spend"] for d_, v_dia in serie_meta.items()
            if anterior["de"] <= d_ <= anterior["ate"]), 2) if serie_meta else 0.0
        ant_meses = {anterior["de"][:7], anterior["ate"][:7]}
        ant_agencia = round(AGENCIA_MENSAL * len(ant_meses), 2)
        ant_total = round(ant_google + ant_meta + ant_agencia, 2)
        anterior["midia"] = {"google": ant_google, "meta": ant_meta,
                             "agencia": ant_agencia, "investimento": ant_total}
        # O investimento atual inclui o ML, que o anterior nao refaz; por isso
        # a comparacao usa as tres fontes que os dois lados tem em comum.
        atual_comparavel = round(
            sum(g["spend"] for g in gasto_linhas) + gasto_meta_periodo + agencia, 2)
        anterior["var_midia"] = {
            "investimento": _var(atual_comparavel, ant_total),
            "google": _var(round(sum(g["spend"] for g in gasto_linhas), 2), ant_google),
            "meta": _var(gasto_meta_periodo, ant_meta),
        }
        anterior["midia"]["atual_comparavel"] = atual_comparavel

    return jsonify({
        "anterior": anterior,
        "ticket_estimativa": {"valor": ticket,
                              "de_vendas_reais": bool(vendas.get("qtd")),
                              "base_qtd": vendas.get("qtd", 0)},
        "de": de, "ate": ate,
        "periodo_efetivo": {"de": ef_de, "ate": ef_ate},
        "gerado_em": leads_bruto.get("gerado_em"),
        "cobertura": cobertura,
        "fora_do_totalk": fora,
        "canais": todos_canais,
        "vendedores": [{"id": vid, "nome": nome} for vid, nome in sorted(
            nomes.items(), key=lambda kv: kv[1])],
        "filtro": {"canal": filtro_canal, "vendedor": filtro_vendedor},
        **agregado,
        "vendas": vendas,
        "conversao": round(100 * vendas["qtd"] / total_leads, 1) if total_leads else None,
        "midia": {
            "investimento": investimento,
            "clicks": cliques,
            "impressions": impressoes,
            # CPC e CTR ficam so no Google: este export do Meta traz conversa
            # iniciada, nao clique, e dividir um pelo outro nao significa nada.
            "cpc": (round(sum(g["spend"] for g in gasto_linhas) / cliques, 2)
                    if cliques else None),
            "ctr": (round(100 * cliques / sum(g["impressions"] for g in gasto_linhas), 2)
                    if sum(g["impressions"] for g in gasto_linhas) else None),
            # Custo por lead e por venda usam o total de leads/vendas do período,
            # não só os que vieram de anúncio: é o custo de mídia por resultado
            # do negócio. Com o gasto do Meta faltando, o número real é maior.
            "custo_por_lead": round(investimento / total_leads, 2) if total_leads else None,
            "custo_por_venda": round(investimento / vendas["qtd"], 2) if vendas["qtd"] else None,
            # Nao chamamos isso de ROAS: o faturamento aqui e o total do
            # periodo, nao o atribuido a anuncio, e falta o gasto do Meta. E
            # "quanto o negocio faturou por real de midia paga", nada mais.
            "faturamento_por_real": (round(vendas["total"] / investimento, 1)
                                     if investimento else None),
            "campanhas": campanhas,
            "por_dia": [{"data": k, "spend": v} for k, v in sorted(gasto_dia.items())],
            "fontes_ausentes": gasto_bruto.get("fontes_ausentes", []),
            "google": {"investimento": round(sum(g["spend"] for g in gasto_linhas), 2),
                       "clicks": cliques,
                       # Campanha que traz gente ate a LOJA (mapa, ligacao) nao
                       # se mede pela regua do site: ela converte em telefone
                       # tocando e gente na porta, que o site nao registra.
                       # Somadas, elas inflavam o custo por venda online e
                       # faziam a campanha local parecer ruim sem ser.
                       "por_objetivo": _google_por_objetivo(gasto_linhas),
                       "por_campanha": _google_por_campanha(gasto_linhas)},
            "meta": meta,
            "meta_rateio": meta_rateio,
            "ml_rateio": ml_rateio,
            # Quando cada fonte foi atualizada pela ultima vez. No plano basico
            # do Windsor so uma conta fica conectada por vez, entao a outra
            # fica congelada — a tela precisa dizer isso.
            "atualizado_em": gasto_bruto.get("atualizado_em") or {},
            "fontes_ausentes": gasto_bruto.get("fontes_ausentes") or [],
            "agencia": {"mensal": AGENCIA_MENSAL, "no_periodo": agencia,
                        "meses": len(meses_agencia)},
        },
    })


@app.route("/api/admin/desempenho")
def api_admin_desempenho():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    vendedores = carregar_vendedores()
    vid = request.args.get("vendedor", "")
    if vid not in vendedores:
        return jsonify({"erro": "Vendedor não encontrado."}), 404
    return jsonify(montar_desempenho(vid, request.args.get("mes"), vendedores))


@app.route("/api/desempenho")
def api_desempenho_vendedor():
    """O mesmo painel, do ponto de vista de quem está logado.

    Uma diferença de propósito: sai a participação no faturamento do time. Com
    ela e o próprio total, o vendedor deduziria quanto a equipe inteira vendeu —
    e o gestor acabou de tirar o ranking do menu dele justamente pra isso não
    ficar exposto. A posição fica: motiva sem entregar número de ninguém."""
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401
    dados = montar_desempenho(vendedor_id, request.args.get("mes"), carregar_vendedores())
    dados["time"] = {"posicao": dados["time"]["posicao"], "de": dados["time"]["de"]}
    return jsonify(dados)


def montar_desempenho(vid, mes, vendedores):
    """Tudo o que dá pra medir de um vendedor sozinho, num mês.

    O que entra aqui é só o que existe em 100% das vendas (data, valor,
    produto). `canal` está preenchido em 4% dos registros e com grafia
    inconsistente, então não vira métrica — mostraria um mix falso. Leads,
    conversas e taxa de conversão não vivem neste portal: a fila do follow-up
    traz só quem não fechou, e sem o total de atendimentos não existe
    denominador. Esse número tem que vir do vendas-insights.
    """
    mes = mes or hoje_br().isoformat()[:7]
    vendas = carregar_vendas_todos(vendedores)
    metas_todas = carregar_metas()

    minhas = [v for v in vendas.values()
              if v["vendedor_id"] == vid and v.get("tipo", "venda") == "venda"]
    bonus = [v for v in vendas.values()
             if v["vendedor_id"] == vid and v.get("tipo") == "bonus"]

    def do_mes(lista, alvo):
        return [v for v in lista if v["data"][:7] == alvo]

    def mes_anterior(alvo):
        ano, m = int(alvo[:4]), int(alvo[5:7])
        return f"{ano - 1}-12" if m == 1 else f"{ano}-{m - 1:02d}"

    def bloco(alvo):
        """Os números de um mês. Serve pro mês escolhido e pro anterior, que é
        o que dá sentido à variação."""
        itens = do_mes(minhas, alvo)
        total = round(sum(valor_liquido(v) for v in itens), 2)
        qtd = len(itens)
        dias = {v["data"] for v in itens}
        return {
            "mes": alvo,
            "total": total,
            "qtd": qtd,
            "ticket": round(total / qtd, 2) if qtd else 0.0,
            "dias_ativos": len(dias),
            "media_dia_ativo": round(total / len(dias), 2) if dias else 0.0,
        }

    atual = bloco(mes)
    anterior = bloco(mes_anterior(mes))

    def variacao(agora, antes):
        if not antes:
            return None          # sem base de comparação: não inventa 100%
        return round(100 * (agora - antes) / antes, 1)

    # ---- ritmo do mês ----
    dias_no_mes = calendar.monthrange(int(mes[:4]), int(mes[5:7]))[1]
    hoje = hoje_br()
    dias_corridos = hoje.day if mes == hoje.isoformat()[:7] else dias_no_mes
    meta_mensal = float(metas_vendedor(vid, metas_todas).get("mensal", 0) or 0)

    # ---- histórico: evolução e metas batidas ----
    por_mes = {}
    for v in minhas:
        chave = v["data"][:7]
        d = por_mes.setdefault(chave, {"total": 0.0, "qtd": 0})
        d["total"] += valor_liquido(v)
        d["qtd"] += 1
    historico = [{
        "mes": k,
        "total": round(d["total"], 2),
        "qtd": d["qtd"],
        "ticket": round(d["total"] / d["qtd"], 2) if d["qtd"] else 0.0,
        # Meta é a de hoje aplicada a todo o histórico: o portal não guarda a
        # meta que valia em cada mês passado. Serve pra tendência, não pra
        # cobrar mês fechado.
        "bateu": bool(meta_mensal) and d["total"] >= meta_mensal,
    } for k, d in sorted(por_mes.items())]

    # ---- dentro do mês ----
    por_dia = {}
    for v in do_mes(minhas, mes):
        por_dia[v["data"]] = por_dia.get(v["data"], 0.0) + valor_liquido(v)
    serie_dia = [{"data": k, "total": round(x, 2)} for k, x in sorted(por_dia.items())]
    melhor_dia = max(serie_dia, key=lambda x: x["total"], default=None)

    itens_mes = do_mes(minhas, mes)
    maior_venda = max(itens_mes, key=lambda v: valor_liquido(v), default=None)

    # Faixas de valor: mostram se o mês veio de muita peça barata ou de poucas
    # caras — duas rotas bem diferentes pro mesmo faturamento.
    FAIXAS = [(0, 200, "até R$ 200"), (200, 500, "R$ 200–500"),
              (500, 1000, "R$ 500–1 mil"), (1000, 3000, "R$ 1–3 mil"),
              (3000, float("inf"), "acima de R$ 3 mil")]
    faixas = []
    for piso, teto, rotulo in FAIXAS:
        dentro = [v for v in itens_mes if piso <= valor_liquido(v) < teto]
        faixas.append({"rotulo": rotulo, "qtd": len(dentro),
                       "total": round(sum(valor_liquido(v) for v in dentro), 2)})

    # Nao existe "top produto" util aqui: produto e texto livre digitado a cada
    # venda e quase nunca se repete igual, entao agrupar por nome so devolveria
    # a maior venda com quantidade 1. O que informa de verdade e a lista das
    # maiores vendas do mes.
    maiores_vendas = [{"produto": (v.get("produto") or "Sem descrição").strip(),
                       "valor": valor_liquido(v), "data": v["data"]}
                      for v in sorted(itens_mes, key=valor_liquido, reverse=True)[:8]]

    # ---- posição no time ----
    totais_time = sorted(
        ((outro, round(sum(valor_liquido(v) for v in vendas.values()
                           if v["vendedor_id"] == outro
                           and v.get("tipo", "venda") == "venda"
                           and v["data"][:7] == mes), 2))
         for outro in vendedores),
        key=lambda kv: kv[1], reverse=True)
    # Conta oculta zerada nao entra no "2º de 5": inflava o tamanho do time.
    totais_time = [(o, t) for o, t in totais_time
                   if t or not vendedores[o].get("oculto") or o == vid]
    posicao = next((i + 1 for i, (outro, _) in enumerate(totais_time) if outro == vid), None)
    total_time = round(sum(t for _, t in totais_time), 2)

    # ---- follow-up ----
    fila = carregar_fila_retomada(vid)
    followup = None
    if fila:
        itens_fila = fila.get("itens", [])
        st = carregar_status_retomada(vid)
        resumo = _resumo_retomada(itens_fila, st)
        trabalhados = resumo["trabalhados"]
        followup = {
            **resumo,
            "pct_trabalhado": round(100 * trabalhados / len(itens_fila)) if itens_fila else 0,
            "pct_resposta": (round(100 * (resumo["respondeu"] + resumo["vendeu"]) / trabalhados)
                             if trabalhados else 0),
        }

    # ---- atendimento (vem do vendas-insights, espelho do Totalk) ----
    # O portal so sabe o que virou venda. Quantos clientes ele atendeu, de que
    # canal vieram e quanto demorou a primeira resposta vive no outro projeto e
    # chega aqui agregado, pela chave insights_<vendedor>.
    insights = ler_json(resolver_pasta_dados() / f"insights_{vid}.json", None)
    atendimento = None
    if insights:
        do_mes_ins = (insights.get("meses") or {}).get(mes)
        if do_mes_ins:
            atend = do_mes_ins.get("atendimentos", 0)
            atendimento = {
                **do_mes_ins,
                "gerado_em": insights.get("gerado_em"),
                # Conversao de verdade: venda lancada no portal dividida pelos
                # clientes que ele atendeu. Nao usamos o "virou_venda" da IA
                # porque o fechamento acontece fora do chat — ela enxerga so uma
                # fracao, e o numero sairia baixo demais.
                "taxa_conversao": (round(100 * atual["qtd"] / atend, 1)
                                   if atend else None),
                "vendas_no_mes": atual["qtd"],
            }

    de_mes, ate_mes = f"{mes}-01", f"{mes}-{dias_no_mes:02d}"
    comissao = calcular_comissao(vid, de_mes, ate_mes, vendedores, vendas)

    return {
        "vendedor": {"id": vid, "nome": vendedores[vid]["nome"],
                     "foto": vendedores[vid].get("foto"),
                     "avatar": vendedores[vid].get("avatar", ""),
                     "percentual": float(vendedores[vid].get("percentual", 0))},
        "mes": mes,
        "atual": atual,
        "anterior": anterior,
        "variacao": {
            "total": variacao(atual["total"], anterior["total"]),
            "qtd": variacao(atual["qtd"], anterior["qtd"]),
            "ticket": variacao(atual["ticket"], anterior["ticket"]),
        },
        "meta": {
            "mensal": meta_mensal,
            "pct": round(100 * atual["total"] / meta_mensal, 1) if meta_mensal else None,
            "falta": round(max(0.0, meta_mensal - atual["total"]), 2) if meta_mensal else None,
            "batidas": sum(1 for h in historico if h["bateu"]),
            "meses_com_venda": len(historico),
        },
        "ritmo": {"dias_no_mes": dias_no_mes, "dias_corridos": dias_corridos,
                  "pct_do_mes": round(100 * dias_corridos / dias_no_mes),
                  "projecao": round(atual["total"] / dias_corridos * dias_no_mes, 2)
                              if dias_corridos else 0.0},
        "comissao": {"valor": comissao["comissao"], "bonus": round(
            sum(v["valor"] for v in do_mes(bonus, mes)), 2)},
        "historico": historico,
        "serie_dia": serie_dia,
        "melhor_dia": melhor_dia,
        "maior_venda": ({"produto": maior_venda.get("produto"),
                         "valor": valor_liquido(maior_venda),
                         "data": maior_venda["data"]} if maior_venda else None),
        "faixas": faixas,
        "maiores_vendas": maiores_vendas,
        "time": {"posicao": posicao, "de": len(totais_time),
                 "total_time": total_time,
                 "participacao": round(100 * atual["total"] / total_time, 1) if total_time else 0},
        "followup": followup,
        "atendimento": atendimento,
    }


@app.route("/api/admin/exportar-mes-xlsx")
def api_admin_exportar_mes_xlsx():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401

    mes = request.args.get("mes", "")
    if len(mes) != 7 or mes[4] != "-":
        return jsonify({"erro": "Mês inválido."}), 400
    de, ate = f"{mes}-01", f"{mes}-31"

    vendedores = carregar_vendedores()
    vendas = carregar_vendas_todos(vendedores)

    wb = Workbook()
    wb.remove(wb.active)
    cabecalho = ["Data", "Produto", "SKU", "Canal", "Valor", "Devolução", "Valor Devolvido", "Valor Líquido"]

    for vid in sorted(vendedores, key=lambda x: vendedores[x]["nome"]):
        lista = [
            v for v in vendas.values()
            if v["vendedor_id"] == vid and de <= v["data"] <= ate and v.get("tipo", "venda") == "venda"
        ]
        lista.sort(key=lambda v: v["data"])
        # Conta oculta sem venda no mes nao vira aba vazia no arquivo que o
        # gestor distribui. Com venda, a aba sai — dinheiro nunca some.
        if not lista and vendedores[vid].get("oculto"):
            continue

        ws = wb.create_sheet(_nome_aba_excel(vendedores[vid]["nome"]))
        ws.append(cabecalho)
        for v in lista:
            dev = v.get("devolucao")
            dev_tipo = ("Total" if dev.get("tipo") == "total" else "Parcial") if dev else ""
            dev_valor = dev.get("valor_devolvido") if dev else None
            ws.append([
                v["data"],
                v["produto"],
                v.get("sku", ""),
                v.get("canal", ""),
                v["valor"],
                dev_tipo,
                dev_valor,
                valor_liquido(v),
            ])
        for coluna, largura in zip("ABCDEFGH", (12, 42, 12, 14, 12, 12, 14, 14)):
            ws.column_dimensions[coluna].width = largura

    if not wb.sheetnames:
        wb.create_sheet("Vendedores").append(cabecalho)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"vendas_{mes}.xlsx",
    )


@app.route("/api/admin/vendedores", methods=["GET"])
def api_admin_listar_vendedores():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    vendedores = carregar_vendedores()
    return jsonify([
        {
            "id": vid,
            "nome": v["nome"],
            "percentual": v.get("percentual", 0),
            "overrides": v.get("overrides", []),
            "foto": v.get("foto"),
            "avatar": v.get("avatar", ""),
            "oculto": bool(v.get("oculto")),
            "perfil": v.get("perfil") or "vendedor",
            "liberacao_retroativa": retroativo_ativo(v),
            "liberacao_retroativa_ate": v.get("liberacao_retroativa_ate") if retroativo_ativo(v) else None,
        }
        for vid, v in sorted(vendedores.items(), key=lambda kv: kv[1]["nome"])
    ])


@app.route("/api/admin/vendedores/<vendedor_id>/liberar-retroativo", methods=["POST"])
def api_admin_liberar_retroativo(vendedor_id):
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    vendedores = carregar_vendedores()
    if vendedor_id not in vendedores:
        return jsonify({"erro": "Vendedor não encontrado."}), 404
    ate = agora_br() + timedelta(minutes=LIBERACAO_RETROATIVA_MINUTOS)
    vendedores[vendedor_id]["liberacao_retroativa_ate"] = ate.isoformat(timespec="seconds")
    salvar_vendedores(vendedores)
    return jsonify({"ok": True, "liberacao_retroativa_ate": vendedores[vendedor_id]["liberacao_retroativa_ate"]})


@app.route("/api/admin/vendedores/<vendedor_id>/liberar-retroativo", methods=["DELETE"])
def api_admin_cancelar_liberacao_retroativo(vendedor_id):
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    vendedores = carregar_vendedores()
    if vendedor_id not in vendedores:
        return jsonify({"erro": "Vendedor não encontrado."}), 404
    vendedores[vendedor_id].pop("liberacao_retroativa_ate", None)
    salvar_vendedores(vendedores)
    return jsonify({"ok": True})


@app.route("/api/admin/vendedores", methods=["POST"])
def api_admin_salvar_vendedor():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    body = request.get_json(force=True)
    vendedor_id = (body.get("id") or "").strip().lower()
    nome = (body.get("nome") or "").strip()
    senha = body.get("senha")
    try:
        percentual = float(body.get("percentual"))
    except (TypeError, ValueError):
        return jsonify({"erro": "Percentual inválido."}), 400
    if not vendedor_id or not nome:
        return jsonify({"erro": "Informe id e nome do vendedor."}), 400
    if percentual < 0 or percentual > 100:
        return jsonify({"erro": "Percentual deve estar entre 0 e 100."}), 400

    overrides = []
    for over in body.get("overrides", []):
        outro_id = (over.get("vendedor_id") or "").strip().lower()
        try:
            outro_percentual = float(over.get("percentual"))
        except (TypeError, ValueError):
            continue
        if not outro_id or outro_id == vendedor_id or outro_percentual <= 0:
            continue
        overrides.append({"vendedor_id": outro_id, "percentual": outro_percentual})

    vendedores = carregar_vendedores()
    existente = vendedores.get(vendedor_id, {})
    if not senha and not existente.get("senha"):
        return jsonify({"erro": "Defina uma senha para o vendedor."}), 400
    vendedores[vendedor_id] = {
        "nome": nome,
        "senha": senha if senha else existente.get("senha"),
        "percentual": percentual,
        "overrides": overrides,
    }
    if existente.get("foto"):
        vendedores[vendedor_id]["foto"] = existente["foto"]
    # Flags que a tela de edicao nao conhece sobrevivem a edicao. Sem isso,
    # editar a senha do caique ressuscitava a conta nas listas (oculto sumia) e
    # editar qualquer vendedor cancelava a liberacao retroativa em silencio.
    for flag in ("oculto", "liberacao_retroativa_ate"):
        if existente.get(flag):
            vendedores[vendedor_id][flag] = existente[flag]
    perfil = (body.get("perfil") or "").strip().lower()
    if perfil in ("vendedor", "expedicao"):
        vendedores[vendedor_id]["perfil"] = perfil
    elif existente.get("perfil") and "perfil" not in body:
        vendedores[vendedor_id]["perfil"] = existente["perfil"]
    # Avatar generico usado quando nao ha foto. E escolha do gestor, nunca
    # deduzida do nome — nome nao diz genero de ninguem.
    avatar = (body.get("avatar") or "").strip().lower()
    if avatar in ("feminino", "masculino"):
        vendedores[vendedor_id]["avatar"] = avatar
    elif existente.get("avatar") and "avatar" not in body:
        vendedores[vendedor_id]["avatar"] = existente["avatar"]
    codigo_recuperacao = (body.get("codigo_recuperacao") or "").strip().upper()
    if codigo_recuperacao:
        if len(codigo_recuperacao) < 6:
            return jsonify({"erro": "O código de recuperação precisa ter pelo menos 6 caracteres."}), 400
        vendedores[vendedor_id]["recuperacao_hash"] = _hash_codigo(codigo_recuperacao)
    elif existente.get("recuperacao_hash"):
        vendedores[vendedor_id]["recuperacao_hash"] = existente["recuperacao_hash"]
    salvar_vendedores(vendedores)
    return jsonify({"ok": True})


@app.route("/api/recuperar-senha-vendedor", methods=["POST"])
def api_recuperar_senha_vendedor():
    if excedeu_tentativas_login("vendedor_recuperacao"):
        return jsonify({"erro": f"Muitas tentativas erradas. Aguarde {LOGIN_JANELA_MINUTOS} minutos e tente de novo."}), 429

    body = request.get_json(force=True)
    vendedor_id = (body.get("vendedor_id") or "").strip().lower()
    codigo = (body.get("codigo") or "").strip().upper()
    nova_senha = body.get("nova_senha") or ""

    vendedores = carregar_vendedores()
    info = vendedores.get(vendedor_id)
    hash_salvo = info.get("recuperacao_hash") if info else None
    if not info or not hash_salvo or _hash_codigo(codigo) != hash_salvo:
        registrar_acesso("vendedor_recuperacao", False, vendedor_id)
        return jsonify({"erro": "Código inválido."}), 401

    if len(nova_senha) < 4:
        return jsonify({"erro": "A nova senha precisa ter pelo menos 4 caracteres."}), 400

    vendedores[vendedor_id]["senha"] = nova_senha
    vendedores[vendedor_id].pop("recuperacao_hash", None)
    salvar_vendedores(vendedores)
    registrar_acesso("vendedor_recuperacao", True, vendedor_id)
    return jsonify({"ok": True})


@app.route("/api/admin/vendedores/<vendedor_id>", methods=["DELETE"])
def api_admin_remover_vendedor(vendedor_id):
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    vendedores = carregar_vendedores()
    if vendedor_id not in vendedores:
        return jsonify({"erro": "Vendedor não encontrado."}), 404
    del vendedores[vendedor_id]
    salvar_vendedores(vendedores)
    return jsonify({"ok": True})


@app.route("/api/admin/metas", methods=["POST"])
def api_admin_salvar_metas():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    body = request.get_json(force=True)
    vendedores = carregar_vendedores()

    def limpar_trio(dados):
        resultado = {}
        for chave in ("diaria", "semanal", "mensal"):
            try:
                resultado[chave] = max(0.0, float(dados.get(chave, 0)))
            except (TypeError, ValueError):
                resultado[chave] = 0.0
        return resultado

    metas = {
        "grupo": limpar_trio(body.get("grupo", {})),
        "vendedores": {
            vid: limpar_trio(body.get("vendedores", {}).get(vid, {}))
            for vid in vendedores
        },
    }
    salvar_metas(metas)
    return jsonify({"ok": True, "metas": metas})


def _supabase_storage_upload(nome_arquivo: str, conteudo: bytes, content_type: str) -> None:
    url = f"{SUPABASE_URL}/storage/v1/object/fotos/{nome_arquivo}"
    req = urllib.request.Request(url, data=conteudo, method="POST", headers={
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "apikey": SUPABASE_SERVICE_KEY,
        "Content-Type": content_type,
        "x-upsert": "true",
    })
    urllib.request.urlopen(req).read()


def _supabase_storage_delete(nome_arquivo: str) -> None:
    url = f"{SUPABASE_URL}/storage/v1/object/fotos/{nome_arquivo}"
    req = urllib.request.Request(url, method="DELETE", headers={
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "apikey": SUPABASE_SERVICE_KEY,
    })
    try:
        urllib.request.urlopen(req).read()
    except urllib.error.HTTPError:
        pass


@app.route("/api/admin/vendedores/<vendedor_id>/foto", methods=["POST"])
def api_admin_upload_foto(vendedor_id):
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    vendedores = carregar_vendedores()
    if vendedor_id not in vendedores:
        return jsonify({"erro": "Vendedor não encontrado."}), 404

    arquivo = request.files.get("foto")
    if not arquivo or not arquivo.filename:
        return jsonify({"erro": "Nenhum arquivo enviado."}), 400
    ext = arquivo.filename.rsplit(".", 1)[-1].lower() if "." in arquivo.filename else ""
    if ext not in EXTENSOES_FOTO_PERMITIDAS:
        return jsonify({"erro": "Formato inválido. Use JPG, PNG ou WEBP."}), 400

    foto_antiga = vendedores[vendedor_id].get("foto")
    nome_arquivo = f"{vendedor_id}.{ext}"

    if SUPABASE_URL:
        if foto_antiga:
            _supabase_storage_delete(foto_antiga)
        _supabase_storage_upload(nome_arquivo, arquivo.read(), arquivo.content_type or "application/octet-stream")
    else:
        FOTOS_DIR.mkdir(parents=True, exist_ok=True)
        if foto_antiga:
            (FOTOS_DIR / foto_antiga).unlink(missing_ok=True)
        arquivo.save(FOTOS_DIR / nome_arquivo)

    vendedores[vendedor_id]["foto"] = nome_arquivo
    salvar_vendedores(vendedores)
    return jsonify({"ok": True, "foto": nome_arquivo})


@app.route("/fotos/<path:filename>")
def servir_foto(filename):
    if SUPABASE_URL:
        return redirect(f"{SUPABASE_URL}/storage/v1/object/public/fotos/{filename}")
    return send_from_directory(FOTOS_DIR, filename)


@app.route("/simulador")
def pagina_simulador():
    return send_from_directory(STATIC_DIR, "simulador.html")


@app.route("/api/admin/simulador/regras", methods=["GET"])
def api_admin_simulador_regras_get():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    regras = carregar_regras_simulador()
    if regras is None:
        return jsonify({"erro": "As regras de desconto ainda não foram importadas."}), 404
    return jsonify(regras)


@app.route("/api/admin/simulador/regras", methods=["PUT"])
def api_admin_simulador_regras_put():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    corpo = request.get_json(silent=True) or {}
    regras_atuais = carregar_regras_simulador() or {}
    for chave in ("desconto_max_pct", "nivel_flexibilidade", "parcelas_max",
                  "valor_minimo_parcelamento", "faixas_tempo", "faixas_valor"):
        if chave in corpo:
            regras_atuais[chave] = corpo[chave]
    regras_atuais["atualizado_em"] = agora_br().isoformat()
    regras_atuais["atualizado_por"] = "gestor"
    salvar_regras_simulador(regras_atuais)
    return jsonify(regras_atuais)


@app.route("/api/admin/simulador/status")
def api_admin_simulador_status():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    return jsonify(status_simulador())


@app.route("/api/admin/simulador/reimportar", methods=["POST"])
def api_admin_simulador_reimportar():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    script = ROOT / "scripts" / "etl_simulador.py"
    resultado = subprocess.run(
        [sys.executable, str(script)],
        capture_output=True, text=True, cwd=str(ROOT), timeout=900,
    )
    if resultado.returncode != 0:
        log = (resultado.stdout[-4000:] + "\n" + resultado.stderr[-4000:])
        return jsonify({"ok": False, "log": log}), 500
    return jsonify({"ok": True, "log": resultado.stdout[-4000:]})


@app.route("/api/simulador/regras")
def api_simulador_regras():
    if not exigir_vendedor():
        return jsonify({"erro": "Não autenticado."}), 401
    regras = carregar_regras_simulador()
    if regras is None:
        return jsonify({"erro": "As regras de desconto ainda não foram sincronizadas."}), 404
    return jsonify(regras)


@app.route("/api/simulador/buscar")
def api_simulador_buscar():
    if not exigir_vendedor():
        return jsonify({"erro": "Não autenticado."}), 401
    q = (request.args.get("q") or "").strip()
    if len(q) < 2:
        return jsonify([])
    return jsonify(buscar_pecas_simulador(q))


@app.route("/api/simulador/simular", methods=["POST"])
def api_simulador_simular():
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401
    corpo = request.get_json(silent=True) or {}
    cod_peca = (corpo.get("cod_peca") or "").strip()
    if not cod_peca:
        return jsonify({"erro": "cod_peca é obrigatório"}), 400

    peca = obter_peca_simulador(cod_peca)
    if not peca:
        return jsonify({"erro": "peça não encontrada"}), 404

    regras = carregar_regras_simulador()
    if regras is None:
        return jsonify({"erro": "As regras de desconto ainda não foram sincronizadas."}), 404

    valor_override = corpo.get("valor_base")
    desconto_escolhido = corpo.get("desconto_pct")
    resultado = calcular_simulacao_peca(peca, valor_override, desconto_escolhido, regras)
    resultado["cod_peca"] = peca["cod_peca"]
    resultado["nome_produto"] = peca["nome_produto"]
    resultado["etiqueta"] = peca["etiqueta"]
    resultado["apelido_veiculo"] = peca["apelido_veiculo"]
    resultado["tipo_peca_rotulo"] = peca["tipo_peca_rotulo"]
    return jsonify(resultado)


@app.route("/api/simulador/simular-rapido", methods=["POST"])
def api_simulador_simular_rapido():
    """Simulação sem buscar peça no catálogo: o vendedor informa valor,
    curva e faixa de tempo em estoque na mão — pensado pra atender o
    cliente rápido no balcão, sem precisar achar o item no sistema."""
    if not exigir_vendedor():
        return jsonify({"erro": "Não autenticado."}), 401
    corpo = request.get_json(silent=True) or {}
    try:
        valor_base = float(corpo.get("valor_base"))
    except (TypeError, ValueError):
        return jsonify({"erro": "valor_base é obrigatório e deve ser numérico"}), 400
    if valor_base <= 0:
        return jsonify({"erro": "valor_base deve ser maior que zero"}), 400

    regras = carregar_regras_simulador()
    if regras is None:
        return jsonify({"erro": "As regras de desconto ainda não foram sincronizadas."}), 404

    curva = (corpo.get("curva") or "").strip()
    if curva not in regras["desconto_max_pct"]:
        return jsonify({"erro": "curva inválida"}), 400

    faixa_tempo_id = (corpo.get("faixa_tempo_id") or "").strip()
    ids_validos = {f["id"] for f in regras["faixas_tempo"]}
    if faixa_tempo_id not in ids_validos:
        return jsonify({"erro": "faixa_tempo_id inválida"}), 400
    dias_representativos = next(f["min_dias"] for f in regras["faixas_tempo"] if f["id"] == faixa_tempo_id)

    desconto_escolhido = corpo.get("desconto_pct")
    resultado = montar_simulacao(valor_base, curva, dias_representativos, desconto_escolhido, regras)
    resultado["dias_em_estoque"] = None  # é uma faixa escolhida à mão, não uma data real
    return jsonify(resultado)


@app.route("/api/simulador/peca/<cod_peca>/data-entrada", methods=["POST"])
def api_simulador_definir_data_entrada(cod_peca):
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401
    corpo = request.get_json(silent=True) or {}
    data_entrada = (corpo.get("data_entrada") or "").strip()
    if not data_entrada:
        return jsonify({"erro": "data_entrada é obrigatória"}), 400
    try:
        data_parseada = datetime.fromisoformat(data_entrada)
    except ValueError:
        return jsonify({"erro": "data inválida, use AAAA-MM-DD"}), 400

    nome = carregar_vendedores().get(vendedor_id, {}).get("nome", vendedor_id)
    ok = definir_data_entrada_simulador(cod_peca, data_parseada.strftime("%Y-%m-%d 00:00:00"), nome)
    if not ok:
        return jsonify({"erro": "peça não encontrada"}), 404
    return jsonify({"ok": True})


# ============================================================
# Retomada — CRM dos clientes que não fecharam
# ============================================================
# A fila não é montada aqui. Ela vem do projeto vendas-insights
# (app/gerar_fila_retomada.py), que lê as conversas do Totalk, classifica cada
# uma com IA e decide quem vale uma segunda tentativa; de lá ela é empurrada
# pra cá por vendas-insights/app/sincronizar_crm.py. Este arquivo só mostra a
# fila do vendedor logado e guarda o que ele marcou.
#
# Duas chaves separadas de propósito: `crm_fila_<id>` é substituída inteira a
# cada sincronização, `crm_status_<id>` é do vendedor e a sincronização nunca
# encosta nela. Se fossem a mesma chave, regerar a fila apagaria o trabalho
# já feito.

STATUS_RETOMADA = {
    "pendente": "Não chamei ainda",
    "chamei": "Chamei, sem resposta",
    "respondeu": "Respondeu",
    "vendeu": "Fechou venda",
    "perdido": "Não vai rolar",
}
STATUS_TRABALHADO = ("chamei", "respondeu", "vendeu", "perdido")

# Quantos clientes a chamar aparecem por vez. O resto fica de fora da tela
# (nao e apagado) — fila gigante desanima e ninguem trabalha.
MAX_FILA_PENDENTES = 50

# Meta de contatos por dia. Fila grande so anda com alvo diario: o vendedor
# precisa saber quando pode parar. Chamou 10 fez o minimo, 20 bateu a meta.
META_CONTATOS_DIA = {"minimo": 10, "bom": 15, "meta": 20}


def _caminho_crm(prefixo: str, vendedor_id: str) -> Path:
    return resolver_pasta_dados() / f"crm_{prefixo}_{vendedor_id}.json"


def carregar_fila_retomada(vendedor_id: str):
    # Padrão None (nunca Path.exists()): em modo banco o arquivo local não
    # existe, e checar o disco faria a fila parecer vazia em produção.
    return ler_json(_caminho_crm("fila", vendedor_id), None)


def carregar_status_retomada(vendedor_id: str) -> dict:
    return ler_json(_caminho_crm("status", vendedor_id), None) or {}


def _contatos_de_hoje(status: dict) -> int:
    """Quantos clientes o vendedor trabalhou hoje. Conta a marcacao, nao o
    cliente: se ele marcou e depois corrigiu, continua sendo um contato feito."""
    hoje = hoje_br().isoformat()
    return sum(1 for m in status.values() if (m.get("em") or "").startswith(hoje))


def _resumo_retomada(itens: list, status: dict) -> dict:
    contagem = {chave: 0 for chave in STATUS_RETOMADA}
    for item in itens:
        atual = (status.get(item["sid"]) or {}).get("status", "pendente")
        contagem[atual if atual in contagem else "pendente"] += 1
    contagem["total"] = len(itens)
    contagem["trabalhados"] = sum(contagem[s] for s in STATUS_TRABALHADO)
    return contagem


# ---------- Mensagem pronta pra cada cliente ----------
# O vendedor não deixa de chamar porque não sabe o que dizer — deixa porque dá
# preguiça de escrever a mesma coisa 50 vezes. Então cada ficha já chega com o
# texto montado, e ele só revisa e envia.
#
# A situação sai de dois sinais que a fila já traz: `gancho` (por que a conversa
# morreu, sempre preenchido) e o assunto da última fala do cliente. O assunto
# manda quando dá pra identificar, porque responder o que ele perguntou converte
# mais que uma reativação genérica; sem assunto, cai no gancho.
#
# `tinha` também pesa: quando a peça exata não era nossa ("Parecida"), a
# mensagem não pode prometer "tenho ela aqui" — isso queima o vendedor na hora.

# Assunto pela última fala do cliente. Ordem importa: a lista é percorrida de
# cima pra baixo e o primeiro que casar vence, então o mais decisivo vem antes.
ASSUNTOS_MSG = [
    ("fechar",  r"vou compr|quero compr|pode separar|vou fechar|realizar a compra|manda o pix|mandar o pix|fazer o pix"),
    ("foto",    r"\bfoto|\bvídeo|\bvideo|imagem|manda uma foto|ver a peç"),
    ("compat",  r"\bserve\b|compatív|compativ|código|codigo|numeraç|original|se encaixa|dá certo no"),
    ("frete",   r"\bfrete|entreg|\bprazo|quanto tempo|quantos dias|chega em|correio|transportadora|sedex"),
    ("terceiro", r"mecânic|mecanic|meu marido|minha esposa|patrão|patrao|meu chefe|confirmaç[aã]o do|falar com o"),
    ("preco",   r"\bpreç|\bpreco|\bvalor|quanto (fica|custa|sai|é)|desconto|mais barat|melhor preç|tá caro|ta caro"),
    ("pensar",  r"vou ver|vou pensar|depois eu|te falo|semana que vem|mês que vem|mes que vem|mais pra frente|qualquer coisa eu"),
]

# Por que a conversa morreu (campo `gancho` da fila).
GANCHOS_MSG = {
    "A conversa parou do nosso lado": "nosso_lado",
    "Respondemos tudo e ele sumiu": "sumiu",
    "Conversou e não fechou": "nao_fechou",
    "Achou caro — cabe negociar": "caro",
    "Travou no frete ou no prazo": "frete",
}

# Como cada situação aparece pro gestor na tela de edição.
ROTULOS_SITUACAO = {
    "fechar": "Já ia comprar",
    "foto": "Pediu foto ou vídeo",
    "compat": "Perguntou se serve no carro dele",
    "frete": "Perguntou frete ou prazo",
    "terceiro": "Ia confirmar com o mecânico",
    "preco": "Falou de preço ou desconto",
    "pensar": "Disse que ia pensar",
    "nosso_lado": "A conversa parou do nosso lado",
    "sumiu": "Respondemos tudo e ele sumiu",
    "nao_fechou": "Conversou e não fechou",
    "caro": "Achou caro",
}

# Os textos evitam de propósito pronome e adjetivo com gênero ("tenho ela
# separada", "ainda está disponível pra você"): a peça vem do texto livre da
# conversa e não dá pra saber o gênero com segurança. Onde precisa de artigo,
# entra {a}, que é calculado por peça. O resto fala "a peça", que é sempre
# feminino e nunca erra.
MODELOS_PADRAO = {
    "saudacao": "Oi {nome}, tudo bem?",
    # Quando a bola ficou com a gente, pedir desculpa é a abertura certa: o
    # cliente não sumiu, nós que não voltamos.
    "saudacao_atraso": "Oi {nome}, tudo bem? Desculpa a demora pra te responder.",
    "corpo": {
        "fechar": "Vi que você ia fechar {a} {peca} e a gente acabou não concluindo. Ainda não vendi essa peça — quer que eu te mande os dados do pix?",
        "foto": "Sobre {a} {peca}: tenho fotos e vídeo aqui. Quer que eu te mande pra você conferir?",
        "compat": "Sobre {a} {peca}: consigo confirmar a compatibilidade pelo número da peça. Me passa o ano e o modelo do carro que eu te garanto se serve.",
        "frete": "Sobre {a} {peca}: consigo te confirmar o prazo e o frete certinho. Me passa seu CEP que eu calculo agora.",
        "terceiro": "Você ia confirmar {a} {peca}. Conseguiu falar com o mecânico? A peça continua aqui.",
        "preco": "Sobre {a} {peca} que você olhou com a gente: consigo ver uma condição melhor pra fechar. Ainda está precisando?",
        "pensar": "Passando pra saber se você chegou a decidir sobre {a} {peca}. Ainda tenho a peça disponível.",
        "nosso_lado": "Sobre {a} {peca} que você procurou: ficou faltando eu te retornar. Ainda está precisando?",
        "sumiu": "Sobre {a} {peca} que a gente conversou: ainda está precisando? A peça continua aqui comigo.",
        "nao_fechou": "Sobre {a} {peca} que você procurou com a gente: ainda tenho aqui. Quer que eu retome o orçamento?",
        "caro": "Sobre {a} {peca}: sei que o valor pesou. Me fala quanto você conseguiria pagar que eu vejo o que dá pra fazer por você.",
    },
    # Trocas pra quando a peça exata não era nossa (`tinha` = "Parecida"):
    # aqui a gente não tem o que prometer, tem o que oferecer.
    "corpo_parecida": {
        "fechar": "Vi que você ia fechar {a} {peca}. Consegui uma opção compatível aqui — quer que eu te mande os detalhes?",
        "terceiro": "Você ia confirmar {a} {peca}. Deu certo com o mecânico? Consigo uma compatível aqui.",
        "pensar": "Passando pra saber se você resolveu {a} {peca}. Se ainda precisar, consigo uma opção compatível.",
        "nosso_lado": "Sobre {a} {peca} que você procurou: ficou faltando eu te retornar. Se ainda precisar, consigo uma opção compatível.",
        "sumiu": "Sobre {a} {peca} que a gente conversou: ainda está precisando? Consigo uma opção compatível aqui.",
        "nao_fechou": "Sobre {a} {peca}: se ainda precisar, consigo uma opção compatível. Quer que eu veja pra você?",
        "caro": "Sobre {a} {peca}: sei que o valor pesou. Me fala quanto cabe no seu orçamento que eu procuro uma opção.",
    },
}

MODELOS_FILE = "crm_modelos"

# Nomes salvos no WhatsApp que não servem pra abrir uma mensagem.
NOMES_RUINS = {"cliente", "contato", "whatsapp", "teste", "novo", "sim", "nao", "não", "ok"}

# Masculinos terminados em -a, que fugiriam da regra pela terminação.
MASCULINOS_EM_A = {"sistema", "problema", "mapa", "dia", "emblema",
                   "paralama", "para-lama", "parabrisa", "para-brisa"}

# Femininos que não terminam em -a — a regra pela terminação chamaria de
# masculino e sairia "o central multimídia".
FEMININAS_EXTRA = {"central", "chave", "luz", "grade", "ponte", "lente", "haste",
                   "torre", "base", "fonte", "corrente", "árvore", "arvore",
                   "hélice", "helice", "cruz", "face", "rede", "sede", "parte",
                   "mangá"}


def carregar_modelos_msg() -> dict:
    salvos = ler_json(resolver_pasta_dados() / f"{MODELOS_FILE}.json", None)
    if not salvos:
        return MODELOS_PADRAO
    # Mescla com o padrão: se um dia entrar uma situação nova no código, ela
    # aparece pro time sem precisar que o gestor salve a tela de novo.
    return {
        "saudacao": salvos.get("saudacao") or MODELOS_PADRAO["saudacao"],
        "saudacao_atraso": salvos.get("saudacao_atraso") or MODELOS_PADRAO["saudacao_atraso"],
        "corpo": {**MODELOS_PADRAO["corpo"], **(salvos.get("corpo") or {})},
        "corpo_parecida": {**MODELOS_PADRAO["corpo_parecida"], **(salvos.get("corpo_parecida") or {})},
    }


def _primeiro_nome(nome: str) -> str:
    """Nome do WhatsApp é o que o cliente quis: tem '.', 'Fn', 'Dede_6cc'. Nesses
    casos é melhor não chamar pelo nome do que chamar errado — a saudação sem
    nome continua natural ('Oi, tudo bem?')."""
    bruto = (nome or "").strip()
    if not bruto:
        return ""
    # Pontuação nas pontas é comum e inofensiva ("Conrado…"); no meio do nome
    # já é apelido de perfil ("Dede_6cc"), e aí é melhor não chamar pelo nome.
    primeiro = re.sub(r"^[^A-Za-zÀ-ÿ]+|[^A-Za-zÀ-ÿ]+$", "", bruto.split()[0])
    valido = re.fullmatch(r"[A-Za-zÀ-ÿ]+(['-][A-Za-zÀ-ÿ]+)?", primeiro)
    if len(primeiro) < 3 or not valido or primeiro.lower() in NOMES_RUINS:
        return ""
    limpo = primeiro
    return limpo if (limpo[:1].isupper() and not limpo.isupper()) else limpo.capitalize()


def _artigo(peca: str) -> str:
    """'o motor', 'a bomba', 'as molas'. Sem isso sai 'a motor Audi Q3', que
    entrega na hora que a mensagem foi feita por máquina."""
    palavras = re.sub(r"[^a-zà-ÿ\s-]", " ", (peca or "").lower()).split()
    if not palavras:
        return "a"
    primeira = palavras[0]
    plural = primeira.endswith("s") and len(primeira) > 3
    base = primeira[:-1] if plural else primeira
    feminino = (base in FEMININAS_EXTRA
                or base.endswith(("ção", "são", "dade", "gem"))
                or (base.endswith("a") and base not in MASCULINOS_EM_A))
    if plural:
        return "as" if feminino else "os"
    return "a" if feminino else "o"


def _peca_curta(peca: str, limite: int = 55) -> str:
    """A fila guarda a peça descrita inteira ('soleira dianteira direita e
    friso/cromado do para-choque traseiro Volkswagen Tiguan R Line 2018/2019').
    Numa mensagem isso não cabe: corta no primeiro separador natural, e quando a
    IA não conseguiu identificar a peça, vira só 'peça'."""
    texto = (peca or "").strip()
    if not texto or re.search(r"não (especificad|identificad|inform)", texto, re.I):
        return "peça"
    if len(texto) > limite:
        for corte in (",", " e "):
            if corte in texto:
                texto = texto.split(corte)[0].strip()
                break
    if len(texto) > limite:
        texto = texto[:limite].rsplit(" ", 1)[0].rstrip(" ,;-/")
    return texto or "peça"


def _assunto_msg(ultimas) -> str:
    """O que o cliente falou por último. Só as falas dele entram na fila, então
    não tem risco de casar com o que a loja escreveu."""
    texto = " ".join(ultimas or []).lower()
    if not texto.strip():
        return ""
    for chave, padrao in ASSUNTOS_MSG:
        if re.search(padrao, texto):
            return chave
    return ""


def _limpar_msg(texto: str) -> str:
    """Sem nome a saudação vira 'Oi , tudo bem?'. Arruma a pontuação solta."""
    texto = re.sub(r"\s+([,.:;!?])", r"\1", texto)
    return re.sub(r"\s{2,}", " ", texto).strip()


def montar_mensagem(item: dict, modelos: dict) -> dict:
    gancho = GANCHOS_MSG.get(item.get("gancho"), "sumiu")
    assunto = _assunto_msg(item.get("ultimas"))
    situacao = assunto or gancho

    tabela = modelos["corpo"]
    if item.get("tinha") != "Sim":
        tabela = {**modelos["corpo"], **modelos["corpo_parecida"]}
    corpo = tabela.get(situacao) or tabela.get(gancho) or modelos["corpo"]["sumiu"]

    # A desculpa pela demora só entra quando a bola ficou mesmo com a gente.
    saudacao = modelos["saudacao_atraso"] if gancho == "nosso_lado" else modelos["saudacao"]

    peca = _peca_curta(item.get("peca"))
    campos = {"nome": _primeiro_nome(item.get("nome")),
              "peca": peca,
              "a": _artigo(peca),
              "dias": item.get("dias", 0)}

    def preencher(txt):
        try:
            return txt.format(**campos)
        except (KeyError, IndexError, ValueError):
            # Texto editado pelo gestor com chave inventada não pode derrubar a
            # fila inteira — melhor mostrar o modelo cru do que não mostrar nada.
            return txt

    return {"texto": _limpar_msg(preencher(saudacao) + " " + preencher(corpo)),
            "situacao": situacao}


def _topo_retomada(itens: list, status: dict, quantos: int = 3) -> list:
    """Os clientes mais quentes que ainda não foram chamados. Mesma ordem da
    fila (ALTA primeiro, depois conversa mais recente, nota desempata), pra o
    painel e o follow-up nunca discordarem sobre quem vem primeiro."""
    pendentes = [i for i in itens
                 if (status.get(i["sid"]) or {}).get("status", "pendente") == "pendente"]
    pendentes.sort(key=lambda x: (x.get("prio") != "ALTA", x.get("dias", 999), -x.get("nota", 0)))
    modelos = carregar_modelos_msg()
    return [{
        "sid": i["sid"],
        "nome": i.get("nome"),
        "peca": i.get("peca"),
        "dias": i.get("dias"),
        "prio": i.get("prio"),
        "link": i.get("link"),
        "gancho": i.get("gancho"),
        "msg": montar_mensagem(i, modelos),
    } for i in pendentes[:quantos]]


@app.route("/follow-up")
@app.route("/retomada")   # endereco antigo, mantido pra nao quebrar link salvo
def pagina_retomada():
    return send_from_directory(STATIC_DIR, "retomada.html")


@app.route("/api/retomada/fila")
def api_retomada_fila():
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401
    fila = carregar_fila_retomada(vendedor_id)
    if not fila:
        return jsonify({"sem_fila": True, "itens": [], "rotulos": STATUS_RETOMADA})
    status = carregar_status_retomada(vendedor_id)
    modelos = carregar_modelos_msg()
    itens = []
    for item in fila.get("itens", []):
        marca = status.get(item["sid"]) or {}
        itens.append({**item, "status": marca.get("status", "pendente"),
                      "marcado_em": marca.get("em"),
                      "msg": montar_mensagem(item, modelos)})
    # Pendente primeiro, e dentro dele prioridade ALTA sempre no topo — nesses a
    # conversa parou do nosso lado, ninguém disse não pro vendedor, e são os que
    # ele tem que chamar antes. Só depois a nota desempata. O que já foi
    # trabalhado desce mas continua na tela, pra ele corrigir a marcação ou
    # voltar num cliente que pediu pra chamar depois.
    itens.sort(key=lambda x: (x["status"] != "pendente", x["prio"] != "ALTA", -x["nota"]))

    # Uma fila de 98 nomes ninguem trabalha — vira lista morta. Entao a tela
    # mostra so os MAX_FILA_PENDENTES mais quentes: prioridade ALTA primeiro,
    # depois os mais recentes (conversa de 5 dias converte mais que a de 30) e
    # a nota da IA desempata. Quem ja foi trabalhado nunca some, pra ele poder
    # corrigir a marcacao. A fila inteira continua guardada; o corte e so aqui.
    pendentes = [i for i in itens if i["status"] == "pendente"]
    trabalhados = [i for i in itens if i["status"] != "pendente"]
    pendentes.sort(key=lambda x: (x["prio"] != "ALTA", x.get("dias", 999), -x["nota"]))
    cortados = max(0, len(pendentes) - MAX_FILA_PENDENTES)
    itens = pendentes[:MAX_FILA_PENDENTES] + trabalhados

    return jsonify({
        "gerado_em": fila.get("gerado_em"),
        "de": fila.get("de"),
        "ate": fila.get("ate"),
        "itens": itens,
        "resumo": _resumo_retomada(fila.get("itens", []), status),
        "rotulos": STATUS_RETOMADA,
        "limite_fila": MAX_FILA_PENDENTES,
        "fora_do_corte": cortados,
        "contatos_hoje": _contatos_de_hoje(status),
        "meta_contatos": META_CONTATOS_DIA,
    })


@app.route("/api/retomada/<sid>/status", methods=["POST"])
def api_retomada_status(sid):
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401
    corpo = request.get_json(silent=True) or {}
    novo = (corpo.get("status") or "").strip()
    if novo not in STATUS_RETOMADA:
        return jsonify({"erro": "status inválido"}), 400
    fila = carregar_fila_retomada(vendedor_id) or {}
    itens = fila.get("itens", [])
    # Só aceita cliente que está na fila DESTE vendedor: sem isso um id chutado
    # entraria no arquivo dele e apareceria no painel do gestor como trabalho.
    if not any(item["sid"] == sid for item in itens):
        return jsonify({"erro": "Este cliente não está na sua fila."}), 404
    status = carregar_status_retomada(vendedor_id)
    if novo == "pendente":
        status.pop(sid, None)
    else:
        status[sid] = {"status": novo, "em": agora_br().isoformat()}
    escrever_json(_caminho_crm("status", vendedor_id), status)
    return jsonify({"ok": True, "resumo": _resumo_retomada(itens, status),
                    "contatos_hoje": _contatos_de_hoje(status)})


@app.route("/api/admin/retomada/modelos", methods=["GET", "POST"])
def api_admin_modelos_msg():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    if request.method == "GET":
        return jsonify({"modelos": carregar_modelos_msg(),
                        "padrao": MODELOS_PADRAO,
                        "rotulos": ROTULOS_SITUACAO})
    corpo = request.get_json(silent=True) or {}
    atual = carregar_modelos_msg()
    novo = {
        "saudacao": (corpo.get("saudacao") or "").strip() or atual["saudacao"],
        "saudacao_atraso": (corpo.get("saudacao_atraso") or "").strip() or atual["saudacao_atraso"],
        "corpo": {},
        "corpo_parecida": {},
    }
    # So aceita situacao que o codigo conhece: chave inventada viraria um modelo
    # que nunca e escolhido, e o gestor acharia que salvou.
    for bloco in ("corpo", "corpo_parecida"):
        enviado = corpo.get(bloco) or {}
        for chave in MODELOS_PADRAO[bloco]:
            texto = (enviado.get(chave) or "").strip()
            if texto:
                novo[bloco][chave] = texto
    escrever_json(resolver_pasta_dados() / f"{MODELOS_FILE}.json", novo)
    return jsonify({"ok": True, "modelos": carregar_modelos_msg()})


@app.route("/api/admin/retomada/resumo")
def api_admin_retomada_resumo():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    linhas, gerado_em, periodo = [], None, {}
    for vendedor_id, vendedor in sorted(carregar_vendedores().items(),
                                        key=lambda kv: kv[1]["nome"]):
        fila = carregar_fila_retomada(vendedor_id)
        if not fila:
            continue
        gerado_em = gerado_em or fila.get("gerado_em")
        periodo = periodo or {"de": fila.get("de"), "ate": fila.get("ate")}
        itens = fila.get("itens", [])
        resumo = _resumo_retomada(itens, carregar_status_retomada(vendedor_id))
        trabalhados = resumo["trabalhados"]
        linhas.append({
            "vendedor_id": vendedor_id,
            "nome": vendedor["nome"],
            **resumo,
            "pct_trabalhado": round(100 * trabalhados / len(itens)) if itens else 0,
            # Entre os que ele chamou, quantos deram sinal de vida. É a medida
            # que interessa: percentual sobre a fila inteira mede só o quanto
            # ele avançou na lista, não se a abordagem funcionou.
            "pct_resposta": (round(100 * (resumo["respondeu"] + resumo["vendeu"])
                                   / trabalhados) if trabalhados else 0),
        })
    return jsonify({"gerado_em": gerado_em, "periodo": periodo,
                    "vendedores": linhas, "rotulos": STATUS_RETOMADA})



# ---------- sincronizador de gasto na nuvem (thread de fundo) ----------
# Google e Meta pelo Windsor, uma vez por dia, sem depender do PC da loja.
try:
    import sincronizador_nuvem as _sn

    def _sn_chave():
        d = ler_json(resolver_pasta_dados() / "segredo_windsor.json", None) or {}
        return d.get("chave")

    def _sn_atual():
        return ler_json(resolver_pasta_dados() / "marketing_gasto.json", None)

    def _sn_gravar(corpo):
        escrever_json(resolver_pasta_dados() / "marketing_gasto.json", corpo)

    _sn.iniciar(_sn_chave, _sn_atual, _sn_gravar)
except Exception as _e:
    print(f"[sinc-nuvem] não subiu: {_e}")


@app.route("/api/admin/perfil-google")
def api_admin_perfil_google():
    """Google Perfil da Empresa: quem acha a loja no Maps/Busca e liga."""
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    d = ler_json(resolver_pasta_dados() / "perfil_google.json", None)
    if not d or not d.get("serie_dia"):
        return jsonify({"sem_dados": True})
    return jsonify(d)


@app.route("/api/admin/sincronizar-perfil", methods=["POST"])
def api_admin_sincronizar_perfil():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    try:
        def ler_atual():
            return ler_json(resolver_pasta_dados() / "perfil_google.json", None)

        def gravar(d):
            escrever_json(resolver_pasta_dados() / "perfil_google.json", d)

        return jsonify({"ok": True,
                        "resumo": _sn.sincronizar_perfil(_sn_chave, ler_atual, gravar)})
    except Exception as e:
        return jsonify({"erro": f"{type(e).__name__}: {e}"}), 502


@app.route("/api/admin/sincronizar-gasto", methods=["POST"])
def api_admin_sincronizar_gasto():
    """Dispara na hora a atualizacao de Google+Meta — o mesmo que o thread
    diario faz as 06:45. Serve pro gestor nao esperar o ciclo quando quer o
    numero fresco agora."""
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    try:
        resumo = _sn.sincronizar_gasto(_sn_chave, _sn_atual, _sn_gravar)
        return jsonify({"ok": True, "resumo": resumo})
    except Exception as e:
        return jsonify({"erro": f"{type(e).__name__}: {e}"}), 502


# ---------- Mercado Livre na nuvem (thread de fundo) ----------
# O refresh_token do ML rotaciona a cada uso, entao a credencial mora no banco
# (segredo_ml) e quem renova grava la antes de tudo — detalhes em
# app/sincronizador_ml_nuvem.py. O script local le da mesma chave.
try:
    import sincronizador_ml_nuvem as _sml

    def _sml_cred():
        return ler_json(resolver_pasta_dados() / "segredo_ml.json", None)

    def _sml_gravar_cred(cred):
        escrever_json(resolver_pasta_dados() / "segredo_ml.json", cred)

    def _sml_atual():
        return ler_json(resolver_pasta_dados() / "ml_conta.json", None)

    def _sml_gravar(pacote):
        escrever_json(resolver_pasta_dados() / "ml_conta.json", pacote)

    _sml.iniciar(_sml_cred, _sml_gravar_cred, _sml_atual, _sml_gravar)
except Exception as _e:
    print(f"[sinc-ml] não subiu: {_e}")


@app.route("/api/admin/sincronizar-ml", methods=["POST"])
def api_admin_sincronizar_ml():
    """Dispara a atualizacao do Mercado Livre na hora."""
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    try:
        resumo = _sml.sincronizar(_sml_cred, _sml_gravar_cred, _sml_atual, _sml_gravar)
        return jsonify({"ok": True, "resumo": resumo})
    except Exception as e:
        return jsonify({"erro": f"{type(e).__name__}: {e}"}), 502


# ---------- monitor de atendimento (thread de fundo) ----------
# Roda dentro do proprio servidor pra nao depender do computador da loja
# ligado — historico completo em app/monitor_atendimento.py. No gunicorn cada
# worker teria o seu; o servico usa 1 worker e as escritas sao idempotentes,
# entao duplicata eventual so custaria chamadas repetidas, nunca dado errado.
try:
    import monitor_atendimento as _mon

    def _mon_token():
        d = ler_json(resolver_pasta_dados() / "segredo_totalk.json", None) or {}
        return d.get("token")

    def _mon_gravar(pacote):
        escrever_json(resolver_pasta_dados() / "atendimento_alerta.json", pacote)

    _mon.iniciar(_mon_token, _mon_gravar)
except Exception as _e:   # o monitor nunca pode derrubar o portal
    print(f"[monitor-atendimento] não subiu: {_e}")


if __name__ == "__main__":
    # use_reloader liga SÓ o recarregador: mudou o código, o servidor reinicia
    # sozinho, sem ninguém precisar lembrar de reiniciar na mão.
    #
    # Continua sem debug=True de propósito. debug traz junto o depurador do
    # Werkzeug, que numa tela de erro abre um console de Python rodando dentro
    # do servidor. Como isto escuta em 0.0.0.0, esse console ficaria ao alcance
    # de qualquer um na rede da loja -- daria pra ler os telefones dos clientes
    # e os segredos do processo.
    #
    # O reloader vigia os módulos Python importados, não a pasta data/. Os JSON
    # que a sincronização reescreve não disparam reinício, e os arquivos de
    # static/ são lidos a cada requisição -- editar a tela não pede reinício.
    app.run(host="0.0.0.0", port=8010, use_reloader=True)
