# -*- coding: utf-8 -*-
"""Area `marketing` do portal — rotas e helpers privados. Extraida do server.py em
03/09/2026 (Fase 4). Texto das funcoes identico ao original; o que e comum
vem do nucleo, importado nominalmente pra ninguem adivinhar de onde vem.
"""
import calendar
import hashlib
import io
import json
import os
import re
import secrets
import sqlite3
import subprocess
import sys
import urllib.error
import urllib.request
import unicodedata
import uuid
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from flask import (
    Flask, g, has_request_context, jsonify, redirect, request, send_file,
    send_from_directory, session,
)
from openpyxl import Workbook
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from nucleo import (
    AREAS,
    EXTENSOES_FOTO_PERMITIDAS,
    FOTOS_DIR,
    LIBERACAO_RETROATIVA_MINUTOS,
    LOGIN_JANELA_MINUTOS,
    ROOT,
    STATIC_DIR,
    SUPABASE_SERVICE_KEY,
    SUPABASE_URL,
    Workbook,
    _hash_codigo,
    _hash_senha,
    _nome_aba_excel,
    _resumo_retomada,
    agora_br,
    app,
    areas_efetivas,
    buscar_pecas_simulador,
    calcular_comissao,
    calcular_simulacao_peca,
    calendar,
    carregar_fila_retomada,
    carregar_metas,
    carregar_regras_simulador,
    carregar_status_retomada,
    carregar_vendas_todos,
    carregar_vendedores,
    date,
    datetime,
    definir_data_entrada_simulador,
    excedeu_tentativas_login,
    exigir_admin,
    exigir_vendedor,
    g,
    hoje_br,
    io,
    jsonify,
    metas_vendedor,
    montar_simulacao,
    obter_peca_simulador,
    redirect,
    registrar_acesso,
    request,
    resolver_pasta_dados,
    retroativo_ativo,
    salvar_metas,
    salvar_regras_simulador,
    salvar_vendedores,
    send_file,
    send_from_directory,
    setor_do_usuario,
    status_simulador,
    subprocess,
    sys,
    timedelta,
    urllib,
    valor_liquido,
)
import nucleo as _nucleo
# O nucleo so define estes num dos modos (Postgres x arquivo), como no
# server.py original. Liga o que existir; o resto fica sem ligar, igual antes.
for _nome in ("ler_json",):
    if hasattr(_nucleo, _nome):
        globals()[_nome] = getattr(_nucleo, _nome)
del _nome, _nucleo

# Canais do GA4 que sao Google Ads pago. "Cross-network" e o Performance Max /
# Demand Gen, que tambem e verba do Google. "Paid Social" NAO entra: e anuncio
# de Meta levando pro site, dinheiro pago mas nao do Google — e representa menos
# de 1% dos leads, entao cai no balde organico com aviso na tela em vez de
# ganhar uma terceira coluna que ninguem usaria.
CANAIS_GOOGLE_ADS = {"Paid Search", "Paid Shopping", "Paid Video",
                     "Cross-network", "Paid Other"}

SITE_PAGO = "Site — Google Ads"

SITE_ORGANICO = "Site — orgânico"

def _dividir_site(linhas):
    """Quebra o canal "Site (produto)" em pago e organico, dia a dia.

    O lead do site chega pelo Totalk, que sabe que a conversa veio de um link de
    produto mas NAO sabe como a pessoa chegou no site. Quem sabe e o GA4, que
    registra o evento de lead com o canal da sessao. Entao a divisao usa a
    proporcao REAL de leads pagos daquele dia (`leads_origem_dia`), nao a
    proporcao de sessoes — sao numeros bem diferentes, e ratear por sessao
    inventaria uma atribuicao que o Analytics ja da de graca.

    Dia sem medicao no GA4 fica como "Site (produto)" mesmo: melhor uma linha
    honestamente nao dividida do que duas divididas por chute.
    """
    ga = ler_json(resolver_pasta_dados() / "analytics_site.json", None) or {}
    por_dia = ga.get("leads_origem_dia") or {}
    if not por_dia:
        return linhas

    saida = []
    for l in linhas:
        canais = por_dia.get(l["data"])
        if not str(l.get("canal") or "").startswith("Site") or not canais:
            saida.append(l)
            continue
        pago = sum(n for c, n in canais.items() if c in CANAIS_GOOGLE_ADS)
        tudo = sum(canais.values())
        if not tudo:
            saida.append(l)
            continue
        fatia = pago / tudo

        # Arredonda uma parte e tira a outra por diferenca: as duas SEMPRE somam
        # o original, entao o total de leads do painel nao muda por causa desta
        # divisao — e o gestor nao ve o numero do topo discordar da tabela.
        def parte(v):
            if isinstance(v, float):
                a_ = round(v * fatia, 2)
                return a_, round(v - a_, 2)
            a_ = int(round(v * fatia))
            return a_, v - a_

        lead_p, lead_o = parte(l["leads"])
        sinal_p, sinal_o = parte(l["sinal"])
        prov_p, prov_o = parte(l.get("provavel", 0))
        for canal, ld, sn, pv in ((SITE_PAGO, lead_p, sinal_p, prov_p),
                                  (SITE_ORGANICO, lead_o, sinal_o, prov_o)):
            if ld or sn or pv:
                saida.append({**l, "canal": canal, "leads": ld,
                              "sinal": sn, "provavel": pv})
    return saida

def _marketing_bruto():
    leads = ler_json(resolver_pasta_dados() / "marketing_leads.json", None) or {}
    gasto = ler_json(resolver_pasta_dados() / "marketing_gasto.json", None) or {}
    if leads.get("linhas"):
        leads = {**leads, "linhas": _dividir_site(leads["linhas"])}
    return leads, gasto

def _recortar(linhas, de, ate, campo_data="data"):
    return [l for l in linhas if de <= l[campo_data] <= ate]

def _agregar_marketing(linhas, vendedores_nome=None):
    """Soma um conjunto de linhas de lead em vários cortes de uma vez."""
    total = {"leads": 0, "sinal": 0, "provavel": 0}
    por_canal, por_dia, por_vendedor = {}, {}, {}
    for l in linhas:
        total["leads"] += l["leads"]
        total["sinal"] += l["sinal"]
        total["provavel"] += l.get("provavel", 0)
        for destino, chave in ((por_canal, l["canal"]), (por_dia, l["data"]),
                               (por_vendedor, l["vendedor"] or "sem_atendente")):
            d = destino.setdefault(chave, {"leads": 0, "sinal": 0, "provavel": 0})
            d["leads"] += l["leads"]
            d["sinal"] += l["sinal"]
            d["provavel"] += l.get("provavel", 0)

    def lista(dic, rotulo, ordenar_por_chave=False):
        itens = [{rotulo: k, **v} for k, v in dic.items()]
        itens.sort(key=(lambda x: x[rotulo]) if ordenar_por_chave
                   else (lambda x: -x["leads"]))
        return itens

    saida = {
        "total": total,
        "por_canal": lista(por_canal, "canal"),
        "por_dia": lista(por_dia, "data", ordenar_por_chave=True),
        "por_vendedor": lista(por_vendedor, "vendedor"),
    }
    if vendedores_nome:
        for item in saida["por_vendedor"]:
            item["nome"] = vendedores_nome.get(item["vendedor"], "Sem atendente")
    return saida

def _vendas_no_periodo(vendedores, de, ate, so_vendedor=None, universo=None):
    """Vendas lançadas no portal — é o que fecha a conta do marketing: leads de
    um lado, venda de verdade do outro.

    `universo` limita a conta a quem realmente aparece no dado de leads. Sem
    isso a Brenda entraria: ela não atende pelo Totalk, então as vendas dela
    somariam no numerador sem nenhum lead no denominador e a conversão sairia
    inflada."""
    vendas = carregar_vendas_todos(vendedores)
    itens = [v for v in vendas.values()
             if v.get("tipo", "venda") == "venda" and de <= v["data"] <= ate
             and (so_vendedor is None or v["vendedor_id"] == so_vendedor)
             and (universo is None or v["vendedor_id"] in universo)]
    total = round(sum(valor_liquido(v) for v in itens), 2)
    por_vendedor = {}
    for v in itens:
        d = por_vendedor.setdefault(v["vendedor_id"], {"qtd": 0, "total": 0.0})
        d["qtd"] += 1
        d["total"] += valor_liquido(v)
    return {
        "qtd": len(itens),
        "total": total,
        "ticket": round(total / len(itens), 2) if itens else 0.0,
        "por_vendedor": {k: {"qtd": d["qtd"], "total": round(d["total"], 2)}
                         for k, d in por_vendedor.items()},
    }

def _periodo_pedido():
    hoje = hoje_br().isoformat()
    de = request.args.get("de") or f"{hoje[:7]}-01"
    ate = request.args.get("ate") or hoje
    return de, ate

def _janela_comparavel(de, ate, cobertura):
    """Leads param na data em que o Totalk foi sincronizado; as vendas seguem
    até hoje. Comparar as duas coisas na janela cheia faz a conversão parecer
    melhor do que é — então a conta usa só o pedaço em que os dois lados
    existem."""
    if not cobertura:
        return de, ate
    return max(de, cobertura["de"]), min(ate, cobertura["ate"])

def _cobertura(linhas):
    """Até quando o espelho do Totalk foi sincronizado. Sem isso, um período
    que passa dessa data mostra queda de leads que é só falta de dado."""
    if not linhas:
        return None
    datas = [l["data"] for l in linhas]
    return {"de": min(datas), "ate": max(datas)}

def _google_por_objetivo(linhas):
    """Separa o gasto do Google entre trazer gente pra LOJA e vender no SITE.

    Linha antiga (antes de 30/08/2026) nao tem o campo `objetivo` — cai em
    "site", que era o unico tipo de campanha que existia ate entao.
    """
    saida = {}
    for g in linhas:
        alvo = g.get("objetivo") or "site"
        d = saida.setdefault(alvo, {"investimento": 0.0, "clicks": 0,
                                    "ligacoes": 0, "conversoes": 0.0})
        d["investimento"] += g.get("spend", 0)
        d["clicks"] += g.get("clicks", 0)
        d["ligacoes"] += g.get("ligacoes", 0)
        d["conversoes"] += g.get("conversoes", 0)
    for d in saida.values():
        d["investimento"] = round(d["investimento"], 2)
        d["conversoes"] = round(d["conversoes"], 2)
    return saida

def _google_por_campanha(linhas):
    """Uma linha por campanha, com tipo e objetivo — pro gestor ver onde o
    dinheiro esta e o que cada uma entrega."""
    saida = {}
    for g in linhas:
        nome = g.get("campanha") or "—"
        d = saida.setdefault(nome, {"campanha": nome, "investimento": 0.0,
                                    "clicks": 0, "ligacoes": 0, "conversoes": 0.0,
                                    "tipo_nome": g.get("tipo_nome") or "",
                                    "objetivo": g.get("objetivo") or "site"})
        d["investimento"] += g.get("spend", 0)
        d["clicks"] += g.get("clicks", 0)
        d["ligacoes"] += g.get("ligacoes", 0)
        d["conversoes"] += g.get("conversoes", 0)
    for d in saida.values():
        d["investimento"] = round(d["investimento"], 2)
        d["conversoes"] = round(d["conversoes"], 2)
        d["cpc"] = round(d["investimento"] / d["clicks"], 2) if d["clicks"] else None
    return sorted(saida.values(), key=lambda x: -x["investimento"])

@app.route("/api/admin/marketing")
def api_marketing_gestor():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    de, ate = _periodo_pedido()
    filtro_canal = request.args.get("canal") or ""
    filtro_vendedor = request.args.get("vendedor") or ""

    leads_bruto, gasto_bruto = _marketing_bruto()
    vendedores = carregar_vendedores()
    # So quem vende entra na tela de marketing. A expedicao usa o mesmo login,
    # mas nao atende lead nenhum — se entrasse aqui apareceria eternamente na
    # lista de "fora do Totalk", como se estivesse deixando de atender.
    nomes = {vid: v["nome"] for vid, v in vendedores.items()
             if (v.get("perfil") or "vendedor") != "expedicao"}

    cobertura = _cobertura(leads_bruto.get("linhas", []))
    ef_de, ef_ate = _janela_comparavel(de, ate, cobertura)

    linhas = _recortar(leads_bruto.get("linhas", []), ef_de, ef_ate)
    # Quem de fato tem lead no periodo — e o universo que pode entrar na
    # conversao. Fora dele a venda nao tem lead correspondente.
    universo = {l["vendedor"] for l in linhas if l["vendedor"]}
    fora = sorted(nomes[vid] for vid in nomes
                  if vid not in universo and not vendedores[vid].get("oculto"))
    if filtro_canal:
        linhas = [l for l in linhas if l["canal"] == filtro_canal]
    if filtro_vendedor:
        linhas = [l for l in linhas if l["vendedor"] == filtro_vendedor]
    agregado = _agregar_marketing(linhas, nomes)

    # canais disponíveis saem do período inteiro, não do recorte — senão filtrar
    # por um canal apagaria os outros da lista e não daria pra voltar
    todos_canais = sorted({l["canal"] for l in _recortar(leads_bruto.get("linhas", []), ef_de, ef_ate)})

    # O investimento segue a janela pedida: gasto de midia nao depende do Totalk.
    gasto_linhas = _recortar(gasto_bruto.get("linhas", []), de, ate)
    investimento = round(sum(g["spend"] for g in gasto_linhas), 2)
    cliques = sum(g["clicks"] for g in gasto_linhas)
    impressoes = sum(g["impressions"] for g in gasto_linhas)

    por_campanha = {}
    for g in gasto_linhas:
        d = por_campanha.setdefault(g["campanha"], {"spend": 0.0, "clicks": 0, "impressions": 0})
        d["spend"] += g["spend"]
        d["clicks"] += g["clicks"]
        d["impressions"] += g["impressions"]
    campanhas = sorted(
        [{"campanha": k, "spend": round(v["spend"], 2), "clicks": v["clicks"],
          "impressions": v["impressions"],
          "cpc": round(v["spend"] / v["clicks"], 2) if v["clicks"] else None,
          "ctr": round(100 * v["clicks"] / v["impressions"], 2) if v["impressions"] else None}
         for k, v in por_campanha.items()],
        key=lambda x: -x["spend"])

    gasto_dia = {}
    for g in gasto_linhas:
        gasto_dia[g["data"]] = round(gasto_dia.get(g["data"], 0.0) + g["spend"], 2)

    # O Meta vem agregado do periodo inteiro do relatorio, nao por dia. Exigir
    # que a janela cobrisse o relatorio inteiro pra somar deixava o investimento
    # dele fora da conta em qualquer mes — que e justamente como o painel e
    # aberto. Entao rateia por dia: o gasto do relatorio dividido pelos dias que
    # ele cobre, multiplicado pelos dias que caem na janela pedida.
    #
    # E aproximacao, e a tela diz isso. Gasto de midia nao e uniforme dia a dia,
    # mas num recorte de semanas o erro e pequeno perto de simplesmente ignorar
    # metade do investimento.
    meta = gasto_bruto.get("meta")
    meta_rateio = None

    # Desde 28/08/2026 o Meta vem do Windsor com serie por dia: soma exata do
    # periodo filtrado, sem rateio. O bloco de rateio abaixo continua pro caso
    # de a fonte voltar a ser o CSV do Gerenciador, que so traz o agregado.
    serie_meta = (meta or {}).get("serie_dia") or {}
    gasto_meta_periodo = 0.0   # nome estavel pra comparacao com o periodo anterior
    if serie_meta:
        dentro = {d: v for d, v in serie_meta.items() if de <= d <= ate}
        if dentro:
            gasto_meta = round(sum(v["spend"] for v in dentro.values()), 2)
            meta_rateio = {
                "de": min(dentro), "ate": max(dentro),
                "dias_dentro": len(dentro), "dias_relatorio": len(dentro),
                "spend": gasto_meta,
                "impressions": sum(v["impressions"] for v in dentro.values()),
                "conversas": sum(v["conversas"] for v in dentro.values()),
                "integral": True,
                "por_dia": True,
            }
            # Facebook x Instagram. A quebra vem dentro de cada dia, entao ela
            # acompanha o filtro de periodo em vez de depender de um total ja
            # fechado. Alcance fica de fora de proposito: e a unica metrica que
            # nao pode ser somada entre as duas, porque a mesma pessoa aparece
            # nas duas e o total viraria gente que nao existe.
            plataformas = {}
            for v_dia in dentro.values():
                for nome, p in (v_dia.get("plataforma") or {}).items():
                    # O Meta devolve um balde "unknown" que vem sempre zerado.
                    # Filtrar pelo nome nao basta — o que importa e nao ter
                    # movimento nenhum, entao a checagem e por valor.
                    if not (p.get("spend") or p.get("conversas")
                            or p.get("clicks")):
                        continue
                    acc = plataformas.setdefault(nome, {"spend": 0.0, "clicks": 0,
                                                        "impressions": 0, "conversas": 0})
                    acc["spend"] = round(acc["spend"] + p.get("spend", 0), 2)
                    acc["clicks"] += p.get("clicks", 0)
                    acc["impressions"] += p.get("impressions", 0)
                    acc["conversas"] += p.get("conversas", 0)
            if plataformas:
                for nome, p in plataformas.items():
                    p["nome"] = {"facebook": "Facebook",
                                 "instagram": "Instagram"}.get(nome, nome.title())
                    p["custo_por_conversa"] = (round(p["spend"] / p["conversas"], 2)
                                               if p["conversas"] else None)
                    p["fatia"] = (round(100 * p["spend"] / gasto_meta, 1)
                                  if gasto_meta else 0)
                meta_rateio["plataformas"] = sorted(
                    plataformas.values(), key=lambda p: -p["spend"])
            investimento = round(investimento + gasto_meta, 2)
            gasto_meta_periodo = gasto_meta
            impressoes += meta_rateio["impressions"]
    elif meta and meta.get("de") and meta.get("ate"):
        ini = max(de, meta["de"])
        fim = min(ate, meta["ate"])
        if ini <= fim:
            dias_relatorio = (date.fromisoformat(meta["ate"])
                              - date.fromisoformat(meta["de"])).days + 1
            dias_dentro = (date.fromisoformat(fim) - date.fromisoformat(ini)).days + 1
            fatia = dias_dentro / dias_relatorio if dias_relatorio else 0
            meta_rateio = {
                "de": ini, "ate": fim,
                "dias_dentro": dias_dentro, "dias_relatorio": dias_relatorio,
                "spend": round(meta["spend"] * fatia, 2),
                "impressions": int(meta.get("impressions", 0) * fatia),
                "conversas": int(meta.get("conversas", 0) * fatia),
                "integral": dias_dentro == dias_relatorio,
            }
            investimento = round(investimento + meta_rateio["spend"], 2)
            gasto_meta_periodo = meta_rateio["spend"]
            impressoes += meta_rateio["impressions"]

    # Product Ads do Mercado Livre, com o MESMO rateio por dia do Meta: o
    # sincronizador entrega uma janela fechada de 30 dias, e daqui sai a fatia
    # que cai no periodo pedido. Aproximacao — e a tela diz isso.
    ml_conta = ler_json(resolver_pasta_dados() / "ml_conta.json", None) or {}
    ml_ads = ml_conta.get("ads")
    ml_rateio = None
    if ml_ads and ml_ads.get("de") and ml_ads.get("ate") and ml_ads.get("investido"):
        ini = max(de, ml_ads["de"])
        fim = min(ate, ml_ads["ate"])
        if ini <= fim:
            dias_relatorio = (date.fromisoformat(ml_ads["ate"])
                              - date.fromisoformat(ml_ads["de"])).days + 1
            dias_dentro = (date.fromisoformat(fim) - date.fromisoformat(ini)).days + 1
            fatia = dias_dentro / dias_relatorio if dias_relatorio else 0
            ml_rateio = {
                "de": ini, "ate": fim,
                "dias_dentro": dias_dentro, "dias_relatorio": dias_relatorio,
                "spend": round(float(ml_ads["investido"]) * fatia, 2),
                "cliques": int((ml_ads.get("cliques") or 0) * fatia),
                "impressions": int((ml_ads.get("impressoes") or 0) * fatia),
                "receita": round(float(ml_ads.get("receita_atribuida") or 0) * fatia, 2),
                "acos": ml_ads.get("acos"),
                "integral": dias_dentro == dias_relatorio,
            }
            investimento = round(investimento + ml_rateio["spend"], 2)
            impressoes += ml_rateio["impressions"]

    # A agencia que gerencia as campanhas custa fixo por mes e entra no
    # investimento, porque gestao e custo de midia tanto quanto o clique.
    # NAO rateia por dia: a nota vem cheia todo mes, independente de quantos
    # dias voce esta olhando. Rateando, um filtro de 01 a 30/08 mostrava
    # R$ 1.838,71 — um valor que nunca foi pago e nao existe em lugar nenhum.
    # Aqui o periodo conta os meses que ele toca, cada um pelo valor cheio.
    # Valor unico declarado aqui: mudou o contrato, muda esta linha.
    AGENCIA_MENSAL = 1900.0

    def meses_tocados(d1: str, d2: str) -> list:
        """Todo mes que o periodo encosta, do primeiro ao ultimo.

        Anda de mes em mes pelo calendario. A versao anterior somava 32 dias de
        cada vez, o que ia acumulando folga: passado mais ou menos um ano e meio
        a conta pulava um mes inteiro.
        """
        ini, fim = date.fromisoformat(d1), date.fromisoformat(d2)
        fora, cursor = [], ini.replace(day=1)
        while cursor <= fim and len(fora) < 240:
            fora.append(cursor.strftime("%Y-%m"))
            cursor = (cursor.replace(day=28) + timedelta(days=4)).replace(day=1)
        return fora or [ini.strftime("%Y-%m")]

    meses_agencia = meses_tocados(de, ate)
    agencia = round(AGENCIA_MENSAL * len(meses_agencia), 2)
    investimento = round(investimento + agencia, 2)

    vendas = _vendas_no_periodo(vendedores, ef_de, ef_ate,
                                so_vendedor=filtro_vendedor or None,
                                universo=universo)
    for item in agregado["por_vendedor"]:
        item["vendas"] = vendas["por_vendedor"].get(item["vendedor"], {"qtd": 0, "total": 0.0})
        item["conversao"] = (round(100 * item["vendas"]["qtd"] / item["leads"], 1)
                             if item["leads"] else None)

    total_leads = agregado["total"]["leads"]

    # Receita provavel por canal. A venda nao carrega o canal de origem — o
    # chat esta no Totalk, a venda esta no portal, e nada liga um ao outro.
    # Entao o unico jeito de saber que canal traz dinheiro e estimar: quantas
    # conversas daquele canal chegaram ao fechamento, vezes o ticket real.
    #
    # O ticket sai das vendas REAIS do proprio periodo, nao de um valor fixo:
    # assim ele acompanha o mes em vez de envelhecer. So cai num padrao quando
    # o periodo nao tem venda lancada suficiente pra sustentar uma media.
    TICKET_PADRAO = 968.0   # calibrado em ago/2026 contra 519 vendas reais
    ticket = (round(vendas["total"] / vendas["qtd"], 2)
              if vendas.get("qtd") else TICKET_PADRAO)
    for item in agregado["por_canal"]:
        item["receita_provavel"] = round(item.get("provavel", 0) * ticket, 2)
    agregado["total"]["receita_provavel"] = round(
        agregado["total"].get("provavel", 0) * ticket, 2)

    # ---- comparacao com o periodo anterior ----
    # A janela anterior tem o mesmo NUMERO DE DIAS COBERTOS, nao o mes
    # calendario anterior. Agosto ainda esta aberto: comparar 28 dias de agosto
    # com 31 de julho inventaria uma queda de 10% que e so calendario. Por isso
    # a base e ef_de/ef_ate, que ja e a janela que tem dado de verdade.
    dias_janela = ((date.fromisoformat(ef_ate) - date.fromisoformat(ef_de)).days + 1
                   if ef_de and ef_ate else 0)
    anterior = None
    if dias_janela > 0:
        ant_ate = (date.fromisoformat(ef_de) - timedelta(days=1)).isoformat()
        ant_de = (date.fromisoformat(ant_ate)
                  - timedelta(days=dias_janela - 1)).isoformat()
        linhas_ant = _recortar(leads_bruto.get("linhas", []), ant_de, ant_ate)
        if filtro_canal:
            linhas_ant = [l for l in linhas_ant if l["canal"] == filtro_canal]
        if filtro_vendedor:
            linhas_ant = [l for l in linhas_ant if l["vendedor"] == filtro_vendedor]
        if linhas_ant:
            ag_ant = _agregar_marketing(linhas_ant, nomes)
            vendas_ant = _vendas_no_periodo(vendedores, ant_de, ant_ate,
                                            so_vendedor=filtro_vendedor or None,
                                            universo=universo)
            # O ticket do periodo anterior e o DELE, nao o de agora: senao a
            # receita "mudaria" so porque o ticket medio mudou.
            ticket_ant = (round(vendas_ant["total"] / vendas_ant["qtd"], 2)
                          if vendas_ant.get("qtd") else ticket)
            for item in ag_ant["por_canal"]:
                item["receita_provavel"] = round(
                    item.get("provavel", 0) * ticket_ant, 2)
            ag_ant["total"]["receita_provavel"] = round(
                ag_ant["total"].get("provavel", 0) * ticket_ant, 2)
            anterior = {
                "de": ant_de, "ate": ant_ate, "dias": dias_janela,
                "total": ag_ant["total"],
                "por_canal": {c["canal"]: c for c in ag_ant["por_canal"]},
                "vendas": vendas_ant,
                "ticket": ticket_ant,
            }

    def _var(agora, antes):
        """Variacao percentual. Sem base anterior devolve None — nao existe
        'subiu 100%' partindo do zero, isso so engana quem le."""
        if not antes:
            return None
        return round(100 * (agora - antes) / antes, 1)

    if anterior:
        a = anterior["total"]
        agregado["total"]["var"] = {
            "leads": _var(agregado["total"]["leads"], a.get("leads")),
            "sinal": _var(agregado["total"]["sinal"], a.get("sinal")),
            "provavel": _var(agregado["total"].get("provavel", 0), a.get("provavel")),
            "receita_provavel": _var(agregado["total"]["receita_provavel"],
                                     a.get("receita_provavel")),
        }
        for item in agregado["por_canal"]:
            ant = anterior["por_canal"].get(item["canal"])
            item["var"] = {
                "leads": _var(item["leads"], (ant or {}).get("leads")),
                "sinal": _var(item["sinal"], (ant or {}).get("sinal")),
                "receita_provavel": _var(item["receita_provavel"],
                                         (ant or {}).get("receita_provavel")),
            } if ant else None
        anterior["var_vendas"] = {
            "qtd": _var(vendas["qtd"], anterior["vendas"]["qtd"]),
            "total": _var(vendas["total"], anterior["vendas"]["total"]),
        }

        # Conversao e taxa, e taxa se compara em PONTO PERCENTUAL, nao em
        # variacao relativa. De 6,3% pra 5,9% a queda e de 0,4 p.p.; dizer
        # "-6,3%" seria verdade aritmetica e leitura errada — parece que
        # despencou quando andou meio ponto. Os dois vao juntos: o p.p. e o
        # numero pra decidir, o relativo fica de contexto.
        leads_ant = anterior["total"].get("leads") or 0
        conv_ant = (round(100 * anterior["vendas"]["qtd"] / leads_ant, 1)
                    if leads_ant else None)
        conv_agora = (round(100 * vendas["qtd"] / total_leads, 1)
                      if total_leads else None)
        anterior["conversao"] = conv_ant
        anterior["var_conversao"] = (
            {"pp": round(conv_agora - conv_ant, 2),
             "relativo": _var(conv_agora, conv_ant)}
            if conv_ant is not None and conv_agora is not None else None)

        # Investimento do periodo anterior. Refeito com as mesmas fontes do
        # atual — Google por dia, Meta por dia, agencia pelos meses tocados —
        # senao comparar gasto contra gasto estaria comparando bases diferentes.
        ant_google = round(sum(g["spend"] for g in _recortar(
            gasto_bruto.get("linhas", []), anterior["de"], anterior["ate"])), 2)
        ant_meta = round(sum(
            v_dia["spend"] for d_, v_dia in serie_meta.items()
            if anterior["de"] <= d_ <= anterior["ate"]), 2) if serie_meta else 0.0
        # Mesma regra do periodo atual. Antes aqui era so {mes do inicio, mes do
        # fim} — no filtro "este ano" isso dava 2 meses contra os 8 do periodo
        # atual, e o comparativo mostrava uma alta de midia que nao existiu.
        ant_agencia = round(
            AGENCIA_MENSAL * len(meses_tocados(anterior["de"], anterior["ate"])), 2)
        ant_total = round(ant_google + ant_meta + ant_agencia, 2)
        anterior["midia"] = {"google": ant_google, "meta": ant_meta,
                             "agencia": ant_agencia, "investimento": ant_total}
        # O investimento atual inclui o ML, que o anterior nao refaz; por isso
        # a comparacao usa as tres fontes que os dois lados tem em comum.
        atual_comparavel = round(
            sum(g["spend"] for g in gasto_linhas) + gasto_meta_periodo + agencia, 2)
        anterior["var_midia"] = {
            "investimento": _var(atual_comparavel, ant_total),
            "google": _var(round(sum(g["spend"] for g in gasto_linhas), 2), ant_google),
            "meta": _var(gasto_meta_periodo, ant_meta),
        }
        anterior["midia"]["atual_comparavel"] = atual_comparavel

    # O periodo pedido pode estar inteiro fora do que o Totalk cobre — a
    # leitura comeca em 28/06/2026, entao maio nao tem lead nenhum. Sem avisar,
    # a tela mostrava 0 lead ao lado de R$ 10.323 de investimento: parecia que
    # o dinheiro foi gasto e ninguem apareceu. E falta de medicao, nao de
    # resultado, e a tela precisa dizer qual dos dois.
    sem_cobertura = bool(ef_de and ef_ate and ef_de > ef_ate)

    return jsonify({
        "anterior": anterior,
        "sem_cobertura": sem_cobertura,
        "ticket_estimativa": {"valor": ticket,
                              "de_vendas_reais": bool(vendas.get("qtd")),
                              "base_qtd": vendas.get("qtd", 0)},
        "de": de, "ate": ate,
        "periodo_efetivo": {"de": ef_de, "ate": ef_ate},
        "gerado_em": leads_bruto.get("gerado_em"),
        "cobertura": cobertura,
        "fora_do_totalk": fora,
        "canais": todos_canais,
        "vendedores": [{"id": vid, "nome": nome} for vid, nome in sorted(
            nomes.items(), key=lambda kv: kv[1])],
        "filtro": {"canal": filtro_canal, "vendedor": filtro_vendedor},
        **agregado,
        "vendas": vendas,
        "conversao": round(100 * vendas["qtd"] / total_leads, 1) if total_leads else None,
        "midia": {
            "investimento": investimento,
            "clicks": cliques,
            "impressions": impressoes,
            # CPC e CTR ficam so no Google: este export do Meta traz conversa
            # iniciada, nao clique, e dividir um pelo outro nao significa nada.
            "cpc": (round(sum(g["spend"] for g in gasto_linhas) / cliques, 2)
                    if cliques else None),
            "ctr": (round(100 * cliques / sum(g["impressions"] for g in gasto_linhas), 2)
                    if sum(g["impressions"] for g in gasto_linhas) else None),
            # Custo por lead e por venda usam o total de leads/vendas do período,
            # não só os que vieram de anúncio: é o custo de mídia por resultado
            # do negócio. Com o gasto do Meta faltando, o número real é maior.
            "custo_por_lead": round(investimento / total_leads, 2) if total_leads else None,
            "custo_por_venda": round(investimento / vendas["qtd"], 2) if vendas["qtd"] else None,
            # Nao chamamos isso de ROAS: o faturamento aqui e o total do
            # periodo, nao o atribuido a anuncio, e falta o gasto do Meta. E
            # "quanto o negocio faturou por real de midia paga", nada mais.
            "faturamento_por_real": (round(vendas["total"] / investimento, 1)
                                     if investimento else None),
            "campanhas": campanhas,
            "por_dia": [{"data": k, "spend": v} for k, v in sorted(gasto_dia.items())],
            "fontes_ausentes": gasto_bruto.get("fontes_ausentes", []),
            "google": {"investimento": round(sum(g["spend"] for g in gasto_linhas), 2),
                       "clicks": cliques,
                       # Campanha que traz gente ate a LOJA (mapa, ligacao) nao
                       # se mede pela regua do site: ela converte em telefone
                       # tocando e gente na porta, que o site nao registra.
                       # Somadas, elas inflavam o custo por venda online e
                       # faziam a campanha local parecer ruim sem ser.
                       "por_objetivo": _google_por_objetivo(gasto_linhas),
                       "por_campanha": _google_por_campanha(gasto_linhas)},
            "meta": meta,
            "meta_rateio": meta_rateio,
            "ml_rateio": ml_rateio,
            # Quando cada fonte foi atualizada pela ultima vez. No plano basico
            # do Windsor so uma conta fica conectada por vez, entao a outra
            # fica congelada — a tela precisa dizer isso.
            "atualizado_em": gasto_bruto.get("atualizado_em") or {},
            "fontes_ausentes": gasto_bruto.get("fontes_ausentes") or [],
            "agencia": {"mensal": AGENCIA_MENSAL, "no_periodo": agencia,
                        "meses": len(meses_agencia)},
        },
    })

@app.route("/api/admin/desempenho")
def api_admin_desempenho():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    vendedores = carregar_vendedores()
    vid = request.args.get("vendedor", "")
    if vid not in vendedores:
        return jsonify({"erro": "Vendedor não encontrado."}), 404
    return jsonify(montar_desempenho(vid, request.args.get("mes"), vendedores))

@app.route("/api/desempenho")
def api_desempenho_vendedor():
    """O mesmo painel, do ponto de vista de quem está logado.

    Uma diferença de propósito: sai a participação no faturamento do time. Com
    ela e o próprio total, o vendedor deduziria quanto a equipe inteira vendeu —
    e o gestor acabou de tirar o ranking do menu dele justamente pra isso não
    ficar exposto. A posição fica: motiva sem entregar número de ninguém."""
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401
    dados = montar_desempenho(vendedor_id, request.args.get("mes"), carregar_vendedores())
    dados["time"] = {"posicao": dados["time"]["posicao"], "de": dados["time"]["de"]}
    return jsonify(dados)

def montar_desempenho(vid, mes, vendedores):
    """Tudo o que dá pra medir de um vendedor sozinho, num mês.

    O que entra aqui é só o que existe em 100% das vendas (data, valor,
    produto). `canal` está preenchido em 4% dos registros e com grafia
    inconsistente, então não vira métrica — mostraria um mix falso. Leads,
    conversas e taxa de conversão não vivem neste portal: a fila do follow-up
    traz só quem não fechou, e sem o total de atendimentos não existe
    denominador. Esse número tem que vir do vendas-insights.
    """
    mes = mes or hoje_br().isoformat()[:7]
    vendas = carregar_vendas_todos(vendedores)
    metas_todas = carregar_metas()

    minhas = [v for v in vendas.values()
              if v["vendedor_id"] == vid and v.get("tipo", "venda") == "venda"]
    bonus = [v for v in vendas.values()
             if v["vendedor_id"] == vid and v.get("tipo") == "bonus"]

    def do_mes(lista, alvo):
        return [v for v in lista if v["data"][:7] == alvo]

    def mes_anterior(alvo):
        ano, m = int(alvo[:4]), int(alvo[5:7])
        return f"{ano - 1}-12" if m == 1 else f"{ano}-{m - 1:02d}"

    def bloco(alvo):
        """Os números de um mês. Serve pro mês escolhido e pro anterior, que é
        o que dá sentido à variação."""
        itens = do_mes(minhas, alvo)
        total = round(sum(valor_liquido(v) for v in itens), 2)
        qtd = len(itens)
        dias = {v["data"] for v in itens}
        return {
            "mes": alvo,
            "total": total,
            "qtd": qtd,
            "ticket": round(total / qtd, 2) if qtd else 0.0,
            "dias_ativos": len(dias),
            "media_dia_ativo": round(total / len(dias), 2) if dias else 0.0,
        }

    atual = bloco(mes)
    anterior = bloco(mes_anterior(mes))

    def variacao(agora, antes):
        if not antes:
            return None          # sem base de comparação: não inventa 100%
        return round(100 * (agora - antes) / antes, 1)

    # ---- ritmo do mês ----
    dias_no_mes = calendar.monthrange(int(mes[:4]), int(mes[5:7]))[1]
    hoje = hoje_br()
    dias_corridos = hoje.day if mes == hoje.isoformat()[:7] else dias_no_mes
    meta_mensal = float(metas_vendedor(vid, metas_todas).get("mensal", 0) or 0)

    # ---- histórico: evolução e metas batidas ----
    por_mes = {}
    for v in minhas:
        chave = v["data"][:7]
        d = por_mes.setdefault(chave, {"total": 0.0, "qtd": 0})
        d["total"] += valor_liquido(v)
        d["qtd"] += 1
    historico = [{
        "mes": k,
        "total": round(d["total"], 2),
        "qtd": d["qtd"],
        "ticket": round(d["total"] / d["qtd"], 2) if d["qtd"] else 0.0,
        # Meta é a de hoje aplicada a todo o histórico: o portal não guarda a
        # meta que valia em cada mês passado. Serve pra tendência, não pra
        # cobrar mês fechado.
        "bateu": bool(meta_mensal) and d["total"] >= meta_mensal,
    } for k, d in sorted(por_mes.items())]

    # ---- dentro do mês ----
    por_dia = {}
    for v in do_mes(minhas, mes):
        por_dia[v["data"]] = por_dia.get(v["data"], 0.0) + valor_liquido(v)
    serie_dia = [{"data": k, "total": round(x, 2)} for k, x in sorted(por_dia.items())]
    melhor_dia = max(serie_dia, key=lambda x: x["total"], default=None)

    itens_mes = do_mes(minhas, mes)
    maior_venda = max(itens_mes, key=lambda v: valor_liquido(v), default=None)

    # Faixas de valor: mostram se o mês veio de muita peça barata ou de poucas
    # caras — duas rotas bem diferentes pro mesmo faturamento.
    FAIXAS = [(0, 200, "até R$ 200"), (200, 500, "R$ 200–500"),
              (500, 1000, "R$ 500–1 mil"), (1000, 3000, "R$ 1–3 mil"),
              (3000, float("inf"), "acima de R$ 3 mil")]
    faixas = []
    for piso, teto, rotulo in FAIXAS:
        dentro = [v for v in itens_mes if piso <= valor_liquido(v) < teto]
        faixas.append({"rotulo": rotulo, "qtd": len(dentro),
                       "total": round(sum(valor_liquido(v) for v in dentro), 2)})

    # Nao existe "top produto" util aqui: produto e texto livre digitado a cada
    # venda e quase nunca se repete igual, entao agrupar por nome so devolveria
    # a maior venda com quantidade 1. O que informa de verdade e a lista das
    # maiores vendas do mes.
    maiores_vendas = [{"produto": (v.get("produto") or "Sem descrição").strip(),
                       "valor": valor_liquido(v), "data": v["data"]}
                      for v in sorted(itens_mes, key=valor_liquido, reverse=True)[:8]]

    # ---- posição no time ----
    totais_time = sorted(
        ((outro, round(sum(valor_liquido(v) for v in vendas.values()
                           if v["vendedor_id"] == outro
                           and v.get("tipo", "venda") == "venda"
                           and v["data"][:7] == mes), 2))
         for outro in vendedores),
        key=lambda kv: kv[1], reverse=True)
    # Conta oculta zerada nao entra no "2º de 5": inflava o tamanho do time.
    totais_time = [(o, t) for o, t in totais_time
                   if t or not vendedores[o].get("oculto") or o == vid]
    posicao = next((i + 1 for i, (outro, _) in enumerate(totais_time) if outro == vid), None)
    total_time = round(sum(t for _, t in totais_time), 2)

    # ---- follow-up ----
    fila = carregar_fila_retomada(vid)
    followup = None
    if fila:
        itens_fila = fila.get("itens", [])
        st = carregar_status_retomada(vid)
        resumo = _resumo_retomada(itens_fila, st)
        trabalhados = resumo["trabalhados"]
        followup = {
            **resumo,
            "pct_trabalhado": round(100 * trabalhados / len(itens_fila)) if itens_fila else 0,
            "pct_resposta": (round(100 * (resumo["respondeu"] + resumo["vendeu"]) / trabalhados)
                             if trabalhados else 0),
        }

    # ---- atendimento (vem do vendas-insights, espelho do Totalk) ----
    # O portal so sabe o que virou venda. Quantos clientes ele atendeu, de que
    # canal vieram e quanto demorou a primeira resposta vive no outro projeto e
    # chega aqui agregado, pela chave insights_<vendedor>.
    insights = ler_json(resolver_pasta_dados() / f"insights_{vid}.json", None)
    atendimento = None
    if insights:
        do_mes_ins = (insights.get("meses") or {}).get(mes)
        if do_mes_ins:
            atend = do_mes_ins.get("atendimentos", 0)
            atendimento = {
                **do_mes_ins,
                "gerado_em": insights.get("gerado_em"),
                # Conversao de verdade: venda lancada no portal dividida pelos
                # clientes que ele atendeu. Nao usamos o "virou_venda" da IA
                # porque o fechamento acontece fora do chat — ela enxerga so uma
                # fracao, e o numero sairia baixo demais.
                "taxa_conversao": (round(100 * atual["qtd"] / atend, 1)
                                   if atend else None),
                "vendas_no_mes": atual["qtd"],
            }

    de_mes, ate_mes = f"{mes}-01", f"{mes}-{dias_no_mes:02d}"
    comissao = calcular_comissao(vid, de_mes, ate_mes, vendedores, vendas)

    return {
        "vendedor": {"id": vid, "nome": vendedores[vid]["nome"],
                     "foto": vendedores[vid].get("foto"),
                     "avatar": vendedores[vid].get("avatar", ""),
                     "percentual": float(vendedores[vid].get("percentual", 0))},
        "mes": mes,
        "atual": atual,
        "anterior": anterior,
        "variacao": {
            "total": variacao(atual["total"], anterior["total"]),
            "qtd": variacao(atual["qtd"], anterior["qtd"]),
            "ticket": variacao(atual["ticket"], anterior["ticket"]),
        },
        "meta": {
            "mensal": meta_mensal,
            "pct": round(100 * atual["total"] / meta_mensal, 1) if meta_mensal else None,
            "falta": round(max(0.0, meta_mensal - atual["total"]), 2) if meta_mensal else None,
            "batidas": sum(1 for h in historico if h["bateu"]),
            "meses_com_venda": len(historico),
        },
        "ritmo": {"dias_no_mes": dias_no_mes, "dias_corridos": dias_corridos,
                  "pct_do_mes": round(100 * dias_corridos / dias_no_mes),
                  "projecao": round(atual["total"] / dias_corridos * dias_no_mes, 2)
                              if dias_corridos else 0.0},
        "comissao": {"valor": comissao["comissao"], "bonus": round(
            sum(v["valor"] for v in do_mes(bonus, mes)), 2)},
        "historico": historico,
        "serie_dia": serie_dia,
        "melhor_dia": melhor_dia,
        "maior_venda": ({"produto": maior_venda.get("produto"),
                         "valor": valor_liquido(maior_venda),
                         "data": maior_venda["data"]} if maior_venda else None),
        "faixas": faixas,
        "maiores_vendas": maiores_vendas,
        "time": {"posicao": posicao, "de": len(totais_time),
                 "total_time": total_time,
                 "participacao": round(100 * atual["total"] / total_time, 1) if total_time else 0},
        "followup": followup,
        "atendimento": atendimento,
    }

@app.route("/api/admin/exportar-mes-xlsx")
def api_admin_exportar_mes_xlsx():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401

    mes = request.args.get("mes", "")
    if len(mes) != 7 or mes[4] != "-":
        return jsonify({"erro": "Mês inválido."}), 400
    de, ate = f"{mes}-01", f"{mes}-31"

    vendedores = carregar_vendedores()
    vendas = carregar_vendas_todos(vendedores)

    wb = Workbook()
    wb.remove(wb.active)
    cabecalho = ["Data", "Produto", "SKU", "Canal", "Valor", "Devolução", "Valor Devolvido", "Valor Líquido"]

    for vid in sorted(vendedores, key=lambda x: vendedores[x]["nome"]):
        lista = [
            v for v in vendas.values()
            if v["vendedor_id"] == vid and de <= v["data"] <= ate and v.get("tipo", "venda") == "venda"
        ]
        lista.sort(key=lambda v: v["data"])
        # Conta oculta sem venda no mes nao vira aba vazia no arquivo que o
        # gestor distribui. Com venda, a aba sai — dinheiro nunca some.
        if not lista and vendedores[vid].get("oculto"):
            continue

        ws = wb.create_sheet(_nome_aba_excel(vendedores[vid]["nome"]))
        ws.append(cabecalho)
        for v in lista:
            dev = v.get("devolucao")
            dev_tipo = ("Total" if dev.get("tipo") == "total" else "Parcial") if dev else ""
            dev_valor = dev.get("valor_devolvido") if dev else None
            ws.append([
                v["data"],
                v["produto"],
                v.get("sku", ""),
                v.get("canal", ""),
                v["valor"],
                dev_tipo,
                dev_valor,
                valor_liquido(v),
            ])
        for coluna, largura in zip("ABCDEFGH", (12, 42, 12, 14, 12, 12, 14, 14)):
            ws.column_dimensions[coluna].width = largura

    if not wb.sheetnames:
        wb.create_sheet("Vendedores").append(cabecalho)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"vendas_{mes}.xlsx",
    )

@app.route("/api/admin/vendedores", methods=["GET"])
def api_admin_listar_vendedores():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    vendedores = carregar_vendedores()
    return jsonify([
        {
            "id": vid,
            "nome": v["nome"],
            "percentual": v.get("percentual", 0),
            "overrides": v.get("overrides", []),
            "foto": v.get("foto"),
            "avatar": v.get("avatar", ""),
            "oculto": bool(v.get("oculto")),
            "desligado_em": v.get("desligado_em") or "",
            # Efetivas, nao o campo cru: a grade precisa mostrar o que de fato
            # vale, senao o primeiro clique apagaria o padrao de quem nunca foi
            # editado — a tela reenvia o que leu.
            "areas": areas_efetivas(v, vid),
            "setor": setor_do_usuario(vid),
            "master": bool(v.get("master")),
            "perfil": v.get("perfil") or "vendedor",
            "liberacao_retroativa": retroativo_ativo(v),
            "liberacao_retroativa_ate": v.get("liberacao_retroativa_ate") if retroativo_ativo(v) else None,
        }
        for vid, v in sorted(vendedores.items(), key=lambda kv: kv[1]["nome"])
    ])

@app.route("/api/admin/vendedores/<vendedor_id>/liberar-retroativo", methods=["POST"])
def api_admin_liberar_retroativo(vendedor_id):
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    vendedores = carregar_vendedores()
    if vendedor_id not in vendedores:
        return jsonify({"erro": "Vendedor não encontrado."}), 404
    ate = agora_br() + timedelta(minutes=LIBERACAO_RETROATIVA_MINUTOS)
    vendedores[vendedor_id]["liberacao_retroativa_ate"] = ate.isoformat(timespec="seconds")
    salvar_vendedores(vendedores)
    return jsonify({"ok": True, "liberacao_retroativa_ate": vendedores[vendedor_id]["liberacao_retroativa_ate"]})

@app.route("/api/admin/vendedores/<vendedor_id>/liberar-retroativo", methods=["DELETE"])
def api_admin_cancelar_liberacao_retroativo(vendedor_id):
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    vendedores = carregar_vendedores()
    if vendedor_id not in vendedores:
        return jsonify({"erro": "Vendedor não encontrado."}), 404
    vendedores[vendedor_id].pop("liberacao_retroativa_ate", None)
    salvar_vendedores(vendedores)
    return jsonify({"ok": True})

@app.route("/api/admin/vendedores", methods=["POST"])
def api_admin_salvar_vendedor():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    body = request.get_json(force=True)
    vendedor_id = (body.get("id") or "").strip().lower()
    nome = (body.get("nome") or "").strip()
    senha = body.get("senha")
    # Conta de gestao nao tem comissao: master nao vende, so controla dado.
    # Antes o cadastro exigia percentual de todo mundo, e criar um usuario
    # master travava num campo que nao se aplica a ele — foi exatamente onde o
    # gestor bateu ao cadastrar o segundo administrador.
    existente_pre = carregar_vendedores().get(vendedor_id, {})
    eh_master = bool(body.get("master", existente_pre.get("master")))
    bruto = body.get("percentual")
    if eh_master and (bruto is None or str(bruto).strip() == ""):
        percentual = 0.0
    else:
        try:
            percentual = float(bruto)
        except (TypeError, ValueError):
            return jsonify({"erro": "Percentual inválido."}), 400
    if not vendedor_id or not nome:
        return jsonify({"erro": "Informe id e nome do vendedor."}), 400
    if percentual < 0 or percentual > 100:
        return jsonify({"erro": "Percentual deve estar entre 0 e 100."}), 400

    overrides = []
    for over in body.get("overrides", []):
        outro_id = (over.get("vendedor_id") or "").strip().lower()
        try:
            outro_percentual = float(over.get("percentual"))
        except (TypeError, ValueError):
            continue
        if not outro_id or outro_id == vendedor_id or outro_percentual <= 0:
            continue
        overrides.append({"vendedor_id": outro_id, "percentual": outro_percentual})

    vendedores = carregar_vendedores()
    existente = vendedores.get(vendedor_id, {})
    if not senha and not existente.get("senha"):
        return jsonify({"erro": "Defina uma senha para o vendedor."}), 400
    vendedores[vendedor_id] = {
        "nome": nome,
        "senha": _hash_senha(senha) if senha else existente.get("senha"),
        "percentual": percentual,
        "overrides": overrides,
    }
    if existente.get("foto"):
        vendedores[vendedor_id]["foto"] = existente["foto"]
    # Flags que a tela de edicao nao conhece sobrevivem a edicao. Sem isso,
    # editar a senha do caique ressuscitava a conta nas listas (oculto sumia) e
    # editar qualquer vendedor cancelava a liberacao retroativa em silencio.
    for flag in ("oculto", "liberacao_retroativa_ate"):
        if existente.get(flag):
            vendedores[vendedor_id][flag] = existente[flag]

    # Desligamento tem DATA, nao e um liga-desliga. "Tudo ate ontem e do
    # Gustavo, de hoje em diante e do Lucas" so se resolve com a data: ela
    # deixa o historico intacto, tira a pessoa das telas de lancamento a
    # partir dali, e ainda diz na tela DESDE QUANDO — um `oculto` booleano
    # faria a pessoa evaporar sem explicar nada.
    # Acesso master. Vem da tela de Permissoes; quem tem isso ve tudo e pode
    # distribuir acesso, entao a marca e explicita e nunca deduzida de outra
    # coisa.
    if "master" in body:
        vendedores[vendedor_id]["master"] = bool(body.get("master"))
    elif existente.get("master"):
        vendedores[vendedor_id]["master"] = True

    # Areas liberadas. Sempre a lista inteira que veio da tela: marcar e
    # desmarcar tem que funcionar, entao nao da pra "preservar se vier vazio".
    if "areas" in body:
        vendedores[vendedor_id]["areas"] = [
            a for a in (body.get("areas") or []) if a in AREAS]
    elif existente.get("areas"):
        vendedores[vendedor_id]["areas"] = existente["areas"]

    if "desligado_em" in body:
        d = (body.get("desligado_em") or "").strip()
        if d:
            try:
                date.fromisoformat(d)
            except ValueError:
                return jsonify({"erro": "Data de desligamento invalida."}), 400
            vendedores[vendedor_id]["desligado_em"] = d
    elif existente.get("desligado_em"):
        vendedores[vendedor_id]["desligado_em"] = existente["desligado_em"]
    perfil = (body.get("perfil") or "").strip().lower()
    if perfil in ("vendedor", "expedicao"):
        vendedores[vendedor_id]["perfil"] = perfil
    elif existente.get("perfil") and "perfil" not in body:
        vendedores[vendedor_id]["perfil"] = existente["perfil"]
    # Avatar generico usado quando nao ha foto. E escolha do gestor, nunca
    # deduzida do nome — nome nao diz genero de ninguem.
    avatar = (body.get("avatar") or "").strip().lower()
    if avatar in ("feminino", "masculino"):
        vendedores[vendedor_id]["avatar"] = avatar
    elif existente.get("avatar") and "avatar" not in body:
        vendedores[vendedor_id]["avatar"] = existente["avatar"]
    codigo_recuperacao = (body.get("codigo_recuperacao") or "").strip().upper()
    if codigo_recuperacao:
        if len(codigo_recuperacao) < 6:
            return jsonify({"erro": "O código de recuperação precisa ter pelo menos 6 caracteres."}), 400
        vendedores[vendedor_id]["recuperacao_hash"] = _hash_codigo(codigo_recuperacao)
    elif existente.get("recuperacao_hash"):
        vendedores[vendedor_id]["recuperacao_hash"] = existente["recuperacao_hash"]
    salvar_vendedores(vendedores)
    return jsonify({"ok": True})

@app.route("/api/recuperar-senha-vendedor", methods=["POST"])
def api_recuperar_senha_vendedor():
    if excedeu_tentativas_login("vendedor_recuperacao"):
        return jsonify({"erro": f"Muitas tentativas erradas. Aguarde {LOGIN_JANELA_MINUTOS} minutos e tente de novo."}), 429

    body = request.get_json(force=True)
    vendedor_id = (body.get("vendedor_id") or "").strip().lower()
    codigo = (body.get("codigo") or "").strip().upper()
    nova_senha = body.get("nova_senha") or ""

    vendedores = carregar_vendedores()
    info = vendedores.get(vendedor_id)
    hash_salvo = info.get("recuperacao_hash") if info else None
    if not info or not hash_salvo or _hash_codigo(codigo) != hash_salvo:
        registrar_acesso("vendedor_recuperacao", False, vendedor_id)
        return jsonify({"erro": "Código inválido."}), 401

    if len(nova_senha) < 4:
        return jsonify({"erro": "A nova senha precisa ter pelo menos 4 caracteres."}), 400

    vendedores[vendedor_id]["senha"] = _hash_senha(nova_senha)
    vendedores[vendedor_id].pop("recuperacao_hash", None)
    salvar_vendedores(vendedores)
    registrar_acesso("vendedor_recuperacao", True, vendedor_id)
    return jsonify({"ok": True})

@app.route("/api/admin/vendedores/<vendedor_id>", methods=["DELETE"])
def api_admin_remover_vendedor(vendedor_id):
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    vendedores = carregar_vendedores()
    if vendedor_id not in vendedores:
        return jsonify({"erro": "Vendedor não encontrado."}), 404
    del vendedores[vendedor_id]
    salvar_vendedores(vendedores)
    return jsonify({"ok": True})

@app.route("/api/admin/metas", methods=["POST"])
def api_admin_salvar_metas():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    body = request.get_json(force=True)
    vendedores = carregar_vendedores()

    def limpar_trio(dados):
        resultado = {}
        for chave in ("diaria", "semanal", "mensal"):
            try:
                resultado[chave] = max(0.0, float(dados.get(chave, 0)))
            except (TypeError, ValueError):
                resultado[chave] = 0.0
        return resultado

    metas = {
        "grupo": limpar_trio(body.get("grupo", {})),
        "vendedores": {
            vid: limpar_trio(body.get("vendedores", {}).get(vid, {}))
            for vid in vendedores
        },
    }
    salvar_metas(metas)
    return jsonify({"ok": True, "metas": metas})

def _supabase_storage_upload(nome_arquivo: str, conteudo: bytes, content_type: str) -> None:
    url = f"{SUPABASE_URL}/storage/v1/object/fotos/{nome_arquivo}"
    req = urllib.request.Request(url, data=conteudo, method="POST", headers={
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "apikey": SUPABASE_SERVICE_KEY,
        "Content-Type": content_type,
        "x-upsert": "true",
    })
    urllib.request.urlopen(req).read()

def _supabase_storage_delete(nome_arquivo: str) -> None:
    url = f"{SUPABASE_URL}/storage/v1/object/fotos/{nome_arquivo}"
    req = urllib.request.Request(url, method="DELETE", headers={
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "apikey": SUPABASE_SERVICE_KEY,
    })
    try:
        urllib.request.urlopen(req).read()
    except urllib.error.HTTPError:
        pass

@app.route("/api/admin/vendedores/<vendedor_id>/foto", methods=["POST"])
def api_admin_upload_foto(vendedor_id):
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    vendedores = carregar_vendedores()
    if vendedor_id not in vendedores:
        return jsonify({"erro": "Vendedor não encontrado."}), 404

    arquivo = request.files.get("foto")
    if not arquivo or not arquivo.filename:
        return jsonify({"erro": "Nenhum arquivo enviado."}), 400
    ext = arquivo.filename.rsplit(".", 1)[-1].lower() if "." in arquivo.filename else ""
    if ext not in EXTENSOES_FOTO_PERMITIDAS:
        return jsonify({"erro": "Formato inválido. Use JPG, PNG ou WEBP."}), 400

    foto_antiga = vendedores[vendedor_id].get("foto")
    nome_arquivo = f"{vendedor_id}.{ext}"

    if SUPABASE_URL:
        if foto_antiga:
            _supabase_storage_delete(foto_antiga)
        _supabase_storage_upload(nome_arquivo, arquivo.read(), arquivo.content_type or "application/octet-stream")
    else:
        FOTOS_DIR.mkdir(parents=True, exist_ok=True)
        if foto_antiga:
            (FOTOS_DIR / foto_antiga).unlink(missing_ok=True)
        arquivo.save(FOTOS_DIR / nome_arquivo)

    vendedores[vendedor_id]["foto"] = nome_arquivo
    salvar_vendedores(vendedores)
    return jsonify({"ok": True, "foto": nome_arquivo})

@app.route("/fotos/<path:filename>")
def servir_foto(filename):
    if SUPABASE_URL:
        return redirect(f"{SUPABASE_URL}/storage/v1/object/public/fotos/{filename}")
    return send_from_directory(FOTOS_DIR, filename)

@app.route("/simulador")
def pagina_simulador():
    return send_from_directory(STATIC_DIR, "simulador.html")

@app.route("/api/admin/simulador/regras", methods=["GET"])
def api_admin_simulador_regras_get():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    regras = carregar_regras_simulador()
    if regras is None:
        return jsonify({"erro": "As regras de desconto ainda não foram importadas."}), 404
    return jsonify(regras)

@app.route("/api/admin/simulador/regras", methods=["PUT"])
def api_admin_simulador_regras_put():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    corpo = request.get_json(silent=True) or {}
    regras_atuais = carregar_regras_simulador() or {}
    for chave in ("desconto_max_pct", "nivel_flexibilidade", "parcelas_max",
                  "valor_minimo_parcelamento", "faixas_tempo", "faixas_valor"):
        if chave in corpo:
            regras_atuais[chave] = corpo[chave]
    regras_atuais["atualizado_em"] = agora_br().isoformat()
    regras_atuais["atualizado_por"] = "gestor"
    salvar_regras_simulador(regras_atuais)
    return jsonify(regras_atuais)

@app.route("/api/admin/simulador/status")
def api_admin_simulador_status():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    return jsonify(status_simulador())

@app.route("/api/admin/simulador/reimportar", methods=["POST"])
def api_admin_simulador_reimportar():
    if not exigir_admin():
        return jsonify({"erro": "Não autenticado."}), 401
    # As planilhas do ERP nunca estiveram no servidor: `data/` esta no
    # .gitignore e elas moram na maquina da loja. Sem esta checagem o botao
    # rodava o ETL num diretorio vazio, que apagava o catalogo e devolvia
    # sucesso. O script tambem trava sozinho agora, mas recusar aqui poupa a
    # ida e diz a coisa certa em vez de mostrar um log de erro.
    bruto = ROOT / "data" / "simulador" / "raw_erp"
    if not any(bruto.glob("relatorio_produtos_76_parte*.xlsx")):
        return jsonify({"ok": False, "erro":
                        "As planilhas do ERP não estão neste servidor — elas ficam "
                        "no computador da loja. Rode a reimportação de lá, que ela "
                        "sobe pro portal sozinha."}), 400

    script = ROOT / "scripts" / "etl_simulador.py"
    resultado = subprocess.run(
        [sys.executable, str(script)],
        capture_output=True, text=True, cwd=str(ROOT), timeout=900,
    )
    if resultado.returncode != 0:
        log = (resultado.stdout[-4000:] + "\n" + resultado.stderr[-4000:])
        return jsonify({"ok": False, "log": log}), 500
    return jsonify({"ok": True, "log": resultado.stdout[-4000:]})

@app.route("/api/simulador/regras")
def api_simulador_regras():
    if not exigir_vendedor():
        return jsonify({"erro": "Não autenticado."}), 401
    regras = carregar_regras_simulador()
    if regras is None:
        return jsonify({"erro": "As regras de desconto ainda não foram sincronizadas."}), 404
    return jsonify(regras)

@app.route("/api/simulador/buscar")
def api_simulador_buscar():
    if not exigir_vendedor():
        return jsonify({"erro": "Não autenticado."}), 401
    q = (request.args.get("q") or "").strip()
    if len(q) < 2:
        return jsonify([])
    return jsonify(buscar_pecas_simulador(q))

@app.route("/api/simulador/simular", methods=["POST"])
def api_simulador_simular():
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401
    corpo = request.get_json(silent=True) or {}
    cod_peca = (corpo.get("cod_peca") or "").strip()
    if not cod_peca:
        return jsonify({"erro": "cod_peca é obrigatório"}), 400

    peca = obter_peca_simulador(cod_peca)
    if not peca:
        return jsonify({"erro": "peça não encontrada"}), 404

    regras = carregar_regras_simulador()
    if regras is None:
        return jsonify({"erro": "As regras de desconto ainda não foram sincronizadas."}), 404

    valor_override = corpo.get("valor_base")
    desconto_escolhido = corpo.get("desconto_pct")
    resultado = calcular_simulacao_peca(peca, valor_override, desconto_escolhido, regras)
    resultado["cod_peca"] = peca["cod_peca"]
    resultado["nome_produto"] = peca["nome_produto"]
    resultado["etiqueta"] = peca["etiqueta"]
    resultado["apelido_veiculo"] = peca["apelido_veiculo"]
    resultado["tipo_peca_rotulo"] = peca["tipo_peca_rotulo"]
    return jsonify(resultado)

@app.route("/api/simulador/simular-rapido", methods=["POST"])
def api_simulador_simular_rapido():
    """Simulação sem buscar peça no catálogo: o vendedor informa valor,
    curva e faixa de tempo em estoque na mão — pensado pra atender o
    cliente rápido no balcão, sem precisar achar o item no sistema."""
    if not exigir_vendedor():
        return jsonify({"erro": "Não autenticado."}), 401
    corpo = request.get_json(silent=True) or {}
    try:
        valor_base = float(corpo.get("valor_base"))
    except (TypeError, ValueError):
        return jsonify({"erro": "valor_base é obrigatório e deve ser numérico"}), 400
    if valor_base <= 0:
        return jsonify({"erro": "valor_base deve ser maior que zero"}), 400

    regras = carregar_regras_simulador()
    if regras is None:
        return jsonify({"erro": "As regras de desconto ainda não foram sincronizadas."}), 404

    curva = (corpo.get("curva") or "").strip()
    if curva not in regras["desconto_max_pct"]:
        return jsonify({"erro": "curva inválida"}), 400

    faixa_tempo_id = (corpo.get("faixa_tempo_id") or "").strip()
    ids_validos = {f["id"] for f in regras["faixas_tempo"]}
    if faixa_tempo_id not in ids_validos:
        return jsonify({"erro": "faixa_tempo_id inválida"}), 400
    dias_representativos = next(f["min_dias"] for f in regras["faixas_tempo"] if f["id"] == faixa_tempo_id)

    desconto_escolhido = corpo.get("desconto_pct")
    resultado = montar_simulacao(valor_base, curva, dias_representativos, desconto_escolhido, regras)
    resultado["dias_em_estoque"] = None  # é uma faixa escolhida à mão, não uma data real
    return jsonify(resultado)

@app.route("/api/simulador/peca/<cod_peca>/data-entrada", methods=["POST"])
def api_simulador_definir_data_entrada(cod_peca):
    vendedor_id = exigir_vendedor()
    if not vendedor_id:
        return jsonify({"erro": "Não autenticado."}), 401
    corpo = request.get_json(silent=True) or {}
    data_entrada = (corpo.get("data_entrada") or "").strip()
    if not data_entrada:
        return jsonify({"erro": "data_entrada é obrigatória"}), 400
    try:
        data_parseada = datetime.fromisoformat(data_entrada)
    except ValueError:
        return jsonify({"erro": "data inválida, use AAAA-MM-DD"}), 400

    nome = carregar_vendedores().get(vendedor_id, {}).get("nome", vendedor_id)
    ok = definir_data_entrada_simulador(cod_peca, data_parseada.strftime("%Y-%m-%d 00:00:00"), nome)
    if not ok:
        return jsonify({"erro": "peça não encontrada"}), 404
    return jsonify({"ok": True})
